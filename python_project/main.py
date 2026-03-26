import gc
import hashlib
import os
import re
import difflib
import functools
import importlib
import html
import json
import logging
import math
import time
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from io import BytesIO
from pathlib import Path
import pandas as pd
pd.options.mode.copy_on_write = True  # PERFORMANCE: deferred copy — makes .copy() nearly free
import streamlit as st
import streamlit.components.v1 as components
import plotly.express as px
import plotly.io as pio
from fpdf import FPDF, XPos, YPos
import map_page_state as _map_page_state
import shared_search_state as _shared_search_state
import src.map_runtime as _map_runtime
import src.map_fragments as _map_fragments
import src.map_page_helpers as _map_page_helpers
import src.page_bundles as _page_bundles
import src.page_detail_bundles as _page_detail_bundles
import src.page_fragments as _page_fragments
import src.ui_runtime as _ui_runtime

# =========================================================
# PERFORMANCE: Pre-compiled regex patterns
# =========================================================
_RE_WHITESPACE = re.compile(r"\s+")
_RE_BILL_PATTERN = re.compile(
    r"\b(HB|SB|HR|SR|HCR|SCR|HJR|SJR)\s*(\d+)\b", re.IGNORECASE
)
_RE_PARENS = re.compile(r"\([^)]*\)")

# =========================================================
# PERFORMANCE: Vectorized display helpers (avoid .apply())
# =========================================================
def _vectorized_person_display(org: pd.Series, last: pd.Series, first: pd.Series) -> pd.Series:
    """Vectorized version of person_display — avoids row-by-row .apply()."""
    org_s = org.fillna("").astype(str).str.strip()
    last_s = last.fillna("").astype(str).str.strip()
    first_s = first.fillna("").astype(str).str.strip()
    # Priority: org if non-empty, then "last, first", then whichever exists
    result = org_s.where(org_s != "", last_s + ", " + first_s)
    # Fix cases where only last or only first
    both = (last_s != "") & (first_s != "")
    only_last = (last_s != "") & (first_s == "")
    only_first = (last_s == "") & (first_s != "")
    neither = (last_s == "") & (first_s == "")
    result = result.where(~(~(org_s != "")) | True, result)  # keep org priority
    # Rebuild for non-org cases
    no_org = org_s == ""
    result = result.where(~no_org, pd.Series("", index=org.index))
    result[no_org & both] = last_s[no_org & both] + ", " + first_s[no_org & both]
    result[no_org & only_last] = last_s[no_org & only_last]
    result[no_org & only_first] = first_s[no_org & only_first]
    result[no_org & neither] = ""
    result[~no_org] = org_s[~no_org]
    return result.str.strip()


def _vectorized_amount_display(exact: pd.Series, low: pd.Series, high: pd.Series, code: pd.Series | None = None) -> pd.Series:
    """Vectorized version of amount_display — avoids row-by-row .apply()."""
    exact_s = exact.fillna("").astype(str).str.strip()
    low_s = low.fillna("").astype(str).str.strip()
    high_s = high.fillna("").astype(str).str.strip()
    code_s = code.fillna("").astype(str).str.strip() if code is not None else pd.Series("", index=exact.index)
    # Priority: exact -> low--high or low -> code -> ""
    result = pd.Series("", index=exact.index)
    has_exact = exact_s != ""
    has_low = low_s != ""
    has_high = high_s != ""
    has_code = code_s != ""
    result = result.where(~has_exact, exact_s)
    need_range = ~has_exact & has_low & has_high
    result[need_range] = low_s[need_range] + "--" + high_s[need_range]
    need_low_only = ~has_exact & has_low & ~has_high
    result[need_low_only] = low_s[need_low_only]
    need_code = ~has_exact & ~has_low & has_code
    result[need_code] = code_s[need_code]
    return result


def _session_base_number_series(s: pd.Series) -> pd.Series:
    base = s.fillna("").astype(str).str.strip().str.extract(r"^(\d+)", expand=False)
    base = base.where(base.str.len() <= 2, base.str[:-1])
    return pd.to_numeric(base, errors="coerce")

# =========================================================
# CONFIG
# =========================================================
DEFAULT_DATA_FILENAME = "TFL Webstite books - combined.parquet"
TEA_ARCGIS_WEBAPP_URL = "https://tea-texas.maps.arcgis.com/apps/webappviewer/index.html?id=51f0c8fa684c4d399d8d182e6edd5d97"
TEA_ARCGIS_SCHOOL_DISTRICT_LAYER_URL = "https://services2.arcgis.com/5MVN2jsqIrNZD4tP/arcgis/rest/services/Map/FeatureServer/0"
TEA_ARCGIS_COUNTY_LAYER_URL = "https://services2.arcgis.com/5MVN2jsqIrNZD4tP/arcgis/rest/services/Counties2019/FeatureServer/0"
CENSUS_ARCGIS_TEXAS_CITY_LAYER_URL = "https://tigerweb.geo.census.gov/arcgis/rest/services/TIGERweb/Places_CouSub_ConCity_SubMCD/MapServer/25"
ARCGIS_GEOCODER_URL = "https://geocode.arcgis.com/arcgis/rest/services/World/GeocodeServer/findAddressCandidates"
TCEQ_WATER_DISTRICTS_LAYER_URL = "https://services2.arcgis.com/LYMgRMwHfrWWEg3s/arcgis/rest/services/TCEQ_Water_Districts/FeatureServer/0"
TCEQ_GROUNDWATER_DISTRICTS_LAYER_URL = "https://services2.arcgis.com/LYMgRMwHfrWWEg3s/arcgis/rest/services/TCEQ_Groundwater_Conservation_Districts/FeatureServer/0"
TEXAS_RMA_LAYER_URL = "https://services.arcgis.com/KTcxiTD9dsQw4r7Z/arcgis/rest/services/Texas_Regional_Mobility_Authority_Boundaries/FeatureServer/0"
TEXAS_HOUSE_DISTRICTS_LAYER_URL = "https://services.arcgis.com/KTcxiTD9dsQw4r7Z/arcgis/rest/services/Texas_State_House_Districts/FeatureServer/0"
TEXAS_SENATE_DISTRICTS_LAYER_URL = "https://services.arcgis.com/KTcxiTD9dsQw4r7Z/arcgis/rest/services/Texas_State_Senate_Districts/FeatureServer/0"
TEXAS_JUNIOR_COLLEGE_LAYER_URL = "https://services1.arcgis.com/hVMNhMnY75fwfIFy/arcgis/rest/services/JuniorCollege_ServiceAreas/FeatureServer/0"
TEXAS_NAVIGATION_DISTRICT_LAYER_URL = "https://services1.arcgis.com/YWG34dhJxrbxQWdF/arcgis/rest/services/Navigation_Districts2/FeatureServer/29"
NCTCOG_TRANSIT_PROVIDERS_LAYER_URL = "https://geospatial.nctcog.org/map/rest/services/Transportation/DFWMaps_Transit/MapServer/10"
TXDOT_SEAPORTS_LAYER_URL = "https://services.arcgis.com/KTcxiTD9dsQw4r7Z/arcgis/rest/services/TxDOT_Seaports/FeatureServer/0"
MAP_BASEMAP_OPTIONS = {
    "Gray Canvas": "gray-vector",
    "Street Detail": "streets-vector",
    "Satellite": "hybrid",
}

# -- Atlas → Forensics / Batch bridge (invisible component relay) --
_TFL_BRIDGE_DIR = Path(__file__).parent / "_atlas_bridge"
_atlas_bridge = components.declare_component("atlas_bridge", path=str(_TFL_BRIDGE_DIR))
_PERSISTENT_HTML_FRAME_DIR = Path(__file__).parent / "_persistent_html_frame"
_persistent_html_frame = components.declare_component(
    "persistent_html_frame",
    path=str(_PERSISTENT_HTML_FRAME_DIR),
)
_TFL_SUBDIVISION_MAP_DIR = Path(__file__).parent / "_tfl_subdivision_map"
_tfl_subdivision_map_component = components.declare_component(
    "tfl_subdivision_map",
    path=str(_TFL_SUBDIVISION_MAP_DIR),
)


def _stable_json_signature(value) -> str:
    return hashlib.sha1(
        json.dumps(
            value,
            ensure_ascii=True,
            sort_keys=True,
            separators=(",", ":"),
            default=str,
        ).encode("utf-8")
    ).hexdigest()


def _clone_session_cache_value(value):
    if isinstance(value, pd.DataFrame):
        return value.copy()
    if isinstance(value, pd.Series):
        return value.copy()
    if isinstance(value, dict):
        return {k: _clone_session_cache_value(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_clone_session_cache_value(v) for v in value]
    if isinstance(value, tuple):
        return tuple(_clone_session_cache_value(v) for v in value)
    return value


def _session_cached_value(cache_key: str, signature: str, builder):
    cached = st.session_state.get(cache_key)
    if isinstance(cached, dict) and cached.get("signature") == signature and "value" in cached:
        return _clone_session_cache_value(cached["value"])
    value = builder()
    st.session_state[cache_key] = {
        "signature": signature,
        "value": _clone_session_cache_value(value),
    }
    return _clone_session_cache_value(value)
SUBDIVISION_TYPE_COLORS = {
    "School District": "#1769AA",
    "County": "#7A3E00",
    "City": "#9E2A2B",
    "Junior College District": "#1E58A5",
    "Groundwater Conservation District": "#0B8F6A",
    "Municipal Utility District": "#5B3FB0",
    "Drainage District": "#CC6B2C",
    "Fresh Water Supply District": "#3382CC",
    "Irrigation District": "#5A8B2D",
    "Levee Improvement District": "#8A6A1F",
    "Municipal Management District": "#7D3FA0",
    "Regional District": "#8A7E24",
    "River Authority": "#0E8791",
    "Soil & Water Control District": "#2E8D73",
    "Special Utility District": "#31688E",
    "Water Improvement District": "#4D6FA9",
    "Water Control & Improvement District": "#2F6DA4",
    "Regional Mobility Authority": "#B08900",
    "Navigation District": "#2C3E50",
    "Transit Authority": "#5A657A",
    "Port Authority": "#3B6F8C",
    "Hospital District": "#9B3E56",
    "Emergency Services District": "#B15C2E",
    "Appraisal District": "#6D5A90",
    "Local Government Corporation": "#4E7A52",
}
WATER_DISTRICT_TYPE_ROOT_PATTERNS = {
    "Municipal Utility District": [r"\bMUNICIPAL\s+UTILITY\s+DISTRICT\b", r"\bDISTRICT\b"],
    "Drainage District": [r"\bDRAINAGE\s+DISTRICT\b", r"\bDISTRICT\b"],
    "Fresh Water Supply District": [r"\bFRESH\s+WATER\s+SUPPLY\s+DISTRICT\b", r"\bDISTRICT\b"],
    "Irrigation District": [r"\bIRRIGATION\s+DISTRICT\b", r"\bDISTRICT\b"],
    "Levee Improvement District": [r"\bLEVEE\s+IMPROVEMENT\s+DISTRICT\b", r"\bDISTRICT\b"],
    "Municipal Management District": [r"\bMUNICIPAL\s+MANAGEMENT\s+DISTRICT\b", r"\bDISTRICT\b"],
    "Regional District": [r"\bREGIONAL\s+DISTRICT\b", r"\bDISTRICT\b"],
    "River Authority": [r"\bRIVER\s+AUTHORITY\b", r"\bAUTHORITY\b"],
    "Soil & Water Control District": [r"\bSOIL\s+(AND\s+)?WATER\s+CONTROL\s+DISTRICT\b", r"\bDISTRICT\b"],
    "Special Utility District": [r"\bSPECIAL\s+UTILITY\s+DISTRICT\b", r"\bDISTRICT\b"],
    "Water Improvement District": [r"\bWATER\s+IMPROVEMENT\s+DISTRICT\b", r"\bDISTRICT\b"],
    "Water Control & Improvement District": [r"\bWATER\s+CONTROL\s+(AND\s+)?IMPROVEMENT\s+DISTRICT\b", r"\bDISTRICT\b"],
}
TRANSIT_AUTHORITY_ROOT_PATTERNS = [
    r"\bMETROPOLITAN\s+TRANSIT\s+AUTHORITY\b",
    r"\bTRANSIT\s+AUTHORITY\b",
    r"\bAREA\s+RAPID\s+TRANSIT\b",
    r"\bREGIONAL\s+TRANSPORTATION\s+AUTHORITY\b",
    r"\bTRANSPORTATION\s+AUTHORITY\b",
]
PORT_AUTHORITY_ROOT_PATTERNS = [
    r"\bPORT\s+AUTHORITY\b",
    r"\bPORT\s+OF\b",
    r"\bNAVIGATION\s+DISTRICT\b",
    r"\bPORT\b",
    r"\bAUTHORITY\b",
    r"\bDISTRICT\b",
]
SPECIAL_NAME_ANCHORED_ENTITY_TYPES = {
    "Hospital District",
    "Emergency Services District",
    "Appraisal District",
    "Local Government Corporation",
    "Transit Authority",
    "Port Authority",
}
COUNTY_BIASED_SPECIAL_ENTITY_TYPES = {
    "Hospital District",
    "Emergency Services District",
    "Appraisal District",
}
CITY_BIASED_SPECIAL_ENTITY_TYPES = {
    "Local Government Corporation",
    "Transit Authority",
    "Port Authority",
}
MAP_DATA_SOURCES = [
    (
        "TEA School District Locator (web app)",
        TEA_ARCGIS_WEBAPP_URL,
        "Reference viewer used for school district context.",
    ),
    (
        "TEA School District boundaries (FeatureServer/0)",
        TEA_ARCGIS_SCHOOL_DISTRICT_LAYER_URL,
        "School district polygons and centroids.",
    ),
    (
        "TEA County boundaries (FeatureServer/0)",
        TEA_ARCGIS_COUNTY_LAYER_URL,
        "County polygons and centroids.",
    ),
    (
        "U.S. Census TIGERweb Texas Places (MapServer/25)",
        CENSUS_ARCGIS_TEXAS_CITY_LAYER_URL,
        "City/place polygons and centroids for Texas (STATE=48).",
    ),
    (
        "TCEQ Water Districts (FeatureServer/0)",
        TCEQ_WATER_DISTRICTS_LAYER_URL,
        "Municipal utility, drainage, fresh water supply, irrigation, levee improvement, municipal management, regional, river authority, soil and water control, special utility, water improvement, and water control and improvement districts.",
    ),
    (
        "TCEQ Groundwater Conservation Districts (FeatureServer/0)",
        TCEQ_GROUNDWATER_DISTRICTS_LAYER_URL,
        "Groundwater conservation district boundaries.",
    ),
    (
        "Texas Regional Mobility Authorities (FeatureServer/0)",
        TEXAS_RMA_LAYER_URL,
        "Regional mobility authority boundaries.",
    ),
    (
        "Texas Junior College Service Areas (FeatureServer/0)",
        TEXAS_JUNIOR_COLLEGE_LAYER_URL,
        "Junior/community college service-area boundaries.",
    ),
    (
        "Texas Navigation Districts (FeatureServer/29)",
        TEXAS_NAVIGATION_DISTRICT_LAYER_URL,
        "Navigation district boundaries.",
    ),
    (
        "NCTCOG Transit Providers (MapServer/10)",
        NCTCOG_TRANSIT_PROVIDERS_LAYER_URL,
        "Transit provider/service-area polygons for the North Central Texas region.",
    ),
    (
        "TxDOT Seaports (FeatureServer/0)",
        TXDOT_SEAPORTS_LAYER_URL,
        "Texas seaport locations and attributes used for port-authority matching.",
    ),
    (
        "ArcGIS World Geocoding Service",
        ARCGIS_GEOCODER_URL,
        "Address geocoding for overlap point lookup plus centroid fallback for special subdivision types without statewide boundary layers.",
    ),
]


def _is_url(path: str) -> bool:
    return path.startswith("http://") or path.startswith("https://")


def _resolve_data_path() -> str:
    # Prefer environment variable, then local fallbacks.
    env_path = os.getenv("DATA_PATH", "").strip()
    if env_path:
        return env_path

    here = Path(__file__).resolve().parent
    candidates = [
        here / DEFAULT_DATA_FILENAME,
        here / "data" / DEFAULT_DATA_FILENAME,
        here.parent / "data" / DEFAULT_DATA_FILENAME,
        here.parent / DEFAULT_DATA_FILENAME,
    ]
    for candidate in candidates:
        if candidate.exists():
            return str(candidate)
    return ""


PATH = _resolve_data_path()

st.set_page_config(page_title="Texas Taxpayer Lobbying Transparency Center", layout="wide")

# =========================================================
# STYLE (unchanged)
# =========================================================
st.markdown(
    """
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;600;700&family=IBM+Plex+Serif:wght@400;600&family=Source+Sans+3:wght@400;600;700&display=swap');
:root{
    --bg: #071627;
    --panel: rgba(255,255,255,0.06);
    --panel2: rgba(255,255,255,0.04);
    --border: rgba(255,255,255,0.10);
    --text: rgba(255,255,255,0.92);
    --muted: rgba(255,255,255,0.70);
    --accent: #1e90ff;
    --accent2: #00e0b8;
    --nav-h: 72px;
    --nav-bg: rgba(6, 16, 30, 0.98);
    --nav-border: rgba(255,255,255,0.08);
    --nav-search-w: 320px;
    --nav-search-h: 38px;
    --space-1: 6px;
    --space-2: 10px;
    --space-3: 16px;
    --space-4: 22px;
    --radius-md: 14px;
    --radius-lg: 18px;
    --shadow-1: 0 10px 25px rgba(0,0,0,0.20);
    --shadow-2: 0 18px 32px rgba(0,0,0,0.28);
}

html, body, [data-testid="stAppViewContainer"]{
    background: radial-gradient(1200px 600px at 20% 15%, rgba(30,144,255,0.16), transparent 60%),
                            radial-gradient(900px 500px at 75% 30%, rgba(0,255,180,0.08), transparent 55%),
                            var(--bg) !important;
    color: var(--text) !important;
    font-family: 'IBM Plex Sans', system-ui, -apple-system, Segoe UI, sans-serif !important;
}

[data-testid="stAppViewContainer"]{
    position: relative;
}
[data-testid="stAppViewContainer"]::before{
    content: "";
    position: fixed;
    inset: 0;
    background-image:
        linear-gradient(transparent 26px, rgba(255,255,255,0.035) 27px),
        linear-gradient(90deg, transparent 26px, rgba(255,255,255,0.035) 27px);
    background-size: 32px 32px;
    opacity: 0.15;
    pointer-events: none;
    z-index: 0;
}
[data-testid="stAppViewContainer"] > div{
    position: relative;
    z-index: 1;
}

[data-testid="stHeader"]{ display: none !important; }
[data-testid="stToolbar"]{ right: 1rem; }
.block-container{
    padding-top: calc(var(--nav-h) + 0.8rem);
    padding-bottom: calc(1rem + env(safe-area-inset-bottom, 0px));
}

h1,h2,h3{ color: var(--text) !important; }
p,li,span,div{ color: var(--text); }

.small-muted{ color: var(--muted); font-size: 0.95rem; }
.hr{ height:1px; background: var(--border); margin: 1rem 0 1.2rem 0; }

.card{
    background: linear-gradient(180deg, var(--panel), var(--panel2));
    border: 1px solid var(--border);
    border-radius: var(--radius-lg);
    padding: 16px 16px 14px 16px;
    box-shadow: var(--shadow-1);
}
div[data-testid="stPlotlyChart"]{
    background: linear-gradient(180deg, rgba(255,255,255,0.05), rgba(255,255,255,0.01));
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: var(--radius-lg);
    padding: 10px 12px 6px 12px;
    box-shadow: var(--shadow-2);
    box-sizing: border-box;
    margin-top: 0.35rem;
}
div[data-testid="stPlotlyChart"] > div{
    border-radius: 14px;
    overflow: hidden;
}
.about-wrap{
    display: flex;
    flex-direction: column;
    gap: 20px;
    width: 100%;
    max-width: 1200px;
    margin: 0 auto;
}
.about-hero{
    position: relative;
    overflow: hidden;
    padding: 24px 24px 20px 24px;
    border-radius: 22px;
    background: linear-gradient(135deg, rgba(0,224,184,0.22), rgba(30,144,255,0.14) 45%, rgba(7,22,39,0.82));
    border: 1px solid rgba(255,255,255,0.14);
    box-shadow: 0 20px 40px rgba(0,0,0,0.35);
    backdrop-filter: blur(6px);
}
.about-hero::before{
    content: "";
    position: absolute;
    inset: -40px -10px auto auto;
    width: 320px;
    height: 320px;
    background: radial-gradient(circle, rgba(30,144,255,0.35), transparent 70%);
    opacity: 0.6;
    pointer-events: none;
}
.about-hero::after{
    content: "";
    position: absolute;
    inset: 0;
    background-image:
        linear-gradient(transparent 24px, rgba(255,255,255,0.03) 25px),
        linear-gradient(90deg, transparent 24px, rgba(255,255,255,0.03) 25px);
    background-size: 32px 32px;
    opacity: 0.18;
    pointer-events: none;
}
.about-hero > *{
    position: relative;
    z-index: 1;
}
.about-hero p{
    max-width: 980px;
}
.about-kicker{
    text-transform: uppercase;
    letter-spacing: 0.18em;
    font-size: 0.68rem;
    color: var(--muted);
    margin-bottom: 8px;
}
.about-title{
    font-size: 2.2rem;
    font-weight: 700;
    margin: 0 0 0.4rem 0;
    text-shadow: 0 6px 16px rgba(0,0,0,0.35);
}
.about-lead{
    font-size: 1.05rem;
    line-height: 1.55;
    margin: 0.2rem 0 0.6rem 0;
}
.about-body{
    color: var(--muted);
    margin: 0 0 0.8rem 0;
}
.about-wrap p{
    line-height: 1.55;
}
.about-meta{
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin-top: 0.6rem;
}
.about-meta .pill{
    background: rgba(7,22,39,0.35);
    border-color: rgba(255,255,255,0.18);
}
.about-shell{
    display: grid;
    grid-template-columns: minmax(0, 0.9fr) minmax(0, 1.1fr);
    gap: 20px;
    align-items: start;
}
.about-sidebar{
    display: flex;
    flex-direction: column;
    gap: 20px;
}
.about-panel{
    padding: 18px 18px 16px 18px;
    border-left: 3px solid rgba(0,224,184,0.55);
    background: linear-gradient(180deg, rgba(255,255,255,0.08), rgba(255,255,255,0.02));
    border-color: rgba(255,255,255,0.12);
    box-shadow: 0 16px 26px rgba(0,0,0,0.26);
    backdrop-filter: blur(4px);
}
.about-panel-head{
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 12px;
    margin-bottom: 0.6rem;
}
.about-panel h3{
    margin: 0;
    font-size: 1.1rem;
}
.about-panel-tag{
    font-size: 0.65rem;
    letter-spacing: 0.18em;
    text-transform: uppercase;
    color: var(--muted);
    border: 1px solid rgba(255,255,255,0.16);
    padding: 3px 8px;
    border-radius: 999px;
    background: rgba(7,22,39,0.35);
}
.about-actions{
    display: grid;
    gap: 8px;
    margin-bottom: 0.6rem;
}
.about-action{
    display: flex;
    gap: 10px;
    align-items: flex-start;
    padding: 8px 10px;
    border-radius: 12px;
    background: rgba(7,22,39,0.4);
    border: 1px solid rgba(255,255,255,0.12);
    box-shadow: inset 0 0 0 1px rgba(255,255,255,0.02);
}
.about-action::before{
    content: "";
    width: 8px;
    height: 8px;
    border-radius: 50%;
    background: var(--accent2);
    margin-top: 0.35rem;
    box-shadow: 0 0 0 3px rgba(0,224,184,0.18);
    flex: 0 0 auto;
}
.about-list-grid{
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 4px 12px;
    margin-top: 0.35rem;
}
.about-note{
    color: var(--muted);
    font-size: 0.95rem;
}
.about-checklist{
    list-style: none;
    padding: 0;
    margin: 0;
}
.about-checklist li{
    position: relative;
    padding-left: 1.4rem;
    margin: 0.35rem 0;
    line-height: 1.45;
}
.about-checklist li::before{
    content: "";
    position: absolute;
    left: 0;
    top: 0.5rem;
    width: 9px;
    height: 9px;
    border-radius: 3px;
    background: rgba(30,144,255,0.85);
    box-shadow: inset 0 0 0 2px rgba(30,144,255,0.15);
}
.about-main{
    display: flex;
    flex-direction: column;
    gap: 20px;
}
.about-section{
    position: relative;
    padding: 18px 18px 16px 26px;
    background: linear-gradient(180deg, rgba(255,255,255,0.07), rgba(255,255,255,0.02));
    border-color: rgba(255,255,255,0.12);
    box-shadow: 0 18px 30px rgba(0,0,0,0.28);
    backdrop-filter: blur(4px);
}
.about-section::before{
    content: "";
    position: absolute;
    left: 14px;
    top: 18px;
    bottom: 18px;
    width: 2px;
    background: linear-gradient(180deg, rgba(30,144,255,0.9), rgba(0,224,184,0.6));
    border-radius: 999px;
}
.about-section-head{
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 0.6rem;
}
.about-section-head h3{
    margin: 0;
    font-size: 1.35rem;
    letter-spacing: -0.01em;
}
.about-section-num{
    font-size: 0.68rem;
    letter-spacing: 0.2em;
    font-weight: 700;
    color: var(--accent2);
    border: 1px solid rgba(0,224,184,0.35);
    border-radius: 999px;
    padding: 4px 8px;
    background: rgba(0,224,184,0.12);
}
.source-grid{
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 14px;
    margin-top: 0.6rem;
}
.source-item{
    position: relative;
    overflow: hidden;
    background: rgba(7,22,39,0.4);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 14px;
    padding: 14px 12px 12px 12px;
    box-shadow: 0 14px 24px rgba(0,0,0,0.24);
    display: flex;
    flex-direction: column;
    gap: 4px;
}
.source-item::before{
    content: "";
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    height: 2px;
    background: linear-gradient(90deg, rgba(0,224,184,0.8), rgba(30,144,255,0.8));
    opacity: 0.7;
}
.source-title{
    font-weight: 700;
    margin-bottom: 0.2rem;
}
.source-text{
    color: var(--muted);
    font-size: 0.93rem;
    margin-bottom: 0.25rem;
}
.source-note{
    color: var(--muted);
    font-size: 0.85rem;
    margin-top: 0.35rem;
}
.source-links{
    display: flex;
    flex-direction: column;
    gap: 4px;
    margin-top: 0.35rem;
}
.video-grid{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
    gap: 16px;
    margin-top: 0.4rem;
}
.video-card{
    background: rgba(7,22,39,0.4);
    border: 1px solid rgba(255,255,255,0.1);
    border-radius: 16px;
    padding: 10px;
    box-shadow: 0 14px 24px rgba(0,0,0,0.24);
}
.video-card.is-active{
    border-color: rgba(30,144,255,0.55);
    box-shadow: 0 0 0 1px rgba(30,144,255,0.35), 0 14px 24px rgba(0,0,0,0.24);
}
.video-embed{
    position: relative;
    padding-top: 56.25%;
    border-radius: 12px;
    overflow: hidden;
    background: rgba(0,0,0,0.25);
}
.video-embed iframe{
    position: absolute;
    inset: 0;
    width: 100%;
    height: 100%;
    border: 0;
}
.tap-hero{
    position: relative;
    overflow: hidden;
    padding: 18px 20px 16px 20px;
    border-radius: 20px;
    background: linear-gradient(135deg, rgba(30,144,255,0.18), rgba(0,224,184,0.12), rgba(7,22,39,0.85));
    border: 1px solid rgba(255,255,255,0.14);
    box-shadow: 0 18px 34px rgba(0,0,0,0.32);
}
.tap-hero::after{
    content: "";
    position: absolute;
    inset: auto -40px -50px -40px;
    height: 120px;
    background: radial-gradient(circle, rgba(30,144,255,0.25), transparent 70%);
    opacity: 0.6;
    pointer-events: none;
}
.tap-hero > *{ position: relative; z-index: 1; }
.tap-hero-kicker{
    text-transform: uppercase;
    letter-spacing: 0.2em;
    font-size: 0.7rem;
    color: var(--muted);
    margin-bottom: 6px;
}
.tap-hero-title{
    font-size: 2rem;
    font-weight: 700;
    margin: 0 0 0.4rem 0;
}
.tap-hero-lead{
    color: var(--muted);
    margin: 0;
    max-width: 860px;
    line-height: 1.55;
}
.tap-feature{
    margin-top: 1rem;
}
.tap-feature-head{
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 16px;
    margin-bottom: 12px;
    flex-wrap: wrap;
}
.tap-feature-kicker{
    text-transform: uppercase;
    letter-spacing: 0.18em;
    font-size: 0.65rem;
    color: var(--muted);
    margin-bottom: 6px;
}
.tap-feature-title{
    font-size: 1.5rem;
    font-weight: 700;
    margin: 0;
}
.tap-feature-summary{
    color: var(--muted);
    margin-top: 6px;
    font-size: 0.95rem;
}
.tap-feature-link{
    color: var(--text);
    text-decoration: none;
    font-weight: 600;
    border: 1px solid rgba(255,255,255,0.2);
    border-radius: 999px;
    padding: 6px 12px;
    background: rgba(7,22,39,0.35);
}
.tap-feature-link:hover{
    border-color: rgba(30,144,255,0.6);
}
.tap-gallery-title{
    margin-top: 1.4rem;
    font-size: 1.1rem;
    font-weight: 700;
    letter-spacing: 0.02em;
}
.tap-thumb{
    display: block;
    border-radius: 12px;
    overflow: hidden;
    margin-bottom: 10px;
    border: 1px solid rgba(255,255,255,0.08);
}
.tap-thumb img{
    width: 100%;
    display: block;
}
.tap-card-title{
    font-weight: 700;
    margin-bottom: 4px;
}
.tap-card-summary{
    color: var(--muted);
    font-size: 0.9rem;
    margin-bottom: 6px;
    line-height: 1.4;
    display: -webkit-box;
    -webkit-line-clamp: 3;
    -webkit-box-orient: vertical;
    overflow: hidden;
}
.tap-card-link{
    color: var(--accent);
    text-decoration: none;
    font-size: 0.85rem;
}
.tap-card-link:hover{
    text-decoration: underline;
}
.about-link{
    color: var(--accent);
    text-decoration: none;
}
.about-link:hover{
    text-decoration: underline;
}
@keyframes about-fade-up{
    from { opacity: 0; transform: translateY(8px); }
    to { opacity: 1; transform: translateY(0); }
}
.fade-up{
    animation: about-fade-up 360ms ease both;
}
.about-hero,
.about-panel,
.about-section{
    animation: about-fade-up 420ms ease both;
}
.about-sidebar .about-panel:nth-child(1){ animation-delay: 60ms; }
.about-sidebar .about-panel:nth-child(2){ animation-delay: 120ms; }
.about-main .about-section:nth-child(1){ animation-delay: 80ms; }
.about-main .about-section:nth-child(2){ animation-delay: 140ms; }
.about-main .about-section:nth-child(3){ animation-delay: 200ms; }
@media (prefers-reduced-motion: reduce){
    .about-hero,
    .about-panel,
    .about-section{
        animation: none;
    }
    .fade-up{
        animation: none;
    }
}
.section-title{
    margin-top: 0.8rem;
    font-size: 1.6rem;
    font-weight: 700;
    letter-spacing: -0.01em;
    min-height: 3.6rem;
    display: flex;
    align-items: flex-end;
}
.section-sub{
    color: var(--muted);
    margin-top: -0.3rem;
    margin-bottom: 0.6rem;
}
.section-caption{
    color: var(--muted);
    font-size: 0.92rem;
    margin-top: 0.2rem;
}
.callout{
    border: 1px solid var(--border);
    background: linear-gradient(180deg, rgba(255,255,255,0.07), rgba(255,255,255,0.02));
    border-radius: var(--radius-md);
    padding: 12px 14px;
    box-shadow: 0 12px 22px rgba(0,0,0,0.22);
}
.callout-title{
    text-transform: uppercase;
    letter-spacing: 0.18em;
    font-size: 0.65rem;
    color: var(--muted);
    margin-bottom: 4px;
}
.callout-body{
    color: var(--text);
    font-size: 0.96rem;
    line-height: 1.5;
}
.geo-hero{
    margin-top: 0.4rem;
    background: linear-gradient(145deg, rgba(30,144,255,0.16), rgba(0,224,184,0.10), rgba(7,22,39,0.9));
    border: 1px solid rgba(255,255,255,0.14);
    box-shadow: 0 18px 30px rgba(0,0,0,0.30);
}
.geo-kicker{
    text-transform: uppercase;
    letter-spacing: 0.18em;
    font-size: 0.65rem;
    color: var(--muted);
    margin-bottom: 0.35rem;
}
.geo-title{
    font-size: 1.3rem;
    font-weight: 700;
    margin-bottom: 0.35rem;
}
.geo-lead{
    color: var(--muted);
    line-height: 1.45;
    margin-bottom: 0.6rem;
}
.geo-step{
    padding: 8px 10px;
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 12px;
    background: rgba(7,22,39,0.42);
    margin-top: 8px;
    line-height: 1.4;
}
.geo-step strong{
    color: var(--accent2);
    font-weight: 700;
}
.geo-note{
    margin-top: 0.4rem;
}
.filter-summary{
    display: flex;
    flex-wrap: wrap;
    align-items: center;
    gap: 6px;
    padding: 10px 12px;
    border: 1px solid var(--border);
    border-radius: var(--radius-md);
    background: rgba(7,22,39,0.55);
    box-shadow: inset 0 0 0 1px rgba(255,255,255,0.02);
}
.filter-summary-label{
    text-transform: uppercase;
    letter-spacing: 0.18em;
    font-size: 0.65rem;
    color: var(--muted);
    margin-right: 4px;
}
#filter-bar-marker + div[data-testid="stHorizontalBlock"],
#filter-summary-marker + div[data-testid="stHorizontalBlock"]{
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 18px;
    padding: 12px 14px;
    background: linear-gradient(180deg, rgba(255,255,255,0.06), rgba(255,255,255,0.02));
    box-shadow: 0 16px 28px rgba(0,0,0,0.24);
    animation: about-fade-up 360ms ease both;
}
#filter-summary-marker + div[data-testid="stHorizontalBlock"]{
    padding: 10px 12px;
}
.pill{
    display:inline-flex;
    align-items:center;
    gap:6px;
    padding: 4px 10px;
    border-radius: 999px;
    border: 1px solid rgba(255,255,255,0.12);
    background: rgba(255,255,255,0.04);
    font-size: 0.8rem;
}
.pill b{ font-weight: 700; }
.pill-list{
    display: flex;
    flex-wrap: wrap;
    gap: 6px;
    margin-top: 6px;
}
.pill.pill-muted{
    color: var(--muted);
    border-color: rgba(255,255,255,0.08);
    background: rgba(255,255,255,0.02);
}
.meta-card{
    margin-top: 0.4rem;
    background: linear-gradient(135deg, rgba(30,144,255,0.12), rgba(0,224,184,0.08), rgba(7,22,39,0.85));
    border: 1px solid rgba(255,255,255,0.16);
    box-shadow: 0 14px 28px rgba(0,0,0,0.26);
}
.meta-title{
    text-transform: uppercase;
    letter-spacing: 0.18em;
    font-size: 0.7rem;
    color: var(--muted);
}
.meta-sub{
    color: var(--muted);
    font-size: 0.9rem;
    margin-top: 4px;
}
.insight-panel{
    display: grid;
    grid-template-columns: minmax(0, 1.4fr) minmax(0, 1fr);
    gap: 16px;
    margin: 10px 0 4px 0;
}
.insight-card{
    background: linear-gradient(160deg, rgba(30,144,255,0.14), rgba(0,224,184,0.08), rgba(7,22,39,0.9));
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 18px;
    padding: 14px 16px;
    box-shadow: 0 16px 28px rgba(0,0,0,0.28);
}
.insight-kicker{
    text-transform: uppercase;
    letter-spacing: 0.18em;
    font-size: 0.65rem;
    color: var(--muted);
    margin-bottom: 6px;
}
.insight-title{
    font-size: 1.25rem;
    font-weight: 700;
    margin: 0 0 0.4rem 0;
}
.insight-list{
    list-style: none;
    padding: 0;
    margin: 0;
}
.insight-list li{
    position: relative;
    padding-left: 1.1rem;
    margin: 0.35rem 0;
    line-height: 1.45;
}
.insight-list li::before{
    content: "";
    position: absolute;
    left: 0;
    top: 0.55rem;
    width: 7px;
    height: 7px;
    border-radius: 2px;
    background: rgba(0,224,184,0.9);
    box-shadow: 0 0 0 2px rgba(0,224,184,0.15);
}
.mini-kpi-grid{
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 10px;
}
.mini-kpi{
    background: rgba(7,22,39,0.5);
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 12px;
    padding: 10px 12px;
}
.mini-kpi .label{
    text-transform: uppercase;
    letter-spacing: 0.14em;
    font-size: 0.62rem;
    color: var(--muted);
    margin-bottom: 4px;
}
.mini-kpi .value{
    font-size: 1.15rem;
    font-weight: 700;
}
.mini-kpi .sub{
    color: var(--muted);
    font-size: 0.82rem;
    margin-top: 4px;
}

.kpi-title{ color: var(--muted); font-size: 0.85rem; margin-bottom: 8px; }
.kpi-value{ font-size: 2.0rem; font-weight: 700; line-height: 1.15; color: var(--text); }
.kpi-sub{ color: var(--muted); font-size: 0.9rem; margin-top: 6px; }

.big-title{
    font-size: 3.0rem;
    font-weight: 800;
    letter-spacing: -0.02em;
    margin: 0.1rem 0 0.2rem 0;
}

.subtitle{
    font-size: 1.2rem;
    color: var(--muted);
    margin-bottom: 1rem;
}

.stTabs [data-baseweb="tab-list"]{
    gap: 8px;
}
.stTabs [data-baseweb="tab"]{
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.07);
    border-radius: 14px;
    padding: 10px 14px;
}
.stTabs [aria-selected="true"]{
    border-color: rgba(30,144,255,0.55) !important;
    background: rgba(30,144,255,0.12) !important;
}

[data-testid="stTextInput"] input,
[data-testid="stTextInput"] textarea{
    background: rgba(255,255,255,0.04) !important;
    border: 1px solid rgba(255,255,255,0.10) !important;
    color: var(--text) !important;
}

[data-testid="stSelectbox"] div[role="combobox"]{
    background: rgba(255,255,255,0.04) !important;
    border: 1px solid rgba(255,255,255,0.10) !important;
}

.chip{
    display:inline-block;
    padding: 6px 10px;
    border-radius: 999px;
    border: 1px solid rgba(255,255,255,0.12);
    background: rgba(255,255,255,0.04);
    font-size: 0.85rem;
    margin-right: 6px;
}

[data-testid="stSidebar"]{
    background: rgba(255,255,255,0.02) !important;
    border-right: 1px solid rgba(255,255,255,0.07);
}

[data-testid="stDataFrame"]{
    border-radius: 16px;
    overflow: hidden;
    border: 1px solid rgba(255,255,255,0.10);
}
div[data-testid="stDataFrame"]{
    background: rgba(7, 22, 39, 0.65);
}

button[kind="primary"]{
    border-radius: 14px !important;
}

/* Force dark text for all selectbox and dropdown items */
[data-testid="stSelectbox"] [data-baseweb="select"] *,
[data-testid="stSelectbox"] [data-baseweb="select"] span,
[data-testid="stSelectbox"] [data-baseweb="select"] div,
[data-testid="stSelectbox"] [data-baseweb="select"] input,
[data-testid="stSelectbox"] [data-baseweb="select"] [role="option"],
[data-testid="stSelectbox"] [data-baseweb="select"] [role="listbox"] *,
[data-baseweb="popover"] ul[role="listbox"] *,
[data-baseweb="popover"] div[role="option"],
[data-baseweb="popover"] div[role="option"] * {
    color: #0b1a2b !important;
    background: white !important;
}

[data-testid="stSelectbox"] [data-baseweb="select"] ::placeholder{
    color: #405264 !important;
}
[data-testid="stSelectbox"] div[role="combobox"] *{
    color: #0b1a2b !important;
}

/* Improve selectbox readability */
[data-testid="stSelectbox"] [data-baseweb="select"] > div,
[data-testid="stSelectbox"] [data-baseweb="select"] span{
    color: #0b1a2b !important;
    font-weight: 700 !important;
    opacity: 1 !important;
}
[data-testid="stSelectbox"] [data-baseweb="select"] ::placeholder{
    color: #2b3c4d !important;
    opacity: 1 !important;
}
[data-baseweb="popover"] div[role="option"]{
    font-weight: 600;
}

/* Custom navigation header */
.custom-nav{
    position: fixed;
    top: 0;
    left: 0;
    right: 0;
    z-index: 1000;
    height: var(--nav-h);
    padding: 0 16px;
    background: linear-gradient(180deg, var(--nav-bg) 0%, rgba(6, 16, 30, 0.94) 70%, rgba(6, 16, 30, 0.9) 100%);
    border-bottom: 1px solid var(--nav-border);
    box-shadow: 0 12px 24px rgba(0,0,0,0.25);
    backdrop-filter: blur(8px);
}
.custom-nav::after{
    content: "";
    position: absolute;
    left: 0;
    right: 0;
    bottom: 0;
    height: 1px;
    background: linear-gradient(90deg, rgba(255,255,255,0.02), rgba(255,255,255,0.16), rgba(255,255,255,0.02));
}
.custom-nav .nav-inner{
    max-width: 1280px;
    margin: 0 auto;
    height: 100%;
    display: flex;
    align-items: center;
    justify-content: flex-start;
    gap: 28px;
    padding-right: 0;
}
.custom-nav .brand{
    display: flex;
    flex-direction: column;
    gap: 2px;
    line-height: 1.05;
    color: var(--text);
    border-left: 3px solid var(--accent);
    padding-left: 12px;
}
.custom-nav .brand-top{
    font-size: 0.9rem;
    letter-spacing: 0.24em;
    text-transform: uppercase;
    opacity: 0.8;
}
.custom-nav .brand-bottom{
    font-size: 1.5rem;
    font-weight: 700;
    letter-spacing: -0.01em;
}
.custom-nav .nav-links{
    display: flex;
    gap: 22px;
    align-items: center;
    flex: 1 1 auto;
    margin-left: 12px;
    padding-right: calc(var(--nav-search-w) + 20px);
    white-space: nowrap;
}
.custom-nav .nav-link{
    position: relative;
    color: var(--muted);
    text-decoration: none;
    font-weight: 600;
    font-size: 0.98rem;
    letter-spacing: 0.01em;
    padding: 10px 2px;
    transition: color 120ms ease;
}
.custom-nav .nav-link::after{
    content: "";
    position: absolute;
    left: 0;
    right: 0;
    bottom: 2px;
    height: 2px;
    background: transparent;
    transition: background 120ms ease;
}
.custom-nav .nav-link:hover{
    color: var(--text);
}
.custom-nav .nav-link.active{
    color: var(--text);
}
.custom-nav .nav-link.active::after{
    background: var(--accent);
}

div[data-testid="stTextInput"]:has(input[aria-label="Nav search"]){
    position: fixed;
    top: 0;
    right: 18px;
    z-index: 1002;
    width: min(var(--nav-search-w), 38vw);
    height: var(--nav-h);
    display: flex;
    align-items: center;
}
div[data-testid="stTextInput"]:has(input[aria-label="Nav search"]) > div{
    width: 100%;
    margin: 0 !important;
}
div[data-testid="stTextInput"]:has(input[aria-label="Nav search"]) input{
    height: var(--nav-search-h) !important;
    border-radius: 999px !important;
    padding: 0 38px 0 14px !important;
    background: rgba(10, 18, 32, 0.75) !important;
    border: 1px solid rgba(255,255,255,0.18) !important;
    box-shadow: inset 0 0 0 1px rgba(255,255,255,0.03);
    color: var(--text) !important;
    background-image: url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' width='16' height='16' viewBox='0 0 24 24' fill='none' stroke='%23b7c2d3' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='11' cy='11' r='7'/><line x1='21' y1='21' x2='16.65' y2='16.65'/></svg>");
    background-repeat: no-repeat;
    background-position: right 12px center;
    background-size: 16px;
}
div[data-testid="stTextInput"]:has(input[aria-label="Nav search"]) input::placeholder{
    color: rgba(255,255,255,0.6);
}

/* Mobile responsive improvements */
@media (max-width: 768px) {
    :root{ --nav-h: 108px; --nav-search-w: 100%; }
    .block-container {
        padding-left: calc(0.5rem + env(safe-area-inset-left, 0px));
        padding-right: calc(0.5rem + env(safe-area-inset-right, 0px));
        padding-top: calc(var(--nav-h) + 3.6rem);
    }
    .section-title { font-size: 1.3rem; min-height: 2.5rem; }
    .big-title { font-size: 2rem; }
    .subtitle { font-size: 1rem; }
    [data-testid="stTextInput"] input { font-size: 16px !important; min-height: 44px !important; }
    [data-testid="stSelectbox"] div[role="combobox"] { font-size: 14px !important; min-height: 44px; }
    [data-testid="stMultiSelect"] div[role="combobox"] { min-height: 44px; }
    button { padding: 0.5rem 1rem !important; font-size: 14px !important; min-height: 44px; }
    .stTabs [data-baseweb="tab"] { padding: 10px 12px !important; font-size: 13px !important; min-height: 40px; }
    .stTabs [data-baseweb="tab-list"]{
        flex-wrap: nowrap;
        overflow-x: auto;
        -webkit-overflow-scrolling: touch;
        gap: 8px;
        padding-bottom: 6px;
    }
    .stTabs [data-baseweb="tab-list"]::-webkit-scrollbar{ display: none; }
    .stTabs [data-baseweb="tab"]{ flex: 0 0 auto; }
    .card{ padding: 12px 12px 10px 12px; border-radius: 16px; }
    .kpi-title{ font-size: 0.78rem; }
    .kpi-value{ font-size: 1.4rem; }
    .kpi-sub{ font-size: 0.85rem; }
    .section-sub{ font-size: 0.9rem; }
    .chip{ font-size: 0.75rem; padding: 4px 8px; margin-right: 4px; }
    [data-testid="stHorizontalBlock"]{ flex-direction: column; }
    [data-testid="column"]{ width: 100% !important; flex: 1 1 100% !important; }
    [data-testid="stDataFrame"]{ overflow-x: auto; -webkit-overflow-scrolling: touch; }
    button[kind="primary"]{ width: 100%; }
    .custom-nav{
        height: var(--nav-h);
        padding: 8px 12px;
    }
    .custom-nav .nav-inner{
        flex-direction: column;
        align-items: flex-start;
        gap: 6px;
        padding-right: 0;
    }
    .custom-nav .nav-links{
        flex-wrap: nowrap;
        gap: 14px;
        overflow-x: auto;
        -webkit-overflow-scrolling: touch;
        padding-right: 0;
        width: 100%;
    }
    .custom-nav .nav-links::-webkit-scrollbar{ display: none; }
    .custom-nav .nav-link{ font-size: 0.9rem; padding: 10px 6px; }
    .custom-nav .brand-top{ font-size: 0.8rem; }
    .custom-nav .brand-bottom{ font-size: 1.2rem; }
    .insight-panel{ grid-template-columns: 1fr; }
    .mini-kpi-grid{ grid-template-columns: repeat(2, minmax(0, 1fr)); }
    .about-shell{ grid-template-columns: 1fr; }
    .about-hero{ padding: 18px 16px 16px 16px; }
    .about-title{ font-size: 1.6rem; }
    .about-panel-head{
        flex-direction: column;
        align-items: flex-start;
    }
    .about-list-grid{ grid-template-columns: 1fr; }
    .source-grid{ grid-template-columns: 1fr; }
    .about-section{ padding-left: 22px; }
    .about-section::before{ left: 12px; }
    .tap-hero{ padding: 16px 14px 14px 14px; }
    .tap-hero-title{ font-size: 1.5rem; }
    .tap-feature-head{ flex-direction: column; align-items: flex-start; }
    .geo-hero{ padding: 14px 12px 12px 12px; }
    .geo-title{ font-size: 1.05rem; }
    .geo-step{ padding: 7px 9px; }
    div[data-testid="stTextInput"]:has(input[aria-label="Nav search"]){
        top: calc(var(--nav-h) + 6px);
        left: calc(12px + env(safe-area-inset-left, 0px));
        right: calc(12px + env(safe-area-inset-right, 0px));
        width: auto;
        height: auto;
    }
    div[data-testid="stTextInput"]:has(input[aria-label="Nav search"]) input{
        width: 100%;
        height: 44px !important;
    }
    div[data-testid="stPlotlyChart"]{ padding: 6px 8px 4px 8px; border-radius: 16px; touch-action: pan-y; }
}
</style>
""",
    unsafe_allow_html=True,
)

st.markdown(
    """
<style>
:root{
    --bg: #0d1724;
    --bg-soft: #132133;
    --surface: #172638;
    --surface-2: #1c2d42;
    --surface-3: #22364f;
    --border: rgba(175, 194, 214, 0.30);
    --border-strong: rgba(198, 214, 231, 0.46);
    --text: #edf3fa;
    --muted: #b6c5d8;
    --accent: #86a7c6;
    --accent-2: #6f92b4;
    --radius-sm: 10px;
    --radius-md: 14px;
    --radius-lg: 18px;
    --shadow-1: 0 8px 18px rgba(2, 9, 16, 0.30);
    --shadow-2: 0 16px 30px rgba(2, 9, 16, 0.36);
}

html, body, [data-testid="stAppViewContainer"]{
    background:
        radial-gradient(900px 360px at 12% -12%, rgba(121, 152, 183, 0.18), transparent 62%),
        radial-gradient(820px 360px at 92% -10%, rgba(78, 108, 139, 0.16), transparent 62%),
        linear-gradient(180deg, #0d1724 0%, #121f2e 50%, #0d1724 100%) !important;
    color: var(--text) !important;
    font-family: 'IBM Plex Sans', 'Source Sans 3', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
}
[data-testid="stAppViewContainer"]::before{
    display: none !important;
}
.block-container{
    max-width: 1360px;
    padding-left: clamp(0.9rem, 1.9vw, 1.6rem);
    padding-right: clamp(0.9rem, 1.9vw, 1.6rem);
}
h1, h2, h3{
    color: var(--text) !important;
}

.card{
    border-radius: var(--radius-md);
    border: 1px solid var(--border);
    background: linear-gradient(180deg, rgba(26, 40, 57, 0.96), rgba(17, 29, 45, 0.94));
    box-shadow: var(--shadow-1);
}
div[data-testid="stPlotlyChart"]{
    border-radius: var(--radius-md);
    border: 1px solid rgba(175, 194, 214, 0.26);
    background: linear-gradient(180deg, rgba(23, 38, 56, 0.95), rgba(15, 26, 40, 0.94));
    box-shadow: var(--shadow-2);
}
[data-testid="stDataFrame"]{
    border-color: rgba(177, 196, 216, 0.30);
    background: rgba(12, 22, 35, 0.86);
    border-radius: var(--radius-md);
}

.custom-nav{
    background: rgba(11, 20, 31, 0.97);
    border-bottom: 1px solid rgba(171, 191, 212, 0.30);
    box-shadow: 0 10px 24px rgba(2, 9, 16, 0.42);
}
.custom-nav .brand{
    border-left-color: var(--accent);
    padding-left: 12px;
}
.custom-nav .brand-top{
    color: var(--muted);
    letter-spacing: 0.18em;
    font-size: 0.64rem;
    font-weight: 600;
}
.custom-nav .brand-bottom{
    font-size: 1.1rem;
    font-weight: 700;
    letter-spacing: 0.01em;
}
.custom-nav .nav-link{
    color: rgba(226, 236, 248, 0.84);
    font-size: 0.91rem;
    font-weight: 600;
}
.custom-nav .nav-link:hover{
    color: #f4f8fc;
}
.custom-nav .nav-link.active{
    color: #f5f9fd;
}
.custom-nav .nav-link.active::after{
    background: var(--accent);
    height: 3px;
}

.policy-hero{
    padding: 20px 22px 18px 22px;
    margin: 0 0 12px 0;
    border: 1px solid var(--border-strong);
    background:
        linear-gradient(128deg, rgba(94, 126, 157, 0.20), rgba(27, 42, 60, 0.92) 58%),
        linear-gradient(180deg, rgba(24, 37, 54, 0.96), rgba(15, 27, 41, 0.94));
}
.policy-kicker{
    text-transform: uppercase;
    letter-spacing: 0.17em;
    font-size: 0.65rem;
    font-weight: 600;
    color: var(--muted);
    margin-bottom: 7px;
}
.policy-title{
    font-family: 'IBM Plex Serif', 'Merriweather', Georgia, serif;
    font-size: clamp(1.76rem, 2.2vw, 2.25rem);
    font-weight: 600;
    line-height: 1.22;
    margin: 0;
}
.policy-subtitle{
    margin: 9px 0 0 0;
    color: var(--muted);
    line-height: 1.56;
    max-width: 940px;
    font-size: 0.99rem;
}
.policy-pill-list{
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin-top: 12px;
}
.policy-pill{
    display: inline-flex;
    align-items: center;
    padding: 4px 11px 5px 11px;
    border-radius: 999px;
    border: 1px solid rgba(151, 180, 209, 0.46);
    background: rgba(97, 128, 160, 0.22);
    color: rgba(236, 243, 251, 0.98);
    font-size: 0.77rem;
    line-height: 1.2;
}

.workspace-links-heading{
    margin: 0.15rem 0 0.35rem 0;
    font-size: 0.72rem;
    letter-spacing: 0.16em;
    text-transform: uppercase;
    color: var(--muted);
    font-weight: 700;
}
.workspace-link-help{
    font-size: 0.82rem;
    line-height: 1.38;
    color: var(--muted);
    margin-top: 0.35rem;
    min-height: 2.45rem;
}

.policy-panel{
    padding: 14px 15px 12px 15px;
    border-radius: 12px;
    border: 1px solid var(--border);
    background: linear-gradient(180deg, rgba(30, 47, 67, 0.62), rgba(15, 27, 41, 0.88));
    margin-bottom: 11px;
}
.policy-panel h3{
    margin: 0 0 7px 0;
    font-size: 1.02rem;
    line-height: 1.32;
}
.policy-panel p{
    margin: 0;
    color: var(--muted);
    line-height: 1.5;
}

.section-title{
    font-family: 'IBM Plex Serif', 'Merriweather', Georgia, serif;
    letter-spacing: 0.01em;
}
.section-sub{
    color: rgba(193, 210, 227, 0.87);
}
.section-caption{
    color: rgba(185, 203, 222, 0.79);
}

.callout{
    border-radius: 12px;
    border-color: rgba(163, 183, 205, 0.30);
    background: linear-gradient(180deg, rgba(28, 45, 65, 0.60), rgba(15, 27, 41, 0.86));
}
.callout-title{
    letter-spacing: 0.15em;
}

.geo-hero{
    margin-top: 0.45rem;
    background:
        linear-gradient(145deg, rgba(88, 120, 150, 0.22), rgba(23, 39, 58, 0.92) 60%),
        linear-gradient(180deg, rgba(20, 34, 51, 0.96), rgba(14, 25, 39, 0.94));
    border: 1px solid rgba(170, 190, 211, 0.31);
    box-shadow: 0 17px 30px rgba(2, 9, 16, 0.42);
}
.geo-title{
    font-family: 'IBM Plex Serif', 'Merriweather', Georgia, serif;
    font-size: 1.2rem;
}
.geo-lead{
    line-height: 1.5;
}

.filter-summary{
    border: 1px solid rgba(166, 187, 208, 0.33);
    border-radius: 12px;
    background: rgba(13, 23, 35, 0.88);
    box-shadow: inset 0 0 0 1px rgba(150, 176, 202, 0.10);
}
.filter-summary-label{
    letter-spacing: 0.15em;
    color: rgba(188, 206, 227, 0.86);
}

.chip{
    border: 1px solid rgba(166, 187, 208, 0.32);
    background: rgba(76, 108, 140, 0.22);
}

.kpi-title{
    color: rgba(191, 208, 226, 0.86);
    font-size: 0.8rem;
    margin-bottom: 7px;
    letter-spacing: 0.02em;
}
.kpi-value{
    font-size: clamp(1.44rem, 2.1vw, 1.9rem);
    font-weight: 700;
    line-height: 1.16;
    color: var(--text);
}
.kpi-sub{
    color: rgba(186, 204, 224, 0.79);
    font-size: 0.85rem;
    margin-top: 5px;
    line-height: 1.36;
}

.insight-panel{
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 12px;
    margin-bottom: 0.9rem;
}
.insight-card{
    border: 1px solid rgba(168, 189, 211, 0.30);
    border-radius: 13px;
    padding: 12px 13px;
    background: linear-gradient(180deg, rgba(31, 48, 68, 0.62), rgba(15, 27, 41, 0.88));
}
.insight-kicker{
    font-size: 0.65rem;
    letter-spacing: 0.16em;
    text-transform: uppercase;
    color: var(--muted);
    margin-bottom: 5px;
}
.insight-title{
    font-family: 'IBM Plex Serif', 'Merriweather', Georgia, serif;
    font-size: 1.05rem;
    margin-bottom: 0.35rem;
}
.insight-list{
    margin: 0.25rem 0 0 1rem;
    padding: 0;
}
.insight-list li{
    margin: 0.22rem 0;
    line-height: 1.45;
    color: var(--muted);
}
.mini-kpi-grid{
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 8px;
}
.mini-kpi{
    border: 1px solid rgba(165, 185, 206, 0.28);
    border-radius: 11px;
    padding: 8px 9px;
    background: rgba(17, 29, 44, 0.72);
}
.mini-kpi .label{
    font-size: 0.72rem;
    letter-spacing: 0.04em;
    color: var(--muted);
    margin-bottom: 3px;
}
.mini-kpi .value{
    font-size: 1.18rem;
    font-weight: 700;
    line-height: 1.12;
    margin-bottom: 2px;
}
.mini-kpi .sub{
    font-size: 0.77rem;
    color: rgba(187, 205, 224, 0.8);
    line-height: 1.3;
}

.app-note{
    border: 1px solid rgba(164, 185, 206, 0.34);
    border-radius: 11px;
    padding: 10px 12px;
    margin: 8px 0 10px 0;
    background: linear-gradient(180deg, rgba(29, 46, 66, 0.62), rgba(14, 24, 37, 0.88));
    color: rgba(198, 214, 233, 0.90);
    font-size: 0.9rem;
    line-height: 1.45;
}
.app-note strong{
    color: var(--text);
}

.handoff-card{
    border: 1px solid rgba(166, 187, 209, 0.34);
    border-radius: 12px;
    padding: 10px 12px 9px 12px;
    margin: 0.55rem 0 0.75rem 0;
    background: linear-gradient(180deg, rgba(32, 49, 69, 0.64), rgba(15, 27, 41, 0.89));
}
.handoff-kicker{
    font-size: 0.64rem;
    letter-spacing: 0.16em;
    text-transform: uppercase;
    color: var(--muted);
    margin-bottom: 4px;
    font-weight: 700;
}
.handoff-title{
    font-size: 0.94rem;
    font-weight: 700;
    margin-bottom: 0.18rem;
}
.handoff-sub{
    color: var(--muted);
    font-size: 0.84rem;
    line-height: 1.4;
    margin-bottom: 0.25rem;
}

.map-legend{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(195px, 1fr));
    gap: 6px;
    margin: 0.4rem 0 0.65rem 0;
}
.map-legend-item{
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 8px;
    border: 1px solid rgba(158, 180, 203, 0.26);
    border-radius: 10px;
    padding: 5px 9px;
    background: rgba(15, 27, 41, 0.78);
    font-size: 0.82rem;
    transition: background 0.18s ease, border-color 0.18s ease;
}
.map-legend-item:hover{
    background: rgba(25, 42, 62, 0.88);
    border-color: rgba(158, 180, 203, 0.44);
}
.map-legend-left{
    display: flex;
    align-items: center;
    gap: 7px;
    min-width: 0;
}
.map-legend-left span{
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.map-legend-chip{
    width: 11px;
    height: 11px;
    border-radius: 50%;
    border: 1px solid rgba(255,255,255,0.26);
    flex-shrink: 0;
}
.map-toolbar-note{
    color: var(--muted);
    font-size: 0.86rem;
    margin: 0.2rem 0 0.42rem 0;
}
.map-tab-banner{
    position: relative;
    overflow: hidden;
    border: 1px solid rgba(255,255,255,0.14);
    border-radius: 16px;
    padding: 13px 16px 11px 16px;
    margin: 0.15rem 0 0.6rem 0;
    background: linear-gradient(135deg, rgba(30,144,255,0.16), rgba(0,224,184,0.08), rgba(10,22,34,0.92));
    box-shadow: 0 12px 24px rgba(0,0,0,0.22);
}
.map-tab-banner::before{
    content: "";
    position: absolute;
    top: 0; left: 0; width: 4px; height: 100%;
    background: linear-gradient(180deg, rgba(30,144,255,0.80), rgba(0,224,184,0.65));
    border-radius: 16px 0 0 16px;
}
.map-tab-banner::after{
    content: "";
    position: absolute;
    inset: 0;
    background-image:
        linear-gradient(transparent 19px, rgba(255,255,255,0.025) 20px),
        linear-gradient(90deg, transparent 19px, rgba(255,255,255,0.025) 20px);
    background-size: 24px 24px;
    opacity: 0.14;
    pointer-events: none;
}
.map-tab-banner > *{
    position: relative;
    z-index: 1;
}
.map-tab-title{
    font-size: 1.08rem;
    font-weight: 700;
    margin: 0;
    letter-spacing: -0.01em;
}
.map-tab-sub{
    color: var(--muted);
    margin-top: 0.2rem;
    line-height: 1.42;
    font-size: 0.86rem;
}
.map-tab-pill-row{
    margin-top: 0.52rem;
    display: flex;
    flex-wrap: wrap;
    gap: 6px;
}
.map-tab-pill{
    display: inline-flex;
    align-items: center;
    border: 1px solid rgba(255,255,255,0.18);
    border-radius: 999px;
    padding: 3px 9px;
    font-size: 0.76rem;
    background: rgba(7,22,39,0.46);
    color: rgba(234,242,248,0.96);
}
.map-controls-shell{
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 16px;
    padding: 10px 12px 8px 12px;
    background: linear-gradient(180deg, rgba(255,255,255,0.06), rgba(255,255,255,0.02));
    box-shadow: 0 14px 26px rgba(0,0,0,0.22);
}
.map-workflow-card{
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 16px;
    padding: 10px 12px;
    background: linear-gradient(145deg, rgba(25, 63, 99, 0.52), rgba(13, 26, 38, 0.90));
}
.map-workflow-title{
    text-transform: uppercase;
    letter-spacing: 0.16em;
    font-size: 0.66rem;
    color: var(--muted);
    margin-bottom: 0.28rem;
}
.map-workflow-step{
    font-size: 0.88rem;
    margin: 0.18rem 0;
    color: rgba(232,241,249,0.96);
}
.map-workflow-step strong{
    color: var(--accent2);
}
.map-context-grid{
    display: grid;
    grid-template-columns: repeat(4, minmax(0, 1fr));
    gap: 8px;
    margin-top: 0.4rem;
}
.map-context-card{
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 12px;
    padding: 8px 10px 8px 12px;
    background: rgba(7,22,39,0.54);
    position: relative;
    overflow: hidden;
}
.map-context-card::before{
    content: "";
    position: absolute;
    top: 0; left: 0; width: 3px; height: 100%;
    background: linear-gradient(180deg, rgba(30,144,255,0.65), rgba(0,224,184,0.50));
    border-radius: 12px 0 0 12px;
}
.map-context-label{
    text-transform: uppercase;
    letter-spacing: 0.14em;
    font-size: 0.58rem;
    color: var(--muted);
}
.map-context-value{
    margin-top: 2px;
    font-size: 0.88rem;
    font-weight: 600;
    color: rgba(242,248,252,0.96);
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.map-stage-shell{
    border: 1px solid rgba(255,255,255,0.10);
    border-radius: 16px;
    padding: 11px 12px 10px 12px;
    background: linear-gradient(180deg, rgba(255,255,255,0.05), rgba(255,255,255,0.015));
    box-shadow: 0 12px 22px rgba(0,0,0,0.2);
    margin: 0.48rem 0 0.72rem 0;
}
.map-stage-title{
    margin: 0;
    font-size: 1rem;
    font-weight: 700;
}
.map-stage-sub{
    color: var(--muted);
    margin-top: 0.2rem;
    line-height: 1.42;
    font-size: 0.9rem;
}
.map-toolbar-strong{
    color: rgba(232,244,251,0.95);
    font-size: 0.87rem;
    margin: 0.2rem 0 0.45rem 0;
}
.map-side-stat-grid{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
    gap: 8px;
}
.map-side-stat{
    border: 1px solid rgba(255,255,255,0.10);
    border-radius: 12px;
    background: rgba(13, 25, 39, 0.84);
    padding: 8px 10px;
}
.map-side-stat-label{
    text-transform: uppercase;
    letter-spacing: 0.14em;
    font-size: 0.62rem;
    color: var(--muted);
}
.map-side-stat-value{
    font-size: 1.02rem;
    font-weight: 700;
    margin-top: 2px;
}
.map-score-chip{
    display: inline-flex;
    align-items: center;
    border-radius: 999px;
    border: 1px solid rgba(255,255,255,0.18);
    padding: 3px 8px;
    font-size: 0.75rem;
    background: rgba(7,22,39,0.44);
    margin-right: 5px;
    margin-top: 4px;
}
.map-batch-rank-shell{
    position: relative;
    border: 1px solid rgba(0,224,184,0.22);
    border-radius: 14px;
    padding: 10px 12px 9px 14px;
    background: linear-gradient(135deg, rgba(0,224,184,0.06) 0%, rgba(11,22,34,0.82) 40%);
    overflow: hidden;
}
.map-batch-rank-shell::before{
    content:""; position:absolute; left:0; top:0; bottom:0; width:3px;
    background:linear-gradient(180deg, rgba(0,224,184,0.65), rgba(30,144,255,0.40));
    border-radius:14px 0 0 14px;
}
.map-flow-grid{
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 12px;
    margin: 0.5rem 0 0.65rem 0;
}
.map-flow-card{
    position: relative;
    border: 1px solid rgba(167, 188, 211, 0.22);
    border-radius: 14px;
    background: linear-gradient(180deg, rgba(30, 48, 68, 0.58), rgba(14, 26, 40, 0.90));
    padding: 11px 13px 10px 13px;
    transition: border-color 0.2s, box-shadow 0.2s, transform 0.15s;
    overflow: hidden;
}
.map-flow-card:hover{
    border-color: rgba(30,144,255,0.35);
    box-shadow: 0 4px 16px rgba(30,144,255,0.08);
    transform: translateY(-1px);
}
.map-flow-kicker{
    text-transform: uppercase;
    letter-spacing: 0.14em;
    font-size: 0.60rem;
    color: var(--muted);
    margin-bottom: 4px;
    display: flex;
    align-items: center;
}
.map-flow-title{
    font-size: 0.96rem;
    font-weight: 700;
    color: rgba(242,248,252,0.96);
}
.map-flow-sub{
    margin-top: 2px;
    color: rgba(193, 209, 227, 0.80);
    font-size: 0.80rem;
    line-height: 1.38;
}
.map-rail-title{
    text-transform: uppercase;
    letter-spacing: 0.15em;
    font-size: 0.64rem;
    color: var(--muted);
    margin: 0.2rem 0 0.42rem 0;
}
.map-control-stack{
    display: grid;
    gap: 0.68rem;
}
.map-inline-note{
    font-size: 0.82rem;
    color: rgba(190, 208, 226, 0.82);
    margin: 0.14rem 0 0.25rem 0;
    line-height: 1.38;
}
.map-insight-grid{
    display: grid;
    grid-template-columns: repeat(4, minmax(0, 1fr));
    gap: 8px;
    margin: 0.35rem 0 0.55rem 0;
}
.map-insight{
    border: 1px solid rgba(165, 186, 208, 0.28);
    border-radius: 12px;
    padding: 8px 10px;
    background: rgba(15, 27, 41, 0.86);
}
.map-insight-label{
    text-transform: uppercase;
    letter-spacing: 0.12em;
    font-size: 0.6rem;
    color: var(--muted);
}
.map-insight-value{
    margin-top: 3px;
    font-size: 1.06rem;
    font-weight: 700;
    line-height: 1.2;
}
.map-insight-sub{
    margin-top: 2px;
    font-size: 0.74rem;
    color: rgba(187, 204, 222, 0.78);
    line-height: 1.35;
}
.map-selected-shell{
    border: 1px solid rgba(170, 192, 215, 0.34);
    border-radius: 13px;
    padding: 9px 11px;
    background: linear-gradient(160deg, rgba(30, 144, 255, 0.16), rgba(0, 224, 184, 0.10), rgba(11, 22, 35, 0.90));
}
.map-selected-title{
    font-size: 0.86rem;
    font-weight: 700;
    margin-bottom: 2px;
}
.map-selected-sub{
    font-size: 0.8rem;
    color: rgba(193, 209, 227, 0.85);
    line-height: 1.38;
}
.map-score-chip.is-high{
    border-color: rgba(66, 210, 162, 0.42);
    background: rgba(22, 103, 80, 0.34);
}
.map-score-chip.is-medium{
    border-color: rgba(227, 192, 108, 0.42);
    background: rgba(102, 80, 32, 0.32);
}
.map-score-chip.is-low{
    border-color: rgba(206, 122, 122, 0.42);
    background: rgba(109, 48, 57, 0.30);
}
.map-score-chip.is-unknown{
    border-color: rgba(173, 192, 214, 0.30);
    background: rgba(32, 47, 67, 0.44);
}
.map-batch-head{
    display: flex;
    flex-wrap: wrap;
    align-items: baseline;
    justify-content: space-between;
    gap: 8px;
    margin-bottom: 0.3rem;
}
.map-batch-head h4{
    margin: 0;
    font-size: 0.96rem;
    letter-spacing: -0.01em;
}
.map-batch-head .meta{
    color: var(--muted);
    font-size: 0.8rem;
}
/* batch status color indicators */
.map-batch-status-ok{
    display: inline-flex; align-items: center;
    padding: 2px 8px; border-radius: 999px; font-size: 0.72rem; font-weight: 600;
    background: rgba(22, 103, 80, 0.30); color: #42d2a2;
    border: 1px solid rgba(66, 210, 162, 0.35);
}
.map-batch-status-fail{
    display: inline-flex; align-items: center;
    padding: 2px 8px; border-radius: 999px; font-size: 0.72rem; font-weight: 600;
    background: rgba(109, 48, 57, 0.28); color: #ce7a7a;
    border: 1px solid rgba(206, 122, 122, 0.35);
}
.map-batch-status-warn{
    display: inline-flex; align-items: center;
    padding: 2px 8px; border-radius: 999px; font-size: 0.72rem; font-weight: 600;
    background: rgba(102, 80, 32, 0.28); color: #e3c06c;
    border: 1px solid rgba(227, 192, 108, 0.35);
}
.map-mission-shell{
    border: 1px solid rgba(171, 192, 214, 0.32);
    border-radius: 14px;
    padding: 10px 14px 10px 14px;
    margin: 0.4rem 0 0.5rem 0;
    background:
        linear-gradient(135deg, rgba(116, 147, 178, 0.16), rgba(19, 33, 48, 0.90) 62%),
        linear-gradient(180deg, rgba(24, 39, 57, 0.92), rgba(14, 25, 39, 0.90));
    box-shadow: 0 12px 22px rgba(2, 9, 16, 0.30);
    position: relative;
    overflow: hidden;
}
.map-mission-shell::before{
    content: "";
    position: absolute;
    top: 0; left: 0; width: 4px; height: 100%;
    background: linear-gradient(180deg, rgba(255, 200, 55, 0.80), rgba(255, 140, 40, 0.60));
    border-radius: 14px 0 0 14px;
}
.map-mission-title{
    text-transform: uppercase;
    letter-spacing: 0.14em;
    font-size: 0.62rem;
    color: rgba(194, 212, 230, 0.90);
    margin-bottom: 3px;
    font-weight: 700;
}
.map-mission-sub{
    color: rgba(211, 225, 239, 0.92);
    font-size: 0.85rem;
    line-height: 1.42;
}
/* evidence rating color coding */
.map-evidence-high{ color: #42d2a2; }
.map-evidence-moderate{ color: #e3c06c; }
.map-evidence-low{ color: #ce7a7a; }
.map-evidence-unknown{ color: rgba(173,192,214,0.82); }
.map-overview-shell{
    position: relative;
    border: 1px solid rgba(170, 191, 214, 0.28);
    border-radius: 14px;
    padding: 12px 14px;
    margin: 0.25rem 0 0.65rem 0;
    background:
        linear-gradient(145deg, rgba(116, 147, 178, 0.14), rgba(18, 31, 45, 0.92) 62%),
        linear-gradient(180deg, rgba(24, 39, 57, 0.92), rgba(14, 25, 39, 0.90));
    box-shadow: 0 14px 26px rgba(2, 9, 16, 0.32);
    overflow: hidden;
}
.map-overview-shell::before{
    content:""; position:absolute; left:0; top:0; bottom:0; width:3px;
    background:linear-gradient(180deg, rgba(30,144,255,0.55), rgba(0,224,184,0.35));
    border-radius:3px 0 0 3px;
}
.map-overview-title{
    text-transform: uppercase;
    letter-spacing: 0.14em;
    font-size: 0.66rem;
    color: rgba(194, 212, 230, 0.90);
    margin-bottom: 4px;
    font-weight: 700;
}
.map-overview-sub{
    color: rgba(214, 227, 240, 0.86);
    font-size: 0.82rem;
    line-height: 1.4;
    margin-bottom: 0.45rem;
}
.map-overview-grid{
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 8px;
}
.map-overview-card{
    position: relative;
    border: 1px solid rgba(167, 188, 210, 0.22);
    border-radius: 11px;
    padding: 8px 10px;
    background: rgba(11, 23, 36, 0.82);
    transition: border-color 0.2s, box-shadow 0.2s;
}
.map-overview-card:hover{
    border-color: rgba(30,144,255,0.35);
    box-shadow: 0 2px 8px rgba(30,144,255,0.08);
}
.map-overview-label{
    text-transform: uppercase;
    letter-spacing: 0.12em;
    font-size: 0.56rem;
    color: rgba(188, 206, 224, 0.80);
}
.map-overview-value{
    margin-top: 3px;
    font-size: 1rem;
    font-weight: 700;
    color: var(--text);
    display: flex;
    align-items: baseline;
    gap: 6px;
}
.map-overview-subtext{
    margin-top: 2px;
    color: rgba(181, 199, 218, 0.72);
    font-size: 0.72rem;
    line-height: 1.32;
}
.map-overview-foot{
    margin-top: 0.48rem;
    color: rgba(201, 216, 233, 0.85);
    font-size: 0.76rem;
    line-height: 1.35;
    padding: 6px 8px;
    border-radius: 8px;
    background: rgba(255,193,7,0.06);
    border: 1px solid rgba(255,193,7,0.15);
}
.map-signal-grid{
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 10px;
    margin: 0.5rem 0 0.75rem 0;
}
.map-signal{
    position: relative;
    border: 1px solid rgba(166, 188, 211, 0.22);
    border-radius: 14px;
    padding: 10px 12px 10px 12px;
    background: linear-gradient(135deg, rgba(13,24,37,0.92) 0%, rgba(18,30,45,0.88) 100%);
    transition: border-color 0.2s, box-shadow 0.2s, transform 0.15s;
    overflow: hidden;
}
.map-signal:hover{
    border-color: rgba(30,144,255,0.40);
    box-shadow: 0 2px 12px rgba(30,144,255,0.10);
    transform: translateY(-1px);
}
.map-signal-icon{
    font-size: 1.05rem;
    margin-bottom: 2px;
    line-height: 1;
}
.map-signal-label{
    text-transform: uppercase;
    letter-spacing: 0.12em;
    font-size: 0.58rem;
    color: rgba(191, 208, 226, 0.82);
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.map-signal-value{
    margin-top: 3px;
    font-size: 1.06rem;
    font-weight: 700;
    line-height: 1.22;
    color: var(--text);
    display: flex;
    align-items: baseline;
    gap: 6px;
    flex-wrap: wrap;
}
.map-signal-sub{
    margin-top: 2px;
    font-size: 0.72rem;
    color: rgba(184, 202, 221, 0.72);
    line-height: 1.34;
}
.map-brief-grid{
    display: grid;
    grid-template-columns: repeat(4, minmax(0, 1fr));
    gap: 10px;
    margin: 0.35rem 0 0.55rem 0;
}
.map-brief-card{
    position: relative;
    border: 1px solid rgba(164, 186, 209, 0.22);
    border-radius: 12px;
    padding: 9px 11px;
    background: linear-gradient(180deg, rgba(28, 45, 65, 0.55), rgba(15, 27, 41, 0.88));
    transition: border-color 0.2s, box-shadow 0.2s;
    overflow: hidden;
}
.map-brief-card:hover{
    border-color: rgba(30,144,255,0.35);
    box-shadow: 0 2px 10px rgba(30,144,255,0.08);
}
.map-brief-label{
    text-transform: uppercase;
    letter-spacing: 0.12em;
    font-size: 0.58rem;
    color: rgba(190, 208, 226, 0.80);
}
.map-brief-value{
    margin-top: 3px;
    font-size: 1.02rem;
    font-weight: 700;
    line-height: 1.2;
}
.map-brief-sub{
    margin-top: 2px;
    font-size: 0.72rem;
    color: rgba(183, 201, 220, 0.72);
    line-height: 1.33;
}
.map-lead-banner{
    border: 1px solid rgba(169, 191, 214, 0.32);
    border-radius: 11px;
    padding: 8px 12px 8px 14px;
    margin: 0.3rem 0 0.45rem 0;
    background: rgba(14, 27, 42, 0.88);
    font-size: 0.84rem;
    color: rgba(213, 226, 240, 0.92);
    position: relative;
    overflow: hidden;
}
.map-lead-banner::before{
    content: "";
    position: absolute;
    top: 0; left: 0; width: 3px; height: 100%;
    background: linear-gradient(180deg, rgba(66,210,162,0.75), rgba(30,144,255,0.60));
    border-radius: 11px 0 0 11px;
}

[data-testid="stTextInput"] input,
[data-testid="stTextInput"] textarea,
[data-testid="stSelectbox"] div[role="combobox"],
[data-testid="stMultiSelect"] div[role="combobox"]{
    border-radius: var(--radius-sm) !important;
    border: 1px solid rgba(162, 185, 209, 0.34) !important;
    background: rgba(15, 27, 41, 0.90) !important;
    color: var(--text) !important;
}
[data-testid="stTextInput"] input::placeholder{
    color: rgba(184, 201, 221, 0.74) !important;
}

button[kind="primary"],
button[kind="secondary"]{
    border-radius: 10px !important;
    border: 1px solid rgba(164, 185, 207, 0.30) !important;
}

.stTabs [data-baseweb="tab"]{
    border-radius: 10px;
    border-color: rgba(160, 182, 205, 0.26);
    background: linear-gradient(180deg, rgba(28, 44, 63, 0.72), rgba(14, 26, 40, 0.88));
    transition: background 0.2s ease, border-color 0.2s ease;
}
.stTabs [data-baseweb="tab"]:hover{
    border-color: rgba(160, 182, 205, 0.45);
    background: linear-gradient(180deg, rgba(38, 56, 78, 0.78), rgba(18, 30, 45, 0.90));
}
.stTabs [aria-selected="true"]{
    border-color: rgba(143, 174, 205, 0.75) !important;
    background: linear-gradient(180deg, rgba(58, 86, 115, 0.62), rgba(19, 33, 49, 0.92)) !important;
    box-shadow: 0 4px 12px rgba(30, 144, 255, 0.12) !important;
}

div[data-testid="stTextInput"]:has(input[aria-label="Nav search"]){
    top: calc(var(--nav-h) + 14px);
}
div[data-testid="stTextInput"]:has(input[aria-label="Nav search"]) input{
    border: 1px solid rgba(173, 194, 216, 0.44) !important;
    background: rgba(14, 24, 36, 0.97) !important;
}

@media (max-width: 950px){
    .insight-panel{
        grid-template-columns: 1fr;
    }
}
@media (max-width: 768px){
    .policy-title{
        font-size: 1.5rem;
    }
    .map-flow-grid{
        grid-template-columns: 1fr;
    }
    .map-insight-grid{
        grid-template-columns: repeat(2, minmax(0, 1fr));
    }
    .map-signal-grid{
        grid-template-columns: 1fr;
    }
    .map-brief-grid{
        grid-template-columns: 1fr;
    }
    .map-overview-grid{
        grid-template-columns: 1fr;
    }
    .policy-hero{
        padding: 16px 14px 14px 14px;
    }
    .policy-subtitle{
        font-size: 0.94rem;
    }
    .kpi-value{
        font-size: 1.3rem;
    }
    .workspace-link-help{
        min-height: 0;
    }
    .mini-kpi-grid{
        grid-template-columns: 1fr;
    }
}
</style>
""",
    unsafe_allow_html=True,
)

# =========================================================
# GLOBAL UX ENHANCEMENTS
# =========================================================
st.markdown(
    """
<style>
/* -- Data-table improvements -- */
[data-testid="stDataFrame"] table tbody tr:nth-child(even) td{
    background: rgba(255,255,255,0.025);
}
[data-testid="stDataFrame"] table tbody tr:hover td{
    background: rgba(30,144,255,0.08) !important;
    transition: background 120ms ease;
}
[data-testid="stDataFrame"] table thead th{
    position: sticky;
    top: 0;
    z-index: 2;
    background: rgba(15,27,42,0.97) !important;
    backdrop-filter: blur(6px);
    border-bottom: 2px solid rgba(160,185,210,0.22);
}

/* -- Smooth scroll behavior -- */
html{ scroll-behavior: smooth; }

/* -- Back-to-top floating button -- */
#tfl-back-to-top{
    position: fixed;
    bottom: 28px;
    right: 28px;
    z-index: 999;
    width: 42px;
    height: 42px;
    border-radius: 50%;
    border: 1px solid rgba(160,185,210,0.35);
    background: rgba(11,20,31,0.92);
    color: rgba(220,235,250,0.90);
    font-size: 18px;
    cursor: pointer;
    display: flex;
    align-items: center;
    justify-content: center;
    opacity: 0;
    pointer-events: none;
    transition: opacity 0.3s ease, transform 0.3s ease, box-shadow 0.2s ease;
    box-shadow: 0 6px 18px rgba(0,0,0,0.28);
    backdrop-filter: blur(6px);
}
#tfl-back-to-top.visible{
    opacity: 1;
    pointer-events: auto;
}
#tfl-back-to-top:hover{
    border-color: rgba(134,167,198,0.65);
    box-shadow: 0 8px 22px rgba(0,0,0,0.35);
    transform: translateY(-2px);
}

/* -- Focus / keyboard accessibility -- */
button:focus-visible,
[data-testid="stSelectbox"] div[role="combobox"]:focus-visible,
[data-testid="stTextInput"] input:focus-visible{
    outline: 2px solid rgba(134,167,198,0.7);
    outline-offset: 2px;
}

/* -- Spinner quality of life -- */
[data-testid="stSpinner"] > div{
    animation: tfl-spinner-fade 0.35s ease both;
}
@keyframes tfl-spinner-fade{
    from{ opacity:0; transform:translateY(4px); }
    to{ opacity:1; transform:translateY(0); }
}

/* -- Toast notification -- */
.tfl-toast{
    position: fixed;
    bottom: 80px;
    right: 28px;
    z-index: 998;
    padding: 10px 18px;
    border-radius: 12px;
    border: 1px solid rgba(160,185,210,0.30);
    background: rgba(11,20,31,0.94);
    color: rgba(220,235,250,0.92);
    font-size: 0.88rem;
    font-family: 'IBM Plex Sans', system-ui, sans-serif;
    box-shadow: 0 8px 22px rgba(0,0,0,0.30);
    backdrop-filter: blur(6px);
    opacity: 0;
    pointer-events: none;
    transform: translateY(8px);
    transition: opacity 0.3s ease, transform 0.3s ease;
}
.tfl-toast.show{
    opacity: 1;
    pointer-events: auto;
    transform: translateY(0);
}

/* -- Print-friendly -- */
@media print{
    .custom-nav,
    #tfl-back-to-top,
    .tfl-toast,
    [data-testid="stSidebar"],
    [data-testid="stToolbar"],
    button[kind="primary"],
    button[kind="secondary"],
    [data-testid="stSpinner"]{ display: none !important; }

    html, body, [data-testid="stAppViewContainer"]{
        background: white !important;
        color: #1a1a1a !important;
    }
    .block-container{ padding-top: 0 !important; max-width: 100% !important; }
    h1, h2, h3{ color: #1a1a1a !important; }
    p, li, span, div{ color: #1a1a1a !important; }
    .card, div[data-testid="stPlotlyChart"]{
        box-shadow: none !important;
        border: 1px solid #ccc !important;
        background: white !important;
    }
    [data-testid="stDataFrame"]{ border-color: #ccc !important; }
    .kpi-title, .section-caption, .callout-title{ color: #555 !important; }
}
</style>

<!-- Back-to-top button -->
<button id="tfl-back-to-top" onclick="window.scrollTo({top:0, behavior:'smooth'})" title="Back to top">&#9650;</button>
<script>
(function(){
  const btn = document.getElementById('tfl-back-to-top');
  if (!btn) return;
  const container = window.parent.document.querySelector('[data-testid="stAppViewContainer"]') || window;
  const target = container.querySelector && container.querySelector('.main') || window;
  const checkScroll = () => {
    const scrollY = window.scrollY || document.documentElement.scrollTop || 0;
    btn.classList.toggle('visible', scrollY > 400);
  };
  window.addEventListener('scroll', checkScroll, {passive:true});
  checkScroll();
})();
</script>
""",
    unsafe_allow_html=True,
)


# =========================================================
# CRASH PROTECTION: Safe page wrapper
# =========================================================
def _safe_page(page_name: str):
    """Decorator that wraps page functions in try/except to prevent full app crashes."""
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except st.runtime.scriptrunner.StopException:
                raise  # Let Streamlit's st.stop() and st.switch_page() propagate
            except Exception as exc:
                logging.exception("Unhandled error in page %s", page_name)
                st.error(
                    f"An unexpected error occurred on the **{page_name}** page. "
                    f"Try refreshing the browser or clearing filters.\n\n"
                    f"Error details: `{type(exc).__name__}: {exc}`"
                )
                # Offer a reset button
                if st.button("Reset and reload", key=f"crash_reset_{page_name}"):
                    for key in list(st.session_state.keys()):
                        del st.session_state[key]
                    st.rerun()
        return wrapper
    return decorator


@_safe_page('About')
def _page_about():
    _run_page_renderer("src.pages.about")


@_safe_page('Media Briefings')
def _page_turn_off_tap():
    _run_page_renderer("src.pages.multimedia")


@_safe_page('Policy Context')
def _page_solutions():
    _run_page_renderer("src.pages.solutions")


@st.cache_data(show_spinner=False, ttl=300, max_entries=4)
def _build_category_chart_data(df: pd.DataFrame) -> pd.DataFrame:
    """Pre-compute expenditure-by-category data for 85th-89th TFL sessions (cached)."""
    cat_base = df.copy()
    cat_base["Session"] = cat_base["Session"].astype(str).str.strip()
    cat_base = ensure_cols(cat_base, {"Client": "", "Low_num": 0.0, "High_num": 0.0, "IsTFL": 0})
    cat_base["IsTFL"] = pd.to_numeric(cat_base["IsTFL"], errors="coerce").fillna(0)
    cat_base = cat_base[cat_base["IsTFL"] == 1]
    cat_base["SessionBase"] = _session_base_number_series(cat_base["Session"])
    cat_base = cat_base[cat_base["SessionBase"].between(85, 89)]
    cat_base = cat_base[cat_base["Client"].fillna("").astype(str).str.strip() != ""]
    if cat_base.empty:
        return pd.DataFrame(columns=["SessionBase", "Category", "Total", "SessionLabel"])
    cat_base["Category"] = cat_base["Client"].map(lambda x: match_entity_type(x)[1])
    cat_base["Low_num"] = pd.to_numeric(cat_base["Low_num"], errors="coerce").fillna(0)
    cat_base["High_num"] = pd.to_numeric(cat_base["High_num"], errors="coerce").fillna(0)
    cat_base["Mid"] = (cat_base["Low_num"] + cat_base["High_num"]) / 2
    cat_group = (
        cat_base.groupby(["SessionBase", "Category"], as_index=False)["Mid"]
        .sum()
        .rename(columns={"Mid": "Total"})
    )
    cat_group["SessionLabel"] = cat_group["SessionBase"].map(_session_base_label)
    return cat_group

@_safe_page('Clients')
def _page_client_lookup():
    _run_page_renderer("src.pages.clients")



def _page_member_lookup():
    _run_page_renderer("src.pages.legislators")


_mp5_miles = _map_page_helpers._mp5_miles
_mp5_method_weight = _map_page_helpers._mp5_method_weight
_mp5_confidence_weight = _map_page_helpers._mp5_confidence_weight
_mp5_priority_from_score = _map_page_helpers._mp5_priority_from_score
_mp5_geocode_badge = _map_page_helpers._mp5_geocode_badge
_build_mp5_css = _map_page_helpers._build_mp5_css


@_safe_page('Map & Address Full')
def _page_map_address_full_pass():
    _run_page_renderer("src.pages.map_address")



@_safe_page("Map & Address Rebuild")
def _page_map_address_rebuild():
    # Route to the full redesign workspace implementation.
    _page_map_address_full_pass()


@_safe_page("Lobbyists")
def _page_lobby_lookup():
    _run_page_renderer("src.pages.lobbyists")

_about_page = st.Page(_page_about, title="Start Here", url_path="about", default=True)
_lobby_page = st.Page(_page_lobby_lookup, title="Lobbyists", url_path="lobbyists")
_client_page = st.Page(_page_client_lookup, title="Clients", url_path="clients")
_map_page = st.Page(_page_map_address_rebuild, title="Map & Address", url_path="map-address")
_member_page = st.Page(_page_member_lookup, title="Legislators", url_path="legislators")
_solutions_page = st.Page(_page_solutions, title="Policy Context", url_path="solutions")
_tap_page = st.Page(_page_turn_off_tap, title="Media Briefings", url_path="multimedia")
_pages = [
    _about_page,
    _lobby_page,
    _client_page,
    _map_page,
    _member_page,
    _solutions_page,
    _tap_page,
]
_active_page = st.navigation(_pages, position="hidden")

def _nav_href(page) -> str:
    url_path = page.url_path
    return "./" if url_path == "" else f"./{url_path}"

def _journey_steps() -> list[tuple[str, str, str, object]]:
    return []

def _render_page_intro(kicker: str, title: str, subtitle: str, pills: list[str] | None = None) -> None:
    kicker_safe = html.escape(kicker or "", quote=True)
    title_safe = html.escape(title or "", quote=True)
    subtitle_safe = html.escape(subtitle or "", quote=True)
    pill_html = ""
    if pills:
        tokens = [f'<span class="policy-pill">{html.escape(str(p), quote=True)}</span>' for p in pills if str(p).strip()]
        if tokens:
            pill_html = f'<div class="policy-pill-list">{"".join(tokens)}</div>'
    st.markdown(
        f"""
<div class="card policy-hero">
  <div class="policy-kicker">{kicker_safe}</div>
  <div class="policy-title">{title_safe}</div>
  <p class="policy-subtitle">{subtitle_safe}</p>
  {pill_html}
</div>
""",
        unsafe_allow_html=True,
    )

def _is_guided_mode() -> bool:
    return False

def _render_journey(current_key: str) -> None:
    return

def _render_workspace_guide(
    question: str,
    steps: list[str] | None = None,
    method_note: str | None = None,
) -> None:
    return

def _render_quickstart(
    page_key: str,
    steps: list[str],
    note: str | None = None,
) -> None:
    return

def _render_evidence_guardrails(
    can_answer: list[str] | None = None,
    cannot_answer: list[str] | None = None,
    next_checks: list[str] | None = None,
) -> None:
    return

def _render_workspace_links(
    key_prefix: str,
    actions: list[tuple[str, object, str]],
) -> None:
    valid_actions = [
        (label, page, help_text)
        for label, page, help_text in actions
        if str(label).strip()
    ]
    if not valid_actions:
        return
    st.markdown('<div class="workspace-links-heading">Continue The Investigation</div>', unsafe_allow_html=True)
    cols = st.columns(len(valid_actions))
    for idx, (label, page, help_text) in enumerate(valid_actions):
        with cols[idx]:
            if st.button(
                label,
                key=f"{key_prefix}_nav_{idx}",
                width="stretch",
                help=help_text,
            ):
                st.switch_page(page)
            if help_text:
                st.markdown(
                    f'<div class="workspace-link-help">{html.escape(help_text, quote=True)}</div>',
                    unsafe_allow_html=True,
                )

_nav_items = [
    (_about_page, "Start Here"),
    (_lobby_page, "Lobbyists"),
    (_client_page, "Clients"),
    (_map_page, "Map & Address"),
    (_member_page, "Legislators"),
    (_solutions_page, "Policy"),
    (_tap_page, "Media"),
]
_nav_links = []
for page, label in _nav_items:
    active = " active" if page == _active_page else ""
    _nav_links.append(
        f'<a class="nav-link{active}" href="{_nav_href(page)}" target="_self">{label}</a>'
    )

st.markdown(
    f"""
<div class="custom-nav">
  <div class="nav-inner">
    <div class="brand">
      <div class="brand-top">Texas Taxpayer Protection</div>
      <div class="brand-bottom">Lobbying Transparency Center</div>
    </div>
    <div class="nav-links">
      {''.join(_nav_links)}
    </div>
  </div>
</div>
""",
    unsafe_allow_html=True,
)

if "nav_search_query" not in st.session_state:
    st.session_state.nav_search_query = ""
if "nav_search_last" not in st.session_state:
    st.session_state.nav_search_last = ""
if "nav_search_trigger" not in st.session_state:
    st.session_state.nav_search_trigger = False
def _nav_submit() -> None:
    st.session_state.nav_search_trigger = True

nav_query_raw = st.text_input(
    "Nav search",
    key="nav_search_query",
    placeholder="Global search: lobbyist, client, legislator, or bill (example: HB 4)",
    label_visibility="collapsed",
    on_change=_nav_submit,
    help="Routes to the best workspace and carries your query forward.",
)
nav_query = nav_query_raw.strip()
nav_search_submitted = False
if nav_query and st.session_state.nav_search_trigger:
    nav_search_submitted = True
    st.session_state.nav_search_last = nav_query
    st.session_state.nav_search_trigger = False
elif not nav_query:
    st.session_state.nav_search_trigger = False
nav_suggest_slot = st.empty()
nav_skip_submit = False

# =========================================================
# HELPERS
# =========================================================
_RE_NONWORD = re.compile(r"[^\w]+", flags=re.UNICODE)
_TITLE_WORDS = {"MR", "MRS", "MS", "MISS", "DR", "HON", "JR", "SR", "II", "III", "IV"}
_RE_TITLE_WORDS = re.compile(r"\b(" + "|".join(_TITLE_WORDS) + r")\b\.?", re.IGNORECASE)
_NICKNAME_MAP = {
    "CHUCK": {"CHARLES"},
    "CHARLIE": {"CHARLES"},
    "CHARLES": {"CHUCK", "CHARLIE"},
}

def norm_name(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).replace("\u00A0", " ").strip().upper()
    return _RE_NONWORD.sub("", s)

def norm_name_series(s: pd.Series) -> pd.Series:
    return (
        s.fillna("")
         .astype(str)
         .str.replace("\u00A0", " ", regex=False)
         .str.strip()
         .str.upper()
         .str.replace(_RE_NONWORD, "", regex=True)
    )

def clean_filer_name_series(s: pd.Series) -> pd.Series:
    s = s.fillna("").astype(str)
    s = s.str.replace(_RE_PARENS, "", regex=True)
    s = s.str.replace(_RE_TITLE_WORDS, "", regex=True)
    s = s.str.replace(_RE_WHITESPACE, " ", regex=True).str.strip()
    return s

def clean_person_name(name: str) -> str:
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return ""
    s = str(name).replace("\u00A0", " ").strip()
    if not s:
        return ""
    s = _RE_PARENS.sub("", s)
    s = _RE_TITLE_WORDS.sub("", s)
    s = _RE_WHITESPACE.sub(" ", s).strip()
    return s

# PERFORMANCE: shared urllib3 connection pool — keeps TCP connections alive
# across the ~10-layer parallel ArcGIS queries instead of opening a new
# socket for every single request.
try:
    import urllib3 as _urllib3
    _ARCGIS_HTTP = _urllib3.PoolManager(
        num_pools=16, maxsize=12, retries=False,
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=_urllib3.Timeout(connect=10, read=30),
    )
except ImportError:
    _ARCGIS_HTTP = None  # type: ignore[assignment]

def _arcgis_get_json(url: str, params: dict | None = None, timeout: int = 30, _retries: int = 3) -> dict:
    target = url
    if params:
        target = f"{url}?{urllib.parse.urlencode(params)}"
    last_exc: Exception | None = None
    for attempt in range(_retries):
        try:
            if _ARCGIS_HTTP is not None:
                resp = _ARCGIS_HTTP.request(
                    "GET", target,
                    timeout=_urllib3.Timeout(connect=10, read=timeout),
                )
                return json.loads(resp.data.decode("utf-8"))
            else:
                req = urllib.request.Request(target, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=timeout) as resp:
                    return json.loads(resp.read().decode("utf-8"))
        except Exception as exc:
            last_exc = exc
            if attempt < _retries - 1:
                time.sleep(min(2 ** attempt, 4))
    logging.warning("ArcGIS request failed after %d attempts: %s — %s", _retries, target[:120], last_exc)
    raise last_exc  # type: ignore[misc]

def _canonical_school_district_name(value: str) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).upper().replace("&", " AND ").replace("/", " ")
    s = re.sub(r"\bC\.?I\.?S\.?D\.?\b", " CONSOLIDATED INDEPENDENT SCHOOL DISTRICT ", s)
    s = re.sub(r"\bI\.?S\.?D\.?\b", " INDEPENDENT SCHOOL DISTRICT ", s)
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _looks_like_school_district_name(value: str) -> bool:
    s = _canonical_school_district_name(value)
    return bool(s) and ("SCHOOL DISTRICT" in s)

@functools.lru_cache(maxsize=2048)
def _school_district_root_key(value: str) -> str:
    s = _canonical_school_district_name(value)
    if not s:
        return ""
    s = re.sub(r"\bTHE\b", " ", s)
    s = re.sub(r"\b(CONSOLIDATED\s+)?INDEPENDENT\s+SCHOOL\s+DISTRICT\b", " ", s)
    s = re.sub(r"\bSCHOOL\s+DISTRICT\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return norm_name(s)

def _canonical_county_name(value: str) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).upper().replace("&", " AND ").replace("/", " ")
    s = re.sub(r"\bCTY\.?\b", " COUNTY ", s)
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _looks_like_county_name(value: str) -> bool:
    s = _canonical_county_name(value)
    return bool(s) and ("COUNTY" in s)

@functools.lru_cache(maxsize=512)
def _county_root_key(value: str) -> str:
    s = _canonical_county_name(value)
    if not s:
        return ""
    s = re.sub(r"\bTHE\b", " ", s)
    s = re.sub(r"\bCOUNTY OF\b", " ", s)
    s = re.sub(r"\bCOMMISSIONERS? COURT\b", " ", s)
    s = re.sub(r"\bCOUNTY\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return norm_name(s)

def _canonical_city_name(value: str) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).upper().replace("&", " AND ").replace("/", " ")
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

@functools.lru_cache(maxsize=4096)
def _canonical_subdivision_text(value: str) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).upper().replace("&", " AND ").replace("/", " ")
    replacements = [
        (r"\bM\.?U\.?D\.?\b", " MUNICIPAL UTILITY DISTRICT "),
        (r"\bW\.?C\.?I\.?D\.?\b", " WATER CONTROL AND IMPROVEMENT DISTRICT "),
        (r"\bW\.?I\.?D\.?\b", " WATER IMPROVEMENT DISTRICT "),
        (r"\bF\.?W\.?S\.?D\.?\b", " FRESH WATER SUPPLY DISTRICT "),
        (r"\bL\.?I\.?D\.?\b", " LEVEE IMPROVEMENT DISTRICT "),
        (r"\bM\.?M\.?D\.?\b", " MUNICIPAL MANAGEMENT DISTRICT "),
        (r"\bS\.?U\.?D\.?\b", " SPECIAL UTILITY DISTRICT "),
        (r"\bS\.?W\.?C\.?D\.?\b", " SOIL AND WATER CONTROL DISTRICT "),
        (r"\bG\.?C\.?D\.?\b", " GROUNDWATER CONSERVATION DISTRICT "),
        (r"\bE\.?S\.?D\.?\b", " EMERGENCY SERVICES DISTRICT "),
        (r"\bR\.?M\.?A\.?\b", " REGIONAL MOBILITY AUTHORITY "),
        (r"\bM\.?T\.?A\.?\b", " METROPOLITAN TRANSIT AUTHORITY "),
        (r"\bD\.?A\.?R\.?T\.?\b", " DALLAS AREA RAPID TRANSIT "),
        (r"\bC\.?A\.?D\.?\b", " APPRAISAL DISTRICT "),
        (r"\bL\.?G\.?C\.?\b", " LOCAL GOVERNMENT CORPORATION "),
        (r"\bCORPERATION\b", " CORPORATION "),
        (r"\bHOSP\.?\s+DIST\.?\b", " HOSPITAL DISTRICT "),
        (r"\bNAV\.?\s+DIST\.?\b", " NAVIGATION DISTRICT "),
        (r"\bDIST\.?\b", " DISTRICT "),
        (r"\bNO\.?\b", " "),
        (r"\bNUMBER\b", " "),
    ]
    for pattern, replacement in replacements:
        s = re.sub(pattern, replacement, s)
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _subdivision_root_from_patterns(value: str, remove_patterns: list[str]) -> str:
    s = _canonical_subdivision_text(value)
    if not s:
        return ""
    s = re.sub(r"\bTHE\b", " ", s)
    for pattern in remove_patterns:
        s = re.sub(pattern, " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return norm_name(s)

@functools.lru_cache(maxsize=2048)
def classify_requested_entity_type(value: str) -> str:
    s = _canonical_subdivision_text(value)
    if not s:
        return ""
    if re.search(r"\b(JUNIOR|COMMUNITY)\s+COLLEGE\b|\bCOLLEGE\s+DISTRICT\b", s):
        return "Junior College District"
    if "HOSPITAL DISTRICT" in s:
        return "Hospital District"
    if "MUNICIPAL UTILITY DISTRICT" in s:
        return "Municipal Utility District"
    if "EMERGENCY SERVICES DISTRICT" in s:
        return "Emergency Services District"
    if "GROUNDWATER CONSERVATION DISTRICT" in s:
        return "Groundwater Conservation District"
    if re.search(r"\bLOCAL\s+GOVERNMENT\s+CORPORATION\b|\bDEVELOPMENT\s+CORPORATION\b", s):
        return "Local Government Corporation"
    if "DRAINAGE DISTRICT" in s:
        return "Drainage District"
    if "FRESH WATER SUPPLY DISTRICT" in s:
        return "Fresh Water Supply District"
    if "IRRIGATION DISTRICT" in s:
        return "Irrigation District"
    if "LEVEE IMPROVEMENT DISTRICT" in s:
        return "Levee Improvement District"
    if "MUNICIPAL MANAGEMENT DISTRICT" in s:
        return "Municipal Management District"
    if "REGIONAL DISTRICT" in s:
        return "Regional District"
    if "RIVER AUTHORITY" in s:
        return "River Authority"
    if re.search(r"\bSOIL\s+(AND\s+)?WATER\s+CONTROL\s+DISTRICT\b", s):
        return "Soil & Water Control District"
    if "SPECIAL UTILITY DISTRICT" in s:
        return "Special Utility District"
    if "WATER IMPROVEMENT DISTRICT" in s:
        return "Water Improvement District"
    if "REGIONAL MOBILITY AUTHORITY" in s:
        return "Regional Mobility Authority"
    if re.search(r"\bWATER\s+CONTROL\s+(AND\s+)?IMPROVEMENT\s+DISTRICT\b", s):
        return "Water Control & Improvement District"
    if "NAVIGATION DISTRICT" in s:
        return "Navigation District"
    if (
        "TRANSIT AUTHORITY" in s
        or "METROPOLITAN TRANSIT AUTHORITY" in s
        or "TRANSPORTATION AUTHORITY" in s
        or re.search(r"\bAREA\s+RAPID\s+TRANSIT\b|\bRAPID\s+TRANSIT\b|\bMASS\s+TRANSIT\b|\bDART\b", s)
        or re.search(r"\bTRANSIT\b", s)
    ):
        return "Transit Authority"
    if "PORT AUTHORITY" in s:
        return "Port Authority"
    if "HOUSING AUTHORITY" in s:
        return "Housing Authority"
    if "APPRAISAL DISTRICT" in s:
        return "Appraisal District"
    return ""

@functools.lru_cache(maxsize=512)
def _canonical_water_district_type(value: str) -> str:
    s = _canonical_subdivision_text(value)
    if not s:
        return ""
    if "MUNICIPAL UTILITY DISTRICT" in s:
        return "Municipal Utility District"
    if "DRAINAGE DISTRICT" in s:
        return "Drainage District"
    if "FRESH WATER SUPPLY DISTRICT" in s:
        return "Fresh Water Supply District"
    if "IRRIGATION DISTRICT" in s:
        return "Irrigation District"
    if "LEVEE IMPROVEMENT DISTRICT" in s:
        return "Levee Improvement District"
    if "MUNICIPAL MANAGEMENT DISTRICT" in s:
        return "Municipal Management District"
    if "REGIONAL DISTRICT" in s:
        return "Regional District"
    if "RIVER AUTHORITY" in s:
        return "River Authority"
    if re.search(r"\bSOIL\s+(AND\s+)?WATER\s+CONTROL\s+DISTRICT\b", s):
        return "Soil & Water Control District"
    if "SPECIAL UTILITY DISTRICT" in s:
        return "Special Utility District"
    if "WATER IMPROVEMENT DISTRICT" in s:
        return "Water Improvement District"
    if re.search(r"\bWATER\s+CONTROL\s+(AND\s+)?IMPROVEMENT\s+DISTRICT\b", s):
        return "Water Control & Improvement District"
    if "NAVIGATION DISTRICT" in s:
        return "Navigation District"
    return ""

def _looks_like_city_name(value: str) -> bool:
    s = _canonical_city_name(value)
    return bool(s) and bool(re.search(r"\b(CITY|TOWN|VILLAGE)\b", s))

def _looks_like_entity_type(value: str, entity_type: str) -> bool:
    return classify_requested_entity_type(value) == str(entity_type).strip()

@functools.lru_cache(maxsize=2048)
def _city_root_key(value: str) -> str:
    s = _canonical_city_name(value)
    if not s:
        return ""
    s = re.sub(r"\bTHE\b", " ", s)
    s = re.sub(r"\b(CITY|TOWN|VILLAGE)\s+OF\b", " ", s)
    s = re.sub(r"\b(CITY|TOWN|VILLAGE)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return norm_name(s)

def _special_entity_root_patterns(entity_type: str) -> list[str]:
    et = str(entity_type).strip()
    if et == "Hospital District":
        return [r"\bHOSPITAL\s+DISTRICT\b", r"\bDISTRICT\b"]
    if et == "Emergency Services District":
        return [r"\bEMERGENCY\s+SERVICES\s+DISTRICT\b", r"\bDISTRICT\b", r"\bE\.?S\.?D\.?\b"]
    if et == "Appraisal District":
        return [r"\bAPPRAISAL\s+DISTRICT\b", r"\bDISTRICT\b", r"\bC\.?A\.?D\.?\b"]
    if et == "Local Government Corporation":
        return [r"\bLOCAL\s+GOVERNMENT\s+CORPORATION\b", r"\bDEVELOPMENT\s+CORPORATION\b", r"\bCORPORATION\b"]
    if et == "Transit Authority":
        return TRANSIT_AUTHORITY_ROOT_PATTERNS
    if et == "Port Authority":
        return PORT_AUTHORITY_ROOT_PATTERNS
    return []

def _anchor_key_variants(value: str) -> set[str]:
    root = norm_name(value)
    if not root:
        return set()

    variants: set[str] = {root}
    no_digits = re.sub(r"\d+", "", root)
    if no_digits:
        variants.add(no_digits)

    no_geo_terms = re.sub(r"(COUNTY|CITY|TOWN|VILLAGE|OF)", "", no_digits)
    no_geo_terms = no_geo_terms.strip()
    if no_geo_terms:
        variants.add(no_geo_terms)
    return {v for v in variants if v}

def _best_lookup_key_for_candidates(
    lookup_keys: tuple[str, ...],
    candidates: set[str],
) -> tuple[str, float]:
    if not lookup_keys or not candidates:
        return "", -1.0

    best_key = ""
    best_score = -1.0
    for candidate in candidates:
        c = str(candidate).strip()
        if not c:
            continue
        for key in lookup_keys:
            k = str(key).strip()
            if not k:
                continue
            score = -1.0
            if c == k:
                score = 1000.0 + float(len(k))
            elif len(c) >= 4 and len(k) >= 4 and (c in k or k in c):
                score = float(min(len(c), len(k)))
            if score > best_score:
                best_score = score
                best_key = k
    return best_key, best_score

def _resolve_special_anchor_keys(
    client_name: str,
    entity_type: str,
    county_lookup_keys: tuple[str, ...],
    city_lookup_keys: tuple[str, ...],
) -> dict:
    candidates: set[str] = set()
    candidates |= _anchor_key_variants(_county_root_key(client_name))
    candidates |= _anchor_key_variants(_city_root_key(client_name))

    root_patterns = _special_entity_root_patterns(entity_type)
    if root_patterns:
        candidates |= _anchor_key_variants(_subdivision_root_from_patterns(client_name, root_patterns))

    county_key, county_score = _best_lookup_key_for_candidates(county_lookup_keys, candidates)
    city_key, city_score = _best_lookup_key_for_candidates(city_lookup_keys, candidates)

    canonical = _canonical_subdivision_text(client_name)
    weighted_county = county_score
    weighted_city = city_score
    if "COUNTY" in canonical:
        weighted_county += 6.0
    if re.search(r"\b(CITY|TOWN|VILLAGE)\b", canonical):
        weighted_city += 6.0
    if entity_type in COUNTY_BIASED_SPECIAL_ENTITY_TYPES:
        weighted_county += 4.0
    if entity_type in CITY_BIASED_SPECIAL_ENTITY_TYPES:
        weighted_city += 3.0

    preferred_scope = ""
    if county_key and (not city_key or weighted_county >= weighted_city):
        preferred_scope = "county"
    elif city_key:
        preferred_scope = "city"

    return {
        "county_key": county_key,
        "city_key": city_key,
        "county_score": county_score,
        "city_score": city_score,
        "preferred_scope": preferred_scope,
    }

def _match_preview(values: list[str], limit: int = 6) -> str:
    if not values:
        return ""
    preview = ", ".join(values[:limit])
    if len(values) > limit:
        return f"{preview}, +{len(values) - limit} more"
    return preview

@st.cache_data(show_spinner=False, ttl=600, max_entries=8)
def _attach_subdivision_spend_totals(matches: pd.DataFrame, client_totals: pd.DataFrame) -> pd.DataFrame:
    out = matches.copy() if isinstance(matches, pd.DataFrame) else pd.DataFrame()
    if out.empty:
        out["low_total"] = pd.Series(dtype=float)
        out["high_total"] = pd.Series(dtype=float)
        return out

    out["low_total"] = 0.0
    out["high_total"] = 0.0
    if not isinstance(client_totals, pd.DataFrame) or client_totals.empty:
        return out

    totals = ensure_cols(client_totals.copy(), {"Client": "", "Low": 0.0, "High": 0.0, "IsTFL": 0})
    totals = totals[totals["IsTFL"] == 1]
    if totals.empty:
        return out

    totals["Client"] = totals["Client"].fillna("").astype(str).str.strip()
    totals = totals[totals["Client"] != ""]
    if totals.empty:
        return out
    totals["Low"] = pd.to_numeric(totals["Low"], errors="coerce").fillna(0.0)
    totals["High"] = pd.to_numeric(totals["High"], errors="coerce").fillna(0.0)
    totals = (
        totals.groupby("Client", as_index=False)
        .agg(Low=("Low", "sum"), High=("High", "sum"))
    )
    spend_lookup = {
        str(r.Client): (float(r.Low), float(r.High))
        for r in totals.itertuples(index=False)
    }

    fallback_clients = pd.Series([[]] * len(out), index=out.index, dtype=object)
    client_series = out.get("match_clients", fallback_clients)
    low_vals: list[float] = []
    high_vals: list[float] = []
    for client_values in client_series.tolist():
        low_total = 0.0
        high_total = 0.0
        if isinstance(client_values, list):
            for raw_client in client_values:
                key = str(raw_client).strip()
                if not key:
                    continue
                low_v, high_v = spend_lookup.get(key, (0.0, 0.0))
                low_total += float(low_v)
                high_total += float(high_v)
        low_vals.append(low_total)
        high_vals.append(high_total)
    out["low_total"] = pd.array(low_vals, dtype="float64")
    out["high_total"] = pd.array(high_vals, dtype="float64")
    out["match_count"] = out.get("match_clients", pd.Series(dtype=object)).map(
        lambda v: len(v) if isinstance(v, list) else 0
    ).astype("int64")
    return out

@st.cache_data(show_spinner=False, ttl=43200, max_entries=2)
def fetch_tea_school_district_centroids() -> pd.DataFrame:
    cols = ["fid", "name", "name2", "name20", "district_code", "district_code_compact", "lon", "lat"]
    rows: list[dict] = []
    page_size = 1000
    offset = 0
    try:
        while True:
            payload = _arcgis_get_json(
                f"{TEA_ARCGIS_SCHOOL_DISTRICT_LAYER_URL}/query",
                params={
                    "where": "1=1",
                    "outFields": "FID,NAME,NAME2,NAME20,DISTRICT,DISTRICT_C",
                    "returnGeometry": "false",
                    "returnCentroid": "true",
                    "outSR": "4326",
                    "orderByFields": "FID ASC",
                    "resultRecordCount": page_size,
                    "resultOffset": offset,
                    "f": "json",
                },
            )
            features = payload.get("features", [])
            if not features:
                break
            for feat in features:
                attrs = feat.get("attributes", {}) or {}
                centroid = feat.get("centroid", {}) or {}
                fid = attrs.get("FID")
                if fid is None:
                    continue
                try:
                    lon = float(centroid.get("x"))
                    lat = float(centroid.get("y"))
                except (TypeError, ValueError):
                    continue
                rows.append(
                    {
                        "fid": int(fid),
                        "name": str(attrs.get("NAME", "")).strip(),
                        "name2": str(attrs.get("NAME2", "")).strip(),
                        "name20": str(attrs.get("NAME20", "")).strip(),
                        "district_code": str(attrs.get("DISTRICT", "")).strip(),
                        "district_code_compact": str(attrs.get("DISTRICT_C", "")).strip(),
                        "lon": lon,
                        "lat": lat,
                    }
                )
            if len(features) < page_size:
                break
            offset += len(features)
    except Exception:
        return pd.DataFrame(columns=cols)
    if not rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(rows, columns=cols)

@st.cache_data(show_spinner=False, ttl=43200, max_entries=2)
def fetch_tea_county_centroids() -> pd.DataFrame:
    cols = ["objectid", "name", "fips", "cntykey", "lon", "lat"]
    rows: list[dict] = []
    try:
        payload = _arcgis_get_json(
            f"{TEA_ARCGIS_COUNTY_LAYER_URL}/query",
            params={
                "where": "1=1",
                "outFields": "OBJECTID,FIPS,CNTYKEY,FENAME",
                "returnGeometry": "false",
                "returnCentroid": "true",
                "outSR": "4326",
                "orderByFields": "OBJECTID ASC",
                "resultRecordCount": 1000,
                "f": "json",
            },
        )
        for feat in payload.get("features", []):
            attrs = feat.get("attributes", {}) or {}
            centroid = feat.get("centroid", {}) or {}
            src_id = attrs.get("OBJECTID")
            if src_id is None:
                continue
            try:
                lon = float(centroid.get("x"))
                lat = float(centroid.get("y"))
            except (TypeError, ValueError):
                continue
            rows.append(
                {
                    "objectid": int(src_id),
                    "name": str(attrs.get("FENAME", "")).strip(),
                    "fips": str(attrs.get("FIPS", "")).strip(),
                    "cntykey": str(attrs.get("CNTYKEY", "")).strip(),
                    "lon": lon,
                    "lat": lat,
                }
            )
    except Exception:
        return pd.DataFrame(columns=cols)
    if not rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(rows, columns=cols)

@st.cache_data(show_spinner=False, ttl=43200, max_entries=2)
def fetch_texas_city_centroids() -> pd.DataFrame:
    cols = ["objectid", "name", "basename", "geoid", "lon", "lat"]
    rows: list[dict] = []
    page_size = 2000
    offset = 0
    try:
        while True:
            payload = _arcgis_get_json(
                f"{CENSUS_ARCGIS_TEXAS_CITY_LAYER_URL}/query",
                params={
                    "where": "STATE='48'",
                    "outFields": "OBJECTID,NAME,BASENAME,GEOID,CENTLON,CENTLAT",
                    "returnGeometry": "false",
                    "orderByFields": "OBJECTID ASC",
                    "resultRecordCount": page_size,
                    "resultOffset": offset,
                    "f": "json",
                },
            )
            features = payload.get("features", [])
            if not features:
                break
            for feat in features:
                attrs = feat.get("attributes", {}) or {}
                src_id = attrs.get("OBJECTID")
                if src_id is None:
                    continue
                try:
                    lon = float(str(attrs.get("CENTLON", "")).strip())
                    lat = float(str(attrs.get("CENTLAT", "")).strip())
                except (TypeError, ValueError):
                    continue
                rows.append(
                    {
                        "objectid": int(src_id),
                        "name": str(attrs.get("NAME", "")).strip(),
                        "basename": str(attrs.get("BASENAME", "")).strip(),
                        "geoid": str(attrs.get("GEOID", "")).strip(),
                        "lon": lon,
                        "lat": lat,
                    }
                )
            if len(features) < page_size:
                break
            offset += len(features)
    except Exception:
        return pd.DataFrame(columns=cols)
    if not rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(rows, columns=cols)

def _build_layer_subdivision_matches(
    tfl_client_names: tuple[str, ...],
    layer_df: pd.DataFrame,
    subdivision_type: str,
    layer_name_cols: list[str],
    layer_code_cols: list[str],
    root_patterns: list[str],
    include_client_fn,
    extra_candidate_builder=None,
    source_name: str = "",
    source_url: str = "",
) -> pd.DataFrame:
    cols = [
        "subdivision_type",
        "subdivision_name",
        "subdivision_code",
        "lon",
        "lat",
        "match_count",
        "match_clients",
        "match_clients_preview",
        "source_name",
        "source_url",
    ]
    if not tfl_client_names or layer_df.empty:
        return pd.DataFrame(columns=cols)

    exact_index: dict[str, set[str]] = {}
    root_index: dict[str, set[str]] = {}
    unique_clients = sorted({str(name).strip() for name in tfl_client_names if str(name).strip()})
    for client in unique_clients:
        canon_key = norm_name(_canonical_subdivision_text(client))
        if canon_key:
            exact_index.setdefault(canon_key, set()).add(client)
        try:
            include_client = bool(include_client_fn(client))
        except Exception:
            include_client = False
        if include_client:
            root_key = _subdivision_root_from_patterns(client, root_patterns)
            if root_key:
                root_index.setdefault(root_key, set()).add(client)
    if not exact_index and not root_index:
        return pd.DataFrame(columns=cols)
    known_root_keys = tuple(root_index.keys())

    out_rows: list[dict] = []
    for row in layer_df.itertuples(index=False):
        names = []
        for col in layer_name_cols:
            v = getattr(row, col, "")
            if v is not None and str(v).strip():
                names.append(str(v).strip())
        if extra_candidate_builder is not None:
            try:
                names.extend(extra_candidate_builder(row) or [])
            except Exception:
                pass
        names = [n for n in names if n]
        if not names:
            continue

        variant_keys = {norm_name(_canonical_subdivision_text(n)) for n in names}
        variant_keys = {k for k in variant_keys if k}
        candidate_root_keys = {_subdivision_root_from_patterns(n, root_patterns) for n in names}
        candidate_root_keys = {k for k in candidate_root_keys if k}
        matched_clients: set[str] = set()
        for key in variant_keys:
            matched_clients |= exact_index.get(key, set())

        for root_key in candidate_root_keys:
            matched_clients |= root_index.get(root_key, set())

        # Conservative fuzzy fallback for near-identical subdivision naming variants.
        if not matched_clients and candidate_root_keys and known_root_keys:
            for candidate_root in candidate_root_keys:
                if len(candidate_root) < 6:
                    continue
                close_roots = difflib.get_close_matches(candidate_root, known_root_keys, n=3, cutoff=0.93)
                for close_root in close_roots:
                    ratio = difflib.SequenceMatcher(None, candidate_root, close_root).ratio()
                    if ratio >= 0.95 or candidate_root in close_root or close_root in candidate_root:
                        matched_clients |= root_index.get(close_root, set())

        if not matched_clients:
            continue

        primary_name = names[0]
        code = ""
        for ccol in layer_code_cols:
            cv = getattr(row, ccol, "")
            if cv is not None and str(cv).strip():
                code = str(cv).strip()
                break

        matched_sorted = sorted(matched_clients)
        out_rows.append(
            {
                "subdivision_type": subdivision_type,
                "subdivision_name": primary_name,
                "subdivision_code": code,
                "lon": float(getattr(row, "lon", 0.0)),
                "lat": float(getattr(row, "lat", 0.0)),
                "match_count": int(len(matched_sorted)),
                "match_clients": matched_sorted,
                "match_clients_preview": _match_preview(matched_sorted),
                "source_name": str(source_name).strip(),
                "source_url": str(source_url).strip(),
            }
        )

    if not out_rows:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(out_rows, columns=cols)
    out = out.sort_values(["match_count", "subdivision_name"], ascending=[False, True])
    return out

@st.cache_data(show_spinner=False, ttl=43200, max_entries=2)
def fetch_tceq_water_district_centroids() -> pd.DataFrame:
    cols = ["district_name", "district_code", "type_code", "type_desc", "lon", "lat"]
    rows: list[dict] = []
    page_size = 2000
    offset = 0
    try:
        while True:
            payload = _arcgis_get_json(
                f"{TCEQ_WATER_DISTRICTS_LAYER_URL}/query",
                params={
                    "where": "1=1",
                    "outFields": "NAME,DISTRICT_ID,TYPE,TYPE_DESCRIPTION",
                    "returnGeometry": "false",
                    "returnCentroid": "true",
                    "outSR": "4326",
                    "orderByFields": "OBJECTID ASC",
                    "resultRecordCount": page_size,
                    "resultOffset": offset,
                    "f": "json",
                },
            )
            features = payload.get("features", [])
            if not features:
                break
            for feat in features:
                attrs = feat.get("attributes", {}) or {}
                centroid = feat.get("centroid", {}) or {}
                name = str(attrs.get("NAME", "")).strip()
                if not name:
                    continue
                try:
                    lon = float(centroid.get("x"))
                    lat = float(centroid.get("y"))
                except (TypeError, ValueError):
                    continue
                rows.append(
                    {
                        "district_name": name,
                        "district_code": str(attrs.get("DISTRICT_ID", "")).strip(),
                        "type_code": str(attrs.get("TYPE", "")).strip(),
                        "type_desc": str(attrs.get("TYPE_DESCRIPTION", "")).strip(),
                        "lon": lon,
                        "lat": lat,
                    }
                )
            if len(features) < page_size:
                break
            offset += len(features)
    except Exception:
        return pd.DataFrame(columns=cols)
    if not rows:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(rows, columns=cols)
    out = (
        out.groupby(["district_name", "district_code", "type_code", "type_desc"], as_index=False)
        .agg(lon=("lon", "mean"), lat=("lat", "mean"))
    )
    return out

@st.cache_data(show_spinner=False, ttl=43200, max_entries=2)
def fetch_tceq_groundwater_district_centroids() -> pd.DataFrame:
    cols = ["district_name", "district_code", "lon", "lat"]
    rows: list[dict] = []
    page_size = 500
    offset = 0
    try:
        while True:
            payload = _arcgis_get_json(
                f"{TCEQ_GROUNDWATER_DISTRICTS_LAYER_URL}/query",
                params={
                    "where": "1=1",
                    "outFields": "DISTNAME,DIST_NUM,SHORTNAM",
                    "returnGeometry": "false",
                    "returnCentroid": "true",
                    "outSR": "4326",
                    "orderByFields": "OBJECTID ASC",
                    "resultRecordCount": page_size,
                    "resultOffset": offset,
                    "f": "json",
                },
            )
            features = payload.get("features", [])
            if not features:
                break
            for feat in features:
                attrs = feat.get("attributes", {}) or {}
                centroid = feat.get("centroid", {}) or {}
                name = str(attrs.get("DISTNAME", "")).strip() or str(attrs.get("SHORTNAM", "")).strip()
                if not name:
                    continue
                try:
                    lon = float(centroid.get("x"))
                    lat = float(centroid.get("y"))
                except (TypeError, ValueError):
                    continue
                rows.append(
                    {
                        "district_name": name,
                        "district_code": str(attrs.get("DIST_NUM", "")).strip(),
                        "lon": lon,
                        "lat": lat,
                    }
                )
            if len(features) < page_size:
                break
            offset += len(features)
    except Exception:
        return pd.DataFrame(columns=cols)
    if not rows:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(rows, columns=cols)
    out = (
        out.groupby(["district_name", "district_code"], as_index=False)
        .agg(lon=("lon", "mean"), lat=("lat", "mean"))
    )
    return out

@st.cache_data(show_spinner=False, ttl=43200, max_entries=2)
def fetch_texas_rma_centroids() -> pd.DataFrame:
    cols = ["district_name", "district_code", "lon", "lat"]
    rows: list[dict] = []
    try:
        payload = _arcgis_get_json(
            f"{TEXAS_RMA_LAYER_URL}/query",
            params={
                "where": "1=1",
                "outFields": "OBJECTID,RMA,Label",
                "returnGeometry": "false",
                "returnCentroid": "true",
                "outSR": "4326",
                "orderByFields": "OBJECTID ASC",
                "resultRecordCount": 1000,
                "f": "json",
            },
        )
        for feat in payload.get("features", []):
            attrs = feat.get("attributes", {}) or {}
            centroid = feat.get("centroid", {}) or {}
            name = str(attrs.get("Label", "")).strip() or str(attrs.get("RMA", "")).strip()
            if not name:
                continue
            try:
                lon = float(centroid.get("x"))
                lat = float(centroid.get("y"))
            except (TypeError, ValueError):
                continue
            rows.append(
                {
                    "district_name": name,
                    "district_code": str(attrs.get("OBJECTID", "")).strip(),
                    "lon": lon,
                    "lat": lat,
                }
            )
    except Exception:
        return pd.DataFrame(columns=cols)
    if not rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(rows, columns=cols)

@st.cache_data(show_spinner=False, ttl=43200, max_entries=2)
def fetch_texas_junior_college_centroids() -> pd.DataFrame:
    cols = ["district_name", "district_code", "name2", "lon", "lat"]
    rows: list[dict] = []
    page_size = 500
    offset = 0
    try:
        while True:
            payload = _arcgis_get_json(
                f"{TEXAS_JUNIOR_COLLEGE_LAYER_URL}/query",
                params={
                    "where": "1=1",
                    "outFields": "OBJECTID,DISTRICT,NAME1,NAME2,NAME3",
                    "returnGeometry": "false",
                    "returnCentroid": "true",
                    "outSR": "4326",
                    "orderByFields": "OBJECTID ASC",
                    "resultRecordCount": page_size,
                    "resultOffset": offset,
                    "f": "json",
                },
            )
            features = payload.get("features", [])
            if not features:
                break
            for feat in features:
                attrs = feat.get("attributes", {}) or {}
                centroid = feat.get("centroid", {}) or {}
                name1 = str(attrs.get("NAME1", "")).strip()
                name2 = str(attrs.get("NAME2", "")).strip()
                name = name1 or name2
                if not name:
                    continue
                try:
                    lon = float(centroid.get("x"))
                    lat = float(centroid.get("y"))
                except (TypeError, ValueError):
                    continue
                rows.append(
                    {
                        "district_name": name,
                        "district_code": str(attrs.get("DISTRICT", "")).strip(),
                        "name2": name2,
                        "lon": lon,
                        "lat": lat,
                    }
                )
            if len(features) < page_size:
                break
            offset += len(features)
    except Exception:
        return pd.DataFrame(columns=cols)
    if not rows:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(rows, columns=cols)
    out = (
        out.groupby(["district_name", "district_code", "name2"], as_index=False)
        .agg(lon=("lon", "mean"), lat=("lat", "mean"))
    )
    return out

@st.cache_data(show_spinner=False, ttl=43200, max_entries=2)
def fetch_texas_navigation_district_centroids() -> pd.DataFrame:
    cols = ["district_name", "district_code", "lon", "lat"]
    rows: list[dict] = []
    page_size = 1000
    offset = 0
    try:
        while True:
            payload = _arcgis_get_json(
                f"{TEXAS_NAVIGATION_DISTRICT_LAYER_URL}/query",
                params={
                    "where": "1=1",
                    "outFields": "OBJECTID,DISTRICT_N",
                    "returnGeometry": "false",
                    "returnCentroid": "true",
                    "outSR": "4326",
                    "orderByFields": "OBJECTID ASC",
                    "resultRecordCount": page_size,
                    "resultOffset": offset,
                    "f": "json",
                },
            )
            features = payload.get("features", [])
            if not features:
                break
            for feat in features:
                attrs = feat.get("attributes", {}) or {}
                centroid = feat.get("centroid", {}) or {}
                name = str(attrs.get("DISTRICT_N", "")).strip()
                if not name:
                    continue
                try:
                    lon = float(centroid.get("x"))
                    lat = float(centroid.get("y"))
                except (TypeError, ValueError):
                    continue
                rows.append(
                    {
                        "district_name": name,
                        "district_code": str(attrs.get("OBJECTID", "")).strip(),
                        "lon": lon,
                        "lat": lat,
                    }
                )
            if len(features) < page_size:
                break
            offset += len(features)
    except Exception:
        return pd.DataFrame(columns=cols)
    if not rows:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(rows, columns=cols)
    out = (
        out.groupby(["district_name"], as_index=False)
        .agg(district_code=("district_code", "min"), lon=("lon", "mean"), lat=("lat", "mean"))
    )
    return out

@st.cache_data(show_spinner=False, ttl=43200, max_entries=2)
def fetch_nctcog_transit_provider_centroids() -> pd.DataFrame:
    cols = ["provider_name", "classification", "district_code", "lon", "lat"]
    rows: list[dict] = []
    page_size = 1000
    offset = 0
    try:
        while True:
            payload = _arcgis_get_json(
                f"{NCTCOG_TRANSIT_PROVIDERS_LAYER_URL}/query",
                params={
                    "where": "1=1",
                    "outFields": "OBJECTID,Name,Classification",
                    "returnGeometry": "true",
                    "outSR": "4326",
                    "orderByFields": "OBJECTID ASC",
                    "resultRecordCount": page_size,
                    "resultOffset": offset,
                    "f": "json",
                },
            )
            features = payload.get("features", [])
            if not features:
                break
            for feat in features:
                attrs = feat.get("attributes", {}) or {}
                geometry = feat.get("geometry", {}) or {}
                name = str(attrs.get("Name", "")).strip()
                if not name:
                    continue
                try:
                    rings = geometry.get("rings", []) or []
                    x_vals = [float(pt[0]) for ring in rings for pt in ring if isinstance(pt, list) and len(pt) >= 2]
                    y_vals = [float(pt[1]) for ring in rings for pt in ring if isinstance(pt, list) and len(pt) >= 2]
                    if not x_vals or not y_vals:
                        continue
                    lon = (min(x_vals) + max(x_vals)) / 2.0
                    lat = (min(y_vals) + max(y_vals)) / 2.0
                except (TypeError, ValueError, IndexError):
                    continue
                rows.append(
                    {
                        "provider_name": name,
                        "classification": str(attrs.get("Classification", "")).strip(),
                        "district_code": str(attrs.get("OBJECTID", "")).strip(),
                        "lon": lon,
                        "lat": lat,
                    }
                )
            if len(features) < page_size:
                break
            offset += len(features)
    except Exception:
        return pd.DataFrame(columns=cols)
    if not rows:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(rows, columns=cols)
    out = (
        out.groupby(["provider_name", "classification", "district_code"], as_index=False)
        .agg(lon=("lon", "mean"), lat=("lat", "mean"))
    )
    return out

@st.cache_data(show_spinner=False, ttl=43200, max_entries=2)
def fetch_txdot_seaport_centroids() -> pd.DataFrame:
    cols = ["port_name", "port_type", "port_code", "lon", "lat"]
    rows: list[dict] = []
    page_size = 1000
    offset = 0
    try:
        while True:
            payload = _arcgis_get_json(
                f"{TXDOT_SEAPORTS_LAYER_URL}/query",
                params={
                    "where": "1=1",
                    "outFields": "OBJECTID,PORT_NM,PORT_TYPE",
                    "returnGeometry": "true",
                    "outSR": "4326",
                    "orderByFields": "OBJECTID ASC",
                    "resultRecordCount": page_size,
                    "resultOffset": offset,
                    "f": "json",
                },
            )
            features = payload.get("features", [])
            if not features:
                break
            for feat in features:
                attrs = feat.get("attributes", {}) or {}
                geometry = feat.get("geometry", {}) or {}
                name = str(attrs.get("PORT_NM", "")).strip()
                if not name:
                    continue
                try:
                    lon = float(geometry.get("x"))
                    lat = float(geometry.get("y"))
                except (TypeError, ValueError):
                    continue
                rows.append(
                    {
                        "port_name": name,
                        "port_type": str(attrs.get("PORT_TYPE", "")).strip(),
                        "port_code": str(attrs.get("OBJECTID", "")).strip(),
                        "lon": lon,
                        "lat": lat,
                    }
                )
            if len(features) < page_size:
                break
            offset += len(features)
    except Exception:
        return pd.DataFrame(columns=cols)
    if not rows:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(rows, columns=cols)
    out = (
        out.groupby(["port_name", "port_type", "port_code"], as_index=False)
        .agg(lon=("lon", "mean"), lat=("lat", "mean"))
    )
    return out

@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def build_tfl_school_district_matches(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    cols = [
        "fid",
        "district_name",
        "district_code",
        "lon",
        "lat",
        "match_count",
        "match_clients",
        "match_clients_preview",
        "source_name",
        "source_url",
    ]
    if not tfl_client_names:
        return pd.DataFrame(columns=cols)
    districts = fetch_tea_school_district_centroids()
    if districts.empty:
        return pd.DataFrame(columns=cols)

    exact_index: dict[str, set[str]] = {}
    root_index: dict[str, set[str]] = {}
    unique_clients = sorted({str(name).strip() for name in tfl_client_names if str(name).strip()})
    for client in unique_clients:
        canon = _canonical_school_district_name(client)
        canon_key = norm_name(canon)
        if canon_key:
            exact_index.setdefault(canon_key, set()).add(client)
        if _looks_like_school_district_name(client):
            root_key = _school_district_root_key(client)
            if root_key:
                root_index.setdefault(root_key, set()).add(client)

    if not exact_index and not root_index:
        return pd.DataFrame(columns=cols)

    out_rows: list[dict] = []
    for row in districts.itertuples(index=False):
        candidates = [row.name20, row.name, row.name2]
        if row.name2:
            candidates.append(f"{row.name2} ISD")
            candidates.append(f"{row.name2} Independent School District")
        variant_keys = set()
        for candidate in candidates:
            canon = _canonical_school_district_name(candidate)
            key = norm_name(canon)
            if key:
                variant_keys.add(key)

        matched_clients: set[str] = set()
        for key in variant_keys:
            matched_clients |= exact_index.get(key, set())
        root_key = _school_district_root_key(row.name20 or row.name or row.name2)
        if root_key:
            matched_clients |= root_index.get(root_key, set())

        if not matched_clients:
            continue
        matched_sorted = sorted(matched_clients)
        preview = ", ".join(matched_sorted[:6])
        if len(matched_sorted) > 6:
            preview = f"{preview}, +{len(matched_sorted) - 6} more"
        out_rows.append(
            {
                "fid": int(row.fid),
                "district_name": row.name20 or row.name or row.name2 or "",
                "district_code": row.district_code or row.district_code_compact or "",
                "lon": float(row.lon),
                "lat": float(row.lat),
                "match_count": int(len(matched_sorted)),
                "match_clients": matched_sorted,
                "match_clients_preview": preview,
                "source_name": "TEA School District boundaries (FeatureServer/0)",
                "source_url": TEA_ARCGIS_SCHOOL_DISTRICT_LAYER_URL,
            }
        )

    if not out_rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(out_rows, columns=cols).sort_values(["match_count", "district_name"], ascending=[False, True])

@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def build_tfl_county_matches(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    cols = [
        "subdivision_type",
        "subdivision_name",
        "subdivision_code",
        "lon",
        "lat",
        "match_count",
        "match_clients",
        "match_clients_preview",
        "source_name",
        "source_url",
    ]
    if not tfl_client_names:
        return pd.DataFrame(columns=cols)
    counties = fetch_tea_county_centroids()
    if counties.empty:
        return pd.DataFrame(columns=cols)

    exact_index: dict[str, set[str]] = {}
    root_index: dict[str, set[str]] = {}
    unique_clients = sorted({str(name).strip() for name in tfl_client_names if str(name).strip()})
    for client in unique_clients:
        canon_key = norm_name(_canonical_county_name(client))
        if canon_key:
            exact_index.setdefault(canon_key, set()).add(client)
        if _looks_like_county_name(client):
            root_key = _county_root_key(client)
            if root_key:
                root_index.setdefault(root_key, set()).add(client)
    if not exact_index and not root_index:
        return pd.DataFrame(columns=cols)

    out_rows: list[dict] = []
    for row in counties.itertuples(index=False):
        candidates = [row.name, f"{row.name} County", f"County of {row.name}"]
        variant_keys = {norm_name(_canonical_county_name(c)) for c in candidates if c}
        variant_keys = {k for k in variant_keys if k}

        matched_clients: set[str] = set()
        for key in variant_keys:
            matched_clients |= exact_index.get(key, set())
        root_key = _county_root_key(f"{row.name} County")
        if root_key:
            matched_clients |= root_index.get(root_key, set())
        if not matched_clients:
            continue

        matched_sorted = sorted(matched_clients)
        out_rows.append(
            {
                "subdivision_type": "County",
                "subdivision_name": f"{row.name} County",
                "subdivision_code": row.fips,
                "lon": float(row.lon),
                "lat": float(row.lat),
                "match_count": int(len(matched_sorted)),
                "match_clients": matched_sorted,
                "match_clients_preview": _match_preview(matched_sorted),
                "source_name": "TEA County boundaries (FeatureServer/0)",
                "source_url": TEA_ARCGIS_COUNTY_LAYER_URL,
            }
        )
    if not out_rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(out_rows, columns=cols).sort_values(["match_count", "subdivision_name"], ascending=[False, True])

@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def build_tfl_city_matches(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    cols = [
        "subdivision_type",
        "subdivision_name",
        "subdivision_code",
        "lon",
        "lat",
        "match_count",
        "match_clients",
        "match_clients_preview",
        "source_name",
        "source_url",
    ]
    if not tfl_client_names:
        return pd.DataFrame(columns=cols)
    cities = fetch_texas_city_centroids()
    if cities.empty:
        return pd.DataFrame(columns=cols)

    exact_index: dict[str, set[str]] = {}
    root_index: dict[str, set[str]] = {}
    unique_clients = sorted({str(name).strip() for name in tfl_client_names if str(name).strip()})
    for client in unique_clients:
        canon_key = norm_name(_canonical_city_name(client))
        if canon_key:
            exact_index.setdefault(canon_key, set()).add(client)
        if _looks_like_city_name(client):
            root_key = _city_root_key(client)
            if root_key:
                root_index.setdefault(root_key, set()).add(client)
    if not exact_index and not root_index:
        return pd.DataFrame(columns=cols)

    out_rows: list[dict] = []
    for row in cities.itertuples(index=False):
        base = row.basename or re.sub(r"\s+(city|town|village)\s*$", "", row.name, flags=re.IGNORECASE).strip()
        if not base:
            continue
        display_name = base
        if not re.search(r"\b(CITY|TOWN|VILLAGE)\b$", display_name, flags=re.IGNORECASE):
            display_name = f"{display_name} City"
        candidates = [base, row.name, f"City of {base}", f"{base} City", f"Town of {base}", f"{base} Town"]
        variant_keys = {norm_name(_canonical_city_name(c)) for c in candidates if c}
        variant_keys = {k for k in variant_keys if k}

        matched_clients: set[str] = set()
        for key in variant_keys:
            matched_clients |= exact_index.get(key, set())
        root_key = _city_root_key(f"City of {base}")
        if root_key:
            matched_clients |= root_index.get(root_key, set())
        if not matched_clients:
            continue

        matched_sorted = sorted(matched_clients)
        out_rows.append(
            {
                "subdivision_type": "City",
                "subdivision_name": display_name,
                "subdivision_code": row.geoid,
                "lon": float(row.lon),
                "lat": float(row.lat),
                "match_count": int(len(matched_sorted)),
                "match_clients": matched_sorted,
                "match_clients_preview": _match_preview(matched_sorted),
                "source_name": "U.S. Census TIGERweb Texas Places (MapServer/25)",
                "source_url": CENSUS_ARCGIS_TEXAS_CITY_LAYER_URL,
            }
        )
    if not out_rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(out_rows, columns=cols).sort_values(["match_count", "subdivision_name"], ascending=[False, True])

@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def build_tfl_transit_authority_matches(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    cols = [
        "subdivision_type",
        "subdivision_name",
        "subdivision_code",
        "lon",
        "lat",
        "match_count",
        "match_clients",
        "match_clients_preview",
        "source_name",
        "source_url",
    ]
    if not tfl_client_names:
        return pd.DataFrame(columns=cols)
    providers = fetch_nctcog_transit_provider_centroids()
    if providers.empty:
        return pd.DataFrame(columns=cols)
    return _build_layer_subdivision_matches(
        tfl_client_names=tfl_client_names,
        layer_df=providers,
        subdivision_type="Transit Authority",
        layer_name_cols=["provider_name", "classification"],
        layer_code_cols=["district_code"],
        root_patterns=TRANSIT_AUTHORITY_ROOT_PATTERNS + [r"\bTRANSIT\b"],
        include_client_fn=lambda client: _looks_like_entity_type(client, "Transit Authority"),
        extra_candidate_builder=lambda row: [
            f"{str(getattr(row, 'provider_name', '')).strip()} Transit Authority",
            f"{str(getattr(row, 'provider_name', '')).strip()} Transportation Authority",
            f"{str(getattr(row, 'provider_name', '')).strip()} Transit",
            "Dallas Area Rapid Transit" if re.search(r"\bDART\b", str(getattr(row, "provider_name", "")), flags=re.IGNORECASE) else "",
        ],
        source_name="NCTCOG Transit Providers (MapServer/10)",
        source_url=NCTCOG_TRANSIT_PROVIDERS_LAYER_URL,
    )

@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def build_tfl_port_authority_matches(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    cols = [
        "subdivision_type",
        "subdivision_name",
        "subdivision_code",
        "lon",
        "lat",
        "match_count",
        "match_clients",
        "match_clients_preview",
        "source_name",
        "source_url",
    ]
    if not tfl_client_names:
        return pd.DataFrame(columns=cols)
    ports = fetch_txdot_seaport_centroids()
    if ports.empty:
        return pd.DataFrame(columns=cols)

    def _port_aliases(row) -> list[str]:
        raw = str(getattr(row, "port_name", "")).strip()
        base = re.sub(r"^\s*PORT\s+OF\s+", "", raw, flags=re.IGNORECASE).strip()
        aliases: list[str] = []
        if raw:
            aliases.extend(
                [
                    raw,
                    f"{raw} Port Authority",
                    f"{raw} Navigation District",
                ]
            )
        if base and base.lower() != raw.lower():
            aliases.extend(
                [
                    f"Port of {base}",
                    f"{base} Port Authority",
                    f"{base} Navigation District",
                    f"{base} Port",
                ]
            )
        if re.search(r"\bNAVIGATION\s+DISTRICT\b", raw, flags=re.IGNORECASE):
            nav_base = re.sub(r"\bNAVIGATION\s+DISTRICT\b", "", raw, flags=re.IGNORECASE).strip(" -")
            if nav_base:
                aliases.extend([f"Port of {nav_base}", f"{nav_base} Port Authority"])
        return [a for a in aliases if str(a).strip()]

    return _build_layer_subdivision_matches(
        tfl_client_names=tfl_client_names,
        layer_df=ports,
        subdivision_type="Port Authority",
        layer_name_cols=["port_name"],
        layer_code_cols=["port_code"],
        root_patterns=PORT_AUTHORITY_ROOT_PATTERNS,
        include_client_fn=lambda client: _looks_like_entity_type(client, "Port Authority"),
        extra_candidate_builder=_port_aliases,
        source_name="TxDOT Seaports (FeatureServer/0)",
        source_url=TXDOT_SEAPORTS_LAYER_URL,
    )

@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def build_tfl_name_anchored_special_matches(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    cols = [
        "subdivision_type",
        "subdivision_name",
        "subdivision_code",
        "lon",
        "lat",
        "match_count",
        "match_clients",
        "match_clients_preview",
        "source_name",
        "source_url",
    ]
    if not tfl_client_names:
        return pd.DataFrame(columns=cols)

    counties = fetch_tea_county_centroids()
    cities = fetch_texas_city_centroids()
    if counties.empty and cities.empty:
        return pd.DataFrame(columns=cols)

    county_lookup: dict[str, dict] = {}
    if not counties.empty:
        for row in counties.itertuples(index=False):
            county_name = str(getattr(row, "name", "")).strip()
            if not county_name:
                continue
            key = _county_root_key(f"{county_name} County")
            if not key or key in county_lookup:
                continue
            county_lookup[key] = {
                "code": str(getattr(row, "fips", "")).strip(),
                "lon": float(getattr(row, "lon", 0.0)),
                "lat": float(getattr(row, "lat", 0.0)),
                "source_name": "Name-anchored county centroid proxy",
                "source_url": TEA_ARCGIS_COUNTY_LAYER_URL,
            }

    city_lookup: dict[str, dict] = {}
    if not cities.empty:
        for row in cities.itertuples(index=False):
            raw_name = str(getattr(row, "name", "")).strip()
            base = str(getattr(row, "basename", "")).strip() or re.sub(
                r"\s+(city|town|village)\s*$", "", raw_name, flags=re.IGNORECASE
            ).strip()
            if not base:
                continue
            display_name = base
            if not re.search(r"\b(CITY|TOWN|VILLAGE)\b$", display_name, flags=re.IGNORECASE):
                display_name = f"{display_name} City"
            key = _city_root_key(display_name)
            if not key or key in city_lookup:
                continue
            city_lookup[key] = {
                "code": str(getattr(row, "geoid", "")).strip(),
                "lon": float(getattr(row, "lon", 0.0)),
                "lat": float(getattr(row, "lat", 0.0)),
                "source_name": "Name-anchored city centroid proxy",
                "source_url": CENSUS_ARCGIS_TEXAS_CITY_LAYER_URL,
            }

    rows: list[dict] = []
    county_lookup_keys = tuple(county_lookup.keys())
    city_lookup_keys = tuple(city_lookup.keys())
    for client in sorted({str(name).strip() for name in tfl_client_names if str(name).strip()}):
        entity_type = classify_requested_entity_type(client)
        if entity_type not in SPECIAL_NAME_ANCHORED_ENTITY_TYPES:
            continue

        anchor = None
        anchor_keys = _resolve_special_anchor_keys(
            client_name=client,
            entity_type=entity_type,
            county_lookup_keys=county_lookup_keys,
            city_lookup_keys=city_lookup_keys,
        )
        county_key = str(anchor_keys.get("county_key", "")).strip()
        city_key = str(anchor_keys.get("city_key", "")).strip()
        preferred_scope = str(anchor_keys.get("preferred_scope", "")).strip()

        if preferred_scope == "county" and county_key in county_lookup:
            anchor = county_lookup[county_key]
        elif preferred_scope == "city" and city_key in city_lookup:
            anchor = city_lookup[city_key]
        elif county_key in county_lookup:
            anchor = county_lookup[county_key]
        elif city_key in city_lookup:
            anchor = city_lookup[city_key]

        if not anchor:
            geocoded = geocode_texas_entity_arcgis(client)
            score = float(geocoded.get("score", 0.0)) if geocoded else 0.0
            if geocoded and score >= 70:
                anchor = {
                    "code": str(geocoded.get("postal", "")).strip(),
                    "lon": float(geocoded.get("lon", 0.0)),
                    "lat": float(geocoded.get("lat", 0.0)),
                    "source_name": "ArcGIS geocoded entity centroid (Texas)",
                    "source_url": ARCGIS_GEOCODER_URL,
                }

        if not anchor:
            continue

        rows.append(
            {
                "subdivision_type": entity_type,
                "subdivision_name": client,
                "subdivision_code": str(anchor.get("code", "")).strip(),
                "lon": float(anchor.get("lon", 0.0)),
                "lat": float(anchor.get("lat", 0.0)),
                "match_count": 1,
                "match_clients": [client],
                "match_clients_preview": client,
                "source_name": str(anchor.get("source_name", "")).strip(),
                "source_url": str(anchor.get("source_url", "")).strip(),
            }
        )

    if not rows:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(rows, columns=cols).drop_duplicates(
        ["subdivision_type", "subdivision_name", "subdivision_code"]
    )
    return out.sort_values(["subdivision_type", "subdivision_name"], ascending=[True, True])

@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def build_tfl_water_district_type_matches(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    cols = [
        "subdivision_type",
        "subdivision_name",
        "subdivision_code",
        "lon",
        "lat",
        "match_count",
        "match_clients",
        "match_clients_preview",
        "source_name",
        "source_url",
    ]
    if not tfl_client_names:
        return pd.DataFrame(columns=cols)
    water = fetch_tceq_water_district_centroids()
    if water.empty:
        return pd.DataFrame(columns=cols)

    water = water.copy()
    water["type_label"] = water["type_desc"].astype(str).map(_canonical_water_district_type)

    parts = []
    for subtype, root_patterns in WATER_DISTRICT_TYPE_ROOT_PATTERNS.items():
        if subtype == "Navigation District":
            # Use the dedicated statewide navigation-district layer for this type.
            continue
        subset = water[water["type_label"].astype(str) == subtype]
        if subset.empty:
            continue
        piece = _build_layer_subdivision_matches(
            tfl_client_names=tfl_client_names,
            layer_df=subset,
            subdivision_type=subtype,
            layer_name_cols=["district_name"],
            layer_code_cols=["district_code"],
            root_patterns=root_patterns,
            include_client_fn=lambda client, target=subtype: _looks_like_entity_type(client, target),
            extra_candidate_builder=None,
            source_name="TCEQ Water Districts (FeatureServer/0)",
            source_url=TCEQ_WATER_DISTRICTS_LAYER_URL,
        )
        if not piece.empty:
            parts.append(piece)
    if not parts:
        return pd.DataFrame(columns=cols)
    return pd.concat(parts, ignore_index=True).sort_values(["subdivision_type", "match_count", "subdivision_name"], ascending=[True, False, True])

@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def build_tfl_groundwater_district_matches(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    districts = fetch_tceq_groundwater_district_centroids()
    if districts.empty:
        return pd.DataFrame(
            columns=[
                "subdivision_type",
                "subdivision_name",
                "subdivision_code",
                "lon",
                "lat",
                "match_count",
                "match_clients",
                "match_clients_preview",
                "source_name",
                "source_url",
            ]
        )
    return _build_layer_subdivision_matches(
        tfl_client_names=tfl_client_names,
        layer_df=districts,
        subdivision_type="Groundwater Conservation District",
        layer_name_cols=["district_name"],
        layer_code_cols=["district_code"],
        root_patterns=[r"\bGROUNDWATER\s+CONSERVATION\s+DISTRICT\b", r"\bDISTRICT\b"],
        include_client_fn=lambda client: _looks_like_entity_type(client, "Groundwater Conservation District"),
        extra_candidate_builder=None,
        source_name="TCEQ Groundwater Conservation Districts (FeatureServer/0)",
        source_url=TCEQ_GROUNDWATER_DISTRICTS_LAYER_URL,
    )

@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def build_tfl_regional_mobility_authority_matches(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    districts = fetch_texas_rma_centroids()
    if districts.empty:
        return pd.DataFrame(
            columns=[
                "subdivision_type",
                "subdivision_name",
                "subdivision_code",
                "lon",
                "lat",
                "match_count",
                "match_clients",
                "match_clients_preview",
                "source_name",
                "source_url",
            ]
        )
    return _build_layer_subdivision_matches(
        tfl_client_names=tfl_client_names,
        layer_df=districts,
        subdivision_type="Regional Mobility Authority",
        layer_name_cols=["district_name"],
        layer_code_cols=["district_code"],
        root_patterns=[r"\bREGIONAL\s+MOBILITY\s+AUTHORITY\b", r"\bAUTHORITY\b", r"\bRMA\b"],
        include_client_fn=lambda client: _looks_like_entity_type(client, "Regional Mobility Authority"),
        extra_candidate_builder=lambda row: [
            str(getattr(row, "district_name", "")).replace("RMA", "Regional Mobility Authority").strip()
        ],
        source_name="Texas Regional Mobility Authorities (FeatureServer/0)",
        source_url=TEXAS_RMA_LAYER_URL,
    )

@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def build_tfl_junior_college_matches(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    districts = fetch_texas_junior_college_centroids()
    if districts.empty:
        return pd.DataFrame(
            columns=[
                "subdivision_type",
                "subdivision_name",
                "subdivision_code",
                "lon",
                "lat",
                "match_count",
                "match_clients",
                "match_clients_preview",
                "source_name",
                "source_url",
            ]
        )
    return _build_layer_subdivision_matches(
        tfl_client_names=tfl_client_names,
        layer_df=districts,
        subdivision_type="Junior College District",
        layer_name_cols=["district_name", "name2"],
        layer_code_cols=["district_code"],
        root_patterns=[r"\bCOMMUNITY\s+COLLEGE\b", r"\bJUNIOR\s+COLLEGE\b", r"\bCOLLEGE\s+DISTRICT\b", r"\bSERVICE\s+AREA\b", r"\bCOLLEGE\b", r"\bDISTRICT\b"],
        include_client_fn=lambda client: _looks_like_entity_type(client, "Junior College District"),
        extra_candidate_builder=lambda row: [
            f"{str(getattr(row, 'district_name', '')).strip()} District",
            f"{str(getattr(row, 'district_name', '')).strip()} Community College District",
        ],
        source_name="Texas Junior College Service Areas (FeatureServer/0)",
        source_url=TEXAS_JUNIOR_COLLEGE_LAYER_URL,
    )

@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def build_tfl_navigation_district_matches(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    districts = fetch_texas_navigation_district_centroids()
    if districts.empty:
        return pd.DataFrame(
            columns=[
                "subdivision_type",
                "subdivision_name",
                "subdivision_code",
                "lon",
                "lat",
                "match_count",
                "match_clients",
                "match_clients_preview",
                "source_name",
                "source_url",
            ]
        )

    def _nav_aliases(row) -> list[str]:
        raw = str(getattr(row, "district_name", "")).strip()
        if not raw:
            return []
        base = re.sub(r"\bNAVIGATION\s+DISTRICT\b", "", raw, flags=re.IGNORECASE).strip(" -")
        aliases = [
            raw,
            f"{raw} Port Authority",
            f"{base} Port Authority" if base else "",
            f"Port of {base}" if base else "",
        ]
        return [a for a in aliases if str(a).strip()]

    return _build_layer_subdivision_matches(
        tfl_client_names=tfl_client_names,
        layer_df=districts,
        subdivision_type="Navigation District",
        layer_name_cols=["district_name"],
        layer_code_cols=["district_code"],
        root_patterns=[r"\bNAVIGATION\s+DISTRICT\b", r"\bPORT\s+AUTHORITY\b", r"\bPORT\s+OF\b", r"\bAUTHORITY\b", r"\bDISTRICT\b"],
        include_client_fn=lambda client: _looks_like_entity_type(client, "Navigation District") or _looks_like_entity_type(client, "Port Authority"),
        extra_candidate_builder=_nav_aliases,
        source_name="Texas Navigation Districts (FeatureServer/29)",
        source_url=TEXAS_NAVIGATION_DISTRICT_LAYER_URL,
    )

def _merge_subdivision_match_rows(df: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "subdivision_type",
        "subdivision_name",
        "subdivision_code",
        "lon",
        "lat",
        "match_count",
        "match_clients",
        "match_clients_preview",
        "source_name",
        "source_url",
    ]
    if df.empty:
        return pd.DataFrame(columns=cols)
    merged: dict[tuple[str, str, str], dict] = {}
    for row in df.itertuples(index=False):
        t = str(getattr(row, "subdivision_type", "")).strip()
        n = str(getattr(row, "subdivision_name", "")).strip()
        c = str(getattr(row, "subdivision_code", "")).strip()
        if not t or not n:
            continue
        key = (t, n, c)
        clients = getattr(row, "match_clients", [])
        client_set = {str(x).strip() for x in clients if str(x).strip()} if isinstance(clients, list) else set()
        source_name = str(getattr(row, "source_name", "")).strip()
        source_url = str(getattr(row, "source_url", "")).strip()
        if key not in merged:
            merged[key] = {
                "subdivision_type": t,
                "subdivision_name": n,
                "subdivision_code": c,
                "lon": float(getattr(row, "lon", 0.0)),
                "lat": float(getattr(row, "lat", 0.0)),
                "match_clients": set(client_set),
                "source_names": {source_name} if source_name else set(),
                "source_urls": {source_url} if source_url else set(),
            }
        else:
            merged[key]["match_clients"].update(client_set)
            if source_name:
                merged[key]["source_names"].add(source_name)
            if source_url:
                merged[key]["source_urls"].add(source_url)

    out_rows = []
    for _, rec in merged.items():
        matched_sorted = sorted(rec["match_clients"])
        out_rows.append(
            {
                "subdivision_type": rec["subdivision_type"],
                "subdivision_name": rec["subdivision_name"],
                "subdivision_code": rec["subdivision_code"],
                "lon": rec["lon"],
                "lat": rec["lat"],
                "match_count": int(len(matched_sorted)),
                "match_clients": matched_sorted,
                "match_clients_preview": _match_preview(matched_sorted),
                "source_name": "; ".join(sorted(rec.get("source_names", set()))),
                "source_url": "; ".join(sorted(rec.get("source_urls", set()))),
            }
        )
    if not out_rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(out_rows, columns=cols).sort_values(["subdivision_type", "match_count", "subdivision_name"], ascending=[True, False, True])

@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def build_tfl_political_subdivision_matches(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    cols = [
        "subdivision_type",
        "subdivision_name",
        "subdivision_code",
        "lon",
        "lat",
        "match_count",
        "match_clients",
        "match_clients_preview",
        "source_name",
        "source_url",
    ]
    if not tfl_client_names:
        return pd.DataFrame(columns=cols)

    # PERFORMANCE: run all build_tfl_* functions concurrently so cold-start
    # ArcGIS fetch calls overlap instead of running sequentially.
    def _school():
        return build_tfl_school_district_matches(tfl_client_names).rename(
            columns={"district_name": "subdivision_name", "district_code": "subdivision_code"}
        ).assign(subdivision_type="School District")

    builders = [
        _school,
        lambda: build_tfl_county_matches(tfl_client_names),
        lambda: build_tfl_city_matches(tfl_client_names),
        lambda: build_tfl_junior_college_matches(tfl_client_names),
        lambda: build_tfl_groundwater_district_matches(tfl_client_names),
        lambda: build_tfl_water_district_type_matches(tfl_client_names),
        lambda: build_tfl_transit_authority_matches(tfl_client_names),
        lambda: build_tfl_port_authority_matches(tfl_client_names),
        lambda: build_tfl_regional_mobility_authority_matches(tfl_client_names),
        lambda: build_tfl_navigation_district_matches(tfl_client_names),
        lambda: build_tfl_name_anchored_special_matches(tfl_client_names),
    ]
    parts: list[pd.DataFrame] = []
    with ThreadPoolExecutor(max_workers=len(builders)) as pool:
        futures = {pool.submit(fn): fn for fn in builders}
        for future in as_completed(futures):
            try:
                result = future.result()
                if isinstance(result, pd.DataFrame) and not result.empty:
                    parts.append(result)
            except Exception:
                pass
    if not parts:
        return pd.DataFrame(columns=cols)
    out = pd.concat(parts, ignore_index=True)
    keep = [c for c in cols if c in out.columns]
    out = out[keep]
    return _merge_subdivision_match_rows(out)

@st.cache_data(show_spinner=False, ttl=86400, max_entries=256)
def geocode_address_arcgis(address: str) -> dict:
    q = str(address).strip()
    if not q:
        return {}
    try:
        payload = _arcgis_get_json(
            ARCGIS_GEOCODER_URL,
            params={
                "SingleLine": q,
                "outFields": "Match_addr,Addr_type,City,Region,RegionAbbr,Postal",
                "maxLocations": 1,
                "f": "json",
            },
            timeout=40,
        )
        candidates = payload.get("candidates", [])
        if not candidates:
            return {}
        best = candidates[0]
        location = best.get("location") or {}
        raw_x, raw_y = location.get("x"), location.get("y")
        if raw_x is None or raw_y is None:
            return {}
        lon, lat = float(raw_x), float(raw_y)
        attrs = best.get("attributes") or {}
        return {
            "input": q,
            "matched_address": str(best.get("address", "")).strip(),
            "score": float(best.get("score", 0.0)),
            "lon": lon,
            "lat": lat,
            "region": str(attrs.get("Region", "")).strip(),
            "region_abbr": str(attrs.get("RegionAbbr", "")).strip(),
            "city": str(attrs.get("City", "")).strip(),
            "postal": str(attrs.get("Postal", "")).strip(),
        }
    except Exception:
        # Don't cache network failures for 24 h — clear this entry so it retries.
        geocode_address_arcgis.clear()
        return {}

@st.cache_data(show_spinner=False, ttl=604800, max_entries=4096)
def geocode_texas_entity_arcgis(entity_name: str) -> dict:
    q = str(entity_name).strip()
    if not q:
        return {}
    candidates_to_try = [f"{q}, Texas", q]
    network_error = False
    for candidate_query in candidates_to_try:
        try:
            payload = _arcgis_get_json(
                ARCGIS_GEOCODER_URL,
                params={
                    "SingleLine": candidate_query,
                    "outFields": "Match_addr,Addr_type,City,Region,RegionAbbr,Postal",
                    "maxLocations": 1,
                    "searchExtent": "-106.65,25.84,-93.51,36.50",
                    "f": "json",
                },
                timeout=35,
            )
            candidates = payload.get("candidates", [])
            if not candidates:
                continue
            best = candidates[0]
            location = best.get("location") or {}
            raw_x, raw_y = location.get("x"), location.get("y")
            if raw_x is None or raw_y is None:
                continue
            lon, lat = float(raw_x), float(raw_y)
            attrs = best.get("attributes") or {}
            region_abbr = str(attrs.get("RegionAbbr", "")).strip().upper()
            if region_abbr and region_abbr != "TX":
                continue
            return {
                "input": q,
                "matched_address": str(best.get("address", "")).strip(),
                "score": float(best.get("score", 0.0)),
                "lon": lon,
                "lat": lat,
                "city": str(attrs.get("City", "")).strip(),
                "region_abbr": region_abbr,
                "postal": str(attrs.get("Postal", "")).strip(),
            }
        except Exception:
            network_error = True
            continue
    if network_error:
        # Don't persist transient failures for 7 days — clear cache entry.
        geocode_texas_entity_arcgis.clear()
    return {}

@st.cache_data(show_spinner=False, ttl=604800, max_entries=8192)
def query_texas_county_for_point(lon: float, lat: float) -> dict:
    try:
        payload = _arcgis_get_json(
            f"{TEA_ARCGIS_COUNTY_LAYER_URL}/query",
            params={
                "geometry": f"{lon},{lat}",
                "geometryType": "esriGeometryPoint",
                "inSR": "4326",
                "spatialRel": "esriSpatialRelIntersects",
                "outFields": "FENAME,FIPS",
                "returnGeometry": "false",
                "f": "json",
            },
        )
        features = payload.get("features", [])
        if not features:
            return {}
        attrs = features[0].get("attributes", {}) or {}
        county_name = str(attrs.get("FENAME", "")).strip()
        fips = str(attrs.get("FIPS", "")).strip()
        return {"county_name": county_name, "county_fips": fips}
    except Exception:
        return {}

@st.cache_data(show_spinner=False, ttl=86400, max_entries=512)
def query_texas_subdivisions_for_point(lon: float, lat: float) -> pd.DataFrame:
    """Query 10 ArcGIS layers in parallel for subdivisions overlapping a point."""
    cols = ["subdivision_type", "subdivision_name", "subdivision_code", "source_name", "source_url"]

    geo_point = f"{lon},{lat}"
    _base_params: dict = {
        "geometry": geo_point,
        "geometryType": "esriGeometryPoint",
        "inSR": "4326",
        "spatialRel": "esriSpatialRelIntersects",
        "returnGeometry": "false",
        "f": "json",
    }

    # -- Helper closures (one per layer) --------------------------
    def _fetch_school_districts() -> list[dict]:
        result: list[dict] = []
        try:
            payload = _arcgis_get_json(
                f"{TEA_ARCGIS_SCHOOL_DISTRICT_LAYER_URL}/query",
                params={**_base_params, "outFields": "NAME,NAME20,DISTRICT,DISTRICT_C"},
            )
            for feat in payload.get("features", []):
                attrs = feat.get("attributes", {}) or {}
                name = str(attrs.get("NAME20", "")).strip() or str(attrs.get("NAME", "")).strip()
                code = str(attrs.get("DISTRICT", "")).strip() or str(attrs.get("DISTRICT_C", "")).strip()
                if name:
                    result.append({"subdivision_type": "School District", "subdivision_name": name,
                                   "subdivision_code": code, "source_name": "TEA School District boundaries (FeatureServer/0)",
                                   "source_url": TEA_ARCGIS_SCHOOL_DISTRICT_LAYER_URL})
        except Exception:
            pass
        return result

    def _fetch_counties() -> list[dict]:
        result: list[dict] = []
        try:
            payload = _arcgis_get_json(
                f"{TEA_ARCGIS_COUNTY_LAYER_URL}/query",
                params={**_base_params, "outFields": "FENAME,FIPS"},
            )
            for feat in payload.get("features", []):
                attrs = feat.get("attributes", {}) or {}
                county_name = str(attrs.get("FENAME", "")).strip()
                if county_name:
                    result.append({"subdivision_type": "County", "subdivision_name": f"{county_name} County",
                                   "subdivision_code": str(attrs.get("FIPS", "")).strip(),
                                   "source_name": "TEA County boundaries (FeatureServer/0)",
                                   "source_url": TEA_ARCGIS_COUNTY_LAYER_URL})
        except Exception:
            pass
        return result

    def _fetch_cities() -> list[dict]:
        result: list[dict] = []
        try:
            payload = _arcgis_get_json(
                f"{CENSUS_ARCGIS_TEXAS_CITY_LAYER_URL}/query",
                params={**_base_params, "where": "STATE='48'", "outFields": "NAME,BASENAME,GEOID"},
            )
            for feat in payload.get("features", []):
                attrs = feat.get("attributes", {}) or {}
                name = str(attrs.get("NAME", "")).strip()
                base = str(attrs.get("BASENAME", "")).strip() or re.sub(
                    r"\s+(city|town|village)\s*$", "", name, flags=re.IGNORECASE
                ).strip()
                if base:
                    display = base
                    if not re.search(r"\b(CITY|TOWN|VILLAGE)\b$", display, flags=re.IGNORECASE):
                        display = f"{display} City"
                    result.append({"subdivision_type": "City", "subdivision_name": display,
                                   "subdivision_code": str(attrs.get("GEOID", "")).strip(),
                                   "source_name": "U.S. Census TIGERweb Texas Places (MapServer/25)",
                                   "source_url": CENSUS_ARCGIS_TEXAS_CITY_LAYER_URL})
        except Exception:
            pass
        return result

    def _fetch_water_districts() -> list[dict]:
        result: list[dict] = []
        try:
            payload = _arcgis_get_json(
                f"{TCEQ_WATER_DISTRICTS_LAYER_URL}/query",
                params={**_base_params, "outFields": "NAME,DISTRICT_ID,TYPE,TYPE_DESCRIPTION"},
            )
            for feat in payload.get("features", []):
                attrs = feat.get("attributes", {}) or {}
                mapped_type = _canonical_water_district_type(str(attrs.get("TYPE_DESCRIPTION", "")).strip())
                if mapped_type == "Navigation District":
                    mapped_type = ""
                if not mapped_type:
                    continue
                name = str(attrs.get("NAME", "")).strip()
                if name:
                    result.append({"subdivision_type": mapped_type, "subdivision_name": name,
                                   "subdivision_code": str(attrs.get("DISTRICT_ID", "")).strip(),
                                   "source_name": "TCEQ Water Districts (FeatureServer/0)",
                                   "source_url": TCEQ_WATER_DISTRICTS_LAYER_URL})
        except Exception:
            pass
        return result

    def _fetch_groundwater() -> list[dict]:
        result: list[dict] = []
        try:
            payload = _arcgis_get_json(
                f"{TCEQ_GROUNDWATER_DISTRICTS_LAYER_URL}/query",
                params={**_base_params, "outFields": "DISTNAME,DIST_NUM,SHORTNAM"},
            )
            for feat in payload.get("features", []):
                attrs = feat.get("attributes", {}) or {}
                name = str(attrs.get("DISTNAME", "")).strip() or str(attrs.get("SHORTNAM", "")).strip()
                if name:
                    result.append({"subdivision_type": "Groundwater Conservation District", "subdivision_name": name,
                                   "subdivision_code": str(attrs.get("DIST_NUM", "")).strip(),
                                   "source_name": "TCEQ Groundwater Conservation Districts (FeatureServer/0)",
                                   "source_url": TCEQ_GROUNDWATER_DISTRICTS_LAYER_URL})
        except Exception:
            pass
        return result

    def _fetch_rma() -> list[dict]:
        result: list[dict] = []
        try:
            payload = _arcgis_get_json(
                f"{TEXAS_RMA_LAYER_URL}/query",
                params={**_base_params, "outFields": "OBJECTID,RMA,Label"},
            )
            for feat in payload.get("features", []):
                attrs = feat.get("attributes", {}) or {}
                name = str(attrs.get("Label", "")).strip() or str(attrs.get("RMA", "")).strip()
                if name:
                    result.append({"subdivision_type": "Regional Mobility Authority", "subdivision_name": name,
                                   "subdivision_code": str(attrs.get("OBJECTID", "")).strip(),
                                   "source_name": "Texas Regional Mobility Authorities (FeatureServer/0)",
                                   "source_url": TEXAS_RMA_LAYER_URL})
        except Exception:
            pass
        return result

    def _fetch_junior_colleges() -> list[dict]:
        result: list[dict] = []
        try:
            payload = _arcgis_get_json(
                f"{TEXAS_JUNIOR_COLLEGE_LAYER_URL}/query",
                params={**_base_params, "outFields": "DISTRICT,NAME1,NAME2"},
            )
            for feat in payload.get("features", []):
                attrs = feat.get("attributes", {}) or {}
                name = str(attrs.get("NAME1", "")).strip() or str(attrs.get("NAME2", "")).strip()
                if name:
                    result.append({"subdivision_type": "Junior College District", "subdivision_name": name,
                                   "subdivision_code": str(attrs.get("DISTRICT", "")).strip(),
                                   "source_name": "Texas Junior College Service Areas (FeatureServer/0)",
                                   "source_url": TEXAS_JUNIOR_COLLEGE_LAYER_URL})
        except Exception:
            pass
        return result

    def _fetch_navigation() -> list[dict]:
        result: list[dict] = []
        try:
            payload = _arcgis_get_json(
                f"{TEXAS_NAVIGATION_DISTRICT_LAYER_URL}/query",
                params={**_base_params, "outFields": "OBJECTID,DISTRICT_N"},
            )
            for feat in payload.get("features", []):
                attrs = feat.get("attributes", {}) or {}
                name = str(attrs.get("DISTRICT_N", "")).strip()
                if name:
                    result.append({"subdivision_type": "Navigation District", "subdivision_name": name,
                                   "subdivision_code": str(attrs.get("OBJECTID", "")).strip(),
                                   "source_name": "Texas Navigation Districts (FeatureServer/29)",
                                   "source_url": TEXAS_NAVIGATION_DISTRICT_LAYER_URL})
        except Exception:
            pass
        return result

    def _fetch_transit() -> list[dict]:
        result: list[dict] = []
        try:
            payload = _arcgis_get_json(
                f"{NCTCOG_TRANSIT_PROVIDERS_LAYER_URL}/query",
                params={**_base_params, "outFields": "OBJECTID,Name,Classification"},
            )
            for feat in payload.get("features", []):
                attrs = feat.get("attributes", {}) or {}
                name = str(attrs.get("Name", "")).strip()
                if name:
                    result.append({"subdivision_type": "Transit Authority", "subdivision_name": name,
                                   "subdivision_code": str(attrs.get("OBJECTID", "")).strip(),
                                   "source_name": "NCTCOG Transit Providers (MapServer/10)",
                                   "source_url": NCTCOG_TRANSIT_PROVIDERS_LAYER_URL})
        except Exception:
            pass
        return result

    def _fetch_seaports() -> list[dict]:
        result: list[dict] = []
        try:
            payload = _arcgis_get_json(
                f"{TXDOT_SEAPORTS_LAYER_URL}/query",
                params={**_base_params, "distance": 25, "units": "esriSRUnit_StatuteMile",
                        "outFields": "OBJECTID,PORT_NM"},
            )
            for feat in payload.get("features", []):
                attrs = feat.get("attributes", {}) or {}
                name = str(attrs.get("PORT_NM", "")).strip()
                if name:
                    result.append({"subdivision_type": "Port Authority", "subdivision_name": name,
                                   "subdivision_code": str(attrs.get("OBJECTID", "")).strip(),
                                   "source_name": "TxDOT Seaports (FeatureServer/0)",
                                   "source_url": TXDOT_SEAPORTS_LAYER_URL})
        except Exception:
            pass
        return result

    # -- Fire all queries concurrently ----------------------------
    fetchers = [
        _fetch_school_districts, _fetch_counties, _fetch_cities,
        _fetch_water_districts, _fetch_groundwater, _fetch_rma,
        _fetch_junior_colleges, _fetch_navigation, _fetch_transit,
        _fetch_seaports,
    ]
    rows: list[dict] = []
    with ThreadPoolExecutor(max_workers=len(fetchers)) as pool:
        futures = {pool.submit(fn): fn.__name__ for fn in fetchers}
        for future in as_completed(futures):
            try:
                rows.extend(future.result())
            except Exception:
                pass

    if not rows:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(rows, columns=cols).drop_duplicates()
    return out.sort_values(["subdivision_type", "subdivision_name"], ascending=[True, True])

@functools.lru_cache(maxsize=4096)
def _subdivision_name_key(subdivision_type: str, subdivision_name: str) -> str:
    t = str(subdivision_type).strip().lower()
    if t == "school district":
        return _school_district_root_key(subdivision_name)
    if t == "county":
        return _county_root_key(subdivision_name)
    if t == "city":
        return _city_root_key(subdivision_name)
    for water_type, root_patterns in WATER_DISTRICT_TYPE_ROOT_PATTERNS.items():
        if t == water_type.lower():
            return _subdivision_root_from_patterns(subdivision_name, root_patterns)
    if t == "transit authority":
        return _subdivision_root_from_patterns(subdivision_name, TRANSIT_AUTHORITY_ROOT_PATTERNS)
    if t == "port authority":
        return _subdivision_root_from_patterns(subdivision_name, PORT_AUTHORITY_ROOT_PATTERNS)
    if t == "hospital district":
        return _subdivision_root_from_patterns(subdivision_name, [r"\bHOSPITAL\s+DISTRICT\b", r"\bDISTRICT\b"])
    if t == "emergency services district":
        return _subdivision_root_from_patterns(subdivision_name, [r"\bEMERGENCY\s+SERVICES\s+DISTRICT\b", r"\bDISTRICT\b", r"\bE\.?S\.?D\.?\b"])
    if t == "appraisal district":
        return _subdivision_root_from_patterns(subdivision_name, [r"\bAPPRAISAL\s+DISTRICT\b", r"\bDISTRICT\b", r"\bC\.?A\.?D\.?\b"])
    if t == "local government corporation":
        return _subdivision_root_from_patterns(
            subdivision_name,
            [r"\bLOCAL\s+GOVERNMENT\s+CORPORATION\b", r"\bDEVELOPMENT\s+CORPORATION\b", r"\bCORPORATION\b"],
        )
    if t == "groundwater conservation district":
        return _subdivision_root_from_patterns(subdivision_name, [r"\bGROUNDWATER\s+CONSERVATION\s+DISTRICT\b", r"\bDISTRICT\b"])
    if t == "regional mobility authority":
        return _subdivision_root_from_patterns(subdivision_name, [r"\bREGIONAL\s+MOBILITY\s+AUTHORITY\b", r"\bAUTHORITY\b", r"\bRMA\b"])
    if t == "junior college district":
        return _subdivision_root_from_patterns(
            subdivision_name,
            [r"\bCOMMUNITY\s+COLLEGE\b", r"\bJUNIOR\s+COLLEGE\b", r"\bCOLLEGE\s+DISTRICT\b", r"\bSERVICE\s+AREA\b", r"\bCOLLEGE\b", r"\bDISTRICT\b"],
        )
    return norm_name(subdivision_name)

def _subdivision_code_key(subdivision_code: str) -> str:
    return norm_name(str(subdivision_code).strip())

def _subdivision_numeric_code_key(subdivision_code: str) -> str:
    digits = re.sub(r"\D+", "", str(subdivision_code).strip())
    if not digits:
        return ""
    stripped = digits.lstrip("0")
    return stripped if stripped else digits

def _prepare_subdivision_match_pool(pool: pd.DataFrame, subdivision_type: str) -> pd.DataFrame:
    if pool.empty:
        return pool
    out = pool
    code_series = (
        out["subdivision_code"].astype(str)
        if "subdivision_code" in out.columns
        else pd.Series([""] * len(out), index=out.index, dtype=object).astype(str)
    )
    name_series = (
        out["subdivision_name"].astype(str)
        if "subdivision_name" in out.columns
        else pd.Series([""] * len(out), index=out.index, dtype=object).astype(str)
    )
    out["_code_key"] = code_series.map(_subdivision_code_key)
    out["_code_numeric_key"] = code_series.map(_subdivision_numeric_code_key)
    out["_name_key"] = name_series.map(
        lambda x: _subdivision_name_key(subdivision_type, x)
    )
    return out

def _pick_overlap_subdivision_matches(
    pool: pd.DataFrame,
    subdivision_type: str,
    subdivision_name: str,
    subdivision_code: str,
) -> tuple[pd.DataFrame, str]:
    if pool.empty:
        return pd.DataFrame(), ""

    code_key = _subdivision_code_key(subdivision_code)
    if code_key and "_code_key" in pool.columns:
        picked = pool[pool["_code_key"].astype(str) == code_key]
        if not picked.empty:
            return picked, "Spatial boundary (code)"

    numeric_code_key = _subdivision_numeric_code_key(subdivision_code)
    if numeric_code_key and "_code_numeric_key" in pool.columns:
        picked = pool[pool["_code_numeric_key"].astype(str) == numeric_code_key]
        if not picked.empty:
            return picked, "Spatial boundary (code)"

    name_key = _subdivision_name_key(subdivision_type, subdivision_name)
    if name_key and "_name_key" in pool.columns:
        picked = pool[pool["_name_key"].astype(str) == name_key]
        if not picked.empty:
            return picked, "Spatial boundary (name)"

        # Conservative fuzzy fallback for minor naming deltas between ArcGIS layers and matched records.
        if len(name_key) >= 6:
            name_pool = pool[pool["_name_key"].astype(str) != ""]
            if not name_pool.empty:
                name_pool["_name_score"] = name_pool["_name_key"].astype(str).map(
                    lambda x: difflib.SequenceMatcher(None, name_key, str(x)).ratio()
                )
                name_pool = name_pool[name_pool["_name_score"] >= 0.90]
                if not name_pool.empty:
                    best_score = float(name_pool["_name_score"].max())
                    picked = name_pool[name_pool["_name_score"] >= max(0.90, best_score - 0.03)]
                    if not picked.empty:
                        return picked.drop(columns=["_name_score"], errors="ignore"), "Spatial boundary (fuzzy)"

    return pd.DataFrame(), ""

def _match_confidence_from_method(match_method: str) -> str:
    m = str(match_method).strip().lower()
    if m in {"spatial boundary (code)", "spatial boundary (name)"}:
        return "High"
    if m == "spatial boundary (fuzzy)":
        return "Medium"
    if m in {"name anchored", "name + geocode context"}:
        return "Low"
    return "Unknown"

def build_address_overlap_spending_rows(
    overlap_subdivisions: pd.DataFrame,
    subdivision_matches: pd.DataFrame,
    tfl_spending: pd.DataFrame,
    *,
    prepared_overlap_pools: dict[str, pd.DataFrame] | None = None,
    spend_lookup: dict[str, dict[str, object]] | None = None,
) -> pd.DataFrame:
    cols = [
        "Subdivision Type",
        "Subdivision",
        "Code",
        "Entity Type",
        "TFL Entity",
        "Match Method",
        "Match Confidence",
        "Map Source",
        "Low",
        "High",
        "Mid",
        "Lobbyists",
    ]
    if overlap_subdivisions.empty or subdivision_matches.empty or tfl_spending.empty:
        return pd.DataFrame(columns=cols)

    spend_lookup_local = dict(spend_lookup or {})
    if not spend_lookup_local:
        spend = tfl_spending.copy()
        spend = ensure_cols(spend, {"Client": "", "Low": 0.0, "High": 0.0, "Lobbyists": 0})
        spend["Client"] = spend["Client"].fillna("").astype(str).str.strip()
        spend = spend[spend["Client"] != ""]
        if spend.empty:
            return pd.DataFrame(columns=cols)
        spend["Low"] = pd.to_numeric(spend["Low"], errors="coerce").fillna(0.0)
        spend["High"] = pd.to_numeric(spend["High"], errors="coerce").fillna(0.0)
        spend["Lobbyists"] = pd.to_numeric(spend["Lobbyists"], errors="coerce").fillna(0).astype(int)
        spend["EntityType"] = spend["Client"].map(classify_requested_entity_type)
        spend = (
            spend.groupby("Client", as_index=False)
            .agg(Low=("Low", "sum"), High=("High", "sum"), Lobbyists=("Lobbyists", "max"), EntityType=("EntityType", "first"))
        )
        spend_lookup_local = {
            str(r.Client): {
                "Low": float(r.Low),
                "High": float(r.High),
                "Lobbyists": int(r.Lobbyists),
                "EntityType": str(r.EntityType).strip(),
            }
            for r in spend.itertuples(index=False)
        }

    rows: list[dict] = []
    existing_keys: set[tuple[str, str, str, str]] = set()
    prepared_overlap_pools = prepared_overlap_pools or {}
    pool_cache: dict[str, pd.DataFrame] = {}
    for overlap in overlap_subdivisions.itertuples(index=False):
        t = str(overlap.subdivision_type).strip()
        n = str(overlap.subdivision_name).strip()
        c = str(overlap.subdivision_code).strip()
        if t in prepared_overlap_pools:
            pool = prepared_overlap_pools.get(t, pd.DataFrame())
        else:
            if t not in pool_cache:
                base_pool = subdivision_matches[subdivision_matches["subdivision_type"].astype(str) == t]
                pool_cache[t] = _prepare_subdivision_match_pool(base_pool, t)
            pool = pool_cache.get(t, pd.DataFrame())
        if pool.empty:
            continue

        picked, spatial_match_method = _pick_overlap_subdivision_matches(pool, t, n, c)
        if picked.empty:
            continue

        matched_clients: set[str] = set()
        picked_source_names = {
            str(v).strip()
            for v in picked.get("source_name", pd.Series(dtype=object)).dropna().astype(str).tolist()
            if str(v).strip()
        }
        picked_source = "; ".join(sorted(picked_source_names))
        for client_list in picked.get("match_clients", pd.Series(dtype=object)).tolist():
            if isinstance(client_list, list):
                matched_clients.update({str(x).strip() for x in client_list if str(x).strip()})

        for client in sorted(matched_clients):
            spend_vals = spend_lookup_local.get(client, {"Low": 0.0, "High": 0.0, "Lobbyists": 0, "EntityType": ""})
            low = float(spend_vals.get("Low", 0.0))
            high = float(spend_vals.get("High", 0.0))
            entity_type = str(spend_vals.get("EntityType", "")).strip()
            method = spatial_match_method or "Spatial boundary (name)"
            rows.append(
                {
                    "Subdivision Type": t,
                    "Subdivision": n,
                    "Code": c,
                    "Entity Type": entity_type,
                    "TFL Entity": client,
                    "Match Method": method,
                    "Match Confidence": _match_confidence_from_method(method),
                    "Map Source": picked_source,
                    "Low": low,
                    "High": high,
                    "Mid": (low + high) / 2,
                    "Lobbyists": int(spend_vals.get("Lobbyists", 0)),
                }
            )
            existing_keys.add((t, n, c, client))

    # Fallback for requested entity types without statewide polygon layers.
    unsupported_types = {
        "Hospital District",
        "Emergency Services District",
        "Local Government Corporation",
        "Transit Authority",
        "Port Authority",
        "Housing Authority",
        "Appraisal District",
    }
    county_lookup = {
        _county_root_key(str(r.subdivision_name)): (str(r.subdivision_name), str(r.subdivision_code))
        for r in overlap_subdivisions.itertuples(index=False)
        if str(r.subdivision_type).strip() == "County" and _county_root_key(str(r.subdivision_name))
    }
    city_lookup = {
        _city_root_key(str(r.subdivision_name)): (str(r.subdivision_name), str(r.subdivision_code))
        for r in overlap_subdivisions.itertuples(index=False)
        if str(r.subdivision_type).strip() == "City" and _city_root_key(str(r.subdivision_name))
    }
    school_lookup = {
        _school_district_root_key(str(r.subdivision_name)): (str(r.subdivision_name), str(r.subdivision_code))
        for r in overlap_subdivisions.itertuples(index=False)
        if str(r.subdivision_type).strip() == "School District" and _school_district_root_key(str(r.subdivision_name))
    }
    county_lookup_keys = tuple(k for k in county_lookup.keys() if k)
    city_lookup_keys = tuple(k for k in city_lookup.keys() if k)

    for client, spend_vals in spend_lookup_local.items():
        entity_type = str(spend_vals.get("EntityType", "")).strip()
        if entity_type not in unsupported_types:
            continue
        low = float(spend_vals.get("Low", 0.0))
        high = float(spend_vals.get("High", 0.0))
        lob = int(spend_vals.get("Lobbyists", 0))

        matched_targets: list[tuple[str, str, str, str, str]] = []
        anchor_keys = _resolve_special_anchor_keys(
            client_name=client,
            entity_type=entity_type,
            county_lookup_keys=county_lookup_keys,
            city_lookup_keys=city_lookup_keys,
        )
        county_key = str(anchor_keys.get("county_key", "")).strip()
        if county_key and county_key in county_lookup:
            n, c = county_lookup[county_key]
            matched_targets.append(("County", n, c, "Name anchored", "Name anchored via overlapping core boundaries"))

        city_key = str(anchor_keys.get("city_key", "")).strip()
        if city_key and city_key in city_lookup:
            n, c = city_lookup[city_key]
            matched_targets.append(("City", n, c, "Name anchored", "Name anchored via overlapping core boundaries"))

        school_key = _school_district_root_key(client)
        if school_key and school_key in school_lookup:
            n, c = school_lookup[school_key]
            matched_targets.append(("School District", n, c, "Name anchored", "Name anchored via overlapping core boundaries"))

        geocoded = geocode_texas_entity_arcgis(client)
        geocode_score = float(geocoded.get("score", 0.0)) if geocoded else 0.0
        if geocoded and geocode_score >= 70:
            try:
                geo_lon = float(geocoded.get("lon", 0.0))
                geo_lat = float(geocoded.get("lat", 0.0))
            except Exception:
                geo_lon = 0.0
                geo_lat = 0.0

            county_info = query_texas_county_for_point(round(geo_lon, 6), round(geo_lat, 6))
            geo_county = str(county_info.get("county_name", "")).strip()
            geo_county_key = _county_root_key(f"{geo_county} County") if geo_county else ""
            if geo_county_key and geo_county_key in county_lookup:
                n, c = county_lookup[geo_county_key]
                matched_targets.append(("County", n, c, "Name + geocode context", "ArcGIS geocoded entity centroid (Texas)"))

            geo_city = str(geocoded.get("city", "")).strip()
            geo_city_key = _city_root_key(f"{geo_city} City") if geo_city else ""
            if geo_city_key and geo_city_key in city_lookup:
                n, c = city_lookup[geo_city_key]
                matched_targets.append(("City", n, c, "Name + geocode context", "ArcGIS geocoded entity centroid (Texas)"))

        for t, n, c, match_method, map_source in matched_targets:
            row_key = (t, n, c, client)
            if row_key in existing_keys:
                continue
            rows.append(
                {
                    "Subdivision Type": t,
                    "Subdivision": n,
                    "Code": c,
                    "Entity Type": entity_type,
                    "TFL Entity": client,
                    "Match Method": match_method,
                    "Match Confidence": _match_confidence_from_method(match_method),
                    "Map Source": map_source,
                    "Low": low,
                    "High": high,
                    "Mid": (low + high) / 2,
                    "Lobbyists": lob,
                }
            )
            existing_keys.add(row_key)

    if not rows:
        return pd.DataFrame(columns=cols)

    out = pd.DataFrame(rows, columns=cols)
    out["_method_order"] = out["Match Method"].map(
        {
            "Spatial boundary (code)": 0,
            "Spatial boundary (name)": 1,
            "Spatial boundary (fuzzy)": 2,
            "Name anchored": 3,
            "Name + geocode context": 4,
        }
    ).fillna(9)
    out = out.sort_values(
        ["_method_order", "Mid", "High", "Low", "Subdivision Type", "Subdivision", "TFL Entity"],
        ascending=[True, False, False, False, True, True, True],
    )
    out = out.drop_duplicates(["Subdivision Type", "Subdivision", "Code", "TFL Entity"], keep="first")
    out = out.drop(columns=["_method_order"], errors="ignore")
    return out

def _subdivision_color_hex(subdivision_type: str) -> str:
    key = str(subdivision_type).strip()
    return SUBDIVISION_TYPE_COLORS.get(key, "#718191")

def _hex_to_rgba(color_hex: str, alpha: float = 0.88) -> list[float]:
    color = str(color_hex).strip().lstrip("#")
    if len(color) != 6 or not re.match(r"^[0-9a-fA-F]{6}$", color):
        return [113, 129, 145, alpha]
    return [int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16), alpha]

def render_subdivision_map_legend(type_counts: dict[str, int]) -> None:
    items = []
    for subtype, count in sorted(type_counts.items(), key=lambda x: (-int(x[1]), str(x[0]))):
        safe_type = html.escape(str(subtype), quote=True)
        safe_count = f"{int(count):,}"
        color = _subdivision_color_hex(str(subtype))
        items.append(
            f"""
<div class="map-legend-item">
  <div class="map-legend-left">
    <span class="map-legend-chip" style="background:{color};"></span>
    <span>{safe_type}</span>
  </div>
  <strong>{safe_count}</strong>
</div>
"""
        )
    if items:
        st.markdown(f'<div class="map-legend">{"".join(items)}</div>', unsafe_allow_html=True)

def build_overlap_map_points(
    overlap_subdivisions: pd.DataFrame,
    subdivision_matches: pd.DataFrame,
    *,
    prepared_overlap_pools: dict[str, pd.DataFrame] | None = None,
) -> pd.DataFrame:
    cols = [
        "subdivision_type",
        "subdivision_name",
        "subdivision_code",
        "lon",
        "lat",
        "match_count",
        "high_total",
        "match_method",
        "source_name",
    ]
    if overlap_subdivisions.empty or subdivision_matches.empty:
        return pd.DataFrame(columns=cols)

    rows: list[dict] = []
    prepared_overlap_pools = prepared_overlap_pools or {}
    pool_cache: dict[str, pd.DataFrame] = {}
    seen_keys: set[tuple[str, str, str]] = set()

    for overlap in overlap_subdivisions.itertuples(index=False):
        subdivision_type = str(overlap.subdivision_type).strip()
        subdivision_name = str(overlap.subdivision_name).strip()
        subdivision_code = str(overlap.subdivision_code).strip()
        if not subdivision_type or not subdivision_name:
            continue

        if subdivision_type in prepared_overlap_pools:
            pool = prepared_overlap_pools.get(subdivision_type, pd.DataFrame())
        else:
            if subdivision_type not in pool_cache:
                base_pool = subdivision_matches[
                    subdivision_matches["subdivision_type"].astype(str) == subdivision_type
                ]
                pool_cache[subdivision_type] = _prepare_subdivision_match_pool(base_pool, subdivision_type)
            pool = pool_cache.get(subdivision_type, pd.DataFrame())
        if pool.empty:
            continue

        picked, match_method = _pick_overlap_subdivision_matches(
            pool,
            subdivision_type,
            subdivision_name,
            subdivision_code,
        )
        if picked.empty:
            continue

        picked["match_count"] = pd.to_numeric(picked.get("match_count", 0), errors="coerce").fillna(0)
        picked = picked.sort_values(["match_count"], ascending=[False])
        best = picked.iloc[0]

        key = (subdivision_type, subdivision_name, subdivision_code)
        if key in seen_keys:
            continue
        seen_keys.add(key)

        rows.append(
            {
                "subdivision_type": subdivision_type,
                "subdivision_name": subdivision_name,
                "subdivision_code": subdivision_code,
                "lon": float(best.get("lon", 0.0)),
                "lat": float(best.get("lat", 0.0)),
                "match_count": int(float(best.get("match_count", 0.0))),
                "high_total": float(best.get("high_total", 0.0)),
                "match_method": match_method or "Spatial boundary (name)",
                "source_name": str(best.get("source_name", "")).strip(),
            }
        )

    # Include name-anchored special-type points when their inferred county/city anchor
    # is present in the overlapping core boundaries for this address.
    overlap_county_keys = {
        _county_root_key(str(r.subdivision_name))
        for r in overlap_subdivisions.itertuples(index=False)
        if str(r.subdivision_type).strip() == "County" and _county_root_key(str(r.subdivision_name))
    }
    overlap_city_keys = {
        _city_root_key(str(r.subdivision_name))
        for r in overlap_subdivisions.itertuples(index=False)
        if str(r.subdivision_type).strip() == "City" and _city_root_key(str(r.subdivision_name))
    }
    if overlap_county_keys or overlap_city_keys:
        special_types = set(SPECIAL_NAME_ANCHORED_ENTITY_TYPES) | {"Housing Authority"}
        special_matches = subdivision_matches[
            subdivision_matches["subdivision_type"].astype(str).isin(special_types)
        ]
        county_lookup_keys = tuple(sorted(overlap_county_keys))
        city_lookup_keys = tuple(sorted(overlap_city_keys))
        for row in special_matches.itertuples(index=False):
            subdivision_type = str(getattr(row, "subdivision_type", "")).strip()
            subdivision_name = str(getattr(row, "subdivision_name", "")).strip()
            subdivision_code = str(getattr(row, "subdivision_code", "")).strip()
            if not subdivision_type or not subdivision_name:
                continue

            clients = getattr(row, "match_clients", [])
            client_list = clients if isinstance(clients, list) else []
            if not client_list:
                client_list = [subdivision_name]

            include_point = False
            for client_name in client_list:
                anchor_keys = _resolve_special_anchor_keys(
                    client_name=str(client_name),
                    entity_type=subdivision_type,
                    county_lookup_keys=county_lookup_keys,
                    city_lookup_keys=city_lookup_keys,
                )
                county_key = str(anchor_keys.get("county_key", "")).strip()
                city_key = str(anchor_keys.get("city_key", "")).strip()
                if (county_key and county_key in overlap_county_keys) or (city_key and city_key in overlap_city_keys):
                    include_point = True
                    break
            if not include_point:
                continue

            key = (subdivision_type, subdivision_name, subdivision_code)
            if key in seen_keys:
                continue
            seen_keys.add(key)

            rows.append(
                {
                    "subdivision_type": subdivision_type,
                    "subdivision_name": subdivision_name,
                    "subdivision_code": subdivision_code,
                    "lon": float(getattr(row, "lon", 0.0)),
                    "lat": float(getattr(row, "lat", 0.0)),
                    "match_count": int(getattr(row, "match_count", 0) or 0),
                    "high_total": float(getattr(row, "high_total", 0.0) or 0.0),
                    "match_method": "Name anchored",
                    "source_name": str(getattr(row, "source_name", "")).strip(),
                }
            )

    if not rows:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(rows, columns=cols).drop_duplicates(
        ["subdivision_type", "subdivision_name", "subdivision_code"]
    )
    return out.sort_values(["subdivision_type", "subdivision_name"], ascending=[True, True])

    if not rows:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(rows, columns=cols).drop_duplicates(
        ["subdivision_type", "subdivision_name", "subdivision_code"]
    )
    return out.sort_values(["subdivision_type", "subdivision_name"], ascending=[True, True])


# =================================================================
# DRAW AREA & SEARCH ADDRESS MAP COMPONENT
# =================================================================
def render_draw_area_search_map(
    height: int = 520,
    basemap: str = "gray-vector",
    map_id: str = "tfl-draw-area-map",
    markers: list[dict] | None = None,
) -> None:
    """Render an ArcGIS JS 4.30 map with drawing tools, address search,
    click-to-reverse-geocode, and an address collector panel.

    Users can:
    * Click anywhere on the map to reverse-geocode that point
    * Use the Search widget to find an address
    * Draw polygon / circle / rectangle areas; all click-points within
      the area are collected
    * Copy discovered addresses from the in-map panel

    The map posts messages (``tfl-draw-address-found``, ``tfl-draw-area-addresses``)
    via ``window.parent.postMessage`` so that the Python layer can listen
    for them (via adjacent Streamlit widgets that manually capture the data).

    Parameters
    ----------
    height : int
        Map container height in pixels.
    basemap : str
        ArcGIS basemap identifier (``"gray-vector"``, ``"streets-vector"``, ``"hybrid"``).
    map_id : str
        HTML element id for the map container (must be unique per page render).
    markers : list[dict] | None
        Optional list of ``{"lat": float, "lon": float, "label": str}`` dicts
        to render pre-existing pins on the map (e.g. docket entity addresses).
    """
    basemap_safe = json.dumps(str(basemap).strip() or "gray-vector")
    markers_json = json.dumps(markers or [], ensure_ascii=True)
    draw_map_signature = _stable_json_signature(
        {
            "map_id": str(map_id).strip(),
            "markers": markers or [],
            "basemap": str(basemap).strip() or "gray-vector",
            "height": int(height),
        }
    )
    draw_map_cache_key = re.sub(r"[^0-9A-Za-z_]+", "_", str(map_id).strip()) or "tfl_draw_area_map"
    arcgis_html = _session_cached_value(
        f"_mp5_draw_map_html_{draw_map_cache_key}_v1",
        draw_map_signature,
        lambda: f"""
<link rel="stylesheet" href="https://js.arcgis.com/4.30/esri/themes/dark/main.css"/>
<style>
  #{map_id} {{ width:100%; height:{height}px; border-radius:14px; overflow:hidden; position:relative; }}

  /* Dark popup */
  .esri-popup__main-container {{
    background:rgba(13,23,36,0.96) !important; color:rgba(220,230,240,0.95) !important;
    border:1px solid rgba(100,140,180,0.22) !important; border-radius:10px !important;
    backdrop-filter:blur(10px) !important; box-shadow:0 8px 32px rgba(0,0,0,0.45) !important;
  }}
  .esri-popup__header-title {{ color:rgba(235,242,250,0.97) !important; font-weight:600 !important; }}
  .esri-popup__content {{ color:rgba(200,215,230,0.92) !important; }}
  .esri-popup__button {{ color:rgba(180,200,220,0.85) !important; }}
  .esri-popup__button:hover {{ color:#fff !important; background:rgba(100,180,255,0.18) !important; }}
  .esri-sketch {{ background:rgba(13,23,36,0.92) !important; border-radius:8px !important; border:1px solid rgba(100,140,180,0.22) !important; }}

  /* Coordinate bar */
  #tfl-draw-coord {{
    position:absolute; bottom:6px; left:50%; transform:translateX(-50%); z-index:90;
    background:rgba(10,20,32,0.88); border:1px solid rgba(100,140,180,0.15);
    border-radius:8px; padding:3px 10px; font-family:'Avenir Next LT Pro',system-ui,sans-serif;
    font-size:10.5px; color:rgba(180,200,220,0.75); white-space:nowrap;
    backdrop-filter:blur(6px); pointer-events:none;
  }}

  /* Address collector panel */
  #tfl-draw-collector {{
    position:absolute; top:12px; right:12px; z-index:95;
    background:rgba(10,20,32,0.94); border:1px solid rgba(30,144,255,0.22);
    border-radius:12px; padding:10px 12px; min-width:240px; max-width:300px;
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:11.5px;
    color:rgba(210,225,240,0.90); backdrop-filter:blur(10px);
    max-height:{height - 40}px; overflow-y:auto;
    box-shadow:0 8px 28px rgba(0,0,0,0.40);
  }}
  #tfl-draw-collector .dc-title {{
    text-transform:uppercase; letter-spacing:0.14em; font-size:8.5px;
    color:rgba(30,144,255,0.82); font-weight:700; margin-bottom:6px;
    display:flex; align-items:center; justify-content:space-between;
  }}
  #tfl-draw-collector .dc-item {{
    display:flex; align-items:flex-start; gap:6px; padding:5px 0;
    border-bottom:1px solid rgba(255,255,255,0.06);
  }}
  #tfl-draw-collector .dc-item:last-child {{ border-bottom:none; }}
  #tfl-draw-collector .dc-num {{
    flex-shrink:0; width:18px; height:18px; border-radius:50%;
    background:rgba(30,144,255,0.18); border:1px solid rgba(30,144,255,0.30);
    display:flex; align-items:center; justify-content:center;
    font-size:9px; font-weight:700; color:rgba(30,144,255,0.90);
  }}
  #tfl-draw-collector .dc-addr {{
    font-size:11px; line-height:1.35; color:rgba(210,230,245,0.85);
  }}
  #tfl-draw-collector .dc-coord {{
    font-size:9px; color:rgba(160,185,210,0.55); margin-top:1px;
  }}
  #tfl-draw-collector .dc-empty {{
    text-align:center; padding:10px 0; color:rgba(180,200,220,0.45); font-size:10.5px;
  }}
  #tfl-draw-collector .dc-actions {{
    display:flex; gap:6px; margin-top:6px;
  }}
  #tfl-draw-collector .dc-btn {{
    flex:1; padding:5px 8px; border-radius:8px; border:1px solid rgba(30,144,255,0.25);
    background:rgba(30,144,255,0.10); color:rgba(30,144,255,0.90); cursor:pointer;
    font-size:10px; font-weight:600; text-align:center; transition:all 0.2s;
  }}
  #tfl-draw-collector .dc-btn:hover {{ background:rgba(30,144,255,0.22); border-color:rgba(30,144,255,0.40); }}
  #tfl-draw-collector .dc-btn.clear {{ background:rgba(255,80,80,0.08); border-color:rgba(255,80,80,0.20); color:rgba(255,120,120,0.85); }}
  #tfl-draw-collector .dc-btn.clear:hover {{ background:rgba(255,80,80,0.18); }}
  #tfl-draw-badge {{
    position:absolute; top:12px; left:12px; z-index:95;
    background:rgba(10,20,32,0.90); border:1px solid rgba(0,224,184,0.22);
    border-radius:10px; padding:6px 12px; font-family:'Avenir Next LT Pro',system-ui,sans-serif;
    font-size:10.5px; color:rgba(0,224,184,0.85); backdrop-filter:blur(8px);
  }}

  /* Loading overlay */
  @keyframes tfl-draw-pulse {{ 0%,100%{{transform:scale(1);opacity:0.92;}} 50%{{transform:scale(1.35);opacity:0.45;}} }}
  #tfl-draw-loading {{
    position:absolute; top:0; left:0; width:100%; height:100%;
    background:rgba(10,16,26,0.92); display:flex; flex-direction:column;
    align-items:center; justify-content:center; z-index:100;
    transition:opacity 0.5s ease;
  }}
  #tfl-draw-loading .ld-dot {{
    width:10px; height:10px; border-radius:50%; background:rgba(30,144,255,0.82);
    animation:tfl-draw-pulse 1.2s ease-in-out infinite;
  }}
  #tfl-draw-loading .ld-text {{
    margin-top:8px; font-family:'Avenir Next LT Pro',system-ui,sans-serif;
    font-size:11px; color:rgba(180,200,220,0.65);
  }}

  /* Pin markers */
  .tfl-pin-marker {{
    width:10px; height:10px; border-radius:50%;
    background:rgba(0,224,184,0.85); border:2px solid rgba(255,255,255,0.70);
    box-shadow:0 2px 8px rgba(0,0,0,0.35);
  }}
</style>

<div id="{map_id}" style="position:relative;">
  <div id="tfl-draw-loading"><div class="ld-dot"></div><div class="ld-text">Initializing map\u2026</div></div>
  <div id="tfl-draw-coord">\u2014</div>
  <div id="tfl-draw-badge">Click map or search to collect addresses</div>
  <div id="tfl-draw-collector">
    <div class="dc-title"><span>&#x1F4CD; Collected Addresses</span><span id="tfl-draw-count">0</span></div>
    <div id="tfl-draw-list"><div class="dc-empty">Click on the map, use Search, or draw an area to collect addresses.</div></div>
    <div class="dc-actions">
      <div class="dc-btn" id="tfl-draw-copy-btn">Copy All</div>
      <div class="dc-btn clear" id="tfl-draw-clear-btn">Clear</div>
    </div>
  </div>
</div>

<script src="https://js.arcgis.com/4.30/"></script>
<script>
  require([
    "esri/Map", "esri/views/MapView", "esri/layers/GraphicsLayer",
    "esri/Graphic", "esri/widgets/Home", "esri/widgets/BasemapToggle",
    "esri/widgets/ScaleBar", "esri/widgets/Compass", "esri/widgets/Fullscreen",
    "esri/widgets/Locate", "esri/widgets/Search", "esri/widgets/Sketch",
    "esri/widgets/Expand", "esri/geometry/geometryEngine",
    "esri/layers/FeatureLayer", "esri/rest/locator"
  ], (Map, MapView, GraphicsLayer, Graphic, Home, BasemapToggle, ScaleBar,
      Compass, Fullscreen, Locate, Search, Sketch, Expand, geometryEngine,
      FeatureLayer, locator) => {{

    const collectedAddresses = [];
    const markersLayer = new GraphicsLayer();
    const sketchLayer = new GraphicsLayer();
    const pinsLayer = new GraphicsLayer();

    /* Reference layers */
    const countyLayer = new FeatureLayer({{
      url: "{TEA_ARCGIS_COUNTY_LAYER_URL}",
      opacity: 0.35, labelsVisible: false, popupEnabled: false,
      renderer: {{ type:"simple", symbol:{{ type:"simple-fill", color:[0,0,0,0], outline:{{ color:[180,200,220,0.3], width:0.6 }} }} }}
    }});
    const cityLayer = new FeatureLayer({{
      url: "{CENSUS_ARCGIS_TEXAS_CITY_LAYER_URL}",
      definitionExpression: "STATE='48'", opacity: 0.30, labelsVisible: false, visible: false, popupEnabled: false,
      renderer: {{ type:"simple", symbol:{{ type:"simple-fill", color:[0,0,0,0], outline:{{ color:[150,190,210,0.25], width:0.5 }} }} }}
    }});
    const districtLayer = new FeatureLayer({{
      url: "{TEA_ARCGIS_SCHOOL_DISTRICT_LAYER_URL}",
      opacity: 0.25, labelsVisible: false, visible: false, popupEnabled: false,
      renderer: {{ type:"simple", symbol:{{ type:"simple-fill", color:[0,0,0,0], outline:{{ color:[100,160,200,0.22], width:0.4 }} }} }}
    }});

    const map = new Map({{
      basemap: {basemap_safe},
      layers: [countyLayer, cityLayer, districtLayer, markersLayer, pinsLayer, sketchLayer]
    }});

    const view = new MapView({{
      container: "{map_id}",
      map, center: [-99.0, 31.2], zoom: 6,
      constraints: {{ minZoom: 4, maxZoom: 18 }},
      popup: {{ dockEnabled: true, dockOptions: {{ breakpoint: false, position: "bottom-left" }} }}
    }});

    /* Pre-existing markers */
    const preMarkers = {markers_json};
    preMarkers.forEach((m, i) => {{
      markersLayer.add(new Graphic({{
        geometry: {{ type: "point", longitude: m.lon, latitude: m.lat }},
        symbol: {{ type: "simple-marker", style: "diamond", size: 11, color: [0,224,184,0.85], outline: {{ color: [255,255,255,0.75], width: 1.5 }} }},
        attributes: {{ label: m.label || "", lat: m.lat, lon: m.lon }},
        popupTemplate: {{ title: m.label || "Marker " + (i+1), content: "Lat: " + m.lat.toFixed(5) + ", Lon: " + m.lon.toFixed(5) }}
      }}));
    }});

    /* Geocoder URL */
    const geocodeUrl = "{ARCGIS_GEOCODER_URL}";

    /* Helpers */
    function updateCollectorUI() {{
      const listEl = document.getElementById("tfl-draw-list");
      const countEl = document.getElementById("tfl-draw-count");
      if (countEl) countEl.textContent = collectedAddresses.length;
      if (!listEl) return;
      if (collectedAddresses.length === 0) {{
        listEl.innerHTML = '<div class="dc-empty">Click on the map, use Search, or draw an area to collect addresses.</div>';
        return;
      }}
      listEl.innerHTML = collectedAddresses.map((a, i) =>
        '<div class="dc-item">'
        + '<div class="dc-num">' + (i + 1) + '</div>'
        + '<div><div class="dc-addr">' + (a.address || "Unknown") + '</div>'
        + '<div class="dc-coord">' + Number(a.lat).toFixed(5) + '\\u00b0 N, ' + Math.abs(a.lon).toFixed(5) + '\\u00b0 W</div>'
        + '</div></div>'
      ).join("");
    }}

    function addAddress(address, lat, lon) {{
      const exists = collectedAddresses.some(a =>
        Math.abs(a.lat - lat) < 0.0001 && Math.abs(a.lon - lon) < 0.0001
      );
      if (exists) return;
      collectedAddresses.push({{ address, lat, lon }});

      /* Drop a pin */
      pinsLayer.add(new Graphic({{
        geometry: {{ type: "point", longitude: lon, latitude: lat }},
        symbol: {{ type: "simple-marker", style: "circle", size: 10, color: [30,144,255,0.85], outline: {{ color: [255,255,255,0.80], width: 1.5 }} }},
        attributes: {{ address, lat, lon }},
        popupTemplate: {{ title: address || "Point", content: "Lat: " + lat.toFixed(5) + ", Lon: " + lon.toFixed(5) }}
      }}));

      updateCollectorUI();

      /* Notify parent */
      try {{
        window.parent.postMessage({{
          type: "tfl-draw-address-found",
          address: address, lat: lat, lon: lon,
          allAddresses: collectedAddresses.slice()
        }}, "*");
      }} catch(e) {{}}
    }}

    function reverseGeocode(lat, lon) {{
      const url = geocodeUrl.replace("findAddressCandidates", "reverseGeocode")
        + "?location=" + lon + "," + lat
        + "&outSR=4326&langCode=en&f=json";
      fetch(url).then(r => r.json()).then(data => {{
        const addr = (data.address && data.address.LongLabel) || (data.address && data.address.ShortLabel) || ("Point: " + lat.toFixed(5) + ", " + lon.toFixed(5));
        addAddress(addr, lat, lon);
      }}).catch(() => {{
        addAddress("Point: " + lat.toFixed(5) + ", " + lon.toFixed(5), lat, lon);
      }});
    }}

    /* Click → reverse geocode */
    view.on("click", (evt) => {{
      if (evt.mapPoint) {{
        reverseGeocode(evt.mapPoint.latitude, evt.mapPoint.longitude);
      }}
    }});

    /* Coordinate readout */
    view.on("pointer-move", (evt) => {{
      const pt = view.toMap(evt);
      const el = document.getElementById("tfl-draw-coord");
      if (pt && el) el.textContent = pt.latitude.toFixed(5) + "\\u00b0 N, " + Math.abs(pt.longitude).toFixed(5) + "\\u00b0 W";
    }});

    /* Widgets */
    const home = new Home({{ view }});
    const basemapToggle = new BasemapToggle({{ view, nextBasemap: {basemap_safe} === "hybrid" ? "gray-vector" : "hybrid" }});
    const scaleBar = new ScaleBar({{ view, unit: "dual" }});
    const compass = new Compass({{ view }});
    const fullscreen = new Fullscreen({{ view }});
    const locate = new Locate({{ view }});

    const search = new Search({{
      view, popupEnabled: true, resultGraphicEnabled: true,
      goToOverride: (view, opts) => view.goTo(opts.target, {{ duration: 800, easing: "ease-in-out" }})
    }});
    search.on("select-result", (evt) => {{
      if (evt.result && evt.result.feature && evt.result.feature.geometry) {{
        const geom = evt.result.feature.geometry;
        addAddress(evt.result.name || "", geom.latitude, geom.longitude);
      }}
    }});

    const sketch = new Sketch({{
      view, layer: sketchLayer, creationMode: "single",
      availableCreateTools: ["polygon", "circle", "rectangle"],
      defaultCreateOptions: {{ mode: "freehand" }},
      visibleElements: {{ selectionTools: {{ "lasso-selection": false, "rectangle-selection": false }}, settingsMenu: false, undoRedoMenu: true }},
      defaultUpdateOptions: {{ tool: "reshape" }}
    }});
    const sketchExpand = new Expand({{
      view, content: sketch, expandIconClass: "esri-icon-polygon",
      expandTooltip: "Draw area to collect addresses", group: "tools"
    }});

    /* Layer toggle */
    const layerDiv = document.createElement("div");
    layerDiv.style.cssText = "background:rgba(13,23,36,0.94);border-radius:8px;padding:10px;font-family:'Avenir Next LT Pro',system-ui,sans-serif;font-size:12px;color:rgba(210,225,240,0.90);min-width:160px;";
    layerDiv.innerHTML = '<div style="font-size:9px;text-transform:uppercase;letter-spacing:0.12em;color:rgba(150,175,200,0.65);font-weight:700;margin-bottom:6px;">Reference Layers</div>'
      + '<label style="display:flex;align-items:center;gap:6px;cursor:pointer;padding:3px 0;"><input type="checkbox" id="tfl-draw-toggle-city" style="accent-color:#28b464;"><span>City Boundaries</span></label>'
      + '<label style="display:flex;align-items:center;gap:6px;cursor:pointer;padding:3px 0;"><input type="checkbox" id="tfl-draw-toggle-district" style="accent-color:#c88c3c;"><span>School Districts</span></label>';
    const layerExpand = new Expand({{ view, content: layerDiv, expandIconClass: "esri-icon-layer-list", expandTooltip: "Toggle reference layers", group: "tools" }});

    view.ui.add(home, "top-left");
    view.ui.add(compass, "top-left");
    view.ui.add(fullscreen, "top-left");
    view.ui.add(locate, "top-left");
    view.ui.add(sketchExpand, "top-left");
    view.ui.add(layerExpand, "top-left");
    view.ui.add(search, "top-right");
    view.ui.add(basemapToggle, "top-right");
    view.ui.add(scaleBar, "bottom-left");

    view.when(() => {{
      const cBox = document.getElementById("tfl-draw-toggle-city");
      const dBox = document.getElementById("tfl-draw-toggle-district");
      if (cBox) cBox.addEventListener("change", () => {{ cityLayer.visible = cBox.checked; }});
      if (dBox) dBox.addEventListener("change", () => {{ districtLayer.visible = dBox.checked; }});
    }});

    /* Sketch complete → reverse-geocode the centroid of the drawn area */
    sketch.on("create", (evt) => {{
      if (evt.state !== "complete") return;
      const geom = evt.graphic.geometry;
      const ext = geom.extent;
      if (!ext) return;

      /* Sample grid of points inside the drawn area for reverse geocoding */
      const cx = ext.center.longitude;
      const cy = ext.center.latitude;
      const dx = (ext.xmax - ext.xmin);
      const dy = (ext.ymax - ext.ymin);
      const SAMPLES = 5;
      const promises = [];

      /* Always geocode the centroid */
      reverseGeocode(cy, cx);

      /* Sample a NxN grid within the extent, but only inside the polygon */
      for (let xi = 0; xi < SAMPLES; xi++) {{
        for (let yi = 0; yi < SAMPLES; yi++) {{
          const px = ext.xmin + (dx * (xi + 0.5) / SAMPLES);
          const py = ext.ymin + (dy * (yi + 0.5) / SAMPLES);
          const testPt = {{ type: "point", longitude: px, latitude: py, spatialReference: {{ wkid: 4326 }} }};
          if (geometryEngine.contains(geom, testPt)) {{
            reverseGeocode(py, px);
          }}
        }}
      }}

      /* Update badge */
      const badge = document.getElementById("tfl-draw-badge");
      if (badge) badge.textContent = "Area scanned — see collected addresses \\u2192";

      /* Post area addresses */
      setTimeout(() => {{
        try {{
          window.parent.postMessage({{
            type: "tfl-draw-area-addresses",
            allAddresses: collectedAddresses.slice()
          }}, "*");
        }} catch(e) {{}}
      }}, 3000);
    }});

    /* Copy all button */
    document.getElementById("tfl-draw-copy-btn").addEventListener("click", () => {{
      if (collectedAddresses.length === 0) return;
      const text = collectedAddresses.map(a => a.address).join("\\n");
      navigator.clipboard.writeText(text).then(() => {{
        const btn = document.getElementById("tfl-draw-copy-btn");
        if (btn) {{ btn.textContent = "Copied!"; setTimeout(() => {{ btn.textContent = "Copy All"; }}, 2000); }}
      }}).catch(() => {{}});

      /* Also post to parent */
      try {{
        window.parent.postMessage({{
          type: "tfl-draw-copy-all",
          allAddresses: collectedAddresses.slice(),
          text: text
        }}, "*");
      }} catch(e) {{}}
    }});

    /* Clear button */
    document.getElementById("tfl-draw-clear-btn").addEventListener("click", () => {{
      collectedAddresses.length = 0;
      pinsLayer.removeAll();
      sketchLayer.removeAll();
      updateCollectorUI();
      const badge = document.getElementById("tfl-draw-badge");
      if (badge) badge.textContent = "Click map or search to collect addresses";
    }});

    /* Loading done */
    view.when(() => {{
      const loader = document.getElementById("tfl-draw-loading");
      if (loader) {{ loader.style.opacity = "0"; setTimeout(() => loader.remove(), 600); }}
      if (preMarkers.length > 0) {{
        view.goTo(markersLayer.graphics.toArray(), {{ padding: {{ top:50, right:50, bottom:50, left:50 }}, duration:1000, easing:"ease-in-out" }}).catch(() => {{}});
      }}
    }});
  }});
</script>
"""
    )
    _persistent_html_frame(
        html=arcgis_html,
        signature=draw_map_signature,
        height=int(height) + 8,
        key=f"mp5_draw_area_map_{draw_map_cache_key}_v1",
        default=None,
    )


def render_address_overlap_arcgis_map(
    lon: float,
    lat: float,
    matched_address: str,
    overlap_points: pd.DataFrame,
    height: int = 440,
    basemap: str = "gray-vector",
) -> None:
    try:
        lon_val = float(lon)
        lat_val = float(lat)
    except Exception:
        return

    # PERFORMANCE: Vectorized point_rows builder — avoids per-row itertuples + html.escape loop
    point_rows: list[dict] = []
    legend_types: dict[str, str] = {}
    if isinstance(overlap_points, pd.DataFrame) and not overlap_points.empty:
        _op = overlap_points.copy()
        for col in ("subdivision_type", "subdivision_name", "subdivision_code", "match_method", "source_name"):
            if col in _op.columns:
                _op[col] = _op[col].fillna("").astype(str).str.strip()
            else:
                _op[col] = ""
        _op["lon"] = pd.to_numeric(_op.get("lon", 0.0), errors="coerce").fillna(0.0)
        _op["lat"] = pd.to_numeric(_op.get("lat", 0.0), errors="coerce").fillna(0.0)
        _op["match_count"] = pd.to_numeric(_op.get("match_count", 0), errors="coerce").fillna(0).astype(int)
        _op["high_total"] = pd.to_numeric(_op.get("high_total", 0.0), errors="coerce").fillna(0.0)
        # Pre-compute colors and escape in bulk
        _color_cache: dict[str, list[float]] = {}
        for st_val in _op["subdivision_type"].unique():
            hex_c = _subdivision_color_hex(st_val)
            _color_cache[st_val] = _hex_to_rgba(hex_c)
            if st_val:
                legend_types[html.escape(st_val, quote=True)] = hex_c
        point_rows = [
            {
                "subdivision_type": html.escape(r.subdivision_type, quote=True),
                "subdivision_name": html.escape(r.subdivision_name, quote=True),
                "subdivision_code": html.escape(r.subdivision_code, quote=True),
                "lon": r.lon,
                "lat": r.lat,
                "match_count": r.match_count,
                "high_total": r.high_total,
                "match_method": html.escape(r.match_method, quote=True),
                "source_name": html.escape(r.source_name, quote=True),
                "color": _color_cache.get(r.subdivision_type, [113, 129, 145, 0.88]),
            }
            for r in _op.itertuples(index=False)
        ]
        del _op

    points_json = json.dumps(point_rows, ensure_ascii=True)
    address_json = json.dumps(
        {
            "lon": lon_val,
            "lat": lat_val,
            "matched_address": html.escape(str(matched_address).strip(), quote=True),
        },
        ensure_ascii=True,
    )
    basemap_safe = json.dumps(str(basemap).strip() or "gray-vector")

    legend_json = json.dumps(
        [{"type": t, "color": c} for t, c in legend_types.items()],
        ensure_ascii=True,
    )
    address_map_signature = _stable_json_signature(
        {
            "address": {
                "lon": lon_val,
                "lat": lat_val,
                "matched_address": str(matched_address).strip(),
            },
            "points": point_rows,
            "basemap": str(basemap).strip() or "gray-vector",
            "height": int(height),
        }
    )
    arcgis_html = _session_cached_value(
        "_mp5_address_overlap_html_v2",
        address_map_signature,
        lambda: f"""
<link rel="preload" href="https://js.arcgis.com/4.30/" as="script"/>
<link rel="stylesheet" href="https://js.arcgis.com/4.30/esri/themes/dark/main.css"/>
<style>
  /* -- Dark popup theme -- */
  .esri-popup__main-container {{
    background: rgba(13,23,36,0.96) !important;
    color: rgba(220,230,240,0.95) !important;
    border: 1px solid rgba(100,140,180,0.22) !important;
    border-radius: 10px !important;
    backdrop-filter: blur(10px) !important;
    box-shadow: 0 8px 32px rgba(0,0,0,0.45) !important;
  }}
  .esri-popup__header-title {{ color: rgba(235,242,250,0.97) !important; font-weight: 600 !important; }}
  .esri-popup__content {{ color: rgba(200,215,230,0.92) !important; }}
  .esri-popup__button {{ color: rgba(180,200,220,0.85) !important; }}
  .esri-popup__button:hover {{ color: #fff !important; background: rgba(100,180,255,0.18) !important; }}
  .esri-popup__pointer-direction {{ background: rgba(13,23,36,0.96) !important; }}

  /* -- Sketch toolbar styling -- */
  .esri-sketch {{ background: rgba(13,23,36,0.92) !important; border-radius: 8px !important; border: 1px solid rgba(100,140,180,0.22) !important; }}

  /* -- Legend — bottom-right, compact -- */
  #tfl-addr-legend {{
    position: absolute; bottom: 36px; right: 12px; z-index: 90;
    background: rgba(10,20,32,0.92); border: 1px solid rgba(100,140,180,0.18);
    border-radius: 10px; padding: 6px 10px; max-width: 210px;
    font-family: 'Avenir Next LT Pro', system-ui, sans-serif; font-size: 10.5px;
    color: rgba(210,225,240,0.90); backdrop-filter: blur(8px);
    max-height: 160px; overflow-y: auto;
  }}
  #tfl-addr-legend .leg-title {{
    text-transform: uppercase; letter-spacing: 0.14em; font-size: 8px;
    color: rgba(150,175,200,0.70); margin-bottom: 3px; font-weight: 700;
  }}
  #tfl-addr-legend .leg-row {{ display: flex; align-items: center; gap: 5px; padding: 1.5px 0; }}
  #tfl-addr-legend .leg-chip {{
    width: 9px; height: 9px; border-radius: 50%; flex-shrink: 0;
    border: 1px solid rgba(255,255,255,0.18);
  }}

  /* -- Loading overlay -- */
  @keyframes tfl-pulse {{ 0%,100% {{ transform:scale(1); opacity:0.92; }} 50% {{ transform:scale(1.35); opacity:0.45; }} }}
  #tfl-addr-loading {{
    position:absolute; top:0; left:0; width:100%; height:100%;
    background:rgba(10,16,26,0.92); display:flex; flex-direction:column;
    align-items:center; justify-content:center; gap:10px;
    z-index:100; border-radius:14px; transition:opacity 0.6s ease;
  }}
  #tfl-addr-loading .ld-spinner {{
    width:30px; height:30px; border:2.5px solid rgba(100,140,180,0.18);
    border-top:2.5px solid rgba(100,180,255,0.80); border-radius:50%;
    animation: tfl-ld-spin 0.8s linear infinite;
  }}
  #tfl-addr-loading .ld-label {{
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:10px;
    color:rgba(160,185,210,0.65); letter-spacing:0.06em;
  }}
  @keyframes tfl-ld-spin {{ 0%{{transform:rotate(0deg)}} 100%{{transform:rotate(360deg)}} }}

  /* -- Coordinate bar -- */
  #tfl-addr-coord {{
    position:absolute; bottom:10px; left:50%; transform:translateX(-50%); z-index:90;
    background:rgba(10,16,26,0.85); border:1px solid rgba(100,140,180,0.15);
    border-radius:6px; padding:2px 10px;
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:10px;
    color:rgba(180,200,220,0.75); pointer-events:none; backdrop-filter:blur(6px);
    white-space:nowrap; letter-spacing:0.04em;
  }}

  /* -- Lasso/selection feedback -- */
  #tfl-addr-sel-info {{
    position:absolute; bottom:36px; left:12px; z-index:90;
    background:rgba(10,16,26,0.92); border:1px solid rgba(0,180,255,0.25);
    border-radius:8px; padding:6px 12px; backdrop-filter:blur(6px);
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:11px;
    color:rgba(200,220,240,0.90); display:none; max-width:280px;
  }}
  #tfl-addr-sel-info .sel-title {{
    font-weight:700; color:rgba(100,200,255,0.95); font-size:10px;
    text-transform:uppercase; letter-spacing:0.08em; margin-bottom:3px;
  }}
  #tfl-addr-adv-btn {{
    position:absolute; top:12px; right:12px; z-index:95;
    border:1px solid rgba(100,180,255,0.32);
    background:rgba(10,20,32,0.92);
    color:rgba(190,220,245,0.92);
    border-radius:8px;
    padding:5px 10px;
    font-family:'Avenir Next LT Pro',system-ui,sans-serif;
    font-size:10px;
    letter-spacing:0.04em;
    cursor:pointer;
    backdrop-filter:blur(6px);
    transition:all .2s ease;
  }}
  #tfl-addr-adv-btn:hover {{
    border-color:rgba(100,180,255,0.52);
    background:rgba(16,34,52,0.95);
  }}
  #tfl-addr-adv-btn[disabled] {{
    opacity:0.7;
    cursor:default;
  }}
</style>
<div style="width:100%;height:{height}px;position:relative;">
  <div id="tfl-address-overlap-map" style="width:100%;height:100%;border-radius:14px;overflow:hidden;"></div>
  <div id="tfl-addr-legend"></div>
  <button id="tfl-addr-adv-btn" type="button">Load advanced tools</button>
  <div id="tfl-addr-loading"><div class="ld-spinner"></div><div class="ld-label">Loading map layers&hellip;</div></div>
  <div id="tfl-addr-coord">&ndash;</div>
  <div id="tfl-addr-sel-info"></div>
</div>
<script src="https://js.arcgis.com/4.30/"></script>
<script>
  const overlapPoints = {points_json};
  const addressPoint = {address_json};
  const baseMapId = {basemap_safe};
  const legendEntries = {legend_json};
  const advBtn = document.getElementById("tfl-addr-adv-btn");
  const scheduleAdvancedLoad = (fn) => {{
    if (typeof window.requestIdleCallback === "function") {{
      window.requestIdleCallback(fn, {{ timeout: 2600 }});
    }} else {{
      setTimeout(fn, 1500);
    }}
  }};

  /* Build on-map legend */
  (function() {{
    const el = document.getElementById("tfl-addr-legend");
    if (!el || legendEntries.length === 0) {{ if (el) el.style.display = "none"; return; }}
    let h = '<div class="leg-title">Subdivision Types</div>';
    for (const e of legendEntries) {{
      h += '<div class="leg-row"><span class="leg-chip" style="background:' + e.color + ';"></span><span>' + e.type + '</span></div>';
    }}
    h += '<div class="leg-row" style="margin-top:2px;"><span class="leg-chip" style="background:#c92234;transform:rotate(45deg);border-radius:2px;"></span><span>Queried Address</span></div>';
    el.innerHTML = h;
  }})();

  require([
    "esri/Map",
    "esri/views/MapView",
    "esri/layers/GraphicsLayer",
    "esri/Graphic",
    "esri/widgets/Home",
    "esri/widgets/ScaleBar",
    "esri/widgets/BasemapToggle"
  ], function(Map, MapView, GraphicsLayer, Graphic, Home, ScaleBar, BasemapToggle) {{
    const map = new Map({{ basemap: baseMapId }});

    const overlapLayer = new GraphicsLayer();
    const addressLayer = new GraphicsLayer();
    const sketchLayer = new GraphicsLayer();
    map.add(overlapLayer);
    map.add(addressLayer);
    map.add(sketchLayer);

    const view = new MapView({{
      container: "tfl-address-overlap-map",
      map,
      center: [addressPoint.lon, addressPoint.lat],
      zoom: 11,
      constraints: {{ minZoom: 5 }},
      popup: {{ dockEnabled: true, dockOptions: {{ position: "bottom-right", breakpoint: false }} }},
      ui: {{ padding: {{ top: 10, right: 10, bottom: 30, left: 10 }} }}
    }});

    const formatUsd = (v) => Number(v||0).toLocaleString("en-US",{{style:"currency",currency:"USD",maximumFractionDigits:0}});
    const maxHigh = overlapPoints.reduce((a,r) => Math.max(a, Number(r.high_total||0)), 0);
    const sz = (v) => {{ if(maxHigh<=0) return 10; const n=Math.max(0,Number(v||0)); return Math.max(9,Math.min(28,9+(Math.log10(n+1)/Math.log10(maxHigh+1))*19)); }};
    const badge = (m) => {{
      const l=(m||"").toLowerCase();
      if(l.includes("spatial")||l.includes("boundary")) return '<span style="display:inline-block;padding:1px 6px;border-radius:5px;font-size:9.5px;font-weight:600;background:rgba(0,200,140,0.18);color:#00c88c;">Spatial</span>';
      if(l.includes("name")||l.includes("anchor")) return '<span style="display:inline-block;padding:1px 6px;border-radius:5px;font-size:9.5px;font-weight:600;background:rgba(255,180,40,0.18);color:#ffb428;">Name-anchored</span>';
      return '<span style="display:inline-block;padding:1px 6px;border-radius:5px;font-size:9.5px;font-weight:600;background:rgba(130,145,160,0.18);color:#8291a0;">Unknown</span>';
    }};

    /* -- Render overlap points and address marker immediately (no layer fetch needed) -- */
    for (const row of overlapPoints) {{
      const g = new Graphic({{
        geometry: {{ type: "point", longitude: row.lon, latitude: row.lat }},
        symbol: {{
          type: "simple-marker", size: sz(row.high_total),
          color: row.color || [113,129,145,0.85],
          outline: {{ color: [255,255,255,0.70], width: 1 }}
        }},
        attributes: row,
        popupTemplate: {{
          title: row.subdivision_name || "Overlapping subdivision",
          content: `<div style="font-family:'Avenir Next LT Pro',system-ui,sans-serif;font-size:12px;line-height:1.5;">
            <div style="margin-bottom:5px;">${{badge(row.match_method)}}</div>
            <table style="border-collapse:collapse;width:100%;">
              <tr><td style="color:#7a94ab;padding:2px 6px 2px 0;font-size:11px;">Type</td><td style="font-weight:600;">${{row.subdivision_type||"N/A"}}</td></tr>
              <tr><td style="color:#7a94ab;padding:2px 6px 2px 0;font-size:11px;">Code</td><td>${{row.subdivision_code||"N/A"}}</td></tr>
              <tr><td style="color:#7a94ab;padding:2px 6px 2px 0;font-size:11px;">Matched clients</td><td style="font-weight:600;">${{row.match_count||0}}</td></tr>
              <tr><td style="color:#7a94ab;padding:2px 6px 2px 0;font-size:11px;">TFL high est.</td><td style="font-weight:600;">${{formatUsd(row.high_total)}}</td></tr>
              <tr><td style="color:#7a94ab;padding:2px 6px 2px 0;font-size:11px;">Method</td><td>${{row.match_method||"N/A"}}</td></tr>
              <tr><td style="color:#7a94ab;padding:2px 6px 2px 0;font-size:11px;">Source</td><td>${{row.source_name||"N/A"}}</td></tr>
            </table>
          </div>`
        }}
      }});
      overlapLayer.add(g);
    }}

    /* Address marker with pulse */
    addressLayer.add(new Graphic({{
      geometry: {{ type: "point", longitude: addressPoint.lon, latitude: addressPoint.lat }},
      symbol: {{ type: "simple-marker", style: "circle", size: 24, color: [201,34,52,0.0], outline: {{ color: [201,34,52,0.50], width: 2 }} }}
    }}));
    addressLayer.add(new Graphic({{
      geometry: {{ type: "point", longitude: addressPoint.lon, latitude: addressPoint.lat }},
      symbol: {{ type: "simple-marker", style: "diamond", size: 14, color: [201,34,52,0.95], outline: {{ color: [255,255,255,0.90], width: 1.8 }} }},
      popupTemplate: {{
        title: "Queried Address",
        content: `<div style="font-family:'Avenir Next LT Pro',system-ui,sans-serif;font-size:12px;">
          <div style="font-weight:600;margin-bottom:3px;">${{addressPoint.matched_address||"Address point"}}</div>
          <div style="color:#7a94ab;font-size:10.5px;">Lat ${{Number(addressPoint.lat).toFixed(5)}}, Lon ${{Number(addressPoint.lon).toFixed(5)}}</div>
        </div>`
      }}
    }}));

    /* -- Essential widgets (loaded in initial bundle) -- */
    const home = new Home({{ view }});
    const basemapToggle = new BasemapToggle({{ view, nextBasemap: baseMapId === "hybrid" ? "gray-vector" : "hybrid" }});
    const scaleBar = new ScaleBar({{ view, unit: "dual" }});
    view.ui.add(home, "top-left");
    view.ui.add(basemapToggle, "top-right");
    view.ui.add(scaleBar, "bottom-left");

    /* Coordinate readout */
    view.on("pointer-move", (evt) => {{
      const pt = view.toMap(evt);
      const el = document.getElementById("tfl-addr-coord");
      if (pt && el) el.textContent = pt.latitude.toFixed(5) + "\u00b0 N, " + Math.abs(pt.longitude).toFixed(5) + "\u00b0 W";
    }});

    /* Hover highlight */
    let hoverHL = null;
    view.on("pointer-move", (evt) => {{
      view.hitTest(evt, {{ include: [overlapLayer] }}).then((r) => {{
        const hit = r.results && r.results.find(x => x.graphic);
        document.getElementById("tfl-address-overlap-map").style.cursor = hit ? "pointer" : "default";
        if (hoverHL) {{ overlapLayer.remove(hoverHL); hoverHL = null; }}
        if (hit && hit.graphic && hit.graphic.geometry) {{
          hoverHL = new Graphic({{
            geometry: hit.graphic.geometry,
            symbol: {{ type: "simple-marker", style: "circle", size: 32, color: [255,255,255,0.0], outline: {{ color: [255,255,255,0.55], width: 2 }} }}
          }});
          overlapLayer.add(hoverHL);
        }}
      }});
    }});

    /* -- Phase 1: View ready — dismiss loading overlay and zoom to graphics -- */
    view.when(() => {{
      const loader = document.getElementById("tfl-addr-loading");
      if (loader) {{ loader.style.opacity = "0"; setTimeout(() => loader.remove(), 600); }}
      const all = [...overlapLayer.graphics.toArray(), ...addressLayer.graphics.toArray()];
      if (all.length > 0) {{
        view.goTo(all, {{ padding: {{ top: 50, right: 50, bottom: 50, left: 50 }}, duration: 1000, easing: "ease-in-out" }}).catch(() => {{}});
      }}

      /* -- Phase 2: Deferred load of reference FeatureLayers and secondary widgets -- */
      let advancedLoaded = false;
      const loadAdvanced = () => {{
        if (advancedLoaded) return;
        advancedLoaded = true;
        if (advBtn) {{
          advBtn.textContent = "Loading advanced tools...";
          advBtn.disabled = true;
        }}
        require([
        "esri/layers/FeatureLayer",
        "esri/widgets/Compass",
        "esri/widgets/Fullscreen",
        "esri/widgets/Search",
        "esri/widgets/Locate",
        "esri/widgets/Sketch",
        "esri/widgets/Expand",
        "esri/geometry/geometryEngine"
      ], function(FeatureLayer, Compass, Fullscreen, Search, Locate, Sketch, Expand, geometryEngine) {{

        /* -- Reference boundary layers (deferred — not needed for initial render) -- */
        const countyLayer = new FeatureLayer({{
          url: "{TEA_ARCGIS_COUNTY_LAYER_URL}",
          outFields: ["FENAME", "FIPS"],
          popupEnabled: false, labelsVisible: false,
          minScale: 2000000,
          labelingInfo: [{{
            labelExpressionInfo: {{ expression: "$feature.FENAME + ' County'" }},
            symbol: {{ type: "text", color: [160, 140, 110, 0.75], haloColor: [13, 23, 36, 0.80], haloSize: 0.8,
              font: {{ size: 11, family: "Avenir Next LT Pro", weight: "600" }} }}
          }}],
          renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [0,0,0,0], outline: {{ color: [145, 111, 63, 0.22], width: 0.6 }} }} }},
          opacity: 0.35
        }});

        const districtLayer = new FeatureLayer({{
          url: "{TEA_ARCGIS_SCHOOL_DISTRICT_LAYER_URL}",
          outFields: ["FID", "NAME20", "DISTRICT"],
          popupEnabled: false, labelsVisible: false,
          minScale: 500000,
          labelingInfo: [{{
            labelExpressionInfo: {{ expression: "$feature.NAME20" }},
            symbol: {{ type: "text", color: [73, 112, 150, 0.65], haloColor: [13, 23, 36, 0.8], haloSize: 0.6,
              font: {{ size: 8, family: "Avenir Next LT Pro", weight: "normal" }} }}
          }}],
          renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [0,0,0,0], outline: {{ color: [30, 144, 255, 0.20], width: 0.5 }} }} }},
          opacity: 0.35
        }});

        const cityLayer = new FeatureLayer({{
          url: "{CENSUS_ARCGIS_TEXAS_CITY_LAYER_URL}",
          outFields: ["NAME", "BASENAME", "GEOID", "STATE"],
          definitionExpression: "STATE = '48'",
          popupEnabled: false, labelsVisible: false,
          minScale: 1000000,
          labelingInfo: [{{
            labelExpressionInfo: {{ expression: "DefaultValue($feature.BASENAME, $feature.NAME)" }},
            symbol: {{ type: "text", color: [165, 100, 105, 0.80], haloColor: [13, 23, 36, 0.76], haloSize: 0.7,
              font: {{ size: 9, family: "Avenir Next LT Pro", weight: "500" }} }}
          }}],
          renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [0,0,0,0], outline: {{ color: [158, 42, 43, 0.12], width: 0.4 }} }} }},
          opacity: 0.20
        }});

        const houseLayer = new FeatureLayer({{
          url: "{TEXAS_HOUSE_DISTRICTS_LAYER_URL}",
          outFields: ["DISTRICT"],
          popupEnabled: true,
          popupTemplate: {{ title: "TX House District {{{{DISTRICT}}}}", content: "Texas House of Representatives District {{{{DISTRICT}}}}" }},
          labelsVisible: false,
          minScale: 2000000,
          labelingInfo: [{{
            labelExpressionInfo: {{ expression: "'HD ' + $feature.DISTRICT" }},
            symbol: {{ type: "text", color: [90, 180, 130, 0.70], haloColor: [13, 23, 36, 0.75], haloSize: 0.6,
              font: {{ size: 8, family: "Avenir Next LT Pro", weight: "600" }} }}
          }}],
          renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [40, 180, 100, 0.03], outline: {{ color: [40, 180, 100, 0.30], width: 0.8 }} }} }},
          opacity: 0.30, visible: false
        }});

        const senateLayer = new FeatureLayer({{
          url: "{TEXAS_SENATE_DISTRICTS_LAYER_URL}",
          outFields: ["DISTRICT"],
          popupEnabled: true,
          popupTemplate: {{ title: "TX Senate District {{{{DISTRICT}}}}", content: "Texas Senate District {{{{DISTRICT}}}}" }},
          labelsVisible: false,
          minScale: 2000000,
          labelingInfo: [{{
            labelExpressionInfo: {{ expression: "'SD ' + $feature.DISTRICT" }},
            symbol: {{ type: "text", color: [180, 130, 90, 0.70], haloColor: [13, 23, 36, 0.75], haloSize: 0.6,
              font: {{ size: 9, family: "Avenir Next LT Pro", weight: "600" }} }}
          }}],
          renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [200, 140, 60, 0.03], outline: {{ color: [200, 140, 60, 0.30], width: 0.8 }} }} }},
          opacity: 0.30, visible: false
        }});

        /* Insert reference layers below overlap graphics */
        map.add(countyLayer, 0);
        map.add(districtLayer, 1);
        map.add(houseLayer, 2);
        map.add(senateLayer, 3);
        map.add(cityLayer, 4);

        /* -- Zoom-dependent label visibility -- */
        const updateLabels = () => {{
          const z = Number(view.zoom || 0);
          countyLayer.labelsVisible = z >= 6;
          cityLayer.labelsVisible = z >= 7;
          districtLayer.labelsVisible = z >= 9;
          houseLayer.labelsVisible = z >= 8;
          senateLayer.labelsVisible = z >= 7;
        }};
        view.watch("zoom", updateLabels);
        updateLabels();

        /* -- Secondary widgets -- */
        const compass = new Compass({{ view }});
        const fullscreen = new Fullscreen({{ view }});
        const locate = new Locate({{ view }});

        const search = new Search({{
          view,
          popupEnabled: true,
          resultGraphicEnabled: true,
          goToOverride: (view, opts) => view.goTo(opts.target, {{ duration: 800, easing: "ease-in-out" }})
        }});
        search.on("select-result", (evt) => {{
          if (evt.result && evt.result.name) {{
            try {{ window.parent.postMessage({{ type: "tfl-map-address-search", address: evt.result.name }}, "*"); }} catch(e) {{}}
          }}
        }});

        const sketch = new Sketch({{
          view,
          layer: sketchLayer,
          creationMode: "single",
          availableCreateTools: ["polygon", "circle", "rectangle"],
          defaultCreateOptions: {{ mode: "freehand" }},
          visibleElements: {{ selectionTools: {{ "lasso-selection": false, "rectangle-selection": false }}, settingsMenu: false, undoRedoMenu: true }},
          defaultUpdateOptions: {{ tool: "reshape" }}
        }});
        const sketchExpand = new Expand({{
          view,
          content: sketch,
          expandIconClass: "esri-icon-polygon",
          expandTooltip: "Draw area for batch analysis",
          group: "tools"
        }});

        const layerDiv = document.createElement("div");
        layerDiv.style.cssText = "background:rgba(13,23,36,0.94);border-radius:8px;padding:10px;font-family:'Avenir Next LT Pro',system-ui,sans-serif;font-size:12px;color:rgba(210,225,240,0.90);min-width:160px;";
        layerDiv.innerHTML = '<div style="font-size:9px;text-transform:uppercase;letter-spacing:0.12em;color:rgba(150,175,200,0.65);font-weight:700;margin-bottom:6px;">Legislative Districts</div>'
          + '<label style="display:flex;align-items:center;gap:6px;cursor:pointer;padding:3px 0;"><input type="checkbox" id="tfl-toggle-house" style="accent-color:#28b464;"><span>TX House Districts</span></label>'
          + '<label style="display:flex;align-items:center;gap:6px;cursor:pointer;padding:3px 0;"><input type="checkbox" id="tfl-toggle-senate" style="accent-color:#c88c3c;"><span>TX Senate Districts</span></label>';
        const layerExpand = new Expand({{
          view,
          content: layerDiv,
          expandIconClass: "esri-icon-layer-list",
          expandTooltip: "Toggle legislative districts",
          group: "tools"
        }});

        view.ui.add(compass, "top-left");
        view.ui.add(fullscreen, "top-left");
        view.ui.add(locate, "top-left");
        view.ui.add(sketchExpand, "top-left");
        view.ui.add(layerExpand, "top-left");
        view.ui.add(search, "top-right");

        /* Wire layer toggles */
        const hBox = document.getElementById("tfl-toggle-house");
        const sBox = document.getElementById("tfl-toggle-senate");
        if (hBox) hBox.addEventListener("change", () => {{ houseLayer.visible = hBox.checked; }});
        if (sBox) sBox.addEventListener("change", () => {{ senateLayer.visible = sBox.checked; }});

        /* Sketch complete → batch analysis info */
        sketch.on("create", (evt) => {{
          if (evt.state !== "complete") return;
          const drawn = evt.graphic.geometry;
          const selInfo = document.getElementById("tfl-addr-sel-info");
          const contained = [];
          overlapLayer.graphics.forEach((g) => {{
            if (g.geometry && geometryEngine.contains(drawn, g.geometry)) {{
              contained.push(g.attributes || {{}});
            }}
          }});
          if (selInfo && contained.length > 0) {{
            const total = contained.reduce((a,r) => a + Number(r.high_total||0), 0);
            const types = [...new Set(contained.map(r => r.subdivision_type).filter(Boolean))];
            selInfo.style.display = "block";
            selInfo.innerHTML = '<div class="sel-title">Area Selection</div>'
              + '<div><strong>' + contained.length + '</strong> subdivision(s) in area</div>'
              + '<div>Combined TFL est.: <strong>' + formatUsd(total) + '</strong></div>'
              + '<div style="font-size:10px;color:rgba(180,200,220,0.65);margin-top:3px;">' + types.join(", ") + '</div>';
            try {{
              const names = contained.map(r => r.subdivision_name).filter(Boolean);
              window.parent.postMessage({{ type: "tfl-map-area-select", count: contained.length, totalHigh: total, names: names, types: types }}, "*");
            }} catch(e) {{}}
          }} else if (selInfo) {{
            selInfo.style.display = "block";
            selInfo.innerHTML = '<div class="sel-title">Area Selection</div><div>No subdivisions in drawn area.</div>';
            setTimeout(() => {{ selInfo.style.display = "none"; }}, 3000);
          }}
        }});
        if (advBtn) {{
          advBtn.textContent = "Advanced tools ready";
          setTimeout(() => {{
            try {{ advBtn.remove(); }} catch (e) {{}}
          }}, 1100);
        }}
      }}); /* end deferred require */
      }};
      if (advBtn) {{
        advBtn.addEventListener("click", () => loadAdvanced(), {{ once: true }});
      }}
      scheduleAdvancedLoad(() => loadAdvanced());
    }}); /* end view.when */
  }});
</script>
"""
    )
    _persistent_html_frame(
        html=arcgis_html,
        signature=address_map_signature,
        height=int(height) + 8,
        key="mp5_address_overlap_map_v2",
        default=None,
    )

def render_tfl_school_district_arcgis_map(matches: pd.DataFrame, height: int = 620, basemap: str = "gray-vector") -> None:
    if matches.empty:
        st.info("No matching school-district clients to plot on the map.")
        return

    payload_rows = []
    for row in matches.itertuples(index=False):
        clients = row.match_clients if isinstance(row.match_clients, list) else []
        safe_clients = [html.escape(str(c), quote=True) for c in clients]
        payload_rows.append(
            {
                "fid": int(row.fid),
                "district_name": html.escape(str(row.district_name), quote=True),
                "district_code": html.escape(str(row.district_code), quote=True),
                "lon": float(row.lon),
                "lat": float(row.lat),
                "match_count": int(row.match_count),
                "high_total": float(getattr(row, "high_total", 0.0) or 0.0),
                "match_clients_preview": html.escape(str(row.match_clients_preview), quote=True),
                "match_clients": safe_clients[:14],
                "extra_count": max(0, len(safe_clients) - 14),
            }
        )
    payload_json = json.dumps(payload_rows, ensure_ascii=True)
    basemap_safe = json.dumps(str(basemap).strip() or "gray-vector")

    total_districts = len(payload_rows)
    total_high = sum(r["high_total"] for r in payload_rows)
    total_high_fmt = f"${total_high:,.0f}"
    arcgis_html = f"""
<link rel="stylesheet" href="https://js.arcgis.com/4.30/esri/themes/dark/main.css"/>
<style>
  /* -- Dark popup theme -- */
  .esri-popup__main-container {{
    background: rgba(13,23,36,0.96) !important;
    color: rgba(220,230,240,0.95) !important;
    border: 1px solid rgba(100,140,180,0.22) !important;
    border-radius: 10px !important;
    backdrop-filter: blur(10px) !important;
    box-shadow: 0 8px 32px rgba(0,0,0,0.45) !important;
  }}
  .esri-popup__header-title {{ color: rgba(235,242,250,0.97) !important; font-weight: 600 !important; }}
  .esri-popup__content {{ color: rgba(200,215,230,0.92) !important; }}
  .esri-popup__button {{ color: rgba(180,200,220,0.85) !important; }}
  .esri-popup__button:hover {{ color: #fff !important; background: rgba(0,224,184,0.18) !important; }}
  .esri-popup__pointer-direction {{ background: rgba(13,23,36,0.96) !important; }}

  .esri-sketch {{ background: rgba(13,23,36,0.92) !important; border-radius: 8px !important; border: 1px solid rgba(100,140,180,0.22) !important; }}

  #tfl-sd-loading {{
    position:absolute; top:0; left:0; width:100%; height:100%;
    background:rgba(10,16,26,0.92); display:flex; flex-direction:column;
    align-items:center; justify-content:center; gap:10px;
    z-index:100; border-radius:14px; transition:opacity 0.6s ease;
  }}
  #tfl-sd-loading .ld-spinner {{
    width:30px; height:30px; border:2.5px solid rgba(100,140,180,0.18);
    border-top:2.5px solid rgba(0,224,184,0.80); border-radius:50%;
    animation: tfl-sd-spin 0.8s linear infinite;
  }}
  #tfl-sd-loading .ld-label {{
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:10px;
    color:rgba(160,185,210,0.65); letter-spacing:0.06em;
  }}
  @keyframes tfl-sd-spin {{ 0%{{transform:rotate(0deg)}} 100%{{transform:rotate(360deg)}} }}
  #tfl-sd-coord {{
    position:absolute; bottom:10px; left:50%; transform:translateX(-50%); z-index:90;
    background:rgba(10,16,26,0.85); border:1px solid rgba(100,140,180,0.15);
    border-radius:6px; padding:2px 10px;
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:10px;
    color:rgba(180,200,220,0.75); pointer-events:none; backdrop-filter:blur(6px);
    white-space:nowrap; letter-spacing:0.04em;
  }}
  #tfl-sd-legend {{
    position:absolute; bottom:36px; right:12px; z-index:90;
    background:rgba(10,20,32,0.92); border:1px solid rgba(100,140,180,0.18);
    border-radius:10px; padding:6px 10px; max-width:210px;
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:10.5px;
    color:rgba(210,225,240,0.90); backdrop-filter:blur(8px);
  }}
  #tfl-sd-legend .leg-title {{
    text-transform:uppercase; letter-spacing:0.14em; font-size:8px;
    color:rgba(150,175,200,0.70); margin-bottom:3px; font-weight:700;
  }}
  #tfl-sd-legend .leg-row {{ display:flex; align-items:center; gap:5px; padding:1.5px 0; }}
  #tfl-sd-legend .leg-chip {{
    width:9px; height:9px; border-radius:50%; flex-shrink:0;
    border:1px solid rgba(255,255,255,0.18);
  }}
  #tfl-sd-sel-info {{
    position:absolute; bottom:36px; left:12px; z-index:90;
    background:rgba(10,16,26,0.92); border:1px solid rgba(0,224,184,0.25);
    border-radius:8px; padding:6px 12px; backdrop-filter:blur(6px);
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:11px;
    color:rgba(200,220,240,0.90); display:none; max-width:280px;
  }}
  #tfl-sd-sel-info .sel-title {{
    font-weight:700; color:rgba(0,224,184,0.95); font-size:10px;
    text-transform:uppercase; letter-spacing:0.08em; margin-bottom:3px;
  }}
</style>
<div style="width:100%;height:{height}px;position:relative;">
  <div id="tfl-arcgis-map" style="width:100%;height:100%;border-radius:14px;overflow:hidden;"></div>
  <div id="tfl-sd-legend">
    <div class="leg-title">Legend</div>
    <div class="leg-row"><span class="leg-chip" style="background:#00e0b8;"></span><span>School District ({total_districts})</span></div>
    <div class="leg-row"><span class="leg-chip" style="background:rgba(145,111,63,0.5);border-color:rgba(145,111,63,0.6);"></span><span style="font-size:10px;color:rgba(180,195,210,0.70);">County boundary</span></div>
    <div class="leg-row"><span class="leg-chip" style="background:rgba(30,144,255,0.3);border-color:rgba(30,144,255,0.5);"></span><span style="font-size:10px;color:rgba(180,195,210,0.70);">District boundary</span></div>
  </div>
  <div id="tfl-sd-loading"><div class="ld-spinner"></div><div class="ld-label">Loading map layers&hellip;</div></div>
  <div id="tfl-sd-coord">&ndash;</div>
  <div id="tfl-sd-sel-info"></div>
</div>
<script src="https://js.arcgis.com/4.30/"></script>
<script>
  const tflPoints = {payload_json};
  const baseMapId = {basemap_safe};
  require([
    "esri/Map",
    "esri/views/MapView",
    "esri/layers/FeatureLayer",
    "esri/layers/GraphicsLayer",
    "esri/Graphic",
    "esri/widgets/Home",
    "esri/widgets/ScaleBar",
    "esri/widgets/BasemapToggle",
    "esri/widgets/Compass",
    "esri/widgets/Fullscreen",
    "esri/widgets/Search",
    "esri/widgets/Locate",
    "esri/widgets/Sketch",
    "esri/widgets/Expand",
    "esri/geometry/geometryEngine"
  ], function(Map, MapView, FeatureLayer, GraphicsLayer, Graphic, Home, ScaleBar, BasemapToggle, Compass, Fullscreen, Search, Locate, Sketch, Expand, geometryEngine) {{
    const map = new Map({{ basemap: baseMapId }});

    const countyLayer = new FeatureLayer({{
      url: "{TEA_ARCGIS_COUNTY_LAYER_URL}",
      outFields: ["FENAME", "FIPS"],
      popupEnabled: false, labelsVisible: false,
      labelingInfo: [{{
        labelExpressionInfo: {{ expression: "$feature.FENAME + ' County'" }},
        symbol: {{ type: "text", color: [160, 140, 110, 0.75], haloColor: [13, 23, 36, 0.80], haloSize: 0.8,
          font: {{ size: 11, family: "Avenir Next LT Pro", weight: "600" }} }}
      }}],
      renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [0,0,0,0], outline: {{ color: [145, 111, 63, 0.22], width: 0.6 }} }} }},
      opacity: 0.35
    }});

    const districtLayer = new FeatureLayer({{
      url: "{TEA_ARCGIS_SCHOOL_DISTRICT_LAYER_URL}",
      outFields: ["FID", "NAME20", "DISTRICT"],
      popupEnabled: false, labelsVisible: false,
      labelingInfo: [{{
        labelExpressionInfo: {{ expression: "$feature.NAME20" }},
        symbol: {{ type: "text", color: [73, 112, 150, 0.65], haloColor: [13, 23, 36, 0.8], haloSize: 0.6,
          font: {{ size: 8, family: "Avenir Next LT Pro", weight: "normal" }} }}
      }}],
      renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [0,0,0,0], outline: {{ color: [30, 144, 255, 0.25], width: 0.6 }} }} }},
      opacity: 0.40
    }});

    /* TX House & Senate district boundaries */
    const houseLayer = new FeatureLayer({{
      url: "{TEXAS_HOUSE_DISTRICTS_LAYER_URL}",
      outFields: ["*"], popupEnabled: true,
      popupTemplate: {{ title: "TX House District {{{{DISTRICT}}}}", content: "Texas House of Representatives District {{{{DISTRICT}}}}" }},
      labelsVisible: false,
      labelingInfo: [{{ labelExpressionInfo: {{ expression: "'HD ' + $feature.DISTRICT" }},
        symbol: {{ type: "text", color: [90, 180, 130, 0.70], haloColor: [13, 23, 36, 0.75], haloSize: 0.6,
          font: {{ size: 8, family: "Avenir Next LT Pro", weight: "600" }} }} }}],
      renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [40, 180, 100, 0.03], outline: {{ color: [40, 180, 100, 0.30], width: 0.8 }} }} }},
      opacity: 0.30, visible: false
    }});
    const senateLayer = new FeatureLayer({{
      url: "{TEXAS_SENATE_DISTRICTS_LAYER_URL}",
      outFields: ["*"], popupEnabled: true,
      popupTemplate: {{ title: "TX Senate District {{{{DISTRICT}}}}", content: "Texas Senate District {{{{DISTRICT}}}}" }},
      labelsVisible: false,
      labelingInfo: [{{ labelExpressionInfo: {{ expression: "'SD ' + $feature.DISTRICT" }},
        symbol: {{ type: "text", color: [180, 130, 90, 0.70], haloColor: [13, 23, 36, 0.75], haloSize: 0.6,
          font: {{ size: 9, family: "Avenir Next LT Pro", weight: "600" }} }} }}],
      renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [200, 140, 60, 0.03], outline: {{ color: [200, 140, 60, 0.30], width: 0.8 }} }} }},
      opacity: 0.30, visible: false
    }});

    map.add(countyLayer);
    map.add(districtLayer);
    map.add(houseLayer);
    map.add(senateLayer);

    const graphics = new GraphicsLayer();
    const sketchLayer = new GraphicsLayer();
    map.add(graphics);
    map.add(sketchLayer);

    const view = new MapView({{
      container: "tfl-arcgis-map",
      map,
      center: [-99.3, 31.1],
      zoom: 5,
      constraints: {{ minZoom: 4 }},
      popup: {{ dockEnabled: true, dockOptions: {{ position: "bottom-right", breakpoint: false }} }},
      ui: {{ padding: {{ top: 10, right: 10, bottom: 30, left: 10 }} }}
    }});

    const formatUsd = (v) => Number(v||0).toLocaleString("en-US",{{style:"currency",currency:"USD",maximumFractionDigits:0}});
    const maxHigh = tflPoints.reduce((a, r) => Math.max(a, Number(r.high_total || 0)), 0);
    const sz = (row) => {{
      if (maxHigh > 0) {{
        const n = Math.max(0, Number(row.high_total || 0));
        return Math.max(8, Math.min(28, 8 + (Math.log10(n+1)/Math.log10(maxHigh+1))*20));
      }}
      return Math.min(26, 8 + Math.log2((row.match_count || 1) + 1) * 5);
    }};

    for (const row of tflPoints) {{
      const clientsHtml = (row.match_clients || []).join(", ");
      const extraHtml = row.extra_count > 0 ? `, +${{row.extra_count}} more` : "";
      const g = new Graphic({{
        geometry: {{ type: "point", longitude: row.lon, latitude: row.lat }},
        symbol: {{
          type: "simple-marker", size: sz(row),
          color: [0, 224, 184, 0.85],
          outline: {{ color: [7, 22, 39, 0.95], width: 1 }}
        }},
        attributes: row,
        popupTemplate: {{
          title: row.district_name || "School District",
          content: `<div style="font-family:'Avenir Next LT Pro',system-ui,sans-serif;font-size:12px;line-height:1.5;">
            <table style="border-collapse:collapse;width:100%;">
              <tr><td style="color:#7a94ab;padding:2px 6px 2px 0;font-size:11px;">District code</td><td style="font-weight:600;">${{row.district_code || "N/A"}}</td></tr>
              <tr><td style="color:#7a94ab;padding:2px 6px 2px 0;font-size:11px;">TFL high est.</td><td style="font-weight:600;">${{formatUsd(row.high_total)}}</td></tr>
              <tr><td style="color:#7a94ab;padding:2px 6px 2px 0;font-size:11px;">Matched clients</td><td style="font-weight:600;">${{row.match_count}}</td></tr>
            </table>
            <div style="margin-top:5px;padding-top:4px;border-top:1px solid rgba(140,160,180,0.20);font-size:11px;">${{clientsHtml}}${{extraHtml}}</div>
          </div>`
        }}
      }});
      graphics.add(g);
    }}

    const updateLabels = () => {{
      const z = Number(view.zoom || 0);
      countyLayer.labelsVisible = z >= 6;
      districtLayer.labelsVisible = z >= 8.5;
      houseLayer.labelsVisible = z >= 8;
      senateLayer.labelsVisible = z >= 7;
    }};
    view.watch("zoom", updateLabels);

    /* -- Widgets -- */
    const home = new Home({{ view }});
    const basemapToggle = new BasemapToggle({{ view, nextBasemap: baseMapId === "hybrid" ? "gray-vector" : "hybrid" }});
    const scaleBar = new ScaleBar({{ view, unit: "dual" }});
    const compass = new Compass({{ view }});
    const fullscreen = new Fullscreen({{ view }});
    const locate = new Locate({{ view }});
    const search = new Search({{
      view, popupEnabled: true, resultGraphicEnabled: true,
      goToOverride: (view, opts) => view.goTo(opts.target, {{ duration: 800, easing: "ease-in-out" }})
    }});

    /* Sketch tool for encircling areas */
    const sketch = new Sketch({{
      view, layer: sketchLayer, creationMode: "single",
      availableCreateTools: ["polygon", "circle", "rectangle"],
      defaultCreateOptions: {{ mode: "freehand" }},
      visibleElements: {{ selectionTools: {{ "lasso-selection": false, "rectangle-selection": false }}, settingsMenu: false, undoRedoMenu: true }},
      defaultUpdateOptions: {{ tool: "reshape" }}
    }});
    const sketchExpand = new Expand({{
      view, content: sketch, expandIconClass: "esri-icon-polygon",
      expandTooltip: "Draw area for batch analysis", group: "tools"
    }});

    /* House/Senate layer toggle */
    const layerDiv = document.createElement("div");
    layerDiv.style.cssText = "background:rgba(13,23,36,0.94);border-radius:8px;padding:10px;font-family:'Avenir Next LT Pro',system-ui,sans-serif;font-size:12px;color:rgba(210,225,240,0.90);min-width:160px;";
    layerDiv.innerHTML = '<div style="font-size:9px;text-transform:uppercase;letter-spacing:0.12em;color:rgba(150,175,200,0.65);font-weight:700;margin-bottom:6px;">Legislative Districts</div>'
      + '<label style="display:flex;align-items:center;gap:6px;cursor:pointer;padding:3px 0;"><input type="checkbox" id="tfl-sd-toggle-house" style="accent-color:#28b464;"><span>TX House Districts</span></label>'
      + '<label style="display:flex;align-items:center;gap:6px;cursor:pointer;padding:3px 0;"><input type="checkbox" id="tfl-sd-toggle-senate" style="accent-color:#c88c3c;"><span>TX Senate Districts</span></label>';
    const layerExpand = new Expand({{
      view, content: layerDiv, expandIconClass: "esri-icon-layer-list",
      expandTooltip: "Toggle legislative districts", group: "tools"
    }});

    view.ui.add(home, "top-left");
    view.ui.add(compass, "top-left");
    view.ui.add(fullscreen, "top-left");
    view.ui.add(locate, "top-left");
    view.ui.add(sketchExpand, "top-left");
    view.ui.add(layerExpand, "top-left");
    view.ui.add(search, "top-right");
    view.ui.add(basemapToggle, "top-right");
    view.ui.add(scaleBar, "bottom-left");

    /* Wire layer toggles */
    view.when(() => {{
      const hBox = document.getElementById("tfl-sd-toggle-house");
      const sBox = document.getElementById("tfl-sd-toggle-senate");
      if (hBox) hBox.addEventListener("change", () => {{ houseLayer.visible = hBox.checked; }});
      if (sBox) sBox.addEventListener("change", () => {{ senateLayer.visible = sBox.checked; }});
    }});

    /* Sketch complete → batch analysis info */
    sketch.on("create", (evt) => {{
      if (evt.state !== "complete") return;
      const drawn = evt.graphic.geometry;
      const selInfo = document.getElementById("tfl-sd-sel-info");
      const contained = [];
      graphics.graphics.forEach((g) => {{
        if (g.geometry && geometryEngine.contains(drawn, g.geometry)) contained.push(g.attributes || {{}});
      }});
      if (selInfo && contained.length > 0) {{
        const total = contained.reduce((a,r) => a + Number(r.high_total||0), 0);
        selInfo.style.display = "block";
        selInfo.innerHTML = '<div class="sel-title">Area Selection</div>'
          + '<div><strong>' + contained.length + '</strong> district(s) in area</div>'
          + '<div>Combined TFL est.: <strong>' + formatUsd(total) + '</strong></div>';
        try {{ window.parent.postMessage({{ type: "tfl-map-area-select", count: contained.length, totalHigh: total }}, "*"); }} catch(e) {{}}
      }} else if (selInfo) {{
        selInfo.style.display = "block";
        selInfo.innerHTML = '<div class="sel-title">Area Selection</div><div>No districts in drawn area.</div>';
        setTimeout(() => {{ selInfo.style.display = "none"; }}, 3000);
      }}
    }});

    /* Coordinate readout */
    view.on("pointer-move", (evt) => {{
      const pt = view.toMap(evt);
      const el = document.getElementById("tfl-sd-coord");
      if (pt && el) el.textContent = pt.latitude.toFixed(5) + "\u00b0 N, " + Math.abs(pt.longitude).toFixed(5) + "\u00b0 W";
    }});

    /* Hover highlight */
    let hoverHL = null;
    view.on("pointer-move", (evt) => {{
      view.hitTest(evt, {{ include: [graphics] }}).then((r) => {{
        const hit = r.results && r.results.find(x => x.graphic);
        document.getElementById("tfl-arcgis-map").style.cursor = hit ? "pointer" : "default";
        if (hoverHL) {{ graphics.remove(hoverHL); hoverHL = null; }}
        if (hit && hit.graphic && hit.graphic.geometry) {{
          hoverHL = new Graphic({{
            geometry: hit.graphic.geometry,
            symbol: {{ type: "simple-marker", style: "circle", size: 30, color: [255,255,255,0.0], outline: {{ color: [0,224,184,0.55], width: 2 }} }}
          }});
          graphics.add(hoverHL);
        }}
      }});
    }});

    view.when(() => {{
      const loader = document.getElementById("tfl-sd-loading");
      if (loader) {{ loader.style.opacity = "0"; setTimeout(() => loader.remove(), 600); }}
      updateLabels();
      if (graphics.graphics.length > 0) {{
        view.goTo(graphics.graphics.toArray(), {{ padding: {{ top: 50, right: 50, bottom: 50, left: 50 }}, duration: 1000, easing: "ease-in-out" }}).catch(() => {{}});
      }}
    }});
  }});
</script>
"""
    components.html(arcgis_html, height=height + 8, scrolling=False)

def render_tfl_subdivision_arcgis_map(
    matches: pd.DataFrame,
    height: int = 640,
    basemap: str = "gray-vector",
) -> None:
    if matches.empty:
        st.info("No matching political-subdivision clients to plot on the map.")
        return

    type_colors = {
        subtype: _hex_to_rgba(color_hex, alpha=0.9)
        for subtype, color_hex in SUBDIVISION_TYPE_COLORS.items()
    }
    type_colors_json = json.dumps(type_colors, ensure_ascii=True)

    # Build hex color map for the on-map legend
    type_hex_colors = {
        subtype: color_hex
        for subtype, color_hex in SUBDIVISION_TYPE_COLORS.items()
    }

    payload_rows = []
    for row in matches.itertuples(index=False):
        clients = row.match_clients if isinstance(row.match_clients, list) else []
        safe_clients = [str(c).strip() for c in clients if str(c).strip()]
        preview = str(getattr(row, "match_clients_preview", "")).strip() or ", ".join(safe_clients[:14])
        payload_rows.append(
            {
                "subdivision_type": html.escape(str(row.subdivision_type), quote=True),
                "subdivision_name": html.escape(str(row.subdivision_name), quote=True),
                "subdivision_code": html.escape(str(row.subdivision_code), quote=True),
                "source_name": html.escape(str(getattr(row, "source_name", "")), quote=True),
                "lon": float(row.lon),
                "lat": float(row.lat),
                "match_count": int(row.match_count),
                "high_total": float(getattr(row, "high_total", 0.0) or 0.0),
                "match_clients_preview": html.escape(preview, quote=True),
                "extra_count": max(0, len(safe_clients) - 14),
            }
        )
    payload_json = json.dumps(payload_rows, ensure_ascii=True)
    basemap_safe = json.dumps(str(basemap).strip() or "gray-vector")

    # Determine which subdivision types are actually present for the legend
    present_types: dict[str, tuple[str, int]] = {}
    for pr in payload_rows:
        st_key = pr["subdivision_type"]
        if st_key:
            if st_key not in present_types:
                present_types[st_key] = (type_hex_colors.get(st_key, "#718191"), 0)
            present_types[st_key] = (present_types[st_key][0], present_types[st_key][1] + 1)
    legend_items_json = json.dumps(
        [{"type": t, "color": c, "count": n} for t, (c, n) in sorted(present_types.items(), key=lambda x: -x[1][1])],
        ensure_ascii=True,
    )

    total_sub = len(payload_rows)
    total_sub_high = sum(r["high_total"] for r in payload_rows)
    total_sub_high_fmt = f"${total_sub_high:,.0f}"
    subdivision_map_signature = _stable_json_signature(
        {
            "renderer_version": 3,
            "points": payload_rows,
            "basemap": str(basemap).strip() or "gray-vector",
            "height": int(height),
        }
    )
    arcgis_html = _session_cached_value(
        "_mp5_subdivision_map_html_v3",
        subdivision_map_signature,
        lambda: f"""
<link rel="stylesheet" href="https://js.arcgis.com/4.30/esri/themes/dark/main.css"/>
<style>
  /* -- Dark popup theme -- */
  .esri-popup__main-container {{
    background: rgba(13,23,36,0.96) !important;
    color: rgba(220,230,240,0.95) !important;
    border: 1px solid rgba(100,140,180,0.22) !important;
    border-radius: 10px !important;
    backdrop-filter: blur(10px) !important;
    box-shadow: 0 8px 32px rgba(0,0,0,0.45) !important;
  }}
  .esri-popup__header-title {{ color: rgba(235,242,250,0.97) !important; font-weight: 600 !important; }}
  .esri-popup__content {{ color: rgba(200,215,230,0.92) !important; }}
  .esri-popup__button {{ color: rgba(180,200,220,0.85) !important; }}
  .esri-popup__button:hover {{ color: #fff !important; background: rgba(100,180,255,0.18) !important; }}
  .esri-popup__pointer-direction {{ background: rgba(13,23,36,0.96) !important; }}

  .esri-sketch {{ background: rgba(13,23,36,0.92) !important; border-radius: 8px !important; border: 1px solid rgba(100,140,180,0.22) !important; }}

  #tfl-sub-legend {{
    position: absolute; bottom: 90px; right: 12px; z-index: 90;
    background: rgba(10,20,32,0.92); border: 1px solid rgba(100,140,180,0.18);
    border-radius: 11px; padding: 0; max-width: 240px; overflow: hidden;
    font-family: 'Avenir Next LT Pro', system-ui, sans-serif; font-size: 10.5px;
    color: rgba(210,225,240,0.90); backdrop-filter: blur(8px);
    transition: max-height 0.3s ease;
  }}
  #tfl-sub-legend .leg-hdr {{
    display: flex; align-items: center; justify-content: space-between;
    padding: 6px 10px 4px 10px; cursor: pointer; user-select: none;
  }}
  #tfl-sub-legend .leg-title {{
    text-transform: uppercase; letter-spacing: 0.14em; font-size: 8px;
    color: rgba(150,175,200,0.70); font-weight: 700;
  }}
  #tfl-sub-legend .leg-toggle {{
    font-size: 12px; color: rgba(150,175,200,0.60); transition: transform 0.25s;
  }}
  #tfl-sub-legend .leg-body {{
    padding: 0 10px 6px 10px; max-height: 180px; overflow-y: auto;
  }}
  #tfl-sub-legend .leg-body::-webkit-scrollbar {{ width:4px; }}
  #tfl-sub-legend .leg-body::-webkit-scrollbar-thumb {{ background:rgba(100,140,180,0.25); border-radius:4px; }}
  #tfl-sub-legend .leg-body::-webkit-scrollbar-track {{ background:transparent; }}
  #tfl-sub-legend .leg-row {{
    display: flex; align-items: center; justify-content: space-between; gap: 5px; padding: 2px 0;
    cursor: pointer; border-radius: 4px; padding-left: 3px; padding-right: 3px;
    transition: background 0.15s, opacity 0.25s;
  }}
  #tfl-sub-legend .leg-row:hover {{ background: rgba(255,255,255,0.06); }}
  #tfl-sub-legend .leg-row.dimmed {{ opacity: 0.28; }}
  #tfl-sub-legend .leg-left {{ display: flex; align-items: center; gap: 5px; min-width: 0; }}
  #tfl-sub-legend .leg-chip {{
    width: 9px; height: 9px; border-radius: 50%; flex-shrink: 0;
    border: 1px solid rgba(255,255,255,0.18);
  }}
  #tfl-sub-legend .leg-label {{ white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
  #tfl-sub-legend .leg-count {{ font-weight: 600; flex-shrink: 0; color: rgba(200,215,230,0.75); font-size: 9.5px; }}
  #tfl-sub-legend.collapsed .leg-body {{ display: none; }}
  #tfl-sub-legend.collapsed .leg-toggle {{ transform: rotate(180deg); }}

  #tfl-sub-loading {{
    position:absolute; top:0; left:0; width:100%; height:100%;
    background:rgba(10,16,26,0.92); display:flex; flex-direction:column;
    align-items:center; justify-content:center; gap:10px;
    z-index:100; border-radius:14px; transition:opacity 0.6s ease;
  }}
  #tfl-sub-loading .ld-spinner {{
    width:30px; height:30px; border:2.5px solid rgba(100,140,180,0.18);
    border-top:2.5px solid rgba(100,180,255,0.80); border-radius:50%;
    animation: tfl-sub-spin 0.8s linear infinite;
  }}
  #tfl-sub-loading .ld-label {{
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:10px;
    color:rgba(160,185,210,0.65); letter-spacing:0.06em;
  }}
  @keyframes tfl-sub-spin {{ 0%{{transform:rotate(0deg)}} 100%{{transform:rotate(360deg)}} }}
  #tfl-sub-coord {{
    position:absolute; bottom:10px; left:50%; transform:translateX(-50%); z-index:90;
    background:rgba(10,16,26,0.85); border:1px solid rgba(100,140,180,0.15);
    border-radius:6px; padding:2px 10px;
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:10px;
    color:rgba(180,200,220,0.75); pointer-events:none; backdrop-filter:blur(6px);
    white-space:nowrap; letter-spacing:0.04em;
  }}
  #tfl-sub-sel-info {{
    position:absolute; top:52px; right:12px; z-index:90;
    background:rgba(10,16,26,0.92); border:1px solid rgba(0,180,255,0.25);
    border-radius:8px; padding:6px 12px; backdrop-filter:blur(6px);
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:11px;
    color:rgba(200,220,240,0.90); display:none; max-width:260px;
  }}
  #tfl-sub-sel-info .sel-title {{
    font-weight:700; color:rgba(100,200,255,0.95); font-size:10px;
    text-transform:uppercase; letter-spacing:0.08em; margin-bottom:3px;
  }}

  /* -- Address Collector Panel -- */
  #tfl-sub-collector {{
    position:absolute; bottom:40px; left:12px; z-index:95;
    background:rgba(10,20,32,0.94); border:1px solid rgba(30,144,255,0.22);
    border-radius:12px; padding:10px 12px; min-width:210px; max-width:260px;
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:11.5px;
    color:rgba(210,225,240,0.90); backdrop-filter:blur(10px);
    max-height:{height - 100}px; overflow-y:auto;
    box-shadow:0 8px 28px rgba(0,0,0,0.40);
    transition: max-height 0.3s ease, opacity 0.3s ease;
  }}
  #tfl-sub-collector::-webkit-scrollbar {{ width:4px; }}
  #tfl-sub-collector::-webkit-scrollbar-thumb {{ background:rgba(30,144,255,0.25); border-radius:4px; }}
  #tfl-sub-collector::-webkit-scrollbar-track {{ background:transparent; }}
  }}
  #tfl-sub-collector.collapsed {{ max-height:32px; overflow:hidden; }}
  #tfl-sub-collector .dc-title {{
    text-transform:uppercase; letter-spacing:0.14em; font-size:8.5px;
    color:rgba(30,144,255,0.82); font-weight:700; margin-bottom:6px;
    display:flex; align-items:center; justify-content:space-between;
    cursor:pointer; user-select:none;
  }}
  #tfl-sub-collector .dc-toggle {{
    font-size:11px; color:rgba(150,175,200,0.60); transition:transform 0.25s;
  }}
  #tfl-sub-collector.collapsed .dc-toggle {{ transform:rotate(180deg); }}
  #tfl-sub-collector .dc-body {{ }}
  #tfl-sub-collector.collapsed .dc-body {{ display:none; }}
  #tfl-sub-collector .dc-item {{
    display:flex; align-items:flex-start; gap:6px; padding:5px 0;
    border-bottom:1px solid rgba(255,255,255,0.06);
  }}
  #tfl-sub-collector .dc-item:last-child {{ border-bottom:none; }}
  #tfl-sub-collector .dc-num {{
    flex-shrink:0; width:18px; height:18px; border-radius:50%;
    background:rgba(30,144,255,0.18); border:1px solid rgba(30,144,255,0.30);
    display:flex; align-items:center; justify-content:center;
    font-size:9px; font-weight:700; color:rgba(30,144,255,0.90);
  }}
  #tfl-sub-collector .dc-addr {{
    font-size:11px; line-height:1.35; color:rgba(210,230,245,0.85);
  }}
  #tfl-sub-collector .dc-coord {{
    font-size:9px; color:rgba(160,185,210,0.55); margin-top:1px;
  }}
  #tfl-sub-collector .dc-empty {{
    text-align:center; padding:10px 0; color:rgba(180,200,220,0.45); font-size:10.5px;
  }}
  #tfl-sub-collector .dc-actions {{
    display:flex; gap:6px; margin-top:6px;
  }}
  #tfl-sub-collector .dc-btn {{
    flex:1; padding:5px 8px; border-radius:8px; border:1px solid rgba(30,144,255,0.25);
    background:rgba(30,144,255,0.10); color:rgba(30,144,255,0.90); cursor:pointer;
    font-size:10px; font-weight:600; text-align:center; transition:all 0.2s;
  }}
  #tfl-sub-collector .dc-btn:hover {{ background:rgba(30,144,255,0.22); border-color:rgba(30,144,255,0.40); }}
  #tfl-sub-collector .dc-btn.clear {{ background:rgba(255,80,80,0.08); border-color:rgba(255,80,80,0.20); color:rgba(255,120,120,0.85); }}
  #tfl-sub-collector .dc-btn.clear:hover {{ background:rgba(255,80,80,0.18); }}
  #tfl-sub-badge {{
    display:none;
  }}

  /* -- Toast Notification System -- */
  #tfl-sub-toast-container {{
    position:absolute; bottom:50px; left:50%; transform:translateX(-50%); z-index:200;
    display:flex; flex-direction:column-reverse; align-items:center; gap:8px;
    pointer-events:none;
  }}
  .tfl-toast {{
    background:rgba(10,20,32,0.95); border:1px solid rgba(30,144,255,0.30);
    border-radius:10px; padding:8px 18px; backdrop-filter:blur(12px);
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:12px;
    color:rgba(210,230,245,0.95); box-shadow:0 6px 24px rgba(0,0,0,0.45);
    display:flex; align-items:center; gap:8px; white-space:nowrap;
    animation: tfl-toast-in 0.35s ease-out forwards;
    pointer-events:auto;
  }}
  .tfl-toast.success {{ border-color:rgba(40,180,100,0.40); }}
  .tfl-toast.success .toast-icon {{ color:#28b464; }}
  .tfl-toast.info {{ border-color:rgba(30,144,255,0.40); }}
  .tfl-toast.info .toast-icon {{ color:#1e90ff; }}
  .tfl-toast.warn {{ border-color:rgba(255,170,50,0.40); }}
  .tfl-toast.warn .toast-icon {{ color:#ffaa32; }}
  .tfl-toast.out {{ animation: tfl-toast-out 0.3s ease-in forwards; }}
  .toast-icon {{ font-size:15px; }}
  @keyframes tfl-toast-in {{ 0%{{opacity:0;transform:translateY(16px);}} 100%{{opacity:1;transform:translateY(0);}} }}
  @keyframes tfl-toast-out {{ 0%{{opacity:1;transform:translateY(0);}} 100%{{opacity:0;transform:translateY(-12px);}} }}

  /* -- Stats ribbon -- */
  #tfl-sub-stats {{
    position:absolute; top:10px; left:56px; z-index:90;
    background:rgba(10,16,26,0.88); border:1px solid rgba(100,140,180,0.15);
    border-radius:20px; padding:4px 16px; backdrop-filter:blur(8px);
    font-family:'Avenir Next LT Pro',system-ui,sans-serif; font-size:11px;
    color:rgba(200,220,240,0.85); pointer-events:none;
    display:flex; align-items:center; gap:14px; white-space:nowrap;
    animation: tfl-toast-in 0.5s ease-out 1.2s both;
  }}
  #tfl-sub-stats .stat-val {{ font-weight:700; color:rgba(100,200,255,0.95); }}
  #tfl-sub-stats .stat-sep {{ color:rgba(100,140,180,0.30); }}

  /* -- Hover tooltip -- */
  #tfl-sub-tooltip {{
    position:absolute; z-index:110; pointer-events:none;
    background:rgba(10,16,26,0.94); border:1px solid rgba(100,180,255,0.22);
    border-radius:8px; padding:5px 10px; backdrop-filter:blur(8px);
    font-family:'Avenir Next LT Pro',system-ui,sans-serif;
    color:rgba(210,230,245,0.92); display:none;
    box-shadow:0 4px 16px rgba(0,0,0,0.35);
    max-width:240px;
  }}
  #tfl-sub-tooltip .tt-name {{ font-size:11.5px; font-weight:600; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
  #tfl-sub-tooltip .tt-val {{ font-size:10px; color:rgba(100,200,255,0.85); margin-top:2px; }}
  #tfl-sub-tooltip .tt-type {{ font-size:9px; color:rgba(150,175,200,0.55); margin-top:1px; }}

  /* -- Address delete button -- */
  #tfl-sub-collector .dc-del {{
    flex-shrink:0; width:16px; height:16px; border-radius:50%; margin-left:auto;
    background:rgba(255,80,80,0.08); border:1px solid rgba(255,80,80,0.20);
    color:rgba(255,120,120,0.70); font-size:10px; line-height:14px;
    text-align:center; cursor:pointer; transition:all 0.2s;
    display:flex; align-items:center; justify-content:center;
  }}
  #tfl-sub-collector .dc-del:hover {{ background:rgba(255,80,80,0.22); color:rgba(255,120,120,1); }}
</style>
<div style="width:100%;height:{height}px;position:relative;">
  <div id="tfl-subdivision-map" style="width:100%;height:100%;border-radius:14px;overflow:hidden;"></div>
  <div id="tfl-sub-legend" class="collapsed">
    <div class="leg-hdr" onclick="this.parentElement.classList.toggle('collapsed')">
      <span class="leg-title">Subdivisions &middot; click to filter</span>
      <span class="leg-toggle">&#9650;</span>
    </div>
    <div class="leg-body" id="tfl-sub-legend-body"></div>
  </div>
  <div id="tfl-sub-loading"><div class="ld-spinner"></div><div class="ld-label">Loading map layers&hellip;</div></div>
  <div id="tfl-sub-coord">&ndash;</div>
  <div id="tfl-sub-sel-info"></div>
  <div id="tfl-sub-badge">Click map or search to collect addresses</div>
  <div id="tfl-sub-toast-container"></div>
  <div id="tfl-sub-stats">
    <span><span class="stat-val" id="tfl-sub-stats-count">{total_sub}</span> subdivisions</span>
    <span class="stat-sep">|</span>
    <span>TFL est. <span class="stat-val" id="tfl-sub-stats-high">{total_sub_high_fmt}</span></span>
  </div>
  <div id="tfl-sub-tooltip"><div class="tt-name"></div><div class="tt-val"></div><div class="tt-type"></div></div>
  <div id="tfl-sub-collector">
    <div class="dc-title" onclick="this.parentElement.classList.toggle('collapsed')">
      <span>&#x1F4CD; Collected Addresses <span id="tfl-sub-addr-count">0</span></span>
      <span class="dc-toggle">&#9650;</span>
    </div>
    <div class="dc-body">
      <div id="tfl-sub-addr-list"><div class="dc-empty">Click on the map, use Search, or draw an area to collect addresses.</div></div>
      <div class="dc-actions">
        <div class="dc-btn" id="tfl-sub-send-forensics-btn">&#x1F50D; Send to Forensics</div>
        <div class="dc-btn" id="tfl-sub-send-batch-btn">&#x1F4E5; Send to Batch</div>
      </div>
      <div class="dc-actions" style="margin-top:4px;">
        <div class="dc-btn clear" id="tfl-sub-clear-btn">Clear</div>
      </div>
    </div>
  </div>
</div>
<script src="https://js.arcgis.com/4.30/"></script>
<script>
  const tflPoints = {payload_json};
  const baseMapId = {basemap_safe};
  const typeColors = {type_colors_json};
  const legendItems = {legend_items_json};

  /* Track hidden types for interactive legend filtering */
  const hiddenTypes = new Set();
  function broadcastAddressPayload(payload) {{
    try {{
      window.parent.postMessage(payload, "*");
    }} catch (_) {{}}
    try {{
      if (window.top && window.top !== window.parent) {{
        window.top.postMessage(payload, "*");
      }}
    }} catch (_) {{}}
    try {{
      const frames = window.parent.frames;
      for (let i = 0; i < frames.length; i++) {{
        try {{ frames[i].postMessage(payload, "*"); }} catch(_) {{}}
      }}
    }} catch (_) {{}}
  }}

  /* Build interactive legend */
  let filterCallback = null;
  (function() {{
    const body = document.getElementById("tfl-sub-legend-body");
    if (!body || legendItems.length === 0) return;
    let h = "";
    for (const e of legendItems) {{
      h += '<div class="leg-row" data-type="' + e.type + '"><div class="leg-left"><span class="leg-chip" style="background:' + e.color + ';"></span><span class="leg-label">' + e.type + '</span></div><span class="leg-count">' + e.count + '</span></div>';
    }}
    body.innerHTML = h;
    body.querySelectorAll(".leg-row").forEach(row => {{
      row.addEventListener("click", () => {{
        const t = row.getAttribute("data-type");
        if (hiddenTypes.has(t)) {{ hiddenTypes.delete(t); row.classList.remove("dimmed"); }}
        else {{ hiddenTypes.add(t); row.classList.add("dimmed"); }}
        if (filterCallback) filterCallback();
      }});
    }});
  }})();

  require([
    "esri/Map",
    "esri/views/MapView",
    "esri/layers/FeatureLayer",
    "esri/layers/GraphicsLayer",
    "esri/Graphic",
    "esri/widgets/Home",
    "esri/widgets/ScaleBar",
    "esri/widgets/BasemapToggle",
    "esri/widgets/Compass",
    "esri/widgets/Fullscreen",
    "esri/widgets/Search",
    "esri/widgets/Locate",
    "esri/widgets/Sketch",
    "esri/widgets/Expand",
    "esri/geometry/geometryEngine"
  ], function(Map, MapView, FeatureLayer, GraphicsLayer, Graphic, Home, ScaleBar, BasemapToggle, Compass, Fullscreen, Search, Locate, Sketch, Expand, geometryEngine) {{
    const map = new Map({{ basemap: baseMapId }});

    /* -- Address collector state -- */
    const collectedAddresses = [];
    const pinsLayer = new GraphicsLayer();
    const geocodeUrl = "{ARCGIS_GEOCODER_URL}";

    const districtLayer = new FeatureLayer({{
      url: "{TEA_ARCGIS_SCHOOL_DISTRICT_LAYER_URL}",
      outFields: ["FID", "NAME20", "DISTRICT"],
      popupEnabled: false, labelsVisible: false,
      labelingInfo: [{{
        labelExpressionInfo: {{ expression: "$feature.NAME20" }},
        symbol: {{ type: "text", color: [73, 112, 150, 0.65], haloColor: [13, 23, 36, 0.80], haloSize: 0.6,
          font: {{ size: 8, family: "Avenir Next LT Pro", weight: "normal" }} }}
      }}],
      renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [0,0,0,0], outline: {{ color: [73, 112, 150, 0.25], width: 0.6 }} }} }},
      opacity: 0.40
    }});

    const countyLayer = new FeatureLayer({{
      url: "{TEA_ARCGIS_COUNTY_LAYER_URL}",
      outFields: ["FENAME", "FIPS"],
      popupEnabled: false, labelsVisible: false,
      labelingInfo: [{{
        labelExpressionInfo: {{ expression: "$feature.FENAME + ' County'" }},
        symbol: {{ type: "text", color: [160, 140, 110, 0.80], haloColor: [13, 23, 36, 0.82], haloSize: 0.9,
          font: {{ size: 12, family: "Avenir Next LT Pro", weight: "600" }} }}
      }}],
      renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [0,0,0,0], outline: {{ color: [145, 111, 63, 0.22], width: 0.6 }} }} }},
      opacity: 0.35
    }});

    const cityLayer = new FeatureLayer({{
      url: "{CENSUS_ARCGIS_TEXAS_CITY_LAYER_URL}",
      outFields: ["NAME", "BASENAME", "GEOID", "STATE"],
      definitionExpression: "STATE = '48'",
      popupEnabled: false, labelsVisible: false,
      labelingInfo: [{{
        labelExpressionInfo: {{ expression: "DefaultValue($feature.BASENAME, $feature.NAME)" }},
        symbol: {{ type: "text", color: [165, 100, 105, 0.80], haloColor: [13, 23, 36, 0.76], haloSize: 0.7,
          font: {{ size: 9, family: "Avenir Next LT Pro", weight: "500" }} }}
      }}],
      renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [0,0,0,0], outline: {{ color: [158, 42, 43, 0.12], width: 0.4 }} }} }},
      opacity: 0.20
    }});

    /* TX House & Senate district boundaries */
    const houseLayer = new FeatureLayer({{
      url: "{TEXAS_HOUSE_DISTRICTS_LAYER_URL}",
      outFields: ["*"], popupEnabled: true,
      popupTemplate: {{ title: "TX House District {{{{DISTRICT}}}}", content: "Texas House of Representatives District {{{{DISTRICT}}}}" }},
      labelsVisible: false,
      labelingInfo: [{{ labelExpressionInfo: {{ expression: "'HD ' + $feature.DISTRICT" }},
        symbol: {{ type: "text", color: [90, 180, 130, 0.70], haloColor: [13, 23, 36, 0.75], haloSize: 0.6,
          font: {{ size: 8, family: "Avenir Next LT Pro", weight: "600" }} }} }}],
      renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [40, 180, 100, 0.03], outline: {{ color: [40, 180, 100, 0.30], width: 0.8 }} }} }},
      opacity: 0.30, visible: false
    }});
    const senateLayer = new FeatureLayer({{
      url: "{TEXAS_SENATE_DISTRICTS_LAYER_URL}",
      outFields: ["*"], popupEnabled: true,
      popupTemplate: {{ title: "TX Senate District {{{{DISTRICT}}}}", content: "Texas Senate District {{{{DISTRICT}}}}" }},
      labelsVisible: false,
      labelingInfo: [{{ labelExpressionInfo: {{ expression: "'SD ' + $feature.DISTRICT" }},
        symbol: {{ type: "text", color: [180, 130, 90, 0.70], haloColor: [13, 23, 36, 0.75], haloSize: 0.6,
          font: {{ size: 9, family: "Avenir Next LT Pro", weight: "600" }} }} }}],
      renderer: {{ type: "simple", symbol: {{ type: "simple-fill", color: [200, 140, 60, 0.03], outline: {{ color: [200, 140, 60, 0.30], width: 0.8 }} }} }},
      opacity: 0.30, visible: false
    }});

    map.add(districtLayer);
    map.add(countyLayer);
    map.add(houseLayer);
    map.add(senateLayer);
    map.add(cityLayer);

    const graphics = new GraphicsLayer();
    const sketchLayer = new GraphicsLayer();
    map.add(graphics);
    map.add(sketchLayer);
    map.add(pinsLayer);

    const view = new MapView({{
      container: "tfl-subdivision-map",
      map,
      center: [-99.3, 31.1],
      zoom: 5,
      constraints: {{ minZoom: 5 }},
      popup: {{ dockEnabled: true, dockOptions: {{ position: "bottom-right", breakpoint: false }} }},
      ui: {{ padding: {{ top: 10, right: 10, bottom: 30, left: 10 }} }}
    }});

    /* -- Address collector helper functions -- */
    function updateCollectorUI() {{
      const listEl = document.getElementById("tfl-sub-addr-list");
      const countEl = document.getElementById("tfl-sub-addr-count");
      if (countEl) countEl.textContent = collectedAddresses.length;
      if (!listEl) return;
      if (collectedAddresses.length === 0) {{
        listEl.innerHTML = '<div class="dc-empty">Click on the map, use Search, or draw an area to collect addresses.</div>';
        return;
      }}
      listEl.innerHTML = collectedAddresses.map((a, i) =>
        '<div class="dc-item">'
        + '<div class="dc-num">' + (i + 1) + '</div>'
        + '<div style="flex:1;min-width:0;"><div class="dc-addr">' + (a.address || "Unknown") + '</div>'
        + '<div class="dc-coord">' + Number(a.lat).toFixed(5) + '\u00b0 N, ' + Math.abs(a.lon).toFixed(5) + '\u00b0 W</div>'
        + '</div>'
        + '<div class="dc-del" onclick="window._tflRemoveAddr(' + i + ')" title="Remove">\u00d7</div>'
        + '</div>'
      ).join("");
    }}

    /* Toast notification system */
    function showToast(message, type) {{
      type = type || "info";
      const icons = {{ success: "\u2713", info: "\u2139\uFE0F", warn: "\u26A0\uFE0F" }};
      const container = document.getElementById("tfl-sub-toast-container");
      if (!container) return;
      const toast = document.createElement("div");
      toast.className = "tfl-toast " + type;
      toast.innerHTML = '<span class="toast-icon">' + (icons[type] || "") + '</span><span>' + message + '</span>';
      container.appendChild(toast);
      setTimeout(() => {{ toast.classList.add("out"); setTimeout(() => toast.remove(), 320); }}, 2600);
    }}

    /* Remove individual address by index */
    function removeAddress(idx) {{
      if (idx < 0 || idx >= collectedAddresses.length) return;
      collectedAddresses.splice(idx, 1);
      /* Rebuild pins layer to match */
      pinsLayer.removeAll();
      collectedAddresses.forEach((a) => {{
        pinsLayer.add(new Graphic({{
          geometry: {{ type: "point", longitude: a.lon, latitude: a.lat }},
          symbol: {{ type: "simple-marker", style: "circle", size: 10, color: [30,144,255,0.85], outline: {{ color: [255,255,255,0.80], width: 1.5 }} }},
          attributes: {{ address: a.address, lat: a.lat, lon: a.lon }},
          popupTemplate: {{ title: a.address || "Point", content: "Lat: " + a.lat.toFixed(5) + ", Lon: " + a.lon.toFixed(5) }}
        }}));
      }});
      updateCollectorUI();
      showToast("Address removed", "info");
    }}
    /* Expose removeAddress globally for inline onclick */
    window._tflRemoveAddr = removeAddress;

    function addAddress(address, lat, lon) {{
      const exists = collectedAddresses.some(a =>
        Math.abs(a.lat - lat) < 0.0001 && Math.abs(a.lon - lon) < 0.0001
      );
      if (exists) {{
        showToast("Address already collected", "warn");
        return;
      }}
      collectedAddresses.push({{ address, lat, lon }});
      pinsLayer.add(new Graphic({{
        geometry: {{ type: "point", longitude: lon, latitude: lat }},
        symbol: {{ type: "simple-marker", style: "circle", size: 12, color: [30,144,255,0.85], outline: {{ color: [255,255,255,0.90], width: 2 }} }},
        attributes: {{ address, lat, lon }},
        popupTemplate: {{ title: address || "Point", content: "Lat: " + lat.toFixed(5) + ", Lon: " + lon.toFixed(5) }}
      }}));
      updateCollectorUI();
      showToast((address || "Point").substring(0, 50) + " added", "success");
      /* Fly to the newly added point */
      view.goTo({{ center: [lon, lat], zoom: Math.max(view.zoom || 10, 12) }}, {{ duration: 700, easing: "ease-in-out" }}).catch(() => {{}});
      try {{
        window.parent.postMessage({{
          type: "tfl-draw-address-found",
          address: address, lat: lat, lon: lon,
          allAddresses: collectedAddresses.slice()
        }}, "*");
      }} catch(e) {{}}
    }}

    function reverseGeocode(lat, lon) {{
      const url = geocodeUrl.replace("findAddressCandidates", "reverseGeocode")
        + "?location=" + lon + "," + lat
        + "&outSR=4326&langCode=en&f=json";
      fetch(url).then(r => r.json()).then(data => {{
        const addr = (data.address && data.address.LongLabel) || (data.address && data.address.ShortLabel) || ("Point: " + lat.toFixed(5) + ", " + lon.toFixed(5));
        addAddress(addr, lat, lon);
      }}).catch(() => {{
        addAddress("Point: " + lat.toFixed(5) + ", " + lon.toFixed(5), lat, lon);
      }});
    }}

    const formatUsd = (v) => Number(v||0).toLocaleString("en-US",{{style:"currency",currency:"USD",maximumFractionDigits:0}});
    const maxHigh = tflPoints.reduce((a, r) => Math.max(a, Number(r.high_total || 0)), 0);
    const sz = (v) => {{
      if (maxHigh <= 0) return 9;
      const n = Math.max(0, Number(v || 0));
      return Math.max(8, Math.min(30, 8 + (Math.log10(n+1)/Math.log10(maxHigh+1))*22));
    }};

    for (const row of tflPoints) {{
      const clientsHtml = row.match_clients_preview || "";
      const extraHtml = row.extra_count > 0 ? `, +${{row.extra_count}} more` : "";
      const g = new Graphic({{
        geometry: {{ type: "point", longitude: row.lon, latitude: row.lat }},
        symbol: {{
          type: "simple-marker", size: sz(row.high_total),
          color: typeColors[row.subdivision_type] || [113, 129, 145, 0.9],
          outline: {{ color: [255, 255, 255, 0.70], width: 1 }}
        }},
        attributes: row,
        popupTemplate: {{
          title: row.subdivision_name || "Political Subdivision",
          content: `<div style="font-family:'Avenir Next LT Pro',system-ui,sans-serif;font-size:12px;line-height:1.5;">
            <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;padding-bottom:6px;border-bottom:1px solid rgba(140,160,180,0.15);">
              <div style="width:10px;height:10px;border-radius:50%;background:${{(Object.values(typeColors).find((_, idx) => Object.keys(typeColors)[idx] === row.subdivision_type) || [113,129,145]).slice(0,3).map(v => typeof v === 'number' ? v : 113).join(',') }};flex-shrink:0;border:1px solid rgba(255,255,255,0.15);"></div>
              <span style="font-size:10px;color:rgba(150,175,200,0.70);text-transform:uppercase;letter-spacing:0.08em;font-weight:600;">${{row.subdivision_type}}</span>
            </div>
            <table style="border-collapse:collapse;width:100%;">
              <tr><td style="color:#7a94ab;padding:3px 8px 3px 0;font-size:11px;">Code</td><td style="font-weight:500;">${{row.subdivision_code || "N/A"}}</td></tr>
              <tr><td style="color:#7a94ab;padding:3px 8px 3px 0;font-size:11px;">TFL high est.</td><td style="font-weight:700;color:rgba(100,200,255,0.95);">${{formatUsd(row.high_total)}}</td></tr>
              <tr><td style="color:#7a94ab;padding:3px 8px 3px 0;font-size:11px;">Source</td><td>${{row.source_name || "N/A"}}</td></tr>
              <tr><td style="color:#7a94ab;padding:3px 8px 3px 0;font-size:11px;">Matched clients</td><td style="font-weight:600;">${{row.match_count}}</td></tr>
            </table>
            <div style="margin-top:6px;padding-top:5px;border-top:1px solid rgba(140,160,180,0.12);font-size:10.5px;line-height:1.6;color:rgba(200,215,230,0.80);">${{clientsHtml}}${{extraHtml}}</div>
          </div>`
        }}
      }});
      graphics.add(g);
    }}

    /* Interactive legend filtering callback */
    filterCallback = () => {{
      let visCount = 0, visHigh = 0;
      graphics.graphics.forEach(g => {{
        if (g.attributes && g.attributes.subdivision_type) {{
          const vis = !hiddenTypes.has(g.attributes.subdivision_type);
          g.visible = vis;
          if (vis) {{ visCount++; visHigh += Number(g.attributes.high_total || 0); }}
        }}
      }});
      /* Update stats ribbon */
      const cEl = document.getElementById("tfl-sub-stats-count");
      const hEl = document.getElementById("tfl-sub-stats-high");
      if (cEl) cEl.textContent = visCount.toLocaleString();
      if (hEl) hEl.textContent = formatUsd(visHigh);
    }};

    const updateLabels = () => {{
      const z = Number(view.zoom || 0);
      countyLayer.labelsVisible = z >= 5;
      cityLayer.labelsVisible = z >= 6.2;
      districtLayer.labelsVisible = z >= 8.5;
      houseLayer.labelsVisible = z >= 8;
      senateLayer.labelsVisible = z >= 7;
    }};
    view.watch("zoom", updateLabels);

    /* -- Widgets -- */
    const home = new Home({{ view }});
    const basemapToggle = new BasemapToggle({{ view, nextBasemap: baseMapId === "hybrid" ? "gray-vector" : "hybrid" }});
    const scaleBar = new ScaleBar({{ view, unit: "dual" }});
    const compass = new Compass({{ view }});
    const fullscreen = new Fullscreen({{ view }});
    const locate = new Locate({{ view }});
    const search = new Search({{
      view, popupEnabled: true, resultGraphicEnabled: true,
      goToOverride: (view, opts) => view.goTo(opts.target, {{ duration: 800, easing: "ease-in-out" }})
    }});
    search.on("select-result", (evt) => {{
      if (evt.result && evt.result.feature && evt.result.feature.geometry) {{
        const geom = evt.result.feature.geometry;
        addAddress(evt.result.name || "", geom.latitude, geom.longitude);
      }}
    }});

    /* Sketch tool for encircling areas */
    const sketch = new Sketch({{
      view, layer: sketchLayer, creationMode: "single",
      availableCreateTools: ["polygon", "circle", "rectangle"],
      defaultCreateOptions: {{ mode: "freehand" }},
      visibleElements: {{ selectionTools: {{ "lasso-selection": false, "rectangle-selection": false }}, settingsMenu: false, undoRedoMenu: true }},
      defaultUpdateOptions: {{ tool: "reshape" }}
    }});
    const sketchExpand = new Expand({{
      view, content: sketch, expandIconClass: "esri-icon-polygon",
      expandTooltip: "Draw area for batch analysis", group: "tools"
    }});

    /* House/Senate layer toggle */
    const layerDiv = document.createElement("div");
    layerDiv.style.cssText = "background:rgba(13,23,36,0.94);border-radius:8px;padding:10px;font-family:'Avenir Next LT Pro',system-ui,sans-serif;font-size:12px;color:rgba(210,225,240,0.90);min-width:160px;";
    layerDiv.innerHTML = '<div style="font-size:9px;text-transform:uppercase;letter-spacing:0.12em;color:rgba(150,175,200,0.65);font-weight:700;margin-bottom:6px;">Legislative Districts</div>'
      + '<label style="display:flex;align-items:center;gap:6px;cursor:pointer;padding:3px 0;"><input type="checkbox" id="tfl-sub-toggle-house" style="accent-color:#28b464;"><span>TX House Districts</span></label>'
      + '<label style="display:flex;align-items:center;gap:6px;cursor:pointer;padding:3px 0;"><input type="checkbox" id="tfl-sub-toggle-senate" style="accent-color:#c88c3c;"><span>TX Senate Districts</span></label>';
    const layerExpand = new Expand({{
      view, content: layerDiv, expandIconClass: "esri-icon-layer-list",
      expandTooltip: "Toggle legislative districts", group: "tools"
    }});

    view.ui.add(home, "top-left");
    view.ui.add(compass, "top-left");
    view.ui.add(fullscreen, "top-left");
    view.ui.add(locate, "top-left");
    view.ui.add(sketchExpand, "top-left");
    view.ui.add(layerExpand, "top-left");
    view.ui.add(search, "top-right");
    view.ui.add(basemapToggle, "top-right");
    view.ui.add(scaleBar, "bottom-left");

    /* Wire layer toggles */
    view.when(() => {{
      const hBox = document.getElementById("tfl-sub-toggle-house");
      const sBox = document.getElementById("tfl-sub-toggle-senate");
      if (hBox) hBox.addEventListener("change", () => {{ houseLayer.visible = hBox.checked; }});
      if (sBox) sBox.addEventListener("change", () => {{ senateLayer.visible = sBox.checked; }});
    }});

    /* Sketch complete → batch analysis info + address scanning */
    sketch.on("create", (evt) => {{
      if (evt.state !== "complete") return;
      const drawn = evt.graphic.geometry;
      const selInfo = document.getElementById("tfl-sub-sel-info");
      const contained = [];
      graphics.graphics.forEach((g) => {{
        if (g.visible !== false && g.geometry && geometryEngine.contains(drawn, g.geometry)) contained.push(g.attributes || {{}});
      }});
      if (selInfo && contained.length > 0) {{
        const total = contained.reduce((a,r) => a + Number(r.high_total||0), 0);
        const types = [...new Set(contained.map(r => r.subdivision_type).filter(Boolean))];
        selInfo.style.display = "block";
        selInfo.innerHTML = '<div class="sel-title">Area Selection</div>'
          + '<div><strong>' + contained.length + '</strong> subdivision(s) in area</div>'
          + '<div>Combined TFL est.: <strong>' + formatUsd(total) + '</strong></div>'
          + '<div style="font-size:10px;color:rgba(180,200,220,0.65);margin-top:3px;">' + types.join(", ") + '</div>';
        try {{ window.parent.postMessage({{ type: "tfl-map-area-select", count: contained.length, totalHigh: total, types: types }}, "*"); }} catch(e) {{}}
      }} else if (selInfo) {{
        selInfo.style.display = "block";
        selInfo.innerHTML = '<div class="sel-title">Area Selection</div><div>No subdivisions in drawn area.</div>';
        setTimeout(() => {{ selInfo.style.display = "none"; }}, 3000);
      }}

      /* Reverse-geocode sampled points within the drawn area */
      const ext = drawn.extent;
      if (ext) {{
        const cx = ext.center.longitude, cy = ext.center.latitude;
        const dx = (ext.xmax - ext.xmin), dy = (ext.ymax - ext.ymin);
        const SAMPLES = 5;
        reverseGeocode(cy, cx);
        for (let xi = 0; xi < SAMPLES; xi++) {{
          for (let yi = 0; yi < SAMPLES; yi++) {{
            const px = ext.xmin + (dx * (xi + 0.5) / SAMPLES);
            const py = ext.ymin + (dy * (yi + 0.5) / SAMPLES);
            const testPt = {{ type: "point", longitude: px, latitude: py, spatialReference: {{ wkid: 4326 }} }};
            if (geometryEngine.contains(drawn, testPt)) {{
              reverseGeocode(py, px);
            }}
          }}
        }}
        const badge = document.getElementById("tfl-sub-badge");
        if (badge) badge.textContent = "Area scanned \u2014 see collected addresses \u2192";
        setTimeout(() => {{
          try {{ window.parent.postMessage({{ type: "tfl-draw-area-addresses", allAddresses: collectedAddresses.slice() }}, "*"); }} catch(e) {{}}
        }}, 3000);
      }}
    }});

    /* Click → reverse geocode for address collection */
    view.on("click", (evt) => {{
      if (evt.mapPoint) {{
        reverseGeocode(evt.mapPoint.latitude, evt.mapPoint.longitude);
      }}
    }});

    /* Coordinate readout */
    view.on("pointer-move", (evt) => {{
      const pt = view.toMap(evt);
      const el = document.getElementById("tfl-sub-coord");
      if (pt && el) el.textContent = pt.latitude.toFixed(5) + "\u00b0 N, " + Math.abs(pt.longitude).toFixed(5) + "\u00b0 W";
    }});

    /* Hover highlight + tooltip */
    let hoverHL = null;
    view.on("pointer-move", (evt) => {{
      const tooltip = document.getElementById("tfl-sub-tooltip");
      view.hitTest(evt, {{ include: [graphics] }}).then((r) => {{
        const hit = r.results && r.results.find(x => x.graphic && x.graphic.attributes && x.graphic.attributes.subdivision_name);
        document.getElementById("tfl-subdivision-map").style.cursor = hit ? "pointer" : "default";
        if (hoverHL) {{ graphics.remove(hoverHL); hoverHL = null; }}
        if (hit && hit.graphic && hit.graphic.geometry) {{
          hoverHL = new Graphic({{
            geometry: hit.graphic.geometry,
            symbol: {{ type: "simple-marker", style: "circle", size: 30, color: [255,255,255,0.0], outline: {{ color: [255,255,255,0.55], width: 2 }} }}
          }});
          graphics.add(hoverHL);
          /* Show tooltip */
          if (tooltip) {{
            const a = hit.graphic.attributes;
            tooltip.querySelector(".tt-name").textContent = a.subdivision_name || "";
            tooltip.querySelector(".tt-val").textContent = formatUsd(a.high_total);
            tooltip.querySelector(".tt-type").textContent = a.subdivision_type || "";
            tooltip.style.display = "block";
            tooltip.style.left = (evt.x + 14) + "px";
            tooltip.style.top = (evt.y - 10) + "px";
          }}
        }} else {{
          if (tooltip) tooltip.style.display = "none";
        }}
      }});
    }});

    /* -- Send to Forensics (first address) -- */
    document.getElementById("tfl-sub-send-forensics-btn").addEventListener("click", () => {{
      if (collectedAddresses.length === 0) {{ showToast("No addresses collected yet", "warn"); return; }}
      const payload = {{
        type: "tfl-send-address",
        action: "forensics",
        address: collectedAddresses[0].address || "",
        addresses: collectedAddresses.map(a => a.address),
        nonce: Date.now()
      }};
      broadcastAddressPayload(payload);
      showToast("Sent to Address Forensics", "success");
    }});

    /* -- Send All to Batch -- */
    document.getElementById("tfl-sub-send-batch-btn").addEventListener("click", () => {{
      if (collectedAddresses.length === 0) {{ showToast("No addresses collected yet", "warn"); return; }}
      const payload = {{
        type: "tfl-send-address",
        action: "batch",
        address: collectedAddresses[0].address || "",
        addresses: collectedAddresses.map(a => a.address),
        nonce: Date.now()
      }};
      broadcastAddressPayload(payload);
      showToast(collectedAddresses.length + " address(es) sent to Batch", "success");
    }});

    /* Clear button */
    document.getElementById("tfl-sub-clear-btn").addEventListener("click", () => {{
      const count = collectedAddresses.length;
      collectedAddresses.length = 0;
      pinsLayer.removeAll();
      sketchLayer.removeAll();
      updateCollectorUI();
      const badge = document.getElementById("tfl-sub-badge");
      if (badge) badge.textContent = "Click map or search to collect addresses";
      const selInfo = document.getElementById("tfl-sub-sel-info");
      if (selInfo) selInfo.style.display = "none";
      if (count > 0) showToast(count + " address(es) cleared", "info");
    }});

    view.when(() => {{
      const loader = document.getElementById("tfl-sub-loading");
      const ldLabel = loader && loader.querySelector(".ld-label");
      if (ldLabel) ldLabel.textContent = "Rendering " + tflPoints.length + " subdivisions\u2026";
      setTimeout(() => {{
        if (ldLabel) ldLabel.textContent = "Almost ready\u2026";
      }}, 600);
      setTimeout(() => {{
        if (loader) {{ loader.style.opacity = "0"; setTimeout(() => loader.remove(), 600); }}
      }}, 900);
      updateLabels();
      if (graphics.graphics.length > 0) {{
        view.goTo(graphics.graphics.toArray(), {{ padding: {{ top: 50, right: 50, bottom: 50, left: 50 }}, duration: 1000, easing: "ease-in-out" }}).catch(() => {{}});
      }}
    }});
  }});
</script>
"""
    )
    _persistent_html_frame(
        html=arcgis_html,
        signature=subdivision_map_signature,
        height=int(height) + 8,
        key="mp5_tfl_subdivision_map_v3",
        default=None,
    )

PRIMARY_PATTERNS = [
    (r"\bmetropolitan transit authority\b", "Transit Authority"),
    (r"\bregional mobility authority\b", "Regional Mobility Authority"),
    (r"\bwater control ?&? improvement district\b", "Water Control & Improvement District"),
    (r"\bmunicipal utility district\b", "Municipal Utility District"),
    (r"\bmud\b", "Municipal Utility District"),
    (r"\bgroundwater conservation district\b", "Groundwater Conservation District"),
    (r"\bhospital district\b", "Hospital District"),
    (r"\bemergency services district\b", "Emergency Services District"),
    (r"\bappraisal district\b", "Appraisal District"),
    (r"\bhousing authority\b", "Housing Authority"),
    (r"\btransit authority\b", "Transit Authority"),
    (r"\bthe league\b|\bleague\b", "League"),
    (r"\briver authority\b", "River Authority"),
    (r"\bnavigation district\b", "Navigation District"),
    (r"\bport authority\b", "Port Authority"),
    (r"\bdrainage district\b", "Drainage District"),
    (r"\b(independent )?school district\b", "Independent School District"),
    (r"\bschool district\b", "Independent School District"),
    (r"(^|\s)isd($|\s|[^a-z])", "Independent School District"),
    (r"\bwcid\b", "Water Control & Improvement District"),
    (r"\bpublic improvement district\b", "Public Improvement District"),
    (r"(^|\s)pid($|\s|[^a-z])", "Public Improvement District"),
    (r"\bmunicipal corporation\b|\blocal government corporation\b|\bcorporation\b", "Local Government Corporation"),
    (r"\bcoalition\b", "Coalition"),
    (r"\bassociation\b", "Association"),
    (r"\bcommittee\b", "Committee"),
    (r"\bfoundation\b", "Foundation"),
    (r"\bcollege\b", "College"),
    (r"\bboard\b", "Board"),
    (r"\bauthority\b", "Authority"),
    (r"\bdistrict\b", "District"),
    (r"\bcity\b", "City"),
    (r"\bcounty\b", "County"),
]

COARSE_CATEGORY = {
    "Independent School District": "Public School Districts",
    "Transit Authority": "Special Districts and Other Authorities",
    "Regional Mobility Authority": "Special Districts and Other Authorities",
    "Water Control & Improvement District": "Special Districts and Other Authorities",
    "Municipal Utility District": "Special Districts and Other Authorities",
    "Groundwater Conservation District": "Special Districts and Other Authorities",
    "Hospital District": "Special Districts and Other Authorities",
    "Emergency Services District": "Special Districts and Other Authorities",
    "Appraisal District": "Special Districts and Other Authorities",
    "Housing Authority": "Special Districts and Other Authorities",
    "River Authority": "Special Districts and Other Authorities",
    "Navigation District": "Special Districts and Other Authorities",
    "Port Authority": "Special Districts and Other Authorities",
    "Drainage District": "Special Districts and Other Authorities",
    "Public Improvement District": "Special Districts and Other Authorities",
    "Local Government Corporation": "Special Districts and Other Authorities",
    "Authority": "Special Districts and Other Authorities",
    "District": "Special Districts and Other Authorities",
    "City": "Cities, Towns, Villages",
    "County": "County",
    "College": "Community and Junior Colleges",
    "Coalition": "Associations",
    "Association": "Associations",
    "Foundation": "Associations",
    "Committee": "Associations",
    "Board": "Associations",
    "League": "Associations",
}

def normalize_entity_name(name: str) -> str:
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return ""
    s = str(name).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s

def match_entity_type(name: str) -> tuple[str, str]:
    s = normalize_entity_name(name)
    for pattern, canonical in PRIMARY_PATTERNS:
        if re.search(pattern, s, flags=re.IGNORECASE):
            coarse = COARSE_CATEGORY.get(canonical, None)
            if canonical == "Independent School District":
                coarse = "Public School Districts"
            if canonical in ("City",):
                coarse = "Cities, Towns, Villages"
            if canonical in ("County",):
                coarse = "County"
            if not coarse:
                coarse = COARSE_CATEGORY.get(canonical, "Special Districts and Other Authorities")
            return canonical, coarse

    if re.search(r"\bschool\b", s):
        return "Independent School District", "Public School Districts"

    if re.search(r"\bcommunity college\b|\bjunior college\b", s):
        return "College", "Community and Junior Colleges"

    if re.search(r"\bcity\b|\btown\b|\bvillage\b", s):
        return "City", "Cities, Towns, Villages"
    if re.search(r"\bcounty\b", s):
        return "County", "County"
    if re.search(r"\bassociation\b|\bcoalition\b|\bfoundation\b|\bcommittee\b|\bboard\b", s):
        return "Association", "Associations"

    return "Other", "Other"

def filter_filer_rows(
    df: pd.DataFrame,
    session: str | None,
    lobbyshort: str,
    name_to_short: dict,
    lobbyist_norms: set[str],
    filerid_to_short: dict | None,
    filer_ids: set[int] | tuple[int, ...] | None = None,
    loose: bool = False,
) -> pd.DataFrame:
    if df.empty:
        return df

    d = df.copy()
    if session is not None:
        d = d[d["Session"].astype(str).str.strip() == str(session)]
    if d.empty:
        return d

    filerid_map = filerid_to_short or {}
    if "FilerID" in d.columns:
        d["FilerID"] = pd.to_numeric(d["FilerID"], errors="coerce").fillna(-1).astype(int)
    elif "filerIdent" in d.columns:
        d["FilerID"] = pd.to_numeric(d["filerIdent"], errors="coerce").fillna(-1).astype(int)
    else:
        d["FilerID"] = -1

    if filerid_map:
        d["FilerShortFromId"] = d["FilerID"].map(filerid_map)
    else:
        d["FilerShortFromId"] = ""

    filer_name = d.get("filerName", pd.Series([""] * len(d)))
    filer_sort = d.get("filerSort", pd.Series([""] * len(d)))
    if isinstance(filer_name, pd.DataFrame):
        filer_name = filer_name.iloc[:, 0]
    if isinstance(filer_sort, pd.DataFrame):
        filer_sort = filer_sort.iloc[:, 0]
    filer_clean = clean_filer_name_series(filer_name)
    d["FilerNormRaw"] = norm_name_series(filer_name)
    d["FilerNormClean"] = norm_name_series(filer_clean)
    d["FilerSortNorm"] = norm_name_series(filer_sort)

    mapped = d["FilerNormRaw"].map(name_to_short)
    mapped = mapped.where(mapped.notna(), d["FilerNormClean"].map(name_to_short))
    mapped = mapped.where(mapped.notna(), d["FilerSortNorm"].map(name_to_short))
    d["FilerShortMapped"] = mapped

    lobbyshort_norm = norm_name(lobbyshort)
    d["FilerIsShort"] = (
        d["FilerNormClean"].eq(lobbyshort_norm) |
        d["FilerNormRaw"].eq(lobbyshort_norm)
    )

    ok = (
        (d["FilerShortFromId"].astype(str) == str(lobbyshort)) |
        (d["FilerShortMapped"].astype(str) == str(lobbyshort)) |
        (d["FilerNormRaw"].isin(lobbyist_norms) if lobbyist_norms else False) |
        (d["FilerNormClean"].isin(lobbyist_norms) if lobbyist_norms else False) |
        (d["FilerSortNorm"].isin(lobbyist_norms) if lobbyist_norms else False) |
        (d["FilerIsShort"])
    )
    if filer_ids:
        filer_ids_set = set()
        for x in filer_ids:
            try:
                if pd.isna(x):
                    continue
            except Exception:
                pass
            try:
                filer_ids_set.add(int(x))
            except Exception:
                try:
                    filer_ids_set.add(int(float(x)))
                except Exception:
                    continue
        if filer_ids_set:
            filer_match = d["FilerID"].isin(filer_ids_set)
            if filer_match.any():
                ok = filer_match
    if loose and not ok.any():
        loose_ok = pd.Series(False, index=d.index)

        if lobbyshort_norm and len(lobbyshort_norm) >= 4:
            loose_ok |= (
                d["FilerNormRaw"].str.contains(lobbyshort_norm, na=False) |
                d["FilerNormClean"].str.contains(lobbyshort_norm, na=False) |
                d["FilerSortNorm"].str.contains(lobbyshort_norm, na=False)
            )

        if lobbyist_norms:
            for n in lobbyist_norms:
                if n and len(n) >= 4:
                    loose_ok |= (
                        d["FilerNormRaw"].str.contains(n, na=False) |
                        d["FilerNormClean"].str.contains(n, na=False) |
                        d["FilerSortNorm"].str.contains(n, na=False)
                    )

        target_last = last_name_norm_from_text(lobbyshort)
        if target_last:
            last_raw = last_name_norm_series(filer_name)
            last_sort = last_name_norm_series(filer_sort)
            loose_ok |= last_raw.eq(target_last) | last_sort.eq(target_last)

        target_init = _last_first_initial_key(lobbyshort)
        if target_init:
            init_raw = filer_name.fillna("").astype(str).map(_last_first_initial_key)
            init_sort = filer_sort.fillna("").astype(str).map(_last_first_initial_key)
            loose_ok |= init_raw.eq(target_init) | init_sort.eq(target_init)

        ok = loose_ok

    return d.loc[ok]

def filter_filer_rows_multi(
    df: pd.DataFrame,
    session: str | None,
    lobbyshorts: list[str],
    name_to_short: dict,
    lobbyist_norms: set[str],
    filerid_to_short: dict | None,
    loose: bool = False,
) -> pd.DataFrame:
    if df.empty or not lobbyshorts:
        return df.iloc[0:0]

    lobbyshorts_set = {str(s).strip() for s in lobbyshorts if str(s).strip()}
    if not lobbyshorts_set:
        return df.iloc[0:0]

    d = df.copy()
    if session is not None:
        d = d[d["Session"].astype(str).str.strip() == str(session)]
    if d.empty:
        return d

    lobbyshort_norms = {norm_name(s) for s in lobbyshorts_set if s}
    norm_to_short = {norm_name(s): s for s in lobbyshorts_set if s}
    filerid_map = filerid_to_short or {}

    if "filerIdent" in d.columns and filerid_map:
        d["FilerID"] = pd.to_numeric(d["filerIdent"], errors="coerce").fillna(-1).astype(int)
        d["FilerShortFromId"] = d["FilerID"].map(filerid_map)
    else:
        d["FilerShortFromId"] = ""

    filer_name = d.get("filerName", pd.Series([""] * len(d)))
    filer_sort = d.get("filerSort", pd.Series([""] * len(d)))
    if isinstance(filer_name, pd.DataFrame):
        filer_name = filer_name.iloc[:, 0]
    if isinstance(filer_sort, pd.DataFrame):
        filer_sort = filer_sort.iloc[:, 0]
    filer_clean = clean_filer_name_series(filer_name)

    d["FilerNormRaw"] = norm_name_series(filer_name)
    d["FilerNormClean"] = norm_name_series(filer_clean)
    d["FilerSortNorm"] = norm_name_series(filer_sort)

    mapped = d["FilerNormRaw"].map(name_to_short)
    mapped = mapped.where(mapped.notna(), d["FilerNormClean"].map(name_to_short))
    mapped = mapped.where(mapped.notna(), d["FilerSortNorm"].map(name_to_short))
    d["FilerShortMapped"] = mapped

    d["FilerIsShort"] = (
        d["FilerNormClean"].isin(lobbyshort_norms) |
        d["FilerNormRaw"].isin(lobbyshort_norms)
    )

    ok = (
        d["FilerShortFromId"].astype(str).isin(lobbyshorts_set) |
        d["FilerShortMapped"].astype(str).isin(lobbyshorts_set) |
        (d["FilerNormRaw"].isin(lobbyist_norms) if lobbyist_norms else False) |
        (d["FilerNormClean"].isin(lobbyist_norms) if lobbyist_norms else False) |
        (d["FilerSortNorm"].isin(lobbyist_norms) if lobbyist_norms else False) |
        d["FilerIsShort"]
    )

    if loose and not ok.any():
        patterns = [re.escape(n) for n in list(lobbyshort_norms) + list(lobbyist_norms) if n and len(n) >= 4]
        if patterns:
            pat = "|".join(patterns)
            loose_ok = (
                d["FilerNormRaw"].str.contains(pat, na=False) |
                d["FilerNormClean"].str.contains(pat, na=False) |
                d["FilerSortNorm"].str.contains(pat, na=False)
            )
            ok = loose_ok

    d = d.loc[ok]
    if d.empty:
        return d

    matched = d["FilerShortFromId"].where(d["FilerShortFromId"].astype(str).isin(lobbyshorts_set), "")
    mapped_short = d["FilerShortMapped"].where(d["FilerShortMapped"].astype(str).isin(lobbyshorts_set), "")
    matched = matched.where(matched.astype(str).str.strip() != "", mapped_short)
    norm_short = d["FilerNormClean"].map(norm_to_short)
    norm_short = norm_short.where(norm_short.notna(), d["FilerNormRaw"].map(norm_to_short))
    matched = matched.where(matched.astype(str).str.strip() != "", norm_short)
    d["MatchedLobbyShort"] = matched.fillna("")
    return d

def last_name_norm_from_text(text: str) -> str:
    if not text:
        return ""
    s = str(text).replace("\u00A0", " ").strip()
    if not s:
        return ""
    if "," in s:
        last = s.split(",", 1)[0].strip()
    else:
        parts = s.split()
        last = parts[-1] if parts else ""
    return norm_name(last)

def last_name_norm_series(s: pd.Series) -> pd.Series:
    if isinstance(s, pd.DataFrame):
        s = s.iloc[:, 0] if s.shape[1] > 0 else pd.Series([], dtype="string")
    if not isinstance(s, pd.Series):
        s = pd.Series(s)
    s = (
        s.fillna("")
         .astype("string")
         .str.replace("\u00A0", " ", regex=False)
         .str.strip()
    )
    comma_mask = s.str.contains(",", na=False)
    last_from_comma = (
        s.where(comma_mask, "")
         .astype("string")
         .str.split(",", n=1)
         .str[0]
         .astype("string")
         .str.strip()
    )
    last_from_space = (
        s.where(~comma_mask, "")
         .astype("string")
         .str.split()
         .str[-1]
         .fillna("")
         .astype("string")
         .str.strip()
    )
    last = last_from_comma.where(comma_mask, last_from_space).fillna("")
    return norm_name_series(last)

def first_name_norm_series(s: pd.Series) -> pd.Series:
    if isinstance(s, pd.DataFrame):
        s = s.iloc[:, 0] if s.shape[1] > 0 else pd.Series([], dtype="string")
    if not isinstance(s, pd.Series):
        s = pd.Series(s)
    s = (
        s.fillna("")
         .astype("string")
         .str.replace("\u00A0", " ", regex=False)
         .str.strip()
    )
    comma_mask = s.str.contains(",", na=False)
    first_from_comma = (
        s.where(comma_mask, "")
         .astype("string")
         .str.split(",", n=1)
         .str[1]
         .fillna("")
         .astype("string")
         .str.strip()
         .str.split()
         .str[0]
         .fillna("")
         .astype("string")
         .str.strip()
    )
    first_from_space = (
        s.where(~comma_mask, "")
         .astype("string")
         .str.split()
         .str[0]
         .fillna("")
         .astype("string")
         .str.strip()
    )
    first = first_from_comma.where(comma_mask, first_from_space).fillna("")
    return norm_name_series(first)

def _last_first_initial_key(name: str) -> str:
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return ""
    s = str(name).replace("\u00A0", " ").strip()
    if not s:
        return ""
    if "," in s:
        last, rest = [p.strip() for p in s.split(",", 1)]
        first = rest
    else:
        toks = s.split()
        if len(toks) < 2:
            return ""
        first, last = toks[0], toks[-1]
    initial = ""
    for ch in first:
        if ch.isalnum():
            initial = ch
            break
    if not last or not initial:
        return ""
    return norm_name(f"{last} {initial}")

def norm_person_variants(user_text: str) -> set[str]:
    if not user_text:
        return set()
    t = clean_person_name(user_text)
    if not t:
        return set()

    if "," in t:
        parts = [p.strip() for p in t.split(",", 1)]
        last = parts[0]
        rest = parts[1] if len(parts) > 1 else ""
        first = rest.split()[0].strip() if rest else ""
    else:
        toks = t.split()
        if len(toks) == 1:
            first, last = "", toks[0]
        else:
            first, last = toks[0], toks[-1]

    variants = {norm_name(t)}
    raw_norm = norm_name(user_text)
    if raw_norm:
        variants.add(raw_norm)
    if first and last:
        variants |= {
            norm_name(f"{first} {last}"),
            norm_name(f"{last}, {first}"),
            norm_name(f"{last} {first}"),
            norm_name(f"{first}{last}"),
            norm_name(f"{last}{first}"),
        }
    return {v for v in variants if v}

def _nickname_variants(first_norm: str) -> set[str]:
    if not first_norm:
        return set()
    variants = {first_norm}
    if first_norm in _NICKNAME_MAP:
        variants |= _NICKNAME_MAP[first_norm]
    for base, nicknames in _NICKNAME_MAP.items():
        if first_norm in nicknames:
            variants.add(base)
            variants |= nicknames
    return {v for v in variants if v}

def norm_person_variants_with_nicknames(user_text: str) -> set[str]:
    variants = norm_person_variants(user_text)
    if not user_text:
        return variants
    t = clean_person_name(user_text)
    if not t:
        return variants

    def _add_nickname_variants(first_val: str, last_val: str) -> None:
        first_norm = norm_name(first_val)
        last_norm = norm_name(last_val)
        if not first_norm or not last_norm:
            return
        for fn in _nickname_variants(first_norm):
            if fn == first_norm:
                continue
            variants.add(f"{fn}{last_norm}")
            variants.add(f"{last_norm}{fn}")

    if "," in t:
        parts = [p.strip() for p in t.split(",", 1)]
        last = parts[0]
        rest = parts[1] if len(parts) > 1 else ""
        first = rest.split()[0].strip() if rest else ""
        _add_nickname_variants(first, last)
    else:
        toks = t.split()
        if len(toks) < 2:
            return variants
        first, last = toks[0], toks[-1]
        _add_nickname_variants(first, last)
        if len(toks) == 2:
            _add_nickname_variants(last, first)
    return {v for v in variants if v}

def person_display(org, last, first) -> str:
    org = "" if pd.isna(org) else str(org).strip()
    last = "" if pd.isna(last) else str(last).strip()
    first = "" if pd.isna(first) else str(first).strip()
    if org:
        return org
    if last and first:
        return f"{last}, {first}"
    return (last or first or "").strip()

def amount_display(exact, low, high, code=None) -> str:
    if pd.notna(exact) and str(exact).strip():
        return str(exact)
    if pd.notna(low) and str(low).strip():
        if pd.notna(high) and str(high).strip():
            return f"{low}--{high}"
        return str(low)
    if pd.notna(code) and str(code).strip():
        return str(code)
    return ""

def ensure_cols(df: pd.DataFrame, cols_with_defaults: dict) -> pd.DataFrame:
    missing = {c: v for c, v in cols_with_defaults.items() if c not in df.columns}
    if not missing:
        return df  # No copy needed when all columns present
    out = df.copy()
    for c, default in missing.items():
        out[c] = default
    return out

_SESSION_BASE_YEAR = 2023
_SESSION_BASE_NUM = 88

def _session_from_year(year_val) -> str:
    try:
        y = int(year_val)
    except Exception:
        return ""
    # Texas regular sessions map odd/even years to the same session.
    # Examples: 2023/2024 -> 88R, 2025/2026 -> 89R.
    session = _SESSION_BASE_NUM + ((y - _SESSION_BASE_YEAR) // 2)
    return f"{session}R"

def _add_session_from_year(df: pd.DataFrame) -> pd.DataFrame:
    if "Session" in df.columns:
        return df
    out = df.copy()
    year_col = None
    for cand in ["applicableYear", "applicable_year", "ApplicableYear", "year", "Year"]:
        if cand in out.columns:
            year_col = cand
            break
    if year_col:
        years = pd.to_numeric(out[year_col], errors="coerce")
        sessions = years.map(_session_from_year)
        out["Session"] = sessions
    else:
        out["Session"] = ""
    return out

def _build_filerid_map(frames: list[tuple[pd.DataFrame, str, str]]) -> dict[int, str]:
    rows = []
    for df, fid_col, short_col in frames:
        if not isinstance(df, pd.DataFrame) or df.empty:
            continue
        if fid_col not in df.columns or short_col not in df.columns:
            continue
        fid = pd.to_numeric(df[fid_col], errors="coerce")
        if fid.isna().all():
            continue
        short = df[short_col].fillna("").astype(str).str.strip()
        tmp = pd.DataFrame({"FilerID": fid, "LobbyShort": short})
        tmp = tmp.dropna(subset=["FilerID"])
        tmp["FilerID"] = tmp["FilerID"].astype(int)
        tmp = tmp[tmp["LobbyShort"].astype(str).str.strip() != ""]
        if not tmp.empty:
            rows.append(tmp)
    if not rows:
        return {}
    all_rows = pd.concat(rows, ignore_index=True)
    counts = (
        all_rows.groupby(["FilerID", "LobbyShort"])
        .size()
        .reset_index(name="n")
        .sort_values(["FilerID", "n"], ascending=[True, False])
        .drop_duplicates("FilerID")
    )
    return dict(zip(counts["FilerID"], counts["LobbyShort"]))

def _tfl_session_for_filter(session_val: str | None, tfl_sessions: set[str]) -> str | None:
    if session_val is None:
        return None
    s = str(session_val).strip()
    if not s:
        return ""
    # Lobby_TFL_Client_All rolls special sessions into the regular session (e.g., 891 -> 89R).
    if s.isdigit() and len(s) >= 3:
        reg = f"{s[:-1]}R"
        if reg in tfl_sessions:
            return reg
    return s



def _compact_ratio_bar(value: float, max_value: float, width: int = 10) -> str:
    try:
        v = float(value)
        max_v = float(max_value)
    except Exception:
        return "-" * max(1, int(width))
    if max_v <= 0:
        return "-" * max(1, int(width))
    span = max(1, int(width))
    ratio = max(0.0, min(1.0, v / max_v))
    filled = int(round(ratio * span))
    return ("#" * filled) + ("-" * max(0, span - filled))

def _priority_cell_style(value: str) -> str:
    label = str(value).strip()
    if label == "Tier 1":
        return "background-color: rgba(242,95,92,0.20); font-weight: 700;"
    if label == "Tier 2":
        return "background-color: rgba(247,178,103,0.20); font-weight: 600;"
    if label == "Tier 3":
        return "background-color: rgba(77,157,224,0.20);"
    return ""

def _safe_style(df: pd.DataFrame):
    try:
        return df.style
    except Exception:
        return None

def _numeric_heatmap_style_fn(
    series: pd.Series,
    *,
    rgb_lo: tuple[int, int, int],
    rgb_hi: tuple[int, int, int],
    alpha_lo: float = 0.06,
    alpha_hi: float = 0.34,
):
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.notna().any():
        min_v = float(numeric.min())
        max_v = float(numeric.max())
    else:
        min_v = 0.0
        max_v = 0.0
    span = max(1e-9, max_v - min_v)

    def _style_one(value) -> str:
        try:
            v = float(value)
        except Exception:
            return ""
        ratio = (v - min_v) / span if span > 0 else 0.0
        ratio = max(0.0, min(1.0, ratio))
        r = int(round(rgb_lo[0] + (rgb_hi[0] - rgb_lo[0]) * ratio))
        g = int(round(rgb_lo[1] + (rgb_hi[1] - rgb_lo[1]) * ratio))
        b = int(round(rgb_lo[2] + (rgb_hi[2] - rgb_lo[2]) * ratio))
        a = alpha_lo + (alpha_hi - alpha_lo) * ratio
        return f"background-color: rgba({r},{g},{b},{a:.3f});"

    return _style_one

def _apply_numeric_heatmap(
    styler,
    df: pd.DataFrame,
    *,
    columns: list[str],
    rgb_lo: tuple[int, int, int],
    rgb_hi: tuple[int, int, int],
):
    out = styler
    for col in columns:
        if col not in df.columns:
            continue
        fn = _numeric_heatmap_style_fn(
            df[col],
            rgb_lo=rgb_lo,
            rgb_hi=rgb_hi,
        )
        out = out.map(fn, subset=[col])
    return out















def _split_authors(text: str) -> list[str]:
    if text is None:
        return []
    s = str(text).strip()
    if not s or s.lower() in {"nan", "none"}:
        return []
    parts = [p.strip() for p in s.split("|")]
    return [p for p in parts if p and p.lower() not in {"nan", "none"}]




def parse_member_name(member_name: str) -> dict:
    t = clean_person_name(member_name)
    if not t:
        return {"full_norm": "", "last_norm": "", "first_norm": "", "first_initial": "", "initial_key": ""}

    if "," in t:
        last, rest = [p.strip() for p in t.split(",", 1)]
        first = rest.split()[0].strip() if rest else ""
    else:
        parts = t.split()
        if len(parts) == 1:
            first, last = "", parts[0]
        else:
            first, last = parts[0], parts[-1]

    first_norm = norm_name(first)
    last_norm = norm_name(last)
    first_initial = norm_name(first[0]) if first else ""
    initial_key = _last_first_initial_key(t)
    return {
        "full_norm": norm_name(t),
        "last_norm": last_norm,
        "first_norm": first_norm,
        "first_initial": first_initial,
        "initial_key": initial_key,
    }

def parse_person_name(person_name: str) -> dict:
    return parse_member_name(person_name)


def member_match_mask(df: pd.DataFrame, member_info: dict) -> pd.Series:
    if df.empty:
        return pd.Series([], dtype=bool)

    org = df.get("recipientNameOrganization", pd.Series([""] * len(df))).fillna("").astype(str)
    last = df.get("recipientNameLast", pd.Series([""] * len(df))).fillna("").astype(str)
    first = df.get("recipientNameFirst", pd.Series([""] * len(df))).fillna("").astype(str)

    org_norm = norm_name_series(org)
    last_norm = norm_name_series(last)
    first_norm = norm_name_series(first)

    last_target = member_info.get("last_norm", "")
    first_target = member_info.get("first_norm", "")
    first_initial = member_info.get("first_initial", "")
    full_norm = member_info.get("full_norm", "")

    mask = pd.Series(False, index=df.index)

    if last_target:
        if first_target:
            first_ok = (first_norm == first_target)
            if first_initial:
                first_ok = first_ok | first_norm.str.startswith(first_initial)
            mask = mask | ((last_norm == last_target) & first_ok)
        else:
            mask = mask | (last_norm == last_target)

        if full_norm:
            mask = mask | org_norm.str.contains(full_norm, na=False)
        elif len(last_target) >= 4:
            mask = mask | org_norm.str.contains(last_target, na=False)
    elif full_norm:
        mask = mask | org_norm.str.contains(full_norm, na=False)

    return mask

def map_filer_to_lobbyshort(df: pd.DataFrame, name_to_short: dict, filerid_to_short: dict | None) -> pd.DataFrame:
    if df.empty:
        return df
    d = df.copy()
    filerid_map = filerid_to_short or {}
    short = pd.Series([""] * len(d), index=d.index)

    if "filerIdent" in d.columns and filerid_map:
        fid = pd.to_numeric(d["filerIdent"], errors="coerce").fillna(-1).astype(int)
        short = fid.map(filerid_map).fillna("")

    filer_name = d.get("filerName", pd.Series([""] * len(d)))
    filer_sort = d.get("filerSort", pd.Series([""] * len(d)))
    if isinstance(filer_name, pd.DataFrame):
        filer_name = filer_name.iloc[:, 0]
    if isinstance(filer_sort, pd.DataFrame):
        filer_sort = filer_sort.iloc[:, 0]

    filer_clean = clean_filer_name_series(filer_name)
    norm_raw = norm_name_series(filer_name)
    norm_clean = norm_name_series(filer_clean)
    norm_sort = norm_name_series(filer_sort)

    mapped = norm_raw.map(name_to_short)
    mapped = mapped.where(mapped.notna(), norm_clean.map(name_to_short))
    mapped = mapped.where(mapped.notna(), norm_sort.map(name_to_short))

    short = short.where(short.astype(str).str.strip() != "", mapped)
    d["LobbyShort"] = short.fillna("")
    return d




AppState = _shared_search_state.AppState
NavSearchBundle = _shared_search_state.NavSearchBundle
NavQueryKey = _shared_search_state.NavQueryKey
_candidate_label = _shared_search_state._candidate_label
build_client_index = _shared_search_state.build_client_index
resolve_client_name = _shared_search_state.resolve_client_name
build_author_bill_index = _shared_search_state.build_author_bill_index
build_member_index = _shared_search_state.build_member_index
resolve_member_name = _shared_search_state.resolve_member_name
build_lobbyist_index = _shared_search_state.build_lobbyist_index
resolve_lobbyshort = _shared_search_state.resolve_lobbyshort
resolve_lobbyshort_from_wit = _shared_search_state.resolve_lobbyshort_from_wit
lobbyist_autocomplete_candidates = _shared_search_state.lobbyist_autocomplete_candidates
format_lobbyist_label = _shared_search_state.format_lobbyist_label
lobby_candidate_key = _shared_search_state.lobby_candidate_key
normalize_bill = _shared_search_state.normalize_bill
is_bill_query = _shared_search_state.is_bill_query
build_nav_search_bundle = _shared_search_state.build_nav_search_bundle
build_nav_search_bundle_cached = _shared_search_state.build_nav_search_bundle_cached
can_reuse_nav_search_bundle = _shared_search_state.can_reuse_nav_search_bundle
build_data_health_table = _page_bundles.build_data_health_table
build_timeline_counts = _page_bundles.build_timeline_counts
bill_position_from_flags = _page_bundles.bill_position_from_flags
build_bills_with_status = _page_bundles.build_bills_with_status
build_policy_mentions = _page_bundles.build_policy_mentions
build_lobby_subject_counts = _page_bundles.build_lobby_subject_counts
build_lobbyist_trend = _page_bundles.build_lobbyist_trend
build_top_clients = _page_bundles.build_top_clients
build_member_activities = _page_bundles.build_member_activities
build_activities = _page_bundles.build_activities
build_activities_multi = _page_bundles.build_activities_multi
build_disclosures = _page_bundles.build_disclosures
build_disclosures_multi = _page_bundles.build_disclosures_multi
_build_tfl_trend_data = _page_bundles._build_tfl_trend_data
_build_top5_tfl_clients = _page_bundles._build_top5_tfl_clients
_build_lobby_display_names = _page_bundles._build_lobby_display_names
build_all_lobbyists_overview_fast = _page_bundles.build_all_lobbyists_overview_fast
_build_filtered_atlas_bundle = _map_runtime._build_filtered_atlas_bundle
_build_ranked_forensics_leads = _map_runtime._build_ranked_forensics_leads
_build_filtered_forensics_bundle = _map_runtime._build_filtered_forensics_bundle

def render_bill_search_results(bill_query: str, session_val: str | None, tfl_session_val: str | None,
                               wit_all: pd.DataFrame, bill_status_all: pd.DataFrame,
                               lobby_tfl_client_all: pd.DataFrame, short_to_names: dict):
    q = normalize_bill(bill_query)
    if not q:
        return False

    d = wit_all.copy()
    d["Session"] = d["Session"].astype(str).str.strip()
    if session_val is not None:
        d = d[d["Session"] == str(session_val)]

    d_bill_norm = d["Bill"].astype(str).str.upper().str.replace(r"\s+", " ", regex=True)
    d = d[d_bill_norm == q]
    total_rows = len(d)
    d = d[d["LobbyShort"].notna() & (d["LobbyShort"].astype(str).str.strip() != "")]
    if "LobbyShortNorm" not in d.columns:
        d["LobbyShortNorm"] = norm_name_series(d["LobbyShort"])

    tfl = lobby_tfl_client_all.copy()
    tfl["Session"] = tfl["Session"].astype(str).str.strip()
    if tfl_session_val is not None:
        tfl = tfl[tfl["Session"] == str(tfl_session_val)]
    tfl = ensure_cols(tfl, {"LobbyShort": ""})
    if "LobbyShortNorm" not in tfl.columns:
        tfl["LobbyShortNorm"] = norm_name_series(tfl["LobbyShort"])
    lobbyshort_set = set(tfl["LobbyShortNorm"].dropna().unique().tolist())
    if lobbyshort_set:
        d = d[d["LobbyShortNorm"].isin(lobbyshort_set)]
        if not tfl.empty:
            norm_to_short = (
                tfl[["LobbyShortNorm", "LobbyShort"]]
                .dropna()
                .drop_duplicates()
                .groupby("LobbyShortNorm")["LobbyShort"]
                .first()
                .to_dict()
            )
            d["LobbyShort"] = d["LobbyShortNorm"].map(norm_to_short).fillna(d["LobbyShort"])
    if d.empty:
        if total_rows > 0:
            st.info(
                f"Found {total_rows} witness-list rows for {q}, but none matched a lobbyist in Texas Ethics Commission filings "
                "for the selected session."
            )
        else:
            st.info("No witness-list rows matched that bill search.")
        return True

    pos = bill_position_from_flags(d)
    bs = bill_status_all.copy()
    bs["Session"] = bs["Session"].astype(str).str.strip()
    if session_val is not None:
        bs = bs[bs["Session"] == str(session_val)]

    merged = pos.merge(bs, on=["Session", "Bill"], how="left")
    merged["Lobbyist"] = merged["LobbyShort"].map(lambda s: _candidate_label(str(s), short_to_names))
    tfl = lobby_tfl_client_all.copy()
    tfl["Session"] = tfl["Session"].astype(str).str.strip()
    if tfl_session_val is not None:
        tfl = tfl[tfl["Session"] == str(tfl_session_val)]
    tfl = ensure_cols(tfl, {"LobbyShort": "", "IsTFL": 0})
    tfl_flag = (
        tfl.groupby("LobbyShort", as_index=False)["IsTFL"]
        .max()
        .rename(columns={"IsTFL": "Has TFL Client"})
    )
    merged = merged.merge(tfl_flag, on="LobbyShort", how="left")
    merged["Has TFL Client"] = merged["Has TFL Client"].fillna(0).astype(int).map({1: "Yes", 0: "No"})

    show_cols = ["Session", "Bill", "Lobbyist", "Has TFL Client", "Position", "Author", "Caption", "Status"]
    show_cols = [c for c in show_cols if c in merged.columns]
    merged["_tfl_sort"] = merged["Has TFL Client"].map({"Yes": 1, "No": 0}).fillna(0)
    view = merged[show_cols + ["_tfl_sort"]].sort_values(
        ["_tfl_sort", "Session", "Bill", "Lobbyist"],
        ascending=[False, True, True, True],
    )
    view = view.drop(columns=["_tfl_sort"])
    st.dataframe(view, width="stretch", height=520, hide_index=True)
    _ = export_dataframe(view, "bill_lobbyists.csv")
    return True

# =========================================================
# FAST MONEY PARSING (vectorized) for Lobby_TFL_Client_All
# =========================================================
_MONEY_RANGE = re.compile(r"(-?\d[\d,]*\.?\d*)\s*(?:-|to)\s*(-?\d[\d,]*\.?\d*)", re.IGNORECASE)

def _to_num_series(s: pd.Series) -> pd.Series:
    """Vectorized $/comma/paren cleanup -> float; blanks->0"""
    s = s.fillna("").astype(str).str.strip()
    neg = s.str.startswith("(") & s.str.endswith(")")
    s = s.str.replace(r"^\(|\)$", "", regex=True)
    s = s.str.replace("$", "", regex=False).str.replace(",", "", regex=False)
    out = pd.to_numeric(s, errors="coerce").fillna(0.0)
    out = out.where(~neg, -out)
    return out

def add_low_high_numeric(df: pd.DataFrame) -> pd.DataFrame:
    """
    Produces Low_num/High_num with minimal Python-level loops.
    Priority:
      1) Low/High columns if present and nonzero
      2) Parse Amount range "1000-5000"
      3) Parse Amount single value
      4) Mirror one side if only one exists
    """
    d = ensure_cols(df, {"Low": 0, "High": 0, "Amount": ""}).copy()

    low = _to_num_series(d["Low"])
    high = _to_num_series(d["High"])

    amt = d["Amount"].fillna("").astype(str).str.strip()
    amt_clean = (
        amt.str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace("\u2013", "-", regex=False)
    )

    # range capture
    rng = amt_clean.str.extract(_MONEY_RANGE)
    rng_lo = pd.to_numeric(rng[0], errors="coerce").fillna(0.0)
    rng_hi = pd.to_numeric(rng[1], errors="coerce").fillna(0.0)

    # single numeric fallback
    single = pd.to_numeric(amt_clean.str.extract(r"(-?\d+(?:\.\d+)?)")[0], errors="coerce").fillna(0.0)

    # If both low/high are zero, use range; else keep existing
    both_zero = (low == 0) & (high == 0)
    low = low.where(~both_zero, rng_lo.where(rng_lo != 0, single))
    high = high.where(~both_zero, rng_hi.where(rng_hi != 0, single))

    # Mirror
    high = high.where(high != 0, low)
    low = low.where(low != 0, high)

    d["Low_num"] = low
    d["High_num"] = high
    return d

# =========================================================
# LOAD WORKBOOK (open once -> much faster)
# =========================================================
def safe_read_excel_xf(xf: pd.ExcelFile, sheet_name: str, cols: list[str]) -> pd.DataFrame:
    try:
        return xf.parse(sheet_name=sheet_name, usecols=cols)
    except Exception:
        try:
            df = xf.parse(sheet_name=sheet_name)
            keep = [c for c in cols if c in df.columns]
            return df[keep]
        except Exception:
            return pd.DataFrame(columns=cols)

@st.cache_data(show_spinner=False)
def _empty_df(cols: list[str]) -> pd.DataFrame:
    return pd.DataFrame(columns=cols)

def read_parquet_cols(path: Path, cols: list[str]) -> pd.DataFrame:
    """PERFORMANCE: read only requested columns via PyArrow, using pf.read()
    directly so we don't re-open the file a second time."""
    try:
        import pyarrow.parquet as pq
        pf = pq.ParquetFile(path)
        available = set(pf.schema.names)
        use_cols = [c for c in cols if c in available]
        tbl = pf.read(columns=use_cols) if use_cols else pf.read()
        return tbl.to_pandas()
    except Exception:
        try:
            df = pd.read_parquet(path)
            keep = [c for c in cols if c in df.columns]
            return df[keep] if keep else df
        except Exception:
            return pd.DataFrame(columns=cols)

@st.cache_resource(show_spinner=False, ttl=3600, max_entries=2)
def load_workbook(path: str) -> dict:
    cfg = {
        "Wit_All": ["session", "bill", "position", "LobbyShort", "name", "org"],
        "Bill_Status_All": ["Session", "Bill", "Authors", "Author", "Caption", "Status"],
        "Fiscal_Impact": ["Session", "Bill", "Version", "EstimatedTwoYearNetImpactGR"],
        "Bill_Sub_All": ["Session", "Bill", "Subject"],
        "Lobby_Sub_All": [
            "Session",
            "legislative_session",
            "Subject Matter",
            "Other Subject Matter Description",
            "Primary Business",
            "FilerID",
            "LobbyShort",
            "lobbyshort",
            "Lobby Name",
            "Unnamed: 0",
        ],
        "Lobbyist_Pol_Funds": [],
        "Lobby_TFL_Client_All": ["Session", "Client", "Lobby Name", "LobbyShort", "IsTFL", "Low", "High", "Amount", "Mid", "FilerID"],
        "Staff_All": ["Session", "session", "Legislator", "member_or_committee", "legislator_name", "Title", "role",
                      "Staffer", "name", "staff_name_last_initial", "lobby name", "source"],
        "LaFood": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst",
                   "restaurantName", "activityDate", "periodStartDt", "activityExactAmount", "activityAmountRangeLow", "activityAmountRangeHigh", "activityAmountCd"],
        "LaEnt": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst",
                  "entertainmentName", "activityDate", "periodStartDt", "activityExactAmount", "activityAmountRangeLow", "activityAmountRangeHigh", "activityAmountCd"],
        "LaTran": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst",
                   "travelPurpose", "transportationTypeDescr", "departureCity", "arrivalCity", "checkInDt", "checkOutDt", "departureDt", "periodStartDt"],
        "LaGift": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst",
                   "activityDescription", "periodStartDt", "activityExactAmount", "activityAmountRangeLow", "activityAmountRangeHigh", "activityAmountCd"],
        "LaEvnt": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst",
                   "activityDescription", "activityDate", "periodStartDt"],
        "LaAwrd": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst",
                   "activityDescription", "periodStartDt", "activityExactAmount", "activityAmountRangeLow", "activityAmountRangeHigh", "activityAmountCd"],
        "LaCvr": ["Session", "filerIdent", "filerName", "filerSort", "filedDt", "periodStartDt", "sourceCategoryCd",
                  "subjectMatterMemo", "docketsMemo", "filerNameOrganization"],
        "LaDock": ["Session", "filerIdent", "filerName", "filerSort", "receivedDt", "periodStartDt", "designationText", "agencyName"],
        "LaI4E": ["Session", "filerIdent", "filerName", "filerSort", "periodStartDt", "onbehalfName",
                  "onbehalfMailingCity", "onbehalfPrimaryPhoneNumber"],
        "LaSub": ["Session", "filerIdent", "filerName", "filerSort", "periodStartDt", "subjectMatterCodeValue", "subjectMatterDescr"],
    }

    base = Path(path)
    if not base.exists():
        return {k: _empty_df(v) for k, v in cfg.items()}
    if base.is_dir():
        parquet_map = {
        "Wit_All": ["Witness_Lists.parquet", "Witness List.parquet", "Witness_List.parquet", "witnesslist.parquet"],
            "Bill_Status_All": "Bill_Status.parquet",
            "Fiscal_Impact": "Fiscal_Notes.parquet",
            "Bill_Sub_All": "Bill_Sub_All.parquet",
            "Lobby_Sub_All": "Lobby.Sub.parquet",
            "Lobbyist_Pol_Funds": "Lobbyist.Pol.Funds.parquet",
            "Lobby_TFL_Client_All": "Lobby_TFL_Client_All.parquet",
            "Staff_All": ["Staff.parquet", "staff.parquet"],
            "LaFood": "LaFood.parquet",
            "LaEnt": "LaEnt.parquet",
            "LaTran": "LaTran.parquet",
            "LaGift": "LaGift.parquet",
            "LaEvnt": "LaEvnt.parquet",
            "LaAwrd": "LaAwrd.parquet",
            "LaCvr": "LaCvr.parquet",
            "LaDock": "LaDock.parquet",
            "LaI4E": "LaI4E.parquet",
            "LaSub": "LaSub.parquet",
        }
        # PERFORMANCE: resolve paths first, then read all parquet files
        # in parallel with ThreadPoolExecutor (was sequential).
        def _resolve_path(key_: str):
            fname = parquet_map.get(key_)
            if not fname:
                return None
            if isinstance(fname, (list, tuple)):
                if key_ == "Wit_All":
                    return [base / c for c in fname if (base / c).exists()] or None
                for c in fname:
                    if (base / c).exists():
                        return base / c
                return None
            cand = base / fname
            return cand if cand.exists() else None

        resolved = {k: _resolve_path(k) for k in cfg}

        def _read_one(key_: str, fpath_, cols_: list[str]) -> tuple[str, pd.DataFrame]:
            if fpath_ is None:
                return (key_, _empty_df(cols_))
            if isinstance(fpath_, list):  # Wit_All multi-file
                frames = []
                for p in fpath_:
                    try:
                        frames.append(read_parquet_cols(p, cols_))
                    except Exception:
                        continue
                if frames:
                    return (key_, pd.concat(frames, ignore_index=True).drop_duplicates())
                return (key_, _empty_df(cols_))
            try:
                return (key_, read_parquet_cols(fpath_, cols_))
            except Exception:
                return (key_, _empty_df(cols_))

        data = {}
        with ThreadPoolExecutor(max_workers=min(len(cfg), 8)) as pool:
            futs = {
                pool.submit(_read_one, k, resolved[k], v): k
                for k, v in cfg.items()
            }
            for fut in as_completed(futs):
                try:
                    k_, df_ = fut.result()
                    data[k_] = df_
                except Exception:
                    k_ = futs[fut]
                    data[k_] = _empty_df(cfg.get(k_, []))
    else:
        xf = pd.ExcelFile(path, engine="openpyxl")  # OPEN ONCE
        data = {k: safe_read_excel_xf(xf, k, v) for k, v in cfg.items()}

    # Normalize parquet schema differences
    wit = data.get("Wit_All")
    if isinstance(wit, pd.DataFrame):
        wit = wit
        if "session" in wit.columns and "Session" not in wit.columns:
            wit = wit.rename(columns={"session": "Session"})
        if "bill" in wit.columns and "Bill" not in wit.columns:
            wit = wit.rename(columns={"bill": "Bill"})
        if "position" in wit.columns:
            pos = wit["position"].fillna("").astype(str).str.upper()
            if "IsFor" not in wit.columns:
                wit["IsFor"] = pos.str.contains(r"\bFOR\b").astype(int)
            if "IsAgainst" not in wit.columns:
                wit["IsAgainst"] = pos.str.contains(r"\bAGAINST\b").astype(int)
            if "IsOn" not in wit.columns:
                wit["IsOn"] = pos.str.contains(r"\bON\b").astype(int)
        if "LobbyShort" not in wit.columns:
            wit["LobbyShort"] = ""
        unnamed = [c for c in wit.columns if str(c).startswith("Unnamed:")]
        if unnamed:
            wit = wit.drop(columns=unnamed)
        data["Wit_All"] = wit

    bs = data.get("Bill_Status_All")
    if isinstance(bs, pd.DataFrame):
        bs = bs
        if "Authors" in bs.columns and "Author" not in bs.columns:
            bs["Author"] = bs["Authors"]
        data["Bill_Status_All"] = bs

    fi = data.get("Fiscal_Impact")
    if isinstance(fi, pd.DataFrame):
        fi = fi
        data["Fiscal_Impact"] = fi

    lt = data.get("Lobby_TFL_Client_All")
    if isinstance(lt, pd.DataFrame):
        lt = lt
        if "IsTFL" not in lt.columns and "TFL?" in lt.columns:
            lt["IsTFL"] = lt["TFL?"].astype(str).str.upper().isin(["Y", "YES", "TRUE", "1"]).astype(int)
        if "IsTFL" in lt.columns:
            lt["IsTFL"] = pd.to_numeric(lt["IsTFL"], errors="coerce").fillna(0).astype(int)
        data["Lobby_TFL_Client_All"] = lt

    staff = data.get("Staff_All")
    if isinstance(staff, pd.DataFrame) and not staff.empty:
        staff = staff
        # Rename session column if needed
        if "session" in staff.columns and "Session" not in staff.columns:
            staff = staff.rename(columns={"session": "Session"})
        # Map staff parquet schema to expected columns
        if "Legislator" not in staff.columns:
            if "legislator_name" in staff.columns:
                leg = staff["legislator_name"].fillna("").astype(str).str.strip()
                if "member_or_committee" in staff.columns:
                    fallback = staff["member_or_committee"].fillna("").astype(str).str.strip()
                    staff["Legislator"] = leg.where(leg != "", fallback)
                else:
                    staff["Legislator"] = leg
            elif "member_or_committee" in staff.columns:
                staff["Legislator"] = staff["member_or_committee"]
            else:
                staff["Legislator"] = ""
        if "Title" not in staff.columns:
            staff["Title"] = staff.get("role", "")
        if "Staffer" not in staff.columns:
            staff["Staffer"] = staff.get("name", staff.get("staff_name_last_initial", ""))
        if "lobby name" not in staff.columns:
            staff["lobby name"] = staff.get("staff_name_last_initial", staff.get("name", ""))
        # Normalized staff name helpers for matching
        staff["StaffNameNorm"] = norm_name_series(staff.get("name", pd.Series(dtype=object)))
        staff["StaffLastInitialNorm"] = norm_name_series(
            staff.get("staff_name_last_initial", staff.get("name", pd.Series(dtype=object)))
        )
        staff["StaffLastNorm"] = last_name_norm_series(
            staff.get("name", staff.get("staff_name_last_initial", pd.Series(dtype=object)))
        )
        # Normalize Session to match app sessions (e.g., 89 -> 89R)
        if "Session" in staff.columns:
            sess = staff["Session"].astype(str).str.strip()
            staff["Session"] = sess.where(~sess.str.fullmatch(r"\d+"), sess + "R")
        data["Staff_All"] = staff

    ls = data.get("Lobby_Sub_All")
    if isinstance(ls, pd.DataFrame):
        ls = ls
        if "Session" not in ls.columns:
            if "legislative_session" in ls.columns:
                ls = ls.rename(columns={"legislative_session": "Session"})
            elif "session" in ls.columns:
                ls = ls.rename(columns={"session": "Session"})
        if "LobbyShort" not in ls.columns:
            if "lobbyshort" in ls.columns:
                ls = ls.rename(columns={"lobbyshort": "LobbyShort"})
            elif "lobby_short" in ls.columns:
                ls = ls.rename(columns={"lobby_short": "LobbyShort"})
        data["Lobby_Sub_All"] = ls

    pf = data.get("Lobbyist_Pol_Funds")
    if isinstance(pf, pd.DataFrame):
        pf = pf
        if "Session" not in pf.columns and "legislative_session" in pf.columns:
            pf = pf.rename(columns={"legislative_session": "Session"})
        if "LobbyShort" not in pf.columns:
            if "lobbyshort" in pf.columns:
                pf = pf.rename(columns={"lobbyshort": "LobbyShort"})
            elif "lobby_short" in pf.columns:
                pf = pf.rename(columns={"lobby_short": "LobbyShort"})
        data["Lobbyist_Pol_Funds"] = pf

    for key in ["LaFood", "LaEnt", "LaTran", "LaGift", "LaEvnt", "LaAwrd", "LaCvr", "LaDock", "LaI4E", "LaSub"]:
        df = data.get(key)
        if isinstance(df, pd.DataFrame):
            data[key] = _add_session_from_year(df)

    # Normalize Session everywhere
    for df in data.values():
        if isinstance(df, pd.DataFrame) and "Session" in df.columns:
            df["Session"] = df["Session"].astype(str).str.strip()

    # Precompute Low_num/High_num once (speed for overview + per-lobbyist)
    if isinstance(data.get("Lobby_TFL_Client_All"), pd.DataFrame) and not data["Lobby_TFL_Client_All"].empty:
        data["Lobby_TFL_Client_All"] = add_low_high_numeric(data["Lobby_TFL_Client_All"])

    # Build mapping from Lobby Name -> LobbyShort (across all sessions)
    lobby_name_rows = []

    def _append_lobby_names(df: pd.DataFrame, name_col: str, short_col: str, fid_col: str) -> None:
        if not isinstance(df, pd.DataFrame) or df.empty:
            return
        if name_col not in df.columns or short_col not in df.columns:
            return
        tmp = df[[name_col, short_col]]
        tmp = tmp.rename(columns={name_col: "Lobby Name", short_col: "LobbyShort"})
        tmp["FilerID"] = df[fid_col] if fid_col in df.columns else pd.NA
        lobby_name_rows.append(tmp)

    _append_lobby_names(data.get("Lobby_TFL_Client_All"), "Lobby Name", "LobbyShort", "FilerID")
    _append_lobby_names(data.get("Lobby_Sub_All"), "Lobby Name", "LobbyShort", "FilerID")
    _append_lobby_names(data.get("Lobbyist_Pol_Funds"), "Lobbyist", "LobbyShort", "FilerID")

    if lobby_name_rows:
        lobby_names = pd.concat(lobby_name_rows, ignore_index=True)
        lobby_names["LobbyShort"] = lobby_names["LobbyShort"].astype(str).str.strip()
        lobby_names["Lobby Name"] = lobby_names["Lobby Name"].astype(str).str.strip()
        lobby_names = lobby_names[(lobby_names["LobbyShort"] != "") & (lobby_names["Lobby Name"] != "")]
        lobby_names = lobby_names.drop_duplicates()
    else:
        lobby_names = pd.DataFrame(columns=["LobbyShort", "Lobby Name", "FilerID"])

    lobbyist_index = build_lobbyist_index(lobby_names)
    lobby_index = lobbyist_index.copy()
    name_to_short = {}
    short_to_names = {}
    known_shorts = set()
    initial_to_short = {}

    if not lobbyist_index.empty:
        known_shorts = set(
            lobbyist_index["LobbyShort"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        tmp = lobbyist_index[["LobbyShort", "Lobby Name"]].dropna()
        tmp["LobbyShort"] = tmp["LobbyShort"].astype(str)
        short_to_names = (
            tmp.groupby("LobbyShort")["Lobby Name"]
            .agg(lambda s: sorted(set(map(str, s)))[:6])
            .to_dict()
        )

        key_frames = []
        for col in ["LobbyNameNorm", "LobbyNameCleanNorm", "LastFirstNorm", "FirstLastNorm", "LastFirstInitialNorm"]:
            if col in lobbyist_index.columns:
                key_frames.append(lobbyist_index[[col, "LobbyShort"]].rename(columns={col: "Key"}))
        if key_frames:
            all_keys = pd.concat(key_frames, ignore_index=True)
            all_keys["Key"] = all_keys["Key"].fillna("").astype(str).str.strip()
            all_keys = all_keys[all_keys["Key"] != ""]
            counts = (
                all_keys.groupby(["Key", "LobbyShort"])
                .size()
                .reset_index(name="n")
                .sort_values(["Key", "n"], ascending=[True, False])
                .drop_duplicates("Key")
            )
            name_to_short = dict(zip(counts["Key"], counts["LobbyShort"]))

        # Map last name + first initial to LobbyShort (helps when names don't match exactly)
        tmp_short = lobbyist_index[["LobbyShort"]].dropna()
        tmp_short["InitialKey"] = tmp_short["LobbyShort"].map(_last_first_initial_key)
        tmp_short = tmp_short[tmp_short["InitialKey"].astype(str).str.strip() != ""]
        if not tmp_short.empty:
            init_counts = (
                tmp_short.groupby(["InitialKey", "LobbyShort"])
                .size()
                .reset_index(name="n")
                .sort_values(["InitialKey", "n"], ascending=[True, False])
                .drop_duplicates("InitialKey")
            )
            initial_to_short = dict(zip(init_counts["InitialKey"], init_counts["LobbyShort"]))

    # Map FilerID -> LobbyShort (used for activity matching)
    filerid_to_short = _build_filerid_map([
        (data.get("Lobby_TFL_Client_All"), "FilerID", "LobbyShort"),
        (data.get("Lobby_Sub_All"), "FilerID", "LobbyShort"),
        (data.get("Lobbyist_Pol_Funds"), "FilerID", "LobbyShort"),
    ])

    # Map witness list names/orgs to LobbyShort where possible
    wit = data.get("Wit_All")
    if isinstance(wit, pd.DataFrame) and not wit.empty:
        wit = wit
        if "LobbyShort" not in wit.columns:
            wit["LobbyShort"] = ""
        name_series = wit.get("name", pd.Series([""] * len(wit))).fillna("").astype(str)
        if "name" in wit.columns:
            wit["NameNorm"] = norm_name_series(name_series)
            wit["NameLastNorm"] = last_name_norm_series(name_series)
            wit["NameFirstNorm"] = first_name_norm_series(name_series)
            wit["NameFirstInitialNorm"] = wit["NameFirstNorm"].str.slice(0, 1)
        if name_to_short:
            name_norm = wit.get("NameNorm", name_series.map(norm_name))
            mapped = name_norm.map(name_to_short)
            if initial_to_short:
                init_key = name_series.map(_last_first_initial_key)
                mapped_init = init_key.map(initial_to_short)
                mapped = mapped.where(mapped.notna() & mapped.astype(str).str.strip().ne(""), mapped_init)
            if "org" in wit.columns:
                org_series = wit.get("org", pd.Series([""] * len(wit))).fillna("").astype(str)
                org_norm = norm_name_series(org_series)
                mapped = mapped.where(mapped.notna() & mapped.astype(str).str.strip().ne(""), org_norm.map(name_to_short))
            blank = wit["LobbyShort"].isna() | (wit["LobbyShort"].astype(str).str.strip() == "")
            wit.loc[blank, "LobbyShort"] = mapped[blank].fillna("")
        data["Wit_All"] = wit

    # Normalize LobbyShort for robust matching (hyphens/case/spacing).
    for key in ["Wit_All", "Lobby_TFL_Client_All", "Lobby_Sub_All"]:
        df = data.get(key)
        if isinstance(df, pd.DataFrame) and "LobbyShort" in df.columns:
            df["LobbyShortNorm"] = norm_name_series(df["LobbyShort"])

    data["name_to_short"] = name_to_short
    data["short_to_names"] = short_to_names
    data["lobby_index"] = lobby_index
    data["lobbyist_index"] = lobbyist_index
    data["known_shorts"] = known_shorts
    data["filerid_to_short"] = filerid_to_short

    # Fill Lobby_Sub_All LobbyShort from FilerID when missing
    ls = data.get("Lobby_Sub_All")
    if isinstance(ls, pd.DataFrame) and not ls.empty and filerid_to_short:
        if "FilerID" in ls.columns and "LobbyShort" in ls.columns:
            ls = ls
            fid = pd.to_numeric(ls["FilerID"], errors="coerce").fillna(-1).astype(int)
            missing = ls["LobbyShort"].isna() | ls["LobbyShort"].astype(str).str.strip().eq("")
            ls.loc[missing, "LobbyShort"] = fid.map(filerid_to_short)
            data["Lobby_Sub_All"] = ls
    gc.collect()
    return data

@st.cache_resource(show_spinner=False, ttl=3600, max_entries=2)
def get_app_state(path: str) -> AppState:
    return _shared_search_state.build_app_state(path, load_workbook(path))


@st.cache_resource(show_spinner=False, ttl=3600, max_entries=2)
def get_map_state(path: str) -> _map_page_state.MapState:
    def _fetch_reference_tables() -> dict[str, pd.DataFrame]:
        return {
            "school_districts": fetch_tea_school_district_centroids(),
            "counties": fetch_tea_county_centroids(),
            "cities": fetch_texas_city_centroids(),
            "water_districts": fetch_tceq_water_district_centroids(),
            "groundwater_districts": fetch_tceq_groundwater_district_centroids(),
            "regional_mobility_authorities": fetch_texas_rma_centroids(),
            "junior_colleges": fetch_texas_junior_college_centroids(),
            "navigation_districts": fetch_texas_navigation_district_centroids(),
            "transit_providers": fetch_nctcog_transit_provider_centroids(),
            "seaports": fetch_txdot_seaport_centroids(),
        }

    def _build_client_matches(
        tfl_client_names: tuple[str, ...],
        _reference_tables: dict[str, pd.DataFrame],
    ) -> pd.DataFrame:
        return build_tfl_political_subdivision_matches(tfl_client_names)

    return _map_page_state.build_map_state_from_sources(
        path,
        load_workbook(path),
        classify_entity_type=classify_requested_entity_type,
        fetch_reference_tables=_fetch_reference_tables,
        build_client_matches=_build_client_matches,
    )


def require_app_state(path: str, *, missing_path_message: str, missing_file_message: str) -> AppState:
    if not path:
        st.error(missing_path_message)
        st.stop()
    if not _is_url(path) and not os.path.exists(path):
        st.error(missing_file_message)
        st.stop()
    return get_app_state(path)


def require_map_state(path: str, *, missing_path_message: str, missing_file_message: str) -> _map_page_state.MapState:
    if not path:
        st.error(missing_path_message)
        st.stop()
    if not _is_url(path) and not os.path.exists(path):
        st.error(missing_file_message)
        st.stop()
    return get_map_state(path)


@st.cache_resource(show_spinner=False, ttl=3600, max_entries=16)
def get_map_atlas_bundle(
    path: str,
    scope: str,
    session_for_filter: str | None,
) -> _map_page_state.AtlasBundle:
    return _map_page_state.build_atlas_bundle(
        get_map_state(path),
        scope=scope,
        session_for_filter=session_for_filter,
        prepare_overlap_pool=_prepare_subdivision_match_pool,
    )


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_client_scope_bundle(path: str, scope: str, session_val: str | None) -> _page_bundles.ClientScopeBundle:
    app_state = get_app_state(path)
    return _page_bundles.build_client_scope_bundle(
        app_state.data["Lobby_TFL_Client_All"],
        session_val,
        scope,
        match_entity_type=match_entity_type,
        build_category_chart_data=_build_category_chart_data,
    )


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_lobby_scope_bundle(path: str, scope: str, session_val: str | None) -> _page_bundles.LobbyScopeBundle:
    app_state = get_app_state(path)
    return _page_bundles.build_lobby_scope_bundle(
        app_state.data["Lobby_TFL_Client_All"],
        session_val,
        scope,
    )


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_member_session_bundle(path: str, session_val: str | None) -> _page_bundles.MemberSessionBundle:
    app_state = get_app_state(path)
    return _page_bundles.build_member_session_bundle(
        app_state.author_bills_all,
        app_state.data["Wit_All"],
        str(session_val or ""),
    )


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_client_workspace_detail_bundle(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
    client_name: str,
) -> _page_detail_bundles.ClientWorkspaceDetailBundle:
    app_state = get_app_state(path)
    return _page_detail_bundles.build_client_workspace_detail_bundle(
        app_state.data,
        name_to_short=app_state.name_to_short,
        session=str(session_val or ""),
        tfl_session_val=tfl_session_val,
        client_name=str(client_name or ""),
    )


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_member_workspace_detail_bundle(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
    member_name: str,
) -> _page_detail_bundles.MemberWorkspaceDetailBundle:
    app_state = get_app_state(path)
    return _page_detail_bundles.build_member_workspace_detail_bundle(
        app_state.data,
        author_bills_all=app_state.author_bills_all,
        name_to_short=app_state.name_to_short,
        short_to_names=app_state.short_to_names,
        session=str(session_val or ""),
        tfl_session_val=tfl_session_val,
        member_name=str(member_name or ""),
    )


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_lobby_workspace_detail_bundle(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
    lobbyshort: str,
    typed_norms_tuple: tuple[str, ...],
    selected_names: tuple[str, ...],
    selected_filer_ids: tuple[int, ...],
) -> _page_detail_bundles.LobbyWorkspaceDetailBundle:
    app_state = get_app_state(path)
    return _page_detail_bundles.build_lobby_workspace_detail_bundle(
        app_state.data,
        name_to_short=app_state.name_to_short,
        short_to_names=app_state.short_to_names,
        session=str(session_val or ""),
        tfl_session_val=tfl_session_val,
        lobbyshort=str(lobbyshort or ""),
        typed_norms_tuple=typed_norms_tuple or tuple(),
        selected_names=selected_names or tuple(),
        selected_filer_ids=selected_filer_ids or tuple(),
    )


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_map_forensics_bundle(
    path: str,
    scope: str,
    session_for_filter: str | None,
    selected_subdivision_signature: str,
) -> _map_runtime.MapForensicsBundle:
    atlas_bundle = get_map_atlas_bundle(path, scope, session_for_filter)
    return _map_runtime.build_map_forensics_bundle(
        atlas_bundle,
        selected_subdivision_signature=selected_subdivision_signature,
    )


WORKBOOK_TABLE_COLUMNS = {
    "Wit_All": ["session", "bill", "position", "LobbyShort", "name", "org"],
    "Bill_Status_All": ["Session", "Bill", "Authors", "Author", "Caption", "Status"],
    "Fiscal_Impact": ["Session", "Bill", "Version", "EstimatedTwoYearNetImpactGR"],
    "Bill_Sub_All": ["Session", "Bill", "Subject"],
    "Lobby_Sub_All": [
        "Session",
        "legislative_session",
        "Subject Matter",
        "Other Subject Matter Description",
        "Primary Business",
        "FilerID",
        "LobbyShort",
        "lobbyshort",
        "Lobby Name",
        "Unnamed: 0",
    ],
    "Lobbyist_Pol_Funds": [],
    "Lobby_TFL_Client_All": ["Session", "Client", "Lobby Name", "LobbyShort", "IsTFL", "Low", "High", "Amount", "Mid", "FilerID"],
    "Staff_All": [
        "Session",
        "session",
        "Legislator",
        "member_or_committee",
        "legislator_name",
        "Title",
        "role",
        "Staffer",
        "name",
        "staff_name_last_initial",
        "lobby name",
        "source",
    ],
    "LaFood": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst", "restaurantName", "activityDate", "periodStartDt", "activityExactAmount", "activityAmountRangeLow", "activityAmountRangeHigh", "activityAmountCd"],
    "LaEnt": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst", "entertainmentName", "activityDate", "periodStartDt", "activityExactAmount", "activityAmountRangeLow", "activityAmountRangeHigh", "activityAmountCd"],
    "LaTran": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst", "travelPurpose", "transportationTypeDescr", "departureCity", "arrivalCity", "checkInDt", "checkOutDt", "departureDt", "periodStartDt"],
    "LaGift": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst", "activityDescription", "periodStartDt", "activityExactAmount", "activityAmountRangeLow", "activityAmountRangeHigh", "activityAmountCd"],
    "LaEvnt": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst", "activityDescription", "activityDate", "periodStartDt"],
    "LaAwrd": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst", "activityDescription", "periodStartDt", "activityExactAmount", "activityAmountRangeLow", "activityAmountRangeHigh", "activityAmountCd"],
    "LaCvr": ["Session", "filerIdent", "filerName", "filerSort", "filedDt", "periodStartDt", "sourceCategoryCd", "subjectMatterMemo", "docketsMemo", "filerNameOrganization"],
    "LaDock": ["Session", "filerIdent", "filerName", "filerSort", "receivedDt", "periodStartDt", "designationText", "agencyName"],
    "LaI4E": ["Session", "filerIdent", "filerName", "filerSort", "periodStartDt", "onbehalfName", "onbehalfMailingCity", "onbehalfPrimaryPhoneNumber"],
    "LaSub": ["Session", "filerIdent", "filerName", "filerSort", "periodStartDt", "subjectMatterCodeValue", "subjectMatterDescr"],
}
PARQUET_FILE_MAP = {
    "Wit_All": ["Witness_Lists.parquet", "Witness List.parquet", "Witness_List.parquet", "witnesslist.parquet"],
    "Bill_Status_All": "Bill_Status.parquet",
    "Fiscal_Impact": "Fiscal_Notes.parquet",
    "Bill_Sub_All": "Bill_Sub_All.parquet",
    "Lobby_Sub_All": "Lobby.Sub.parquet",
    "Lobbyist_Pol_Funds": "Lobbyist.Pol.Funds.parquet",
    "Lobby_TFL_Client_All": "Lobby_TFL_Client_All.parquet",
    "Staff_All": ["Staff.parquet", "staff.parquet"],
    "LaFood": "LaFood.parquet",
    "LaEnt": "LaEnt.parquet",
    "LaTran": "LaTran.parquet",
    "LaGift": "LaGift.parquet",
    "LaEvnt": "LaEvnt.parquet",
    "LaAwrd": "LaAwrd.parquet",
    "LaCvr": "LaCvr.parquet",
    "LaDock": "LaDock.parquet",
    "LaI4E": "LaI4E.parquet",
    "LaSub": "LaSub.parquet",
}
BASE_APP_STATE_TABLE_KEYS = (
    "Lobby_TFL_Client_All",
    "Bill_Status_All",
    "Wit_All",
    "Lobby_Sub_All",
    "Lobbyist_Pol_Funds",
    "Staff_All",
)
DETAIL_TABLE_KEYS = (
    "Fiscal_Impact",
    "Bill_Sub_All",
    "LaFood",
    "LaEnt",
    "LaTran",
    "LaGift",
    "LaEvnt",
    "LaAwrd",
    "LaCvr",
    "LaDock",
    "LaI4E",
    "LaSub",
)
SESSION_SCOPED_TABLE_KEYS = (
    "Wit_All",
    "Bill_Status_All",
    "Lobby_Sub_All",
    "Fiscal_Impact",
    "Bill_Sub_All",
    "LaFood",
    "LaEnt",
    "LaTran",
    "LaGift",
    "LaEvnt",
    "LaAwrd",
    "LaCvr",
    "LaDock",
    "LaI4E",
    "LaSub",
)
ALL_WORKBOOK_TABLE_KEYS = tuple(WORKBOOK_TABLE_COLUMNS.keys())


def _normalize_loaded_table(table_key: str, df: pd.DataFrame) -> pd.DataFrame:
    data = df.copy()
    if table_key == "Wit_All":
        if "session" in data.columns and "Session" not in data.columns:
            data = data.rename(columns={"session": "Session"})
        if "bill" in data.columns and "Bill" not in data.columns:
            data = data.rename(columns={"bill": "Bill"})
        if "position" in data.columns:
            pos = data["position"].fillna("").astype(str).str.upper()
            if "IsFor" not in data.columns:
                data["IsFor"] = pos.str.contains(r"\bFOR\b").astype(int)
            if "IsAgainst" not in data.columns:
                data["IsAgainst"] = pos.str.contains(r"\bAGAINST\b").astype(int)
            if "IsOn" not in data.columns:
                data["IsOn"] = pos.str.contains(r"\bON\b").astype(int)
        if "LobbyShort" not in data.columns:
            data["LobbyShort"] = ""
        unnamed = [column for column in data.columns if str(column).startswith("Unnamed:")]
        if unnamed:
            data = data.drop(columns=unnamed)
    elif table_key == "Bill_Status_All":
        if "Authors" in data.columns and "Author" not in data.columns:
            data["Author"] = data["Authors"]
    elif table_key == "Lobby_TFL_Client_All":
        if "IsTFL" not in data.columns and "TFL?" in data.columns:
            data["IsTFL"] = data["TFL?"].astype(str).str.upper().isin(["Y", "YES", "TRUE", "1"]).astype(int)
        if "IsTFL" in data.columns:
            data["IsTFL"] = pd.to_numeric(data["IsTFL"], errors="coerce").fillna(0).astype(int)
        data = add_low_high_numeric(data)
    elif table_key == "Staff_All":
        if "session" in data.columns and "Session" not in data.columns:
            data = data.rename(columns={"session": "Session"})
        if "Legislator" not in data.columns:
            if "legislator_name" in data.columns:
                leg = data["legislator_name"].fillna("").astype(str).str.strip()
                if "member_or_committee" in data.columns:
                    fallback = data["member_or_committee"].fillna("").astype(str).str.strip()
                    data["Legislator"] = leg.where(leg != "", fallback)
                else:
                    data["Legislator"] = leg
            elif "member_or_committee" in data.columns:
                data["Legislator"] = data["member_or_committee"]
            else:
                data["Legislator"] = ""
        if "Title" not in data.columns:
            data["Title"] = data.get("role", "")
        if "Staffer" not in data.columns:
            data["Staffer"] = data.get("name", data.get("staff_name_last_initial", ""))
        if "lobby name" not in data.columns:
            data["lobby name"] = data.get("staff_name_last_initial", data.get("name", ""))
        data["StaffNameNorm"] = norm_name_series(data.get("name", pd.Series(dtype=object)))
        data["StaffLastInitialNorm"] = norm_name_series(data.get("staff_name_last_initial", data.get("name", pd.Series(dtype=object))))
        data["StaffLastNorm"] = last_name_norm_series(data.get("name", data.get("staff_name_last_initial", pd.Series(dtype=object))))
        if "Session" in data.columns:
            sess = data["Session"].astype(str).str.strip()
            data["Session"] = sess.where(~sess.str.fullmatch(r"\d+"), sess + "R")
    elif table_key == "Lobby_Sub_All":
        if "Session" not in data.columns:
            if "legislative_session" in data.columns:
                data = data.rename(columns={"legislative_session": "Session"})
            elif "session" in data.columns:
                data = data.rename(columns={"session": "Session"})
        if "LobbyShort" not in data.columns:
            if "lobbyshort" in data.columns:
                data = data.rename(columns={"lobbyshort": "LobbyShort"})
            elif "lobby_short" in data.columns:
                data = data.rename(columns={"lobby_short": "LobbyShort"})
    elif table_key == "Lobbyist_Pol_Funds":
        if "Session" not in data.columns and "legislative_session" in data.columns:
            data = data.rename(columns={"legislative_session": "Session"})
        if "LobbyShort" not in data.columns:
            if "lobbyshort" in data.columns:
                data = data.rename(columns={"lobbyshort": "LobbyShort"})
            elif "lobby_short" in data.columns:
                data = data.rename(columns={"lobby_short": "LobbyShort"})
    elif table_key in {"LaFood", "LaEnt", "LaTran", "LaGift", "LaEvnt", "LaAwrd", "LaCvr", "LaDock", "LaI4E", "LaSub"}:
        data = _add_session_from_year(data)

    if "Session" in data.columns:
        data["Session"] = data["Session"].fillna("").astype(str).str.strip()
    return data


def _postprocess_table_for_state(table_key: str, df: pd.DataFrame) -> pd.DataFrame:
    data = _normalize_loaded_table(table_key, df)
    if table_key == "Wit_All":
        return _shared_search_state._ensure_session_key_column(_shared_search_state._ensure_witness_search_columns(data))
    if table_key == "Bill_Status_All":
        return _shared_search_state._ensure_session_key_column(data)
    if table_key == "Lobby_TFL_Client_All":
        return _shared_search_state._ensure_lobby_client_lookup_columns(data)
    if table_key == "Staff_All":
        return _shared_search_state._ensure_session_key_column(_shared_search_state._ensure_staff_search_columns(data))
    return _shared_search_state._ensure_session_key_column(data)


def _resolve_table_source(path: str, table_key: str):
    base = Path(path)
    if not base.exists():
        return None
    if not base.is_dir():
        return ("excel", table_key)

    filename = PARQUET_FILE_MAP.get(table_key)
    if not filename:
        return None
    if isinstance(filename, (list, tuple)):
        matches = [base / candidate for candidate in filename if (base / candidate).exists()]
        return matches or None
    candidate = base / filename
    return candidate if candidate.exists() else None


def _read_table_source(path: str, table_key: str, columns: list[str]) -> pd.DataFrame:
    source = _resolve_table_source(path, table_key)
    if source is None:
        return _empty_df(columns)
    if isinstance(source, tuple) and source and source[0] == "excel":
        try:
            xf = pd.ExcelFile(path, engine="openpyxl")
            return safe_read_excel_xf(xf, table_key, columns)
        except Exception:
            return _empty_df(columns)
    if isinstance(source, list):
        frames = []
        for item in source:
            try:
                frames.append(read_parquet_cols(item, columns))
            except Exception:
                continue
        return pd.concat(frames, ignore_index=True).drop_duplicates() if frames else _empty_df(columns)
    try:
        return read_parquet_cols(source, columns)
    except Exception:
        return _empty_df(columns)


def _table_keys_tuple(keys: tuple[str, ...] | list[str]) -> tuple[str, ...]:
    ordered: list[str] = []
    seen: set[str] = set()
    for key in keys:
        if key in WORKBOOK_TABLE_COLUMNS and key not in seen:
            ordered.append(key)
            seen.add(key)
    return tuple(ordered)


@st.cache_resource(show_spinner=False, ttl=3600, max_entries=128)
def _load_table_resource(path: str, table_key: str) -> pd.DataFrame:
    columns = WORKBOOK_TABLE_COLUMNS.get(table_key, [])
    raw = _read_table_source(path, table_key, columns)
    return _normalize_loaded_table(table_key, raw)


def get_app_tables(path: str, keys: tuple[str, ...] | list[str]) -> dict[str, pd.DataFrame]:
    return {
        key: _load_table_resource(path, key).copy()
        for key in _table_keys_tuple(keys)
    }


def _filter_table_by_session(df: pd.DataFrame, session_val: str | None) -> pd.DataFrame:
    session = str(session_val or "").strip()
    if not session or not isinstance(df, pd.DataFrame) or df.empty:
        return df.copy()
    if "SessionKey" in df.columns:
        series = df["SessionKey"].fillna("").astype(str)
        return df[series == session].copy()
    if "Session" in df.columns:
        series = df["Session"].fillna("").astype(str).str.strip()
        return df[series == session].copy()
    if "session" in df.columns:
        series = df["session"].fillna("").astype(str).str.strip()
        return df[series == session].copy()
    return df.copy()


@st.cache_data(show_spinner=False, ttl=300, max_entries=32)
def get_workspace_table_overlays(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
) -> dict[str, pd.DataFrame]:
    session = str(session_val or "").strip()
    overlays: dict[str, pd.DataFrame] = {}
    for key, df in get_app_tables(path, SESSION_SCOPED_TABLE_KEYS).items():
        overlays[key] = _filter_table_by_session(df, session)
    return overlays


def _normalized_manifest_column_count(table_key: str, source_columns: list[str]) -> int:
    sample = pd.DataFrame({column: pd.Series(dtype="object") for column in source_columns})
    return int(len(_postprocess_table_for_state(table_key, sample).columns))


def _read_manifest_probe(path: str, table_key: str, source_columns: list[str]) -> pd.DataFrame:
    probe_columns = []
    for candidate in ("Session", "session", "applicableYear", "LobbyShort", "lobbyshort"):
        if candidate in source_columns and candidate not in probe_columns:
            probe_columns.append(candidate)
    if not probe_columns:
        return pd.DataFrame()
    return _read_table_source(path, table_key, probe_columns)


@st.cache_data(show_spinner=False, ttl=3600, max_entries=4)
def get_table_manifest(path: str) -> dict[str, dict[str, object]]:
    manifest: dict[str, dict[str, object]] = {}
    base = Path(path)
    for table_key in ALL_WORKBOOK_TABLE_KEYS:
        source = _resolve_table_source(path, table_key)
        source_columns: list[str] = []
        row_count = 0
        if source is None:
            manifest[table_key] = {
                "rows": 0,
                "cols": int(len(_postprocess_table_for_state(table_key, _empty_df(WORKBOOK_TABLE_COLUMNS.get(table_key, []))).columns)),
                "has_session": False,
                "empty": True,
                "sessions": 0,
                "lobby_count": 0,
            }
            continue

        if isinstance(source, tuple) and source and source[0] == "excel":
            try:
                import openpyxl

                workbook = openpyxl.load_workbook(base, read_only=True, data_only=True)
                if table_key in workbook.sheetnames:
                    sheet = workbook[table_key]
                    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                    source_columns = [str(value) for value in header if value is not None]
                    row_count = max(int(sheet.max_row or 0) - 1, 0)
            except Exception:
                source_columns = list(WORKBOOK_TABLE_COLUMNS.get(table_key, []))
                row_count = 0
        else:
            try:
                import pyarrow.parquet as pq

                sources = source if isinstance(source, list) else [source]
                row_total = 0
                column_names: set[str] = set()
                for item in sources:
                    pf = pq.ParquetFile(item)
                    row_total += int(pf.metadata.num_rows or 0)
                    column_names.update(str(name) for name in pf.schema.names)
                source_columns = [column for column in WORKBOOK_TABLE_COLUMNS.get(table_key, []) if column in column_names]
                if not source_columns:
                    source_columns = sorted(column_names)
                row_count = row_total
            except Exception:
                probe = _read_table_source(path, table_key, WORKBOOK_TABLE_COLUMNS.get(table_key, []))
                source_columns = list(probe.columns)
                row_count = int(len(probe))

        probe = _normalize_loaded_table(table_key, _read_manifest_probe(path, table_key, source_columns))
        session_count = int(probe["Session"].dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique()) if "Session" in probe.columns else 0
        lobby_count = int(probe["LobbyShort"].dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique()) if "LobbyShort" in probe.columns else 0
        manifest[table_key] = {
            "rows": int(row_count),
            "cols": _normalized_manifest_column_count(table_key, source_columns),
            "has_session": bool("Session" in probe.columns),
            "empty": int(row_count) == 0,
            "sessions": session_count,
            "lobby_count": lobby_count,
        }
    return manifest


@st.cache_resource(show_spinner=False, ttl=3600, max_entries=2)
def load_workbook(path: str) -> dict:
    data = get_app_tables(path, ALL_WORKBOOK_TABLE_KEYS)
    data["table_manifest"] = get_table_manifest(path)
    return data


@st.cache_resource(show_spinner=False, ttl=3600, max_entries=2)
def get_app_state(path: str) -> AppState:
    workbook = get_app_tables(path, BASE_APP_STATE_TABLE_KEYS)
    workbook["table_manifest"] = get_table_manifest(path)
    return _shared_search_state.build_app_state(path, workbook)


@st.cache_data(show_spinner=False, ttl=3600, max_entries=1)
def _fetch_map_reference_tables_cached() -> dict[str, pd.DataFrame]:
    return {
        "school_districts": fetch_tea_school_district_centroids(),
        "counties": fetch_tea_county_centroids(),
        "cities": fetch_texas_city_centroids(),
        "water_districts": fetch_tceq_water_district_centroids(),
        "groundwater_districts": fetch_tceq_groundwater_district_centroids(),
        "regional_mobility_authorities": fetch_texas_rma_centroids(),
        "junior_colleges": fetch_texas_junior_college_centroids(),
        "navigation_districts": fetch_texas_navigation_district_centroids(),
        "transit_providers": fetch_nctcog_transit_provider_centroids(),
        "seaports": fetch_txdot_seaport_centroids(),
    }


@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def _build_map_client_matches_cached(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    return build_tfl_political_subdivision_matches(tfl_client_names)


@st.cache_resource(show_spinner=False, ttl=3600, max_entries=2)
def get_map_state(path: str) -> _map_page_state.MapState:
    def _build_client_matches(
        tfl_client_names: tuple[str, ...],
        _reference_tables: dict[str, pd.DataFrame],
    ) -> pd.DataFrame:
        return _build_map_client_matches_cached(tfl_client_names)

    return _map_page_state.build_map_state_from_sources(
        path,
        get_app_tables(path, ("Lobby_TFL_Client_All",)),
        classify_entity_type=classify_requested_entity_type,
        fetch_reference_tables=_fetch_map_reference_tables_cached,
        build_client_matches=_build_client_matches,
    )


def require_app_state(path: str, *, missing_path_message: str, missing_file_message: str) -> AppState:
    if not path:
        st.error(missing_path_message)
        st.stop()
    if not _is_url(path) and not os.path.exists(path):
        st.error(missing_file_message)
        st.stop()
    return get_app_state(path)


def require_map_state(path: str, *, missing_path_message: str, missing_file_message: str) -> _map_page_state.MapState:
    if not path:
        st.error(missing_path_message)
        st.stop()
    if not _is_url(path) and not os.path.exists(path):
        st.error(missing_file_message)
        st.stop()
    return get_map_state(path)


@st.cache_resource(show_spinner=False, ttl=3600, max_entries=16)
def get_map_atlas_bundle(
    path: str,
    scope: str,
    session_for_filter: str | None,
) -> _map_page_state.AtlasBundle:
    return _map_page_state.build_atlas_bundle(
        get_map_state(path),
        scope=scope,
        session_for_filter=session_for_filter,
        prepare_overlap_pool=_prepare_subdivision_match_pool,
    )


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_client_scope_bundle(path: str, scope: str, session_val: str | None) -> _page_bundles.ClientScopeBundle:
    app_state = get_app_state(path)
    return _page_bundles.build_client_scope_bundle(
        app_state.data["Lobby_TFL_Client_All"],
        session_val,
        scope,
        match_entity_type=match_entity_type,
        build_category_chart_data=_build_category_chart_data,
    )


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_lobby_scope_bundle(path: str, scope: str, session_val: str | None) -> _page_bundles.LobbyScopeBundle:
    app_state = get_app_state(path)
    return _page_bundles.build_lobby_scope_bundle(
        app_state.data["Lobby_TFL_Client_All"],
        session_val,
        scope,
    )


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_member_session_bundle(path: str, session_val: str | None) -> _page_bundles.MemberSessionBundle:
    app_state = get_app_state(path)
    witness_rows = _filter_table_by_session(app_state.data["Wit_All"], session_val)
    return _page_bundles.build_member_session_bundle(
        app_state.author_bills_all,
        witness_rows,
        str(session_val or ""),
    )


def _workspace_data(path: str, session_val: str | None, tfl_session_val: str | None) -> dict[str, object]:
    data = dict(get_app_state(path).data)
    data.update(get_workspace_table_overlays(path, session_val, tfl_session_val))
    return data


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_client_workspace_detail_bundle(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
    client_name: str,
) -> _page_detail_bundles.ClientWorkspaceDetailBundle:
    app_state = get_app_state(path)
    return _page_detail_bundles.build_client_workspace_detail_bundle(
        _workspace_data(path, session_val, tfl_session_val),
        name_to_short=app_state.name_to_short,
        session=str(session_val or ""),
        tfl_session_val=tfl_session_val,
        client_name=str(client_name or ""),
    )


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_member_workspace_detail_bundle(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
    member_name: str,
) -> _page_detail_bundles.MemberWorkspaceDetailBundle:
    app_state = get_app_state(path)
    return _page_detail_bundles.build_member_workspace_detail_bundle(
        _workspace_data(path, session_val, tfl_session_val),
        author_bills_all=app_state.author_bills_all,
        name_to_short=app_state.name_to_short,
        short_to_names=app_state.short_to_names,
        session=str(session_val or ""),
        tfl_session_val=tfl_session_val,
        member_name=str(member_name or ""),
    )


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_lobby_workspace_detail_bundle(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
    lobbyshort: str,
    typed_norms_tuple: tuple[str, ...],
    selected_names: tuple[str, ...],
    selected_filer_ids: tuple[int, ...],
) -> _page_detail_bundles.LobbyWorkspaceDetailBundle:
    app_state = get_app_state(path)
    return _page_detail_bundles.build_lobby_workspace_detail_bundle(
        _workspace_data(path, session_val, tfl_session_val),
        name_to_short=app_state.name_to_short,
        short_to_names=app_state.short_to_names,
        session=str(session_val or ""),
        tfl_session_val=tfl_session_val,
        lobbyshort=str(lobbyshort or ""),
        typed_norms_tuple=typed_norms_tuple or tuple(),
        selected_names=selected_names or tuple(),
        selected_filer_ids=selected_filer_ids or tuple(),
    )


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def get_map_forensics_bundle(
    path: str,
    scope: str,
    session_for_filter: str | None,
    selected_subdivision_signature: str,
) -> _map_runtime.MapForensicsBundle:
    atlas_bundle = get_map_atlas_bundle(path, scope, session_for_filter)
    return _map_runtime.build_map_forensics_bundle(
        atlas_bundle,
        selected_subdivision_signature=selected_subdivision_signature,
    )


_ui_runtime.configure_helpers(**globals())

fmt_usd = _ui_runtime.fmt_usd
_shorten_text = _ui_runtime._shorten_text
render_pill_list = _ui_runtime.render_pill_list
_current_filter_parts = _ui_runtime._current_filter_parts
_export_context_label = _ui_runtime._export_context_label
_export_filename = _ui_runtime._export_filename
export_dataframe = _ui_runtime.export_dataframe
require_columns = _ui_runtime.require_columns
reset_filters = _ui_runtime.reset_filters
_remember_recent_search = _ui_runtime._remember_recent_search
reset_client_filters = _ui_runtime.reset_client_filters
reset_member_filters = _ui_runtime.reset_member_filters
_remember_recent_client_search = _ui_runtime._remember_recent_client_search
_remember_recent_member_search = _ui_runtime._remember_recent_member_search
_ordinal = _ui_runtime._ordinal
_session_base_label = _ui_runtime._session_base_label
_session_label = _ui_runtime._session_label
_session_long_label = _ui_runtime._session_long_label
_session_range_label = _ui_runtime._session_range_label
_default_session_from_list = _ui_runtime._default_session_from_list
_slugify = _ui_runtime._slugify
_clean_options = _ui_runtime._clean_options
PDF_CHART_ERROR_KEY = _ui_runtime.PDF_CHART_ERROR_KEY
PLOTLY_CONFIG = _ui_runtime.PLOTLY_CONFIG
CHART_COLORS = _ui_runtime.CHART_COLORS
FUNDING_COLOR_MAP = _ui_runtime.FUNDING_COLOR_MAP
OPPOSITION_COLOR_MAP = _ui_runtime.OPPOSITION_COLOR_MAP
TREND_COLOR_MAP = _ui_runtime.TREND_COLOR_MAP
_apply_plotly_layout = _ui_runtime._apply_plotly_layout
_render_pdf_report_section = _ui_runtime._render_pdf_report_section


_PAGE_FRAGMENT_HELPER_KEYS = (
    "CHART_COLORS",
    "FUNDING_COLOR_MAP",
    "OPPOSITION_COLOR_MAP",
    "PLOTLY_CONFIG",
    "TEA_ARCGIS_WEBAPP_URL",
    "TREND_COLOR_MAP",
    "_apply_plotly_layout",
    "_clean_options",
    "_client_page",
    "_last_first_initial_key",
    "_lobby_page",
    "_map_page",
    "_member_page",
    "_session_base_label",
    "_session_label",
    "_shorten_text",
    "_solutions_page",
    "bill_position_from_flags",
    "build_activities",
    "build_activities_multi",
    "build_bills_with_status",
    "build_disclosures",
    "build_disclosures_multi",
    "build_lobby_subject_counts",
    "build_lobbyist_trend",
    "build_member_activities",
    "build_policy_mentions",
    "build_timeline_counts",
    "build_top_clients",
    "ensure_cols",
    "export_dataframe",
    "first_name_norm_series",
    "fmt_usd",
    "get_app_state",
    "get_app_tables",
    "get_client_scope_bundle",
    "get_client_workspace_detail_bundle",
    "get_lobby_scope_bundle",
    "get_lobby_workspace_detail_bundle",
    "get_member_session_bundle",
    "get_member_workspace_detail_bundle",
    "last_name_norm_from_text",
    "last_name_norm_series",
    "norm_name",
    "norm_name_series",
    "norm_person_variants",
    "norm_person_variants_with_nicknames",
    "parse_member_name",
    "parse_person_name",
    "pd",
    "px",
    "render_pill_list",
    "require_columns",
)
_MAP_FRAGMENT_HELPER_KEYS = (
    "MAP_BASEMAP_OPTIONS",
    "PATH",
    "ThreadPoolExecutor",
    "_atlas_bridge",
    "_build_filtered_atlas_bundle",
    "_build_filtered_forensics_bundle",
    "_map_runtime",
    "_mp5_confidence_weight",
    "_mp5_geocode_badge",
    "_mp5_method_weight",
    "_mp5_miles",
    "_session_cached_value",
    "_stable_json_signature",
    "_tfl_session_for_filter",
    "as_completed",
    "build_address_overlap_spending_rows",
    "build_overlap_map_points",
    "classify_requested_entity_type",
    "export_dataframe",
    "fmt_usd",
    "geocode_address_arcgis",
    "get_map_atlas_bundle",
    "get_map_forensics_bundle",
    "get_map_state",
    "pd",
    "px",
    "query_texas_subdivisions_for_point",
    "render_address_overlap_arcgis_map",
    "render_draw_area_search_map",
    "render_subdivision_map_legend",
    "render_tfl_subdivision_arcgis_map",
)
_CLIENT_WORKSPACE_CTX_KEYS = (
    "Bill_Status_All",
    "Bill_Sub_All",
    "Fiscal_Impact",
    "LaCvr",
    "LaDock",
    "LaI4E",
    "LaSub",
    "Lobby_Sub_All",
    "Lobby_TFL_Client_All",
    "Staff_All",
    "Wit_All",
    "all_clients",
    "all_stats",
    "client_scope_bundle",
    "data",
    "name_to_short",
    "tfl_session_val",
)
_MEMBER_WORKSPACE_CTX_KEYS = (
    "Lobby_TFL_Client_All",
    "Staff_All",
    "Wit_All",
    "all_leg_stats",
    "all_legislators",
    "author_bills_all",
    "data",
    "name_to_short",
    "short_to_names",
    "tfl_session_val",
)
_MAP_WORKSPACE_CTX_KEYS = (
    "_atlas_label",
    "_docket_label",
    "_forensics_label",
    "_open_client",
    "_render_cross_context_banner",
    "atlas_bundle",
    "subdivision_matches",
    "tfl_spend",
    "total_high",
)


def _build_fragment_ctx(keys: tuple[str, ...], values: dict[str, object]) -> dict[str, object]:
    return {key: values[key] for key in keys if key in values}


def _subset_globals(keys: tuple[str, ...]) -> dict[str, object]:
    scope = globals()
    return {key: scope[key] for key in keys if key in scope}

@functools.lru_cache(maxsize=None)
def _page_renderer_module(module_name: str):
    return importlib.import_module(module_name)


def _run_page_renderer(module_name: str, ctx: dict[str, object] | None = None) -> None:
    module = _page_renderer_module(module_name)
    helper_keys = tuple(getattr(module, "HELPER_KEYS", ()))
    if helper_keys:
        module.configure_helpers(**_subset_globals(helper_keys))
    module.render_page(ctx or {})


def _page_fragment_helpers() -> dict[str, object]:
    return _subset_globals(_PAGE_FRAGMENT_HELPER_KEYS)


def _map_fragment_helpers() -> dict[str, object]:
    return _subset_globals(_MAP_FRAGMENT_HELPER_KEYS)


_page_fragments.configure_page_fragment_helpers(**_page_fragment_helpers())
_map_fragments.configure_map_fragment_helpers(**_map_fragment_helpers())


_NAV_SEARCH_BUNDLE_KEY = "_nav_search_bundle_v1"
_NAV_SEARCH_QUERY_KEY = "_nav_search_bundle_query_v1"


def _remember_nav_search_bundle(bundle: NavSearchBundle) -> None:
    st.session_state[_NAV_SEARCH_BUNDLE_KEY] = bundle
    st.session_state[_NAV_SEARCH_QUERY_KEY] = bundle.normalized_query


def _cached_nav_search_bundle(query: str) -> NavSearchBundle | None:
    bundle = st.session_state.get(_NAV_SEARCH_BUNDLE_KEY)
    if can_reuse_nav_search_bundle(query, bundle):
        return bundle
    return None


DATA_SOURCE_LABELS = {
    "Wit_All": "Texas Legislature Online (Witness lists)",
    "Bill_Status_All": "Texas Legislature Online (Bill status)",
    "Fiscal_Impact": "Texas Legislature Online (Fiscal notes)",
    "Bill_Sub_All": "Texas Legislature Online (Bill subjects)",
    "Lobby_Sub_All": "Texas Ethics Commission (Subject matter filings)",
    "Lobby_TFL_Client_All": "Texas Ethics Commission (Lobbyist filings and compensation)",
    "Staff_All": "House Research Organization (Legislative staff lists)",
    "LaFood": "Texas Ethics Commission (Activity: Food)",
    "LaEnt": "Texas Ethics Commission (Activity: Entertainment)",
    "LaTran": "Texas Ethics Commission (Activity: Travel)",
    "LaGift": "Texas Ethics Commission (Activity: Gifts)",
    "LaEvnt": "Texas Ethics Commission (Activity: Events)",
    "LaAwrd": "Texas Ethics Commission (Activity: Awards)",
    "LaCvr": "Texas Ethics Commission (Disclosure: Coverage)",
    "LaDock": "Texas Ethics Commission (Disclosure: Docket)",
    "LaI4E": "Texas Ethics Commission (Disclosure: On Behalf)",
    "LaSub": "Texas Ethics Commission (Disclosure: Subject Matter)",
}

def _source_label(key: str) -> str:
    return DATA_SOURCE_LABELS.get(key, key)

@st.cache_data(show_spinner=False, ttl=3600, hash_funcs={dict: id})
def data_health_table(data: dict) -> pd.DataFrame:
    return _page_bundles.build_data_health_table(data or {}, DATA_SOURCE_LABELS)

# =========================================================
# ACTIVITIES (unchanged logic, still cached)
# =========================================================




nav_bundle = None
if nav_query and len(nav_query) >= 2 and PATH and (_is_url(PATH) or os.path.exists(PATH)):
    nav_bundle = build_nav_search_bundle_cached(NavQueryKey(nav_query), get_app_state(PATH))
    _remember_nav_search_bundle(nav_bundle)

nav_suggestions = list(nav_bundle.nav_suggestions) if nav_bundle else []
nav_suggestion_map = dict(nav_bundle.nav_suggestion_map) if nav_bundle else {}

if nav_suggestions:
    nav_pick = nav_suggest_slot.selectbox(
        "Nav suggestions",
        ["Select a match..."] + nav_suggestions,
        index=0,
        key="nav_suggestions_select",
        label_visibility="collapsed",
    )
    if nav_pick in nav_suggestion_map:
        nav_skip_submit = True
        target, value = nav_suggestion_map[nav_pick]
        nav_value = value if isinstance(value, str) else (value.get("name", "") or value.get("lobbyshort", ""))
        st.session_state.nav_search_query = nav_value
        st.session_state.nav_search_last = nav_value
        if target == "client":
            st.session_state.client_query = value
            st.session_state.client_query_input = value
            if _active_page != _client_page:
                st.switch_page(_client_page)
                st.stop()
        elif target == "member":
            st.session_state.member_query = value
            st.session_state.member_query_input = value
            if _active_page != _member_page:
                st.switch_page(_member_page)
                st.stop()
        else:
            sel = value if isinstance(value, dict) else {"lobbyshort": value, "name": value, "label": value, "filerid": None}
            sel_name = sel.get("name", "") or sel.get("lobbyshort", "")
            st.session_state.search_query = sel_name
            st.session_state.lobby_match_query = sel_name
            if sel.get("label"):
                st.session_state.lobby_match_select = sel.get("label")
            if sel.get("filerid") is not None:
                try:
                    st.session_state.lobby_filerid = int(sel.get("filerid"))
                except Exception:
                    st.session_state.lobby_filerid = sel.get("filerid")
            if sel.get("lobbyshort"):
                st.session_state.lobbyshort = sel.get("lobbyshort")
            if _active_page != _lobby_page:
                st.switch_page(_lobby_page)
                st.stop()
else:
    nav_suggest_slot.empty()

if nav_search_submitted and not nav_skip_submit:
    if nav_query:
        nav_bundle = _cached_nav_search_bundle(nav_query)
        if nav_bundle is None:
            nav_bundle = build_nav_search_bundle_cached(
                NavQueryKey(nav_query),
                require_app_state(
                    PATH,
                    missing_path_message="Data path not configured. Set the DATA_PATH environment variable.",
                    missing_file_message="Data path not found. Set DATA_PATH or place the parquet file in ./data.",
                ),
            )
            _remember_nav_search_bundle(nav_bundle)

        if nav_bundle.bill_query:
            st.session_state.search_query = nav_bundle.bill_query
            st.session_state.lobbyshort = ""
            if _active_page != _lobby_page:
                st.switch_page(_lobby_page)
                st.stop()
        else:
            resolved_client = nav_bundle.resolved_client
            client_suggestions = list(nav_bundle.client_suggestions)
            resolved_member = nav_bundle.resolved_member
            member_suggestions = list(nav_bundle.member_suggestions)
            lobby_candidates = list(nav_bundle.lobby_candidates)
            resolved_lobby = nav_bundle.resolved_lobby
            resolved_lobby_filer = nav_bundle.resolved_lobby_filer
            resolved_lobby_name = nav_bundle.resolved_lobby_name
            lobby_suggestions = list(nav_bundle.lobby_suggestions)

            target_page = _lobby_page
            if resolved_client:
                target_page = _client_page
                st.session_state.client_query = resolved_client
                st.session_state.client_query_input = resolved_client
            elif resolved_member:
                target_page = _member_page
                st.session_state.member_query = resolved_member
                st.session_state.member_query_input = resolved_member
            elif resolved_lobby:
                target_page = _lobby_page
                st.session_state.search_query = resolved_lobby_name or nav_query
                if resolved_lobby_filer is not None:
                    st.session_state.lobby_filerid = resolved_lobby_filer
                if lobby_candidates:
                    st.session_state.lobby_match_query = st.session_state.search_query
                    st.session_state.lobby_match_select = lobby_candidates[0].get("label", "")
            else:
                if "," in nav_query and member_suggestions:
                    target_page = _member_page
                elif client_suggestions and not member_suggestions:
                    target_page = _client_page
                elif member_suggestions and not client_suggestions:
                    target_page = _member_page
                elif client_suggestions:
                    target_page = _client_page
                elif member_suggestions:
                    target_page = _member_page
                elif lobby_suggestions:
                    target_page = _lobby_page

                if target_page == _client_page:
                    st.session_state.client_query = nav_query
                    st.session_state.client_query_input = nav_query
                elif target_page == _member_page:
                    st.session_state.member_query = nav_query
                    st.session_state.member_query_input = nav_query
                else:
                    st.session_state.search_query = nav_query

            if target_page != _active_page:
                st.switch_page(target_page)
                st.stop()

# Render the active page after nav-search routing completes.
_active_page.run()
st.stop()
