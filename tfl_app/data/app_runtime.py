from __future__ import annotations

import hashlib
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd

try:
    import streamlit as st
except ModuleNotFoundError:  # pragma: no cover - import smoke fallback
    class _CacheStub:
        def __call__(self, *decorator_args, **decorator_kwargs):
            if decorator_args and callable(decorator_args[0]) and len(decorator_args) == 1 and not decorator_kwargs:
                func = decorator_args[0]
                func.clear = lambda: None
                return func

            def decorator(func):
                func.clear = lambda: None
                return func

            return decorator

    class _StreamlitStub:
        cache_data = _CacheStub()
        cache_resource = _CacheStub()

        @staticmethod
        def error(*args, **kwargs) -> None:
            return None

        @staticmethod
        def stop() -> None:
            raise RuntimeError("streamlit stop")

    st = _StreamlitStub()

import tfl_app.map.state as _map_page_state
import tfl_app.search.state as _shared_search_state
import tfl_app.map.runtime as _map_runtime
import tfl_app.bundles.page_bundles as _page_bundles
import tfl_app.bundles.page_detail_bundles as _page_detail_bundles
import tfl_app.data.catalog as _catalog
import tfl_app.data.loaders as _loaders
from tfl_app.shared.sessions import add_session_from_year as _add_session_from_year
from tfl_app.shared.sessions import session_from_year as _session_from_year
from tfl_app.shared.sessions import tfl_session_for_filter as _tfl_session_for_filter
from tfl_app.map.geo_runtime import (
    build_tfl_political_subdivision_matches,
    classify_requested_entity_type,
)
from tfl_app.map.reference_runtime import (
    fetch_nctcog_transit_provider_centroids,
    get_reference_snapshot_version,
    fetch_tceq_groundwater_district_centroids,
    fetch_tceq_water_district_centroids,
    fetch_tea_county_centroids,
    fetch_tea_school_district_centroids,
    fetch_texas_city_centroids,
    fetch_texas_junior_college_centroids,
    fetch_texas_navigation_district_centroids,
    fetch_texas_rma_centroids,
    fetch_txdot_seaport_centroids,
)


def _is_url(path: str) -> bool:
    return str(path or "").startswith(("http://", "https://"))


_MONEY_RANGE = re.compile(r"(-?\d[\d,]*\.?\d*)\s*(?:-|to)\s*(-?\d[\d,]*\.?\d*)", re.IGNORECASE)
def _to_num_series(series: pd.Series) -> pd.Series:
    cleaned = series.fillna("").astype(str).str.strip()
    neg = cleaned.str.startswith("(") & cleaned.str.endswith(")")
    cleaned = cleaned.str.replace(r"^\(|\)$", "", regex=True)
    cleaned = cleaned.str.replace("$", "", regex=False).str.replace(",", "", regex=False)
    out = pd.to_numeric(cleaned, errors="coerce").fillna(0.0)
    return out.where(~neg, -out)


def add_low_high_numeric(df: pd.DataFrame) -> pd.DataFrame:
    data = _page_bundles.ensure_cols(df, {"Low": 0, "High": 0, "Amount": ""}).copy()

    low = _to_num_series(data["Low"])
    high = _to_num_series(data["High"])

    amount = data["Amount"].fillna("").astype(str).str.strip()
    amount_clean = (
        amount.str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace("\u2013", "-", regex=False)
    )
    rng = amount_clean.str.extract(_MONEY_RANGE)
    rng_low = pd.to_numeric(rng[0], errors="coerce").fillna(0.0)
    rng_high = pd.to_numeric(rng[1], errors="coerce").fillna(0.0)
    single = pd.to_numeric(amount_clean.str.extract(r"(-?\d+(?:\.\d+)?)")[0], errors="coerce").fillna(0.0)

    both_zero = (low == 0) & (high == 0)
    low = low.where(~both_zero, rng_low.where(rng_low != 0, single))
    high = high.where(~both_zero, rng_high.where(rng_high != 0, single))
    high = high.where(high != 0, low)
    low = low.where(low != 0, high)

    data["Low_num"] = low
    data["High_num"] = high
    return data
def safe_read_excel_xf(xf: pd.ExcelFile, sheet_name: str, cols: list[str]) -> pd.DataFrame:
    try:
        return xf.parse(sheet_name=sheet_name, usecols=cols)
    except Exception:
        try:
            df = xf.parse(sheet_name=sheet_name)
            keep = [column for column in cols if column in df.columns]
            return df[keep]
        except Exception:
            return pd.DataFrame(columns=cols)


@st.cache_data(show_spinner=False)
def _empty_df(cols: list[str]) -> pd.DataFrame:
    return pd.DataFrame(columns=cols)


def read_parquet_cols(path: Path, cols: list[str]) -> pd.DataFrame:
    try:
        import pyarrow.parquet as pq

        parquet_file = pq.ParquetFile(path)
        available = set(parquet_file.schema.names)
        use_cols = [column for column in cols if column in available]
        table = parquet_file.read(columns=use_cols) if use_cols else parquet_file.read()
        return table.to_pandas()
    except Exception:
        try:
            df = pd.read_parquet(path)
            keep = [column for column in cols if column in df.columns]
            return df[keep] if keep else df
        except Exception:
            return pd.DataFrame(columns=cols)


WORKBOOK_TABLE_COLUMNS = {
    "Wit_All": ["session", "bill", "position", "LobbyShort", "name", "org"],
    "Bill_Status_All": ["Session", "Bill", "Authors", "Author", "Caption", "Status", "Link"],
    "Fiscal_Impact": ["Session", "Bill", "Version", "EstimatedTwoYearNetImpactGR"],
    "Bill_Sub_All": ["Session", "Bill", "Subject"],
    "Lobby_Sub_All": [
        "Session",
        "session",
        "legislative_session",
        "Subject Matter",
        "Other Subject Matter Description",
        "Primary Business",
        "FilerID",
        "LobbyShort",
        "lobbyshort",
        "Lobby Name",
        "Unnamed: 0",
    ],
    "Lobbyist_Pol_Funds": [],
    "Lobby_TFL_Client_All": ["Session", "Client", "Lobby Name", "LobbyShort", "IsTFL", "Low", "High", "Amount", "Mid", "FilerID"],
    "Staff_All": [
        "Session",
        "session",
        "Legislator",
        "member_or_committee",
        "legislator_name",
        "Title",
        "role",
        "Staffer",
        "name",
        "staff_name_last_initial",
        "lobby name",
        "source",
    ],
    "LaFood": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst", "restaurantName", "activityDate", "periodStartDt", "activityExactAmount", "activityAmountRangeLow", "activityAmountRangeHigh", "activityAmountCd"],
    "LaEnt": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst", "entertainmentName", "activityDate", "periodStartDt", "activityExactAmount", "activityAmountRangeLow", "activityAmountRangeHigh", "activityAmountCd"],
    "LaTran": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst", "travelPurpose", "transportationTypeDescr", "departureCity", "arrivalCity", "checkInDt", "checkOutDt", "departureDt", "periodStartDt"],
    "LaGift": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst", "activityDescription", "periodStartDt", "activityExactAmount", "activityAmountRangeLow", "activityAmountRangeHigh", "activityAmountCd"],
    "LaEvnt": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst", "activityDescription", "activityDate", "periodStartDt"],
    "LaAwrd": ["Session", "applicableYear", "filerIdent", "filerName", "filerSort", "recipientNameOrganization", "recipientNameLast", "recipientNameFirst", "activityDescription", "periodStartDt", "activityExactAmount", "activityAmountRangeLow", "activityAmountRangeHigh", "activityAmountCd"],
    "LaCvr": ["Session", "filerIdent", "filerName", "filerSort", "filedDt", "periodStartDt", "sourceCategoryCd", "subjectMatterMemo", "docketsMemo", "filerNameOrganization"],
    "LaDock": ["Session", "filerIdent", "filerName", "filerSort", "receivedDt", "periodStartDt", "designationText", "agencyName"],
    "LaI4E": ["Session", "filerIdent", "filerName", "filerSort", "periodStartDt", "onbehalfName", "onbehalfMailingCity", "onbehalfPrimaryPhoneNumber"],
    "LaSub": ["Session", "filerIdent", "filerName", "filerSort", "periodStartDt", "subjectMatterCodeValue", "subjectMatterDescr"],
}
PARQUET_FILE_MAP = {
    "Wit_All": ["Witness_Lists.parquet", "Witness List.parquet", "Witness_List.parquet", "witnesslist.parquet"],
    "Bill_Status_All": "Bill_Status.parquet",
    "Fiscal_Impact": "Fiscal_Notes.parquet",
    "Bill_Sub_All": "Bill_Sub_All.parquet",
    "Lobby_Sub_All": "Lobby.Sub.parquet",
    "Lobbyist_Pol_Funds": "Lobbyist.Pol.Funds.parquet",
    "Lobby_TFL_Client_All": "Lobby_TFL_Client_All.parquet",
    "Staff_All": ["Staff.parquet", "staff.parquet"],
    "LaFood": "LaFood.parquet",
    "LaEnt": "LaEnt.parquet",
    "LaTran": "LaTran.parquet",
    "LaGift": "LaGift.parquet",
    "LaEvnt": "LaEvnt.parquet",
    "LaAwrd": "LaAwrd.parquet",
    "LaCvr": "LaCvr.parquet",
    "LaDock": "LaDock.parquet",
    "LaI4E": "LaI4E.parquet",
    "LaSub": "LaSub.parquet",
}
BASE_APP_STATE_TABLE_KEYS = (
    "Wit_All",
    "Bill_Status_All",
    "Lobby_TFL_Client_All",
    "Lobby_Sub_All",
    "Lobbyist_Pol_Funds",
)
APP_STATE_BOOTSTRAP_COLUMNS = {
    "Wit_All": ["Session", "session"],
    "Bill_Status_All": ["Session", "Bill", "Authors", "Author", "Caption", "Status", "Link", "Chamber"],
    "Lobby_TFL_Client_All": ["Session", "Client", "Lobby Name", "LobbyShort", "IsTFL", "FilerID"],
    "Lobby_Sub_All": ["Session", "session", "legislative_session", "Lobby Name", "LobbyShort", "lobbyshort", "FilerID"],
    "Lobbyist_Pol_Funds": ["Session", "legislative_session", "Lobbyist", "LobbyShort", "lobbyshort", "FilerID"],
}
CLIENT_DETAIL_TABLE_KEYS = (
    "Wit_All",
    "Bill_Status_All",
    "Lobby_Sub_All",
    "Fiscal_Impact",
    "Bill_Sub_All",
    "LaFood",
    "LaEnt",
    "LaTran",
    "LaGift",
    "LaEvnt",
    "LaAwrd",
    "LaCvr",
    "LaDock",
    "LaI4E",
    "LaSub",
)
MEMBER_DETAIL_TABLE_KEYS = (
    "Wit_All",
    "LaFood",
    "LaEnt",
    "LaTran",
    "LaGift",
    "LaEvnt",
    "LaAwrd",
)
LOBBY_DETAIL_TABLE_KEYS = (
    "Wit_All",
    "Bill_Status_All",
    "Lobby_Sub_All",
    "Fiscal_Impact",
    "Bill_Sub_All",
    "LaFood",
    "LaEnt",
    "LaTran",
    "LaGift",
    "LaEvnt",
    "LaAwrd",
    "LaCvr",
    "LaDock",
    "LaI4E",
    "LaSub",
)
SESSION_SCOPED_TABLE_KEYS = (
    "Wit_All",
    "Bill_Status_All",
    "Lobby_Sub_All",
    "Fiscal_Impact",
    "Bill_Sub_All",
    "LaFood",
    "LaEnt",
    "LaTran",
    "LaGift",
    "LaEvnt",
    "LaAwrd",
    "LaCvr",
    "LaDock",
    "LaI4E",
    "LaSub",
)
FILER_NORMALIZED_TABLE_KEYS = (
    "LaFood",
    "LaEnt",
    "LaTran",
    "LaGift",
    "LaEvnt",
    "LaAwrd",
    "LaCvr",
    "LaDock",
    "LaI4E",
    "LaSub",
)
ALL_WORKBOOK_TABLE_KEYS = tuple(WORKBOOK_TABLE_COLUMNS.keys())

# Public constants stay on this facade module, but their definitions now live
# in the split catalog module.
WORKBOOK_TABLE_COLUMNS = _catalog.WORKBOOK_TABLE_COLUMNS
PARQUET_FILE_MAP = _catalog.PARQUET_FILE_MAP
BASE_APP_STATE_TABLE_KEYS = _catalog.BASE_APP_STATE_TABLE_KEYS
APP_STATE_BOOTSTRAP_COLUMNS = _catalog.APP_STATE_BOOTSTRAP_COLUMNS
CLIENT_DETAIL_TABLE_KEYS = _catalog.CLIENT_DETAIL_TABLE_KEYS
MEMBER_DETAIL_TABLE_KEYS = _catalog.MEMBER_DETAIL_TABLE_KEYS
LOBBY_DETAIL_TABLE_KEYS = _catalog.LOBBY_DETAIL_TABLE_KEYS
SESSION_SCOPED_TABLE_KEYS = _catalog.SESSION_SCOPED_TABLE_KEYS
FILER_NORMALIZED_TABLE_KEYS = _catalog.FILER_NORMALIZED_TABLE_KEYS
ALL_WORKBOOK_TABLE_KEYS = _catalog.ALL_WORKBOOK_TABLE_KEYS


@dataclass(frozen=True)
class SessionOverlayBundle:
    session: str | None
    tables: dict[str, pd.DataFrame]
    witness_search: pd.DataFrame


def _fingerprint_paths(paths: list[Path], *, seed: str) -> str:
    digest = hashlib.sha1(str(seed or "").encode("utf-8"))
    for item in sorted({path.resolve() for path in paths if isinstance(path, Path)}, key=lambda path: str(path)):
        try:
            stat = item.stat()
            payload = f"{item}|{int(stat.st_size)}|{int(getattr(stat, 'st_mtime_ns', int(stat.st_mtime * 1_000_000_000)))}"
        except Exception:
            payload = str(item)
        digest.update(payload.encode("utf-8"))
    return digest.hexdigest()


def _normalize_loaded_table(table_key: str, df: pd.DataFrame) -> pd.DataFrame:
    data = df.copy()
    if table_key == "Wit_All":
        if "session" in data.columns and "Session" not in data.columns:
            data = data.rename(columns={"session": "Session"})
        if "bill" in data.columns and "Bill" not in data.columns:
            data = data.rename(columns={"bill": "Bill"})
        if "position" in data.columns:
            position = data["position"].fillna("").astype(str).str.upper()
            if "IsFor" not in data.columns:
                data["IsFor"] = position.str.contains(r"\bFOR\b").astype(int)
            if "IsAgainst" not in data.columns:
                data["IsAgainst"] = position.str.contains(r"\bAGAINST\b").astype(int)
            if "IsOn" not in data.columns:
                data["IsOn"] = position.str.contains(r"\bON\b").astype(int)
        if "LobbyShort" not in data.columns:
            data["LobbyShort"] = ""
        unnamed = [column for column in data.columns if str(column).startswith("Unnamed:")]
        if unnamed:
            data = data.drop(columns=unnamed)
    elif table_key == "Bill_Status_All":
        if "Authors" in data.columns and "Author" not in data.columns:
            data["Author"] = data["Authors"]
    elif table_key == "Lobby_TFL_Client_All":
        if "IsTFL" not in data.columns and "TFL?" in data.columns:
            data["IsTFL"] = data["TFL?"].astype(str).str.upper().isin(["Y", "YES", "TRUE", "1"]).astype(int)
        if "IsTFL" in data.columns:
            data["IsTFL"] = pd.to_numeric(data["IsTFL"], errors="coerce").fillna(0).astype(int)
        data = add_low_high_numeric(data)
    elif table_key == "Staff_All":
        if "session" in data.columns and "Session" not in data.columns:
            data = data.rename(columns={"session": "Session"})
        if "Legislator" not in data.columns:
            if "legislator_name" in data.columns:
                leg = data["legislator_name"].fillna("").astype(str).str.strip()
                if "member_or_committee" in data.columns:
                    fallback = data["member_or_committee"].fillna("").astype(str).str.strip()
                    data["Legislator"] = leg.where(leg != "", fallback)
                else:
                    data["Legislator"] = leg
            elif "member_or_committee" in data.columns:
                data["Legislator"] = data["member_or_committee"]
            else:
                data["Legislator"] = ""
        if "Title" not in data.columns:
            data["Title"] = data.get("role", "")
        if "Staffer" not in data.columns:
            data["Staffer"] = data.get("name", data.get("staff_name_last_initial", ""))
        if "lobby name" not in data.columns:
            data["lobby name"] = data.get("staff_name_last_initial", data.get("name", ""))
        data["StaffNameNorm"] = _shared_search_state.norm_name_series(data.get("name", pd.Series(dtype=object)))
        data["StaffLastInitialNorm"] = _shared_search_state.norm_name_series(
            data.get("staff_name_last_initial", data.get("name", pd.Series(dtype=object)))
        )
        data["StaffLastNorm"] = _shared_search_state.last_name_norm_series(
            data.get("name", data.get("staff_name_last_initial", pd.Series(dtype=object)))
        )
        if "Session" in data.columns:
            session = data["Session"].astype(str).str.strip()
            data["Session"] = session.where(~session.str.fullmatch(r"\d+"), session + "R")
    elif table_key == "Lobby_Sub_All":
        if "Session" not in data.columns:
            if "legislative_session" in data.columns:
                data = data.rename(columns={"legislative_session": "Session"})
            elif "session" in data.columns:
                data = data.rename(columns={"session": "Session"})
        if "LobbyShort" not in data.columns:
            if "lobbyshort" in data.columns:
                data = data.rename(columns={"lobbyshort": "LobbyShort"})
            elif "lobby_short" in data.columns:
                data = data.rename(columns={"lobby_short": "LobbyShort"})
    elif table_key == "Lobbyist_Pol_Funds":
        if "Session" not in data.columns and "legislative_session" in data.columns:
            data = data.rename(columns={"legislative_session": "Session"})
        if "LobbyShort" not in data.columns:
            if "lobbyshort" in data.columns:
                data = data.rename(columns={"lobbyshort": "LobbyShort"})
            elif "lobby_short" in data.columns:
                data = data.rename(columns={"lobby_short": "LobbyShort"})
    elif table_key in {"LaFood", "LaEnt", "LaTran", "LaGift", "LaEvnt", "LaAwrd", "LaCvr", "LaDock", "LaI4E", "LaSub"}:
        data = _add_session_from_year(data)

    if "Session" in data.columns:
        data["Session"] = data["Session"].fillna("").astype(str).str.strip()
    return data


def _ensure_filer_base_columns(df: pd.DataFrame) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df
    required = {"FilerID", "FilerNormRaw", "FilerNormClean", "FilerSortNorm"}
    if required.issubset(set(df.columns)):
        return df
    data = df.copy()
    if "FilerID" not in data.columns:
        if "filerIdent" in data.columns:
            data["FilerID"] = pd.to_numeric(data["filerIdent"], errors="coerce").fillna(-1).astype(int)
        else:
            data["FilerID"] = -1
    else:
        data["FilerID"] = pd.to_numeric(data["FilerID"], errors="coerce").fillna(-1).astype(int)
    filer_name = data.get("filerName", pd.Series("", index=data.index))
    filer_sort = data.get("filerSort", pd.Series("", index=data.index))
    if isinstance(filer_name, pd.DataFrame):
        filer_name = filer_name.iloc[:, 0]
    if isinstance(filer_sort, pd.DataFrame):
        filer_sort = filer_sort.iloc[:, 0]
    if "FilerNormRaw" not in data.columns:
        data["FilerNormRaw"] = _shared_search_state.norm_name_series(filer_name)
    if "FilerNormClean" not in data.columns:
        filer_clean = _shared_search_state.clean_filer_name_series(filer_name)
        data["FilerNormClean"] = _shared_search_state.norm_name_series(filer_clean)
    if "FilerSortNorm" not in data.columns:
        data["FilerSortNorm"] = _shared_search_state.norm_name_series(filer_sort)
    return data


def _postprocess_table_for_state(table_key: str, df: pd.DataFrame) -> pd.DataFrame:
    data = _normalize_loaded_table(table_key, df)
    if table_key == "Wit_All":
        return _shared_search_state._ensure_session_key_column(_shared_search_state._ensure_witness_search_columns(data))
    if table_key == "Bill_Status_All":
        return _shared_search_state._ensure_session_key_column(data)
    if table_key == "Lobby_TFL_Client_All":
        return _shared_search_state._ensure_lobby_client_lookup_columns(data)
    if table_key == "Staff_All":
        return _shared_search_state._ensure_session_key_column(_shared_search_state._ensure_staff_search_columns(data))
    if table_key == "Lobby_Sub_All":
        data = _shared_search_state._ensure_session_key_column(data)
        if "LobbyShort" in data.columns:
            data["LobbyShort"] = data["LobbyShort"].fillna("").astype(str).str.strip()
            data["LobbyShortNorm"] = _shared_search_state.norm_name_series(data["LobbyShort"])
        return data
    if table_key in FILER_NORMALIZED_TABLE_KEYS:
        return _shared_search_state._ensure_session_key_column(_ensure_filer_base_columns(data))
    return _shared_search_state._ensure_session_key_column(data)


def _resolve_table_source(path: str, table_key: str):
    base = Path(path)
    if not base.exists():
        return None
    if not base.is_dir():
        return ("excel", table_key)

    filename = PARQUET_FILE_MAP.get(table_key)
    if not filename:
        return None
    if isinstance(filename, (list, tuple)):
        matches = [base / candidate for candidate in filename if (base / candidate).exists()]
        return matches or None
    candidate = base / filename
    return candidate if candidate.exists() else None


@st.cache_data(show_spinner=False, max_entries=8)
def get_dataset_version(path: str) -> str:
    if not path:
        return ""
    if _is_url(path):
        return hashlib.sha1(str(path).strip().encode("utf-8")).hexdigest()
    base = Path(path)
    if not base.exists():
        return hashlib.sha1(str(base).encode("utf-8")).hexdigest()
    if base.is_file():
        return _fingerprint_paths([base], seed=str(base.resolve()))
    resolved_paths: list[Path] = []
    for table_key in ALL_WORKBOOK_TABLE_KEYS:
        source = _resolve_table_source(path, table_key)
        if isinstance(source, Path):
            resolved_paths.append(source)
        elif isinstance(source, list):
            resolved_paths.extend(item for item in source if isinstance(item, Path))
    if not resolved_paths:
        try:
            resolved_paths = [item for item in base.rglob("*") if item.is_file()]
        except Exception:
            resolved_paths = []
    return _fingerprint_paths(resolved_paths, seed=str(base.resolve()))


def _read_table_source(path: str, table_key: str, columns: list[str]) -> pd.DataFrame:
    source = _resolve_table_source(path, table_key)
    if source is None:
        return _empty_df(columns)
    if isinstance(source, tuple) and source and source[0] == "excel":
        try:
            xf = pd.ExcelFile(path, engine="openpyxl")
            return safe_read_excel_xf(xf, table_key, columns)
        except Exception:
            return _empty_df(columns)
    if isinstance(source, list):
        frames = []
        for item in source:
            try:
                frames.append(read_parquet_cols(item, columns))
            except Exception:
                continue
        return pd.concat(frames, ignore_index=True).drop_duplicates() if frames else _empty_df(columns)
    try:
        return read_parquet_cols(source, columns)
    except Exception:
        return _empty_df(columns)


def _table_keys_tuple(keys: tuple[str, ...] | list[str]) -> tuple[str, ...]:
    ordered: list[str] = []
    seen: set[str] = set()
    for key in keys:
        if key in WORKBOOK_TABLE_COLUMNS and key not in seen:
            ordered.append(key)
            seen.add(key)
    return tuple(ordered)


@st.cache_resource(show_spinner=False, max_entries=128)
def _load_table_resource(path: str, table_key: str, data_version: str) -> pd.DataFrame:
    del data_version
    columns = WORKBOOK_TABLE_COLUMNS.get(table_key, [])
    raw = _read_table_source(path, table_key, columns)
    return _postprocess_table_for_state(table_key, raw)


def get_app_table(path: str, table_key: str, *, copy: bool = True) -> pd.DataFrame:
    data = _load_table_resource(path, table_key, get_dataset_version(path))
    return data.copy() if copy else data


def get_app_tables(path: str, keys: tuple[str, ...] | list[str], *, copy: bool = True) -> dict[str, pd.DataFrame]:
    return {
        key: get_app_table(path, key, copy=copy)
        for key in _table_keys_tuple(keys)
    }


def get_app_table_readonly(path: str, table_key: str) -> pd.DataFrame:
    return get_app_table(path, table_key, copy=False)


def get_app_tables_readonly(path: str, keys: tuple[str, ...] | list[str]) -> dict[str, pd.DataFrame]:
    return {
        key: get_app_table_readonly(path, key)
        for key in _table_keys_tuple(keys)
    }


def _dedupe_columns(columns: list[str] | tuple[str, ...]) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()
    for column in columns:
        key = str(column or "").strip()
        if key and key not in seen:
            ordered.append(key)
            seen.add(key)
    return ordered


@st.cache_resource(show_spinner=False, max_entries=64)
def _load_projected_table_resource(path: str, table_key: str, columns_key: tuple[str, ...], data_version: str) -> pd.DataFrame:
    del data_version
    raw = _read_table_source(path, table_key, list(columns_key))
    return _postprocess_table_for_state(table_key, raw)


def _get_app_state_bootstrap_tables(path: str, data_version: str, *, copy: bool = True) -> dict[str, pd.DataFrame]:
    tables: dict[str, pd.DataFrame] = {}
    for table_key in _table_keys_tuple(BASE_APP_STATE_TABLE_KEYS):
        columns = tuple(_dedupe_columns(APP_STATE_BOOTSTRAP_COLUMNS.get(table_key, WORKBOOK_TABLE_COLUMNS.get(table_key, []))))
        frame = _load_projected_table_resource(path, table_key, columns, data_version)
        tables[table_key] = frame.copy() if copy else frame
    tables["table_manifest"] = _get_table_manifest_cached(path, data_version)
    return tables


def _filter_table_by_session(df: pd.DataFrame, session_val: str | None, *, copy: bool = True) -> pd.DataFrame:
    session = str(session_val or "").strip()
    if not session or not isinstance(df, pd.DataFrame) or df.empty:
        return df.copy() if copy else df
    if "SessionKey" in df.columns:
        session_key = df["SessionKey"]
        if session_key.hasnans:
            session_key = session_key.fillna("")
        out = df.loc[session_key == session]
        return out.copy() if copy else out
    if "Session" in df.columns:
        session_col = df["Session"].fillna("").astype(str).str.strip()
        out = df.loc[session_col == session]
        return out.copy() if copy else out
    if "session" in df.columns:
        session_col = df["session"].fillna("").astype(str).str.strip()
        out = df.loc[session_col == session]
        return out.copy() if copy else out
    return df.copy() if copy else df


def _ensure_filer_lookup_columns(
    df: pd.DataFrame,
    *,
    name_to_short: dict[str, str],
    filerid_to_short: dict[int, str],
) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df

    required = {"FilerID", "FilerShortFromId", "FilerNormRaw", "FilerNormClean", "FilerSortNorm", "FilerShortMapped"}
    if required.issubset(set(df.columns)):
        return df

    data = _ensure_filer_base_columns(df)

    if "FilerShortFromId" not in data.columns:
        data["FilerShortFromId"] = data["FilerID"].map(filerid_to_short).fillna("") if filerid_to_short else ""

    filer_name = data.get("filerName", pd.Series("", index=data.index))
    filer_sort = data.get("filerSort", pd.Series("", index=data.index))
    if isinstance(filer_name, pd.DataFrame):
        filer_name = filer_name.iloc[:, 0]
    if isinstance(filer_sort, pd.DataFrame):
        filer_sort = filer_sort.iloc[:, 0]

    if "FilerNormRaw" not in data.columns:
        data["FilerNormRaw"] = _shared_search_state.norm_name_series(filer_name)
    if "FilerNormClean" not in data.columns:
        filer_clean = _shared_search_state.clean_filer_name_series(filer_name)
        data["FilerNormClean"] = _shared_search_state.norm_name_series(filer_clean)
    if "FilerSortNorm" not in data.columns:
        data["FilerSortNorm"] = _shared_search_state.norm_name_series(filer_sort)
    if "FilerShortMapped" not in data.columns:
        mapped = data["FilerNormRaw"].map(name_to_short)
        mapped = mapped.where(mapped.notna(), data["FilerNormClean"].map(name_to_short))
        mapped = mapped.where(mapped.notna(), data["FilerSortNorm"].map(name_to_short))
        data["FilerShortMapped"] = mapped.fillna("")
    return data


@st.cache_resource(show_spinner=False, max_entries=16)
def _get_witness_table_resource(path: str, data_version: str) -> pd.DataFrame:
    app_state = _get_app_state_cached(path, data_version)
    witness = get_app_table_readonly(path, "Wit_All")
    return _shared_search_state._fill_missing_witness_lobbyshorts(
        witness,
        name_to_short=app_state.name_to_short,
    )


@st.cache_resource(show_spinner=False, max_entries=16)
def _get_witness_search_table_resource(path: str, data_version: str) -> pd.DataFrame:
    return _shared_search_state._ensure_witness_search_columns(
        _get_witness_table_resource(path, data_version)
    )


@st.cache_resource(show_spinner=False, max_entries=16)
def _get_lobby_sub_lookup_table_resource(path: str, data_version: str) -> pd.DataFrame:
    app_state = _get_app_state_cached(path, data_version)
    lobby_sub = get_app_table_readonly(path, "Lobby_Sub_All")
    if lobby_sub.empty:
        return lobby_sub

    data = lobby_sub.copy()
    if "LobbyShort" not in data.columns:
        data["LobbyShort"] = ""
    else:
        data["LobbyShort"] = data["LobbyShort"].fillna("").astype(str).str.strip()
    if app_state.filerid_to_short and "FilerID" in data.columns:
        fid = pd.to_numeric(data["FilerID"], errors="coerce").fillna(-1).astype(int)
        missing = data["LobbyShort"].eq("")
        if missing.any():
            data.loc[missing, "LobbyShort"] = fid.loc[missing].map(app_state.filerid_to_short).fillna("")
    if "LobbyShortNorm" not in data.columns:
        data["LobbyShortNorm"] = _shared_search_state.norm_name_series(data["LobbyShort"])
    return data


@st.cache_resource(show_spinner=False, max_entries=64)
def _get_filer_lookup_table_resource(path: str, table_key: str, data_version: str) -> pd.DataFrame:
    app_state = _get_app_state_cached(path, data_version)
    rows = get_app_table_readonly(path, table_key)
    return _ensure_filer_lookup_columns(
        rows,
        name_to_short=app_state.name_to_short,
        filerid_to_short=app_state.filerid_to_short,
    )


@st.cache_data(show_spinner=False, max_entries=64)
def _get_witness_rows_for_session(
    path: str,
    data_version: str,
    session_val: str | None,
    *,
    include_name_columns: bool,
) -> pd.DataFrame:
    witness_table = (
        _get_witness_search_table_resource(path, data_version)
        if include_name_columns
        else _get_witness_table_resource(path, data_version)
    )
    witness_rows = _filter_table_by_session(
        witness_table,
        session_val,
        copy=True,
    )
    return witness_rows


@st.cache_data(show_spinner=False, max_entries=64)
def _get_lobby_sub_rows_for_session(path: str, data_version: str, session_val: str | None) -> pd.DataFrame:
    lobby_sub = _filter_table_by_session(
        _get_lobby_sub_lookup_table_resource(path, data_version),
        session_val,
        copy=True,
    )
    return lobby_sub


@st.cache_data(show_spinner=False, max_entries=128)
def _get_filer_rows_for_session(path: str, table_key: str, data_version: str, session_val: str | None) -> pd.DataFrame:
    rows = _filter_table_by_session(
        _get_filer_lookup_table_resource(path, table_key, data_version),
        session_val,
        copy=True,
    )
    return rows


@st.cache_data(show_spinner=False, max_entries=32)
def _get_session_overlay_bundle(
    path: str,
    data_version: str,
    session_val: str | None,
) -> SessionOverlayBundle:
    session = str(session_val or "").strip()
    tables: dict[str, pd.DataFrame] = {}
    witness_rows = _get_witness_rows_for_session(path, data_version, session, include_name_columns=False)
    witness_search_rows = _get_witness_rows_for_session(path, data_version, session, include_name_columns=True)
    tables["Wit_All"] = witness_rows
    tables["Lobby_Sub_All"] = _get_lobby_sub_rows_for_session(path, data_version, session)
    for table_key in FILER_NORMALIZED_TABLE_KEYS:
        tables[table_key] = _get_filer_rows_for_session(path, table_key, data_version, session)
    for table_key in SESSION_SCOPED_TABLE_KEYS:
        if table_key in tables:
            continue
        tables[table_key] = _filter_table_by_session(
            get_app_table_readonly(path, table_key),
            session,
            copy=True,
        )
    return SessionOverlayBundle(
        session=session or None,
        tables=tables,
        witness_search=witness_search_rows,
    )


@st.cache_data(show_spinner=False, max_entries=64)
def _get_workspace_table_overlays_for_keys(
    path: str,
    data_version: str,
    session_val: str | None,
    keys: tuple[str, ...],
) -> dict[str, pd.DataFrame]:
    overlay_bundle = _get_session_overlay_bundle(path, data_version, session_val)
    overlays: dict[str, pd.DataFrame] = {}
    for key in _table_keys_tuple(keys):
        table = overlay_bundle.tables.get(key)
        if isinstance(table, pd.DataFrame):
            overlays[key] = table.copy()
            continue
        overlays[key] = _filter_table_by_session(
            get_app_table(path, key, copy=False),
            overlay_bundle.session,
            copy=True,
        )
    return overlays


@st.cache_data(show_spinner=False, max_entries=32)
def get_workspace_table_overlays(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
) -> dict[str, pd.DataFrame]:
    del tfl_session_val
    return _get_workspace_table_overlays_for_keys(path, get_dataset_version(path), session_val, SESSION_SCOPED_TABLE_KEYS)


def _normalized_manifest_column_count(table_key: str, source_columns: list[str]) -> int:
    sample = pd.DataFrame({column: pd.Series(dtype="object") for column in source_columns})
    return int(len(_postprocess_table_for_state(table_key, sample).columns))


def _read_manifest_probe(path: str, table_key: str, source_columns: list[str]) -> pd.DataFrame:
    probe_columns: list[str] = []
    for candidate in ("Session", "session", "applicableYear", "LobbyShort", "lobbyshort"):
        if candidate in source_columns and candidate not in probe_columns:
            probe_columns.append(candidate)
    if not probe_columns:
        return pd.DataFrame()
    return _read_table_source(path, table_key, probe_columns)


def _empty_manifest_entry(table_key: str) -> dict[str, object]:
    return {
        "rows": 0,
        "cols": int(len(_postprocess_table_for_state(table_key, _empty_df(WORKBOOK_TABLE_COLUMNS.get(table_key, []))).columns)),
        "has_session": False,
        "empty": True,
        "sessions": 0,
        "lobby_count": 0,
    }


@st.cache_data(show_spinner=False, max_entries=128)
def _get_table_manifest_entry(path: str, table_key: str, data_version: str) -> dict[str, object]:
    del data_version
    base = Path(path)
    source = _resolve_table_source(path, table_key)
    source_columns: list[str] = []
    row_count = 0
    if source is None:
        return _empty_manifest_entry(table_key)

    if isinstance(source, tuple) and source and source[0] == "excel":
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(base, read_only=True, data_only=True)
            if table_key in workbook.sheetnames:
                sheet = workbook[table_key]
                header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                source_columns = [str(value) for value in header if value is not None]
                row_count = max(int(sheet.max_row or 0) - 1, 0)
        except Exception:
            source_columns = list(WORKBOOK_TABLE_COLUMNS.get(table_key, []))
    else:
        try:
            import pyarrow.parquet as pq

            sources = source if isinstance(source, list) else [source]
            row_total = 0
            column_names: set[str] = set()
            for item in sources:
                parquet_file = pq.ParquetFile(item)
                row_total += int(parquet_file.metadata.num_rows or 0)
                column_names.update(str(name) for name in parquet_file.schema.names)
            source_columns = [column for column in WORKBOOK_TABLE_COLUMNS.get(table_key, []) if column in column_names]
            if not source_columns:
                source_columns = sorted(column_names)
            row_count = row_total
        except Exception:
            probe = _read_table_source(path, table_key, WORKBOOK_TABLE_COLUMNS.get(table_key, []))
            source_columns = list(probe.columns)
            row_count = int(len(probe))

    probe = _normalize_loaded_table(table_key, _read_manifest_probe(path, table_key, source_columns))
    session_count = int(probe["Session"].dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique()) if "Session" in probe.columns else 0
    lobby_count = int(probe["LobbyShort"].dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique()) if "LobbyShort" in probe.columns else 0
    return {
        "rows": int(row_count),
        "cols": _normalized_manifest_column_count(table_key, source_columns),
        "has_session": bool("Session" in probe.columns),
        "empty": int(row_count) == 0,
        "sessions": session_count,
        "lobby_count": lobby_count,
    }


@st.cache_data(show_spinner=False, max_entries=4)
def _get_table_manifest_cached(path: str, data_version: str) -> dict[str, dict[str, object]]:
    return {
        table_key: dict(_get_table_manifest_entry(path, table_key, data_version))
        for table_key in ALL_WORKBOOK_TABLE_KEYS
    }


def get_table_manifest(path: str) -> dict[str, dict[str, object]]:
    return _get_table_manifest_cached(path, get_dataset_version(path))


get_table_manifest.clear = getattr(_get_table_manifest_cached, "clear", lambda: None)


@st.cache_resource(show_spinner=False, max_entries=2)
def _load_workbook_cached(path: str, data_version: str) -> dict[str, object]:
    del data_version
    data = get_app_tables(path, ALL_WORKBOOK_TABLE_KEYS, copy=True)
    data["table_manifest"] = get_table_manifest(path)
    return data


def load_workbook(path: str) -> dict[str, object]:
    return _load_workbook_cached(path, get_dataset_version(path))


load_workbook.clear = getattr(_load_workbook_cached, "clear", lambda: None)

# Rebind the facade to the split loader implementations after the legacy local
# definitions above so subsequent runtime calls resolve through the new module.
_is_url = _loaders._is_url
add_low_high_numeric = _loaders.add_low_high_numeric
safe_read_excel_xf = _loaders.safe_read_excel_xf
_empty_df = _loaders._empty_df
read_parquet_cols = _loaders.read_parquet_cols
_fingerprint_paths = _loaders._fingerprint_paths
_normalize_loaded_table = _loaders._normalize_loaded_table
_postprocess_table_for_state = _loaders._postprocess_table_for_state
_resolve_table_source = _loaders._resolve_table_source
get_dataset_version = _loaders.get_dataset_version
_read_table_source = _loaders._read_table_source
_table_keys_tuple = _loaders._table_keys_tuple
_dedupe_columns = _loaders._dedupe_columns
_normalized_manifest_column_count = _loaders._normalized_manifest_column_count
_read_manifest_probe = _loaders._read_manifest_probe
_empty_manifest_entry = _loaders._empty_manifest_entry
_get_table_manifest_entry = _loaders._get_table_manifest_entry
_get_table_manifest_cached = _loaders._get_table_manifest_cached
get_table_manifest = _loaders.get_table_manifest


@st.cache_resource(show_spinner=False, max_entries=2)
def _get_app_state_cached(path: str, data_version: str) -> _shared_search_state.AppState:
    return _shared_search_state.build_app_state(
        path,
        _get_app_state_bootstrap_tables(path, data_version, copy=False),
        data_version=data_version,
    )


def get_app_state(path: str) -> _shared_search_state.AppState:
    return _get_app_state_cached(path, get_dataset_version(path))


get_app_state.clear = getattr(_get_app_state_cached, "clear", lambda: None)


@st.cache_data(show_spinner=False, max_entries=1)
def _fetch_map_reference_tables_cached(reference_version: str) -> dict[str, pd.DataFrame]:
    del reference_version
    return {
        "school_districts": fetch_tea_school_district_centroids(),
        "counties": fetch_tea_county_centroids(),
        "cities": fetch_texas_city_centroids(),
        "water_districts": fetch_tceq_water_district_centroids(),
        "groundwater_districts": fetch_tceq_groundwater_district_centroids(),
        "regional_mobility_authorities": fetch_texas_rma_centroids(),
        "junior_colleges": fetch_texas_junior_college_centroids(),
        "navigation_districts": fetch_texas_navigation_district_centroids(),
        "transit_providers": fetch_nctcog_transit_provider_centroids(),
        "seaports": fetch_txdot_seaport_centroids(),
    }


@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def _build_map_client_matches_cached(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    return build_tfl_political_subdivision_matches(tfl_client_names)


@st.cache_data(show_spinner=False, ttl=3600, max_entries=16)
def _build_map_client_edges_cached(tfl_client_names: tuple[str, ...]) -> pd.DataFrame:
    matches = _build_map_client_matches_cached(tfl_client_names)
    return _map_page_state.build_client_subdivision_edges(matches)


@st.cache_resource(show_spinner=False, max_entries=2)
def _get_map_state_cached(path: str, data_version: str, reference_version: str) -> _map_page_state.MapState:
    del data_version
    return _map_page_state.build_map_state_from_sources(
        path,
        get_app_tables(path, ("Lobby_TFL_Client_All",), copy=False),
        classify_entity_type=classify_requested_entity_type,
        fetch_reference_tables=lambda: _fetch_map_reference_tables_cached(reference_version),
    )


def get_map_state(path: str) -> _map_page_state.MapState:
    return _get_map_state_cached(
        path,
        get_dataset_version(path),
        get_reference_snapshot_version(),
    )


get_map_state.clear = getattr(_get_map_state_cached, "clear", lambda: None)


def require_app_state(
    path: str,
    *,
    missing_path_message: str,
    missing_file_message: str,
) -> _shared_search_state.AppState:
    if not path:
        st.error(missing_path_message)
        st.stop()
    if not _is_url(path) and not os.path.exists(path):
        st.error(missing_file_message)
        st.stop()
    return get_app_state(path)


def require_map_state(
    path: str,
    *,
    missing_path_message: str,
    missing_file_message: str,
) -> _map_page_state.MapState:
    if not path:
        st.error(missing_path_message)
        st.stop()
    if not _is_url(path) and not os.path.exists(path):
        st.error(missing_file_message)
        st.stop()
    return get_map_state(path)


@st.cache_resource(show_spinner=False, max_entries=16)
def _get_map_atlas_bundle_cached(
    path: str,
    scope: str,
    session_for_filter: str | None,
    data_version: str,
    reference_version: str,
) -> _map_page_state.AtlasBundle:
    map_state = _get_map_state_cached(path, data_version, reference_version)
    active_tfl_clients = _map_page_state.resolve_active_tfl_clients(
        map_state,
        scope=scope,
        session_for_filter=session_for_filter,
    )
    client_edges = _build_map_client_edges_cached(active_tfl_clients)
    return _map_page_state.build_atlas_bundle(
        map_state,
        scope=scope,
        session_for_filter=session_for_filter,
        client_subdivision_edges_all=client_edges,
    )


def get_map_atlas_bundle(
    path: str,
    scope: str,
    session_for_filter: str | None,
) -> _map_page_state.AtlasBundle:
    return _get_map_atlas_bundle_cached(
        path,
        scope,
        session_for_filter,
        get_dataset_version(path),
        get_reference_snapshot_version(),
    )


get_map_atlas_bundle.clear = getattr(_get_map_atlas_bundle_cached, "clear", lambda: None)


def _build_client_scope_stats(overview: pd.DataFrame) -> dict[str, Any]:
    if overview.empty:
        return {}
    return {
        "total_clients": int(overview["Client"].nunique()),
        "tfl_clients": int((overview["IsTFL"] == 1).sum()),
        "private_clients": int((overview["IsTFL"] == 0).sum()),
        "tfl_low_total": float(overview.loc[overview["IsTFL"] == 1, "Low"].sum()),
        "tfl_high_total": float(overview.loc[overview["IsTFL"] == 1, "High"].sum()),
        "pri_low_total": float(overview.loc[overview["IsTFL"] == 0, "Low"].sum()),
        "pri_high_total": float(overview.loc[overview["IsTFL"] == 0, "High"].sum()),
    }


def _build_lobby_scope_stats(all_pivot: pd.DataFrame) -> dict[str, Any]:
    if all_pivot.empty:
        return {}
    return {
        "total_lobbyists": int(all_pivot["LobbyShort"].nunique()),
        "has_tfl": int(all_pivot["Has_TFL"].sum()),
        "only_private": int(all_pivot["Only_Private"].sum()),
        "only_tfl": int(all_pivot["Only_TFL"].sum()),
        "mixed": int(all_pivot["Mixed"].sum()),
        "tfl_low_total": float(all_pivot["Low_TFL"].sum()),
        "tfl_high_total": float(all_pivot["High_TFL"].sum()),
        "pri_low_total": float(all_pivot["Low_Private"].sum()),
        "pri_high_total": float(all_pivot["High_Private"].sum()),
    }


@st.cache_data(show_spinner=False, max_entries=16)
def _get_client_scope_bundle_cached(
    path: str,
    scope: str,
    session_val: str | None,
    data_version: str,
) -> _page_bundles.ClientScopeBundle:
    app_state = _get_app_state_cached(path, data_version)
    overview = app_state.client_scope_overview_all
    session = str(session_val or "").strip()
    if scope == "This Session" and session:
        overview = app_state.client_scope_overview_by_session
        overview = overview[overview["SessionKey"].astype(str) == session].drop(columns=["SessionKey"], errors="ignore")
    overview = overview.reset_index(drop=True).copy()
    return _page_bundles.ClientScopeBundle(
        overview=overview,
        stats=_build_client_scope_stats(overview),
        category_chart_data=app_state.client_category_chart_data,
    )


def get_client_scope_bundle(path: str, scope: str, session_val: str | None) -> _page_bundles.ClientScopeBundle:
    return _get_client_scope_bundle_cached(path, scope, session_val, get_dataset_version(path))


get_client_scope_bundle.clear = getattr(_get_client_scope_bundle_cached, "clear", lambda: None)


@st.cache_data(show_spinner=False, max_entries=16)
def _get_lobby_scope_bundle_cached(path: str, scope: str, session_val: str | None, data_version: str) -> _page_bundles.LobbyScopeBundle:
    app_state = _get_app_state_cached(path, data_version)
    session = str(session_val or "").strip()
    all_pivot = app_state.lobby_scope_pivot_all
    top_clients = app_state.lobby_scope_top_clients_all
    if scope == "This Session" and session:
        all_pivot = app_state.lobby_scope_pivot_by_session
        all_pivot = all_pivot[all_pivot["SessionKey"].astype(str) == session].drop(columns=["SessionKey"], errors="ignore")
        top_clients = app_state.lobby_scope_top_clients_by_session
        top_clients = top_clients[top_clients["SessionKey"].astype(str) == session].drop(columns=["SessionKey"], errors="ignore")
    all_pivot = all_pivot.reset_index(drop=True).copy()
    top_clients = top_clients.reset_index(drop=True).copy()
    return _page_bundles.LobbyScopeBundle(
        all_pivot=all_pivot,
        all_stats=_build_lobby_scope_stats(all_pivot),
        trend_group=app_state.lobby_scope_trend_group,
        top_clients=top_clients,
        lobby_display=app_state.lobby_display,
    )


def get_lobby_scope_bundle(path: str, scope: str, session_val: str | None) -> _page_bundles.LobbyScopeBundle:
    return _get_lobby_scope_bundle_cached(
        path,
        scope,
        session_val,
        get_dataset_version(path),
    )


get_lobby_scope_bundle.clear = getattr(_get_lobby_scope_bundle_cached, "clear", lambda: None)


@st.cache_data(show_spinner=False, max_entries=16)
def _get_member_session_bundle_cached(path: str, session_val: str | None, data_version: str) -> _page_bundles.MemberSessionBundle:
    app_state = _get_app_state_cached(path, data_version)
    witness_rows = _get_witness_rows_for_session(path, data_version, session_val, include_name_columns=False)
    return _page_bundles.build_member_session_bundle(
        app_state.author_bills_all,
        witness_rows,
        str(session_val or ""),
    )


def get_member_session_bundle(path: str, session_val: str | None) -> _page_bundles.MemberSessionBundle:
    return _get_member_session_bundle_cached(path, session_val, get_dataset_version(path))


get_member_session_bundle.clear = getattr(_get_member_session_bundle_cached, "clear", lambda: None)


def _detail_base_data(app_state: _shared_search_state.AppState, path: str) -> dict[str, object]:
    return {
        "Lobby_TFL_Client_All": get_app_table_readonly(path, "Lobby_TFL_Client_All"),
        "Staff_All": get_app_table_readonly(path, "Staff_All"),
        "filerid_to_short": app_state.filerid_to_short,
    }


def _client_workspace_data(
    app_state: _shared_search_state.AppState,
    path: str,
    session_val: str | None,
) -> dict[str, object]:
    data = _detail_base_data(app_state, path)
    data.update(_get_workspace_table_overlays_for_keys(path, app_state.data_version, session_val, CLIENT_DETAIL_TABLE_KEYS))
    return data


def _member_workspace_data(
    app_state: _shared_search_state.AppState,
    path: str,
    session_val: str | None,
) -> dict[str, object]:
    data = _detail_base_data(app_state, path)
    data.update(_get_workspace_table_overlays_for_keys(path, app_state.data_version, session_val, MEMBER_DETAIL_TABLE_KEYS))
    return data


def get_witness_name_match_table(path: str, session_val: str | None) -> pd.DataFrame:
    overlay_bundle = _get_session_overlay_bundle(path, get_dataset_version(path), session_val)
    return overlay_bundle.witness_search.copy()


get_witness_name_match_table.clear = getattr(_get_witness_rows_for_session, "clear", lambda: None)
if hasattr(_get_session_overlay_bundle, "clear"):
    get_witness_name_match_table.clear = _get_session_overlay_bundle.clear


def _lobby_workspace_data(
    app_state: _shared_search_state.AppState,
    path: str,
    session_val: str | None,
    *,
    include_witness_name_columns: bool,
) -> dict[str, object]:
    data = _detail_base_data(app_state, path)
    overlays = _get_workspace_table_overlays_for_keys(path, app_state.data_version, session_val, LOBBY_DETAIL_TABLE_KEYS)
    if include_witness_name_columns:
        overlays["Wit_All"] = get_witness_name_match_table(path, session_val)
    data.update(overlays)
    return data


@st.cache_data(show_spinner=False, max_entries=16)
def _get_client_workspace_detail_bundle_cached(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
    client_name: str,
    data_version: str,
) -> _page_detail_bundles.ClientWorkspaceDetailBundle:
    app_state = _get_app_state_cached(path, data_version)
    return _page_detail_bundles.build_client_workspace_detail_bundle(
        _client_workspace_data(app_state, path, session_val),
        name_to_short=app_state.name_to_short,
        session=str(session_val or ""),
        tfl_session_val=tfl_session_val,
        client_name=str(client_name or ""),
    )


def get_client_workspace_detail_bundle(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
    client_name: str,
) -> _page_detail_bundles.ClientWorkspaceDetailBundle:
    return _get_client_workspace_detail_bundle_cached(
        path,
        session_val,
        tfl_session_val,
        client_name,
        get_dataset_version(path),
    )


get_client_workspace_detail_bundle.clear = getattr(_get_client_workspace_detail_bundle_cached, "clear", lambda: None)


@st.cache_data(show_spinner=False, max_entries=16)
def _get_member_workspace_detail_bundle_cached(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
    member_name: str,
    data_version: str,
) -> _page_detail_bundles.MemberWorkspaceDetailBundle:
    app_state = _get_app_state_cached(path, data_version)
    return _page_detail_bundles.build_member_workspace_detail_bundle(
        _member_workspace_data(app_state, path, session_val),
        author_bills_all=app_state.author_bills_all,
        name_to_short=app_state.name_to_short,
        short_to_names=app_state.short_to_names,
        initial_to_short=app_state.initial_to_short,
        session=str(session_val or ""),
        tfl_session_val=tfl_session_val,
        member_name=str(member_name or ""),
    )


def get_member_workspace_detail_bundle(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
    member_name: str,
) -> _page_detail_bundles.MemberWorkspaceDetailBundle:
    return _get_member_workspace_detail_bundle_cached(
        path,
        session_val,
        tfl_session_val,
        member_name,
        get_dataset_version(path),
    )


get_member_workspace_detail_bundle.clear = getattr(_get_member_workspace_detail_bundle_cached, "clear", lambda: None)


@st.cache_data(show_spinner=False, max_entries=16)
def _get_lobby_workspace_detail_bundle_cached(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
    lobbyshort: str,
    typed_norms_tuple: tuple[str, ...],
    selected_names: tuple[str, ...],
    selected_filer_ids: tuple[int, ...],
    data_version: str,
) -> _page_detail_bundles.LobbyWorkspaceDetailBundle:
    app_state = _get_app_state_cached(path, data_version)
    return _page_detail_bundles.build_lobby_workspace_detail_bundle(
        _lobby_workspace_data(
            app_state,
            path,
            session_val,
            include_witness_name_columns=bool(selected_names),
        ),
        name_to_short=app_state.name_to_short,
        short_to_names=app_state.short_to_names,
        session=str(session_val or ""),
        tfl_session_val=tfl_session_val,
        lobbyshort=str(lobbyshort or ""),
        typed_norms_tuple=typed_norms_tuple or tuple(),
        selected_names=selected_names or tuple(),
        selected_filer_ids=selected_filer_ids or tuple(),
    )


def get_lobby_workspace_detail_bundle(
    path: str,
    session_val: str | None,
    tfl_session_val: str | None,
    lobbyshort: str,
    typed_norms_tuple: tuple[str, ...],
    selected_names: tuple[str, ...],
    selected_filer_ids: tuple[int, ...],
) -> _page_detail_bundles.LobbyWorkspaceDetailBundle:
    return _get_lobby_workspace_detail_bundle_cached(
        path,
        session_val,
        tfl_session_val,
        lobbyshort,
        typed_norms_tuple,
        selected_names,
        selected_filer_ids,
        get_dataset_version(path),
    )


get_lobby_workspace_detail_bundle.clear = getattr(_get_lobby_workspace_detail_bundle_cached, "clear", lambda: None)


@st.cache_data(show_spinner=False, max_entries=16)
def _get_map_forensics_bundle_cached(
    path: str,
    scope: str,
    session_for_filter: str | None,
    selected_subdivision_signature: str,
    data_version: str,
    reference_version: str,
) -> _map_runtime.MapForensicsBundle:
    atlas_bundle = _get_map_atlas_bundle_cached(path, scope, session_for_filter, data_version, reference_version)
    return _map_runtime.build_map_forensics_bundle(
        atlas_bundle,
        selected_subdivision_signature=selected_subdivision_signature,
    )


def get_map_forensics_bundle(
    path: str,
    scope: str,
    session_for_filter: str | None,
    selected_subdivision_signature: str,
) -> _map_runtime.MapForensicsBundle:
    data_version = get_dataset_version(path)
    reference_version = get_reference_snapshot_version()
    return _get_map_forensics_bundle_cached(
        path,
        scope,
        session_for_filter,
        selected_subdivision_signature,
        data_version,
        reference_version,
    )


get_map_forensics_bundle.clear = getattr(_get_map_forensics_bundle_cached, "clear", lambda: None)


__all__ = [
    "get_app_table",
    "get_app_tables",
    "get_app_table_readonly",
    "get_app_tables_readonly",
    "get_workspace_table_overlays",
    "get_witness_name_match_table",
    "get_table_manifest",
    "load_workbook",
    "get_app_state",
    "get_map_state",
    "require_app_state",
    "require_map_state",
    "get_map_atlas_bundle",
    "get_client_scope_bundle",
    "get_lobby_scope_bundle",
    "get_member_session_bundle",
    "get_client_workspace_detail_bundle",
    "get_member_workspace_detail_bundle",
    "get_lobby_workspace_detail_bundle",
    "get_map_forensics_bundle",
]

