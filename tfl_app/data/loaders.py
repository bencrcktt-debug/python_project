from __future__ import annotations

import hashlib
import re
from pathlib import Path

import pandas as pd

try:
    import streamlit as st
except ModuleNotFoundError:  # pragma: no cover - import smoke fallback
    class _CacheStub:
        def __call__(self, *decorator_args, **decorator_kwargs):
            if decorator_args and callable(decorator_args[0]) and len(decorator_args) == 1 and not decorator_kwargs:
                func = decorator_args[0]
                func.clear = lambda: None
                return func

            def decorator(func):
                func.clear = lambda: None
                return func

            return decorator

    class _StreamlitStub:
        cache_data = _CacheStub()

    st = _StreamlitStub()

import tfl_app.search.state as _shared_search_state
import tfl_app.bundles.page_bundles as _page_bundles
from tfl_app.shared.sessions import add_session_from_year as _add_session_from_year
from tfl_app.data.catalog import ALL_WORKBOOK_TABLE_KEYS, FILER_NORMALIZED_TABLE_KEYS, PARQUET_FILE_MAP, WORKBOOK_TABLE_COLUMNS


def _is_url(path: str) -> bool:
    return str(path or "").startswith(("http://", "https://"))


_MONEY_RANGE = re.compile(r"(-?\d[\d,]*\.?\d*)\s*(?:-|to)\s*(-?\d[\d,]*\.?\d*)", re.IGNORECASE)


def _to_num_series(series: pd.Series) -> pd.Series:
    cleaned = series.fillna("").astype(str).str.strip()
    neg = cleaned.str.startswith("(") & cleaned.str.endswith(")")
    cleaned = cleaned.str.replace(r"^\(|\)$", "", regex=True)
    cleaned = cleaned.str.replace("$", "", regex=False).str.replace(",", "", regex=False)
    out = pd.to_numeric(cleaned, errors="coerce").fillna(0.0)
    return out.where(~neg, -out)


def add_low_high_numeric(df: pd.DataFrame) -> pd.DataFrame:
    data = _page_bundles.ensure_cols(df, {"Low": 0, "High": 0, "Amount": ""}).copy()

    low = _to_num_series(data["Low"])
    high = _to_num_series(data["High"])

    amount = data["Amount"].fillna("").astype(str).str.strip()
    amount_clean = (
        amount.str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace("\u2013", "-", regex=False)
    )
    rng = amount_clean.str.extract(_MONEY_RANGE)
    rng_low = pd.to_numeric(rng[0], errors="coerce").fillna(0.0)
    rng_high = pd.to_numeric(rng[1], errors="coerce").fillna(0.0)
    single = pd.to_numeric(amount_clean.str.extract(r"(-?\d+(?:\.\d+)?)")[0], errors="coerce").fillna(0.0)

    both_zero = (low == 0) & (high == 0)
    low = low.where(~both_zero, rng_low.where(rng_low != 0, single))
    high = high.where(~both_zero, rng_high.where(rng_high != 0, single))
    high = high.where(high != 0, low)
    low = low.where(low != 0, high)

    data["Low_num"] = low
    data["High_num"] = high
    return data


def safe_read_excel_xf(xf: pd.ExcelFile, sheet_name: str, cols: list[str]) -> pd.DataFrame:
    try:
        return xf.parse(sheet_name=sheet_name, usecols=cols)
    except Exception:
        try:
            df = xf.parse(sheet_name=sheet_name)
            keep = [column for column in cols if column in df.columns]
            return df[keep]
        except Exception:
            return pd.DataFrame(columns=cols)


@st.cache_data(show_spinner=False)
def _empty_df(cols: list[str]) -> pd.DataFrame:
    return pd.DataFrame(columns=cols)


def read_parquet_cols(path: Path, cols: list[str]) -> pd.DataFrame:
    try:
        import pyarrow.parquet as pq

        parquet_file = pq.ParquetFile(path)
        available = set(parquet_file.schema.names)
        use_cols = [column for column in cols if column in available]
        table = parquet_file.read(columns=use_cols) if use_cols else parquet_file.read()
        return table.to_pandas()
    except Exception:
        try:
            df = pd.read_parquet(path)
            keep = [column for column in cols if column in df.columns]
            return df[keep] if keep else df
        except Exception:
            return pd.DataFrame(columns=cols)


def _fingerprint_paths(paths: list[Path], *, seed: str) -> str:
    digest = hashlib.sha1(str(seed).encode("utf-8"))
    for candidate in sorted((Path(path) for path in paths), key=lambda item: str(item)):
        try:
            stat = candidate.stat()
            digest.update(str(candidate.resolve()).encode("utf-8"))
            digest.update(str(int(stat.st_mtime_ns)).encode("utf-8"))
            digest.update(str(int(stat.st_size)).encode("utf-8"))
        except Exception:
            digest.update(str(candidate).encode("utf-8"))
    return digest.hexdigest()


def _normalize_loaded_table(table_key: str, df: pd.DataFrame) -> pd.DataFrame:
    data = df.copy()
    if table_key == "Wit_All":
        if "Session" not in data.columns:
            session_source = data.get("session", pd.Series("", index=data.index))
            data["Session"] = session_source.fillna("").astype(str).str.strip()
        if "Bill" not in data.columns:
            bill_source = data.get("bill", pd.Series("", index=data.index))
            data["Bill"] = bill_source.fillna("").astype(str).str.upper().str.replace(r"\s+", " ", regex=True).str.strip()
        if "Position" not in data.columns:
            position_source = data.get("position", pd.Series("", index=data.index)).fillna("").astype(str).str.strip().str.upper()
            data["Position"] = position_source.str.title()
        if "LobbyShort" not in data.columns:
            data["LobbyShort"] = data.get("lobbyshort", pd.Series("", index=data.index)).fillna("").astype(str).str.strip()
        if "name" in data.columns and "WitnessName" not in data.columns:
            data["WitnessName"] = data["name"].fillna("").astype(str).str.strip()
        if "org" in data.columns and "Organization" not in data.columns:
            data["Organization"] = data["org"].fillna("").astype(str).str.strip()
        position_series = data.get("Position", pd.Series("", index=data.index)).fillna("").astype(str).str.upper()
        data["IsFor"] = position_series.str.contains(r"\bFOR\b", na=False).astype(int)
        data["IsAgainst"] = position_series.str.contains("AGAINST", na=False).astype(int)
        data["IsOn"] = position_series.str.contains(r"\bON\b", na=False).astype(int)
        return data
    if table_key == "Bill_Status_All":
        if "Authors" not in data.columns and "Author" in data.columns:
            data["Authors"] = data["Author"]
        if "Author" not in data.columns and "Authors" in data.columns:
            data["Author"] = data["Authors"].fillna("").astype(str).str.split("|").str[0].fillna("").str.strip()
        defaults = {
            "Session": "",
            "Bill": "",
            "Authors": "",
            "Author": "",
            "Caption": "",
            "Status": "",
            "Link": "",
            "Chamber": "",
        }
        return _page_bundles.ensure_cols(data, defaults)
    if table_key == "Lobby_TFL_Client_All":
        defaults = {
            "Session": "",
            "Client": "",
            "Lobby Name": "",
            "LobbyShort": "",
            "IsTFL": 0,
            "Low": 0,
            "High": 0,
            "Amount": "",
            "Mid": 0.0,
            "FilerID": None,
        }
        data = _page_bundles.ensure_cols(data, defaults)
        data = add_low_high_numeric(data)
        if "Mid" not in data.columns or data["Mid"].isna().all():
            data["Mid"] = (data["Low_num"] + data["High_num"]) / 2
        data["Client"] = data["Client"].fillna("").astype(str).str.strip()
        data["Lobby Name"] = data["Lobby Name"].fillna("").astype(str).str.strip()
        data["LobbyShort"] = data["LobbyShort"].fillna("").astype(str).str.strip()
        data["Session"] = data["Session"].fillna("").astype(str).str.strip()
        data["IsTFL"] = pd.to_numeric(data["IsTFL"], errors="coerce").fillna(0).astype(int)
        data["ClientNorm"] = _shared_search_state.norm_name_series(data["Client"])
        data["LobbyNameNorm"] = _shared_search_state.norm_name_series(data["Lobby Name"])
        data["LobbyShortNorm"] = _shared_search_state.norm_name_series(data["LobbyShort"])
        return data
    if table_key == "Lobby_Sub_All":
        if "Session" not in data.columns:
            if "legislative_session" in data.columns:
                data["Session"] = data["legislative_session"].fillna("").astype(str).str.strip()
            elif "session" in data.columns:
                data["Session"] = data["session"].fillna("").astype(str).str.strip()
        if "LobbyShort" not in data.columns and "lobbyshort" in data.columns:
            data["LobbyShort"] = data["lobbyshort"].fillna("").astype(str).str.strip()
        defaults = {
            "Session": "",
            "Subject Matter": "",
            "Other Subject Matter Description": "",
            "Primary Business": "",
            "FilerID": None,
            "LobbyShort": "",
            "Lobby Name": "",
        }
        data = _page_bundles.ensure_cols(data, defaults)
        data["LobbyShort"] = data["LobbyShort"].fillna("").astype(str).str.strip()
        data["Lobby Name"] = data["Lobby Name"].fillna("").astype(str).str.strip()
        return data
    if table_key == "Staff_All":
        data = _add_session_from_year(data)
        defaults = {
            "Session": "",
            "Legislator": "",
            "Title": "",
            "Staffer": "",
            "source": "",
        }
        data = _page_bundles.ensure_cols(data, defaults)
        if "Staffer" not in df.columns:
            staff_source = data.get("name", data.get("staff_name_last_initial", pd.Series("", index=data.index)))
            data["Staffer"] = staff_source.fillna("").astype(str).str.strip()
        if "Legislator" not in df.columns:
            legislator_source = data.get("legislator_name", data.get("member_or_committee", pd.Series("", index=data.index)))
            data["Legislator"] = legislator_source.fillna("").astype(str).str.strip()
        return data
    if table_key in FILER_NORMALIZED_TABLE_KEYS:
        data = _add_session_from_year(data)
        defaults = {
            "Session": "",
            "filerIdent": None,
            "filerName": "",
            "filerSort": "",
            "periodStartDt": "",
        }
        data = _page_bundles.ensure_cols(data, defaults)
        if "FilerID" not in data.columns:
            data["FilerID"] = pd.to_numeric(data["filerIdent"], errors="coerce")
        return data
    return data


def _ensure_filer_base_columns(df: pd.DataFrame) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df
    data = df
    if "FilerID" not in data.columns:
        if "filerIdent" in data.columns:
            data = data.copy()
            data["FilerID"] = pd.to_numeric(data["filerIdent"], errors="coerce")
        else:
            data = data.copy()
            data["FilerID"] = pd.Series([-1] * len(data), index=data.index)
    else:
        data = data.copy()
        data["FilerID"] = pd.to_numeric(data["FilerID"], errors="coerce")
    if "filerName" not in data.columns:
        data["filerName"] = ""
    if "filerSort" not in data.columns:
        data["filerSort"] = ""
    return data


def _postprocess_table_for_state(table_key: str, df: pd.DataFrame) -> pd.DataFrame:
    data = _normalize_loaded_table(table_key, df)
    if table_key == "Wit_All":
        return _shared_search_state._ensure_session_key_column(_shared_search_state._ensure_witness_search_columns(data))
    if table_key == "Bill_Status_All":
        return _shared_search_state._ensure_session_key_column(data)
    if table_key == "Lobby_TFL_Client_All":
        return _shared_search_state._ensure_lobby_client_lookup_columns(data)
    if table_key == "Staff_All":
        return _shared_search_state._ensure_session_key_column(_shared_search_state._ensure_staff_search_columns(data))
    if table_key == "Lobby_Sub_All":
        data = _shared_search_state._ensure_session_key_column(data)
        if "LobbyShort" in data.columns:
            data["LobbyShort"] = data["LobbyShort"].fillna("").astype(str).str.strip()
            data["LobbyShortNorm"] = _shared_search_state.norm_name_series(data["LobbyShort"])
        return data
    if table_key in FILER_NORMALIZED_TABLE_KEYS:
        return _shared_search_state._ensure_session_key_column(_ensure_filer_base_columns(data))
    return _shared_search_state._ensure_session_key_column(data)


def _resolve_table_source(path: str, table_key: str):
    base = Path(path)
    if not base.exists():
        return None
    if not base.is_dir():
        return ("excel", table_key)

    filename = PARQUET_FILE_MAP.get(table_key)
    if not filename:
        return None
    if isinstance(filename, (list, tuple)):
        matches = [base / candidate for candidate in filename if (base / candidate).exists()]
        return matches or None
    candidate = base / filename
    return candidate if candidate.exists() else None


@st.cache_data(show_spinner=False, max_entries=8)
def get_dataset_version(path: str) -> str:
    if not path:
        return ""
    if _is_url(path):
        return hashlib.sha1(str(path).strip().encode("utf-8")).hexdigest()
    base = Path(path)
    if not base.exists():
        return hashlib.sha1(str(base).encode("utf-8")).hexdigest()
    if base.is_file():
        return _fingerprint_paths([base], seed=str(base.resolve()))
    resolved_paths: list[Path] = []
    for table_key in ALL_WORKBOOK_TABLE_KEYS:
        source = _resolve_table_source(path, table_key)
        if isinstance(source, Path):
            resolved_paths.append(source)
        elif isinstance(source, list):
            resolved_paths.extend(item for item in source if isinstance(item, Path))
    if not resolved_paths:
        try:
            resolved_paths = [item for item in base.rglob("*") if item.is_file()]
        except Exception:
            resolved_paths = []
    return _fingerprint_paths(resolved_paths, seed=str(base.resolve()))


def _read_table_source(path: str, table_key: str, columns: list[str]) -> pd.DataFrame:
    source = _resolve_table_source(path, table_key)
    if source is None:
        return _empty_df(columns)
    if isinstance(source, tuple) and source and source[0] == "excel":
        try:
            xf = pd.ExcelFile(path, engine="openpyxl")
            return safe_read_excel_xf(xf, table_key, columns)
        except Exception:
            return _empty_df(columns)
    if isinstance(source, list):
        frames = []
        for item in source:
            try:
                frames.append(read_parquet_cols(item, columns))
            except Exception:
                continue
        return pd.concat(frames, ignore_index=True).drop_duplicates() if frames else _empty_df(columns)
    try:
        return read_parquet_cols(source, columns)
    except Exception:
        return _empty_df(columns)


def _table_keys_tuple(keys: tuple[str, ...] | list[str]) -> tuple[str, ...]:
    ordered: list[str] = []
    seen: set[str] = set()
    for key in keys:
        if key in WORKBOOK_TABLE_COLUMNS and key not in seen:
            ordered.append(key)
            seen.add(key)
    return tuple(ordered)


def _dedupe_columns(columns: list[str] | tuple[str, ...]) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()
    for column in columns:
        key = str(column or "").strip()
        if key and key not in seen:
            ordered.append(key)
            seen.add(key)
    return ordered


def _normalized_manifest_column_count(table_key: str, source_columns: list[str]) -> int:
    sample = pd.DataFrame({column: pd.Series(dtype="object") for column in source_columns})
    return int(len(_postprocess_table_for_state(table_key, sample).columns))


def _read_manifest_probe(path: str, table_key: str, source_columns: list[str]) -> pd.DataFrame:
    probe_columns: list[str] = []
    for candidate in ("Session", "session", "applicableYear", "LobbyShort", "lobbyshort"):
        if candidate in source_columns and candidate not in probe_columns:
            probe_columns.append(candidate)
    if not probe_columns:
        return pd.DataFrame()
    return _read_table_source(path, table_key, probe_columns)


def _empty_manifest_entry(table_key: str) -> dict[str, object]:
    return {
        "rows": 0,
        "cols": int(len(_postprocess_table_for_state(table_key, _empty_df(WORKBOOK_TABLE_COLUMNS.get(table_key, []))).columns)),
        "has_session": False,
        "empty": True,
        "sessions": 0,
        "lobby_count": 0,
    }


@st.cache_data(show_spinner=False, max_entries=128)
def _get_table_manifest_entry(path: str, table_key: str, data_version: str) -> dict[str, object]:
    del data_version
    base = Path(path)
    source = _resolve_table_source(path, table_key)
    source_columns: list[str] = []
    row_count = 0
    if source is None:
        return _empty_manifest_entry(table_key)

    if isinstance(source, tuple) and source and source[0] == "excel":
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(base, read_only=True, data_only=True)
            if table_key in workbook.sheetnames:
                sheet = workbook[table_key]
                header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                source_columns = [str(value) for value in header if value is not None]
                row_count = max(int(sheet.max_row or 0) - 1, 0)
        except Exception:
            source_columns = list(WORKBOOK_TABLE_COLUMNS.get(table_key, []))
            probe = _read_manifest_probe(path, table_key, source_columns)
            row_count = int(len(probe))
    else:
        sources = source if isinstance(source, list) else [source]
        column_names: list[str] = []
        for item in sources:
            try:
                import pyarrow.parquet as pq

                parquet = pq.ParquetFile(item)
                if not column_names:
                    column_names = list(parquet.schema.names)
                row_count += int(parquet.metadata.num_rows or 0)
            except Exception:
                if not column_names:
                    try:
                        sample = pd.read_parquet(item)
                        column_names = list(sample.columns)
                        row_count += int(len(sample))
                    except Exception:
                        continue
        if column_names:
            source_columns = [column for column in WORKBOOK_TABLE_COLUMNS.get(table_key, []) if column in column_names]
        else:
            source_columns = list(WORKBOOK_TABLE_COLUMNS.get(table_key, []))
        if row_count == 0:
            probe = _read_table_source(path, table_key, WORKBOOK_TABLE_COLUMNS.get(table_key, []))
            row_count = int(len(probe))

    if not source_columns:
        source_columns = list(WORKBOOK_TABLE_COLUMNS.get(table_key, []))

    normalized_cols = _normalized_manifest_column_count(table_key, source_columns)
    probe = _read_manifest_probe(path, table_key, source_columns)
    has_session = bool(
        ("Session" in source_columns)
        or ("session" in source_columns)
        or ("SessionKey" in probe.columns)
        or ("Session" in probe.columns)
    )
    session_count = 0
    if not probe.empty:
        session_probe = _shared_search_state._ensure_session_key_column(_normalize_loaded_table(table_key, probe.copy()))
        if "SessionKey" in session_probe.columns:
            session_count = int(session_probe["SessionKey"].dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique())
    lobby_count = 0
    if not probe.empty:
        lobby_probe = _normalize_loaded_table(table_key, probe.copy())
        if "LobbyShort" in lobby_probe.columns:
            lobby_count = int(
                lobby_probe["LobbyShort"]
                .dropna()
                .astype(str)
                .str.strip()
                .replace("", pd.NA)
                .dropna()
                .nunique()
            )
    return {
        "rows": int(row_count),
        "cols": int(normalized_cols),
        "has_session": bool(has_session),
        "empty": bool(row_count == 0),
        "sessions": int(session_count),
        "lobby_count": int(lobby_count),
    }


@st.cache_data(show_spinner=False, max_entries=4)
def _get_table_manifest_cached(path: str, data_version: str) -> dict[str, dict[str, object]]:
    return {
        table_key: _get_table_manifest_entry(path, table_key, data_version)
        for table_key in ALL_WORKBOOK_TABLE_KEYS
    }


def get_table_manifest(path: str) -> dict[str, dict[str, object]]:
    return _get_table_manifest_cached(path, get_dataset_version(path))


get_table_manifest.clear = getattr(_get_table_manifest_cached, "clear", lambda: None)
